from re import findall
from re import sub

from os.path import expanduser
from os.path import basename
from os.path import join
from os.path import dirname

from traceback import format_exc

from math import log10, floor

from ast import literal_eval

from regex import search

from lxml import etree

import sys

if sys.version_info >= (3, 12):
    from PySide6.QtWidgets import QMainWindow
    from PySide6.QtCore import Signal
else:
    from PySide2.QtWidgets import QMainWindow
    from PySide2.QtCore import Signal

from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from .xml_processing import delete_first_line
from .xml_processing import replace_special_characters
from .xml_processing import linearize_xml
from .unit_list import unit_list

FILEPATH = dirname(__file__)


class UnrecognizedUnit(Exception):
    pass


class UnitTable():
    def __init__(self, main_window: QMainWindow = None, progress: Signal = None, console: Signal = None) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")
        self.conversion_factors = None
        self.extracted_numbers = None
        self.main_window = main_window
        self.progress = progress
        self.console = console

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.
        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            # xml_content = linearize_xml(xml_content)
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
            # xml_content = sub("<!DOCTYPE.*?\n", "", xml_content)
            # xml_content = sub("<!ENTITY.*?>\n", "", xml_content)
            # xml_content = sub("]>\n", "", xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def extract_conversion_factors(self) -> str:
        """Function saves the part containing the conversion factors
        and removes it from the actual file.
        Returns:
            str: A string representation of the XML content with the 'Units of Measure' topic removed.
        """
        xml_content = self.replace_entities()
        root = etree.fromstring(xml_content)
        topic = root.xpath("//topic[title='Units of Measure']")[0]
        self.conversion_factors = etree.tostring(topic).decode()
        topic.getparent().remove(topic)
        xml_content = etree.tostring(root).decode()
        return xml_content

    def extract_pageblocks(self) -> list:
        """Extracts page blocks from the XML content using XPath queries.
        Returns:
            list: A list of page blocks in string format.
        """
        xml_content = self.extract_conversion_factors()

        root = etree.fromstring(xml_content)
        pgblk_list = root.findall(".//pgblk")
        pgblk_list = [etree.tostring(pgblk).decode() for pgblk in pgblk_list]
        return pgblk_list

    def remove_x_and_super(self, line: str) -> str:
        """Removes the 'x 10' and super tags from the given string. (If contained)
        Args:
            line (str): A string with 'x 10' and super tags.
        Returns:
            str: The string with 'x 10' and super tags removed.
        """
        if search(r"x \d+\<super\>-?\d\</super\> ", line):
            line = sub(r"x \d+\<super\>-?\d\</super\> ", "", line)
        if search(r"(\<super\>)(-?\d)(\<\/super\>)", line):
            line = sub(r"(\<super\>)(-?\d)(\<\/super\>)", r"\g<2>", line)
        return line

    def extract_numbers_with_units(self):
        """This function extracts numerical values with units from a list of page blocks using regular expressions.
        The function loops through each page block, searches for values with units based on specific patterns, and extracts the matches.
        The extracted values are stored in a list that is returned at the end of the function.
        Returns:
            list: A list of values with units.
        """
        pgblk_list = self.extract_pageblocks()

        for ind, line in enumerate(pgblk_list):
            pgblk_list[ind] = pgblk_list[ind].split("\n")
        # Regular expression pattern to match values with units
        values = []
        for ind, pgblk in enumerate(pgblk_list):
            pgblk_nbr = search(r'(pgblknbr=")(\d+)(")', pgblk[0]).group(2)
            for ind, line in enumerate(pgblk):
                line_matches = []
                # Case 1: Between 791 and 1154 Nm
                if search(r"-?[0-9]+\.?[0-9]+ and -?[0-9]+\.?[0-9]+", line):
                    line = self.remove_x_and_super(line)
                    for elem in unit_list:
                        pattern = r"-?[0-9]+\.?[0-9]+ and -?[0-9]+\.?[0-9]+ " + elem
                        matches = findall(pattern, line)
                        line_matches, values = self.check_for_substrings(
                            matches, line_matches, values, pgblk_nbr)
                # Case 2: Torque the screws (1-20) to 2 +0.5 Nm
                # elif "orque the" in line:
                elif search(r"-?\d+\.?\d* (?:&#177;|&#xb1;|\+|-){1}\d+\.?\d*", line):
                    line = self.remove_x_and_super(line)
                    for elem in unit_list:
                        pattern = r"-?\d+\.?\d* (?:&#177;|&#xb1;|\+|-){1}\d+\.?\d* " + elem
                        matches = findall(pattern, line)
                        line_matches, values = self.check_for_substrings(
                            matches, line_matches, values, pgblk_nbr)
                # Case 3: 3.73 to 3.85 kg
                elif search(r"-?\d+\.?\d* to -?\d+\.?\d*", line):
                    line = self.remove_x_and_super(line)
                    for elem in unit_list:
                        pattern = r"-?\d+\.?\d* to -?\d+\.?\d* " + elem
                        matches = findall(pattern, line)
                        line_matches, values = self.check_for_substrings(
                            matches, line_matches, values, pgblk_nbr)
                # Case 4: Normal case
                else:
                    line = self.remove_x_and_super(line)
                    for elem in unit_list:
                        pattern = r"(&#177;?|-?)([0-9]+\.?[0-9]* )" + \
                            rf"({elem})" + r"(?![a-zA-Z])"
                        matches = findall(pattern, line)
                        line_matches, values = self.check_for_substrings(
                            matches, line_matches, values, pgblk_nbr)
        for i, elem in enumerate(values):
            if type(values[i][0]) == tuple:
                values[i] = (values[i][0][0] + values[i][0]
                             [1] + values[i][0][2], values[i][1])
        self.extracted_numbers = values
        return values

    def check_for_substrings(self, matches: list, line_matches: list, values: list, pgblk_nbr: str) -> tuple:
        """Some units are substrings of others (i.e. l -> lb -> lbf -> lbf.ft) therefore,
        we would have unwanted outputs.
        This function prevents this "bug" and only adds entries from matches to
        line_matches and values if they are not substrings of other entries in line_matches.
        Args:
            matches (list): list of all matches found in a line (includes unwanted entrys)
            line_matches (list): list for storing the right units (excludes unwanted entrys)
            values (list): list containing all unit values
        Returns:
            tuple: _description_
        """
        for entry in matches:
            if not any([entry in value for value in line_matches]):
                line_matches.append(entry)
                values.append((entry, pgblk_nbr))
        return line_matches, values

    def individualize_values(self) -> list:
        """Extracts unique values with units from the given list of values.

        Returns:
            list: A list of unique values with units.
        """
        values = self.extract_numbers_with_units()
        unique_values = []
        for elem in values:
            if elem not in unique_values:
                unique_values.append(elem)
        return unique_values

    def get_conversion_factors(self) -> dict:
        """This method extracts conversion factors from XML data and returns them as a dictionary
        where the key is the original unit and the value is the conversion factor and the resulting
        unit (e.g {'g': '0.03527 oz'})
        Returns:
            dict: A dictionary containing the conversion factors extracted from the XML data.
        """
        topic = self.conversion_factors
        topic = linearize_xml(topic)
        topic = sub(r"\<\?.*?\?\> *", "", topic)

        blocks = findall(r"\<defdata\>.*?\</defdata\>", topic)
        term_list = []
        def_list = []
        for elem in blocks:
            term = search(r"(\<term\>)(.*?)(\</term\>)", elem).group(2)
            if search(r"(\<def\>)(.*?)(\</def\>)", elem):
                defdata = search(r"(\<def\>)(.*?)(\</def\>)", elem).group(2)
            else:
                defdata = None
            if defdata is not None:
                term_list.append(term)
                def_list.append(defdata)
        conversion_dict = {}
        for key, value in zip(term_list, def_list):
            if key not in conversion_dict:
                conversion_dict[key] = value
            else:
                conversion_dict[key] = (conversion_dict[key], value)
        for key, value in conversion_dict.items():
            if type(value) is not tuple:
                conversion_dict[key] = value.strip()

        if "kN" not in conversion_dict:
            conversion_dict['kN'] = "225 lbf"
        if "daN" not in conversion_dict:
            conversion_dict['daN'] = "10 N"
        if "Pa" not in conversion_dict:
            conversion_dict['Pa'] = "0.000145037 psi"
        if "&#x3bc;m" not in conversion_dict:
            conversion_dict['&#x3bc;m'] = "39.370078740158 &#x3bc;in."
        return conversion_dict

    def calculate_conversion(self, value: str) -> str:
        """Calculate conversion of a given value string to a different
        unit using the conversion factors from dictionary created in
        get_conversion_factors method.
        Args:
            value (str): The string value to be converted.
        Returns:
            str: The converted value in the new unit. If an error occurs during the conversion process, it returns the string "Error".
        """
        conversion_dict = self.get_conversion_factors()
        split_value = value.split()
        if len(split_value) == 0:
            return value
        original_unit = split_value[-1]
        if original_unit not in conversion_dict:
            if self.main_window is not None:
                self.console.emit(f'Unit "{original_unit}" not found in the Units of Measure and Conversion Factors Table. Please Check the XML file.')
                return value
            raise UnrecognizedUnit(f"It seems like the xml uses rarely used unit ({original_unit}). Check if that's the case and \
                    contact the developer, to add the unit to the program. Conversion Dict: {conversion_dict}".replace("                    ", ""))
        # Cosmin - changed to isinstance
        if not isinstance(conversion_dict[original_unit], tuple):
            try:
                conversion = float(conversion_dict[original_unit].split()[0])
                new_unit = conversion_dict[original_unit].split()[-1]
                if new_unit == "(US)":
                    new_unit = "gal(US)"
            except KeyError:
                return "Error"
            if len(findall(r"\d+\.?\d* ", value)) == 2:
                # Operations for "to" "and" and "operator" case
                if original_unit == "&#176;C":
                    new_unit = "&#176;F"
                    numbers = findall(r"-?\d+\.?\d* ", value)
                    numbers = [''.join(string.split()) for string in numbers]
                    if "&#177;" in value:
                        value = sub(
                            r"(-?\d+\.?\d* )(.*?)(-?\d+\.?\d* )",
                            f"{1.8 * float(numbers[0]) + 32}" + r" \g<2>" + f"{1.8 * float(numbers[1])} ",
                            value)
                    else:
                        value = sub(
                            r"(-?\d+\.?\d* )(.*?)(-?\d+\.?\d* )",
                            f"{1.8 * float(numbers[0]) + 32}" + r" \g<2>" + f"{1.8 * float(numbers[1]) + 32} ",
                            value)
                    value = sub(r"(-?\d+\.?\d*)(.*?)(-?\d+\.?\d* )(.*)",
                                r"\1\2\3" + new_unit, value)
                    value = sub("  ", " ", value)
                elif original_unit == "&#176;F":
                    new_unit = "&#176;C"
                    numbers = findall(r"-?\d+\.?\d* ", value)
                    numbers = [''.join(string.split()) for string in numbers]
                    if "&#177;" in value:
                        value = sub(
                            r"(-?\d+\.?\d* )(.*?)(-?\d+\.?\d* )",
                            f"{0.556 * (float(numbers[0]) - 32)}" + r" \g<2>" + f"{0.556 * float(numbers[1])} ",
                            value)
                    else:
                        value = sub(
                            r"(-?\d+\.?\d* )(.*?)(-?\d+\.?\d* )",
                            f"{0.556 * (float(numbers[0]) - 32)}" + r" \g<2>" + f"{0.556 * (float(numbers[1]) - 32)} ",
                            value)
                    value = sub(r"(-?\d+\.?\d*)(.*?)(-?\d+\.?\d* )(.*)",
                                r"\1\2\3" + new_unit, value)
                else:
                    numbers = findall(r"-?\d+\.?\d* ", value)
                    numbers = [''.join(string.split()) for string in numbers]
                    value = sub(
                        r"(-?\d+\.?\d* )(.*?)(-?\d+\.?\d* )", f"{float(numbers[0]) * conversion}" + r" \g<2>" + f"{float(numbers[1]) * conversion} ",
                        value)
                    value = sub(r"(-?\d+\.?\d*)(.*?)(-?\d+\.?\d* )(.*)",
                                r"\1\2\3" + new_unit, value)
                    value = sub("  ", " ", value)
                return value
            else:
                # Operations for standard case
                if original_unit == "&#176;C":
                    new_unit = "&#176;F"
                    numbers = findall(r"-?\d+\.?\d* ", value)
                    numbers = [''.join(string.split()) for string in numbers]
                    if "&#177;" in value:
                        value = sub(r"(.*?)(-?\d+\.?\d* )(.*)", r"\g<1>" + f"{1.8 * float(numbers[0])} " + new_unit, value)
                    else:
                        value = sub(r"(.*?)(-?\d+\.?\d* )(.*)", r"\g<1>" + f"{1.8 * float(numbers[0]) + 32} " + new_unit, value)

                elif original_unit == "&#176;F":
                    new_unit = "&#176;C"
                    numbers = findall(r"-?\d+\.?\d* ", value)
                    numbers = [''.join(string.split()) for string in numbers]
                    if "&#177;" in value:
                        value = sub(r"(.*?)(-?\d+\.?\d* )(.*)", r"\g<1>" + f"{0.556 * (float(numbers[0]))} " + new_unit, value)
                    else:
                        value = sub(r"(.*?)(-?\d+\.?\d* )(.*)", r"\g<1>" + f"{0.556 * (float(numbers[0]) - 32)} " + new_unit, value)
                else:
                    numbers = findall(r"-?\d+\.?\d* ", value)
                    numbers = [''.join(string.split()) for string in numbers]
                    value = sub(r"(.*?)(-?\d+\.?\d* )(.*)", r"\g<1>" + f"{float(numbers[0]) * conversion} " + new_unit, value)
                return value
        else:
            # Operations for tuple case
            try:
                conversion1 = float(
                    conversion_dict[original_unit][0].split()[0])
                new_unit1 = conversion_dict[original_unit][0].split()[-1]
                conversion2 = float(
                    conversion_dict[original_unit][1].split()[0])
                new_unit2 = conversion_dict[original_unit][1].split()[-1]
            except KeyError:
                return "Error"
            if len(findall(r"\d+\.?\d* ", value)) == 2:
                # Operations for "to" "and" and "operator" case (tuple)
                numbers = findall(r"-?\d+\.?\d* ", value)
                numbers = [''.join(string.split()) for string in numbers]
                value1 = sub(
                    r"(-?\d+\.?\d* )(.*?)(-?\d+\.?\d* )",
                    f"{float(numbers[0]) * conversion1}" + r" \g<2>" + f"{float(numbers[1]) * conversion1} ", value)
                value1 = value1.rsplit(" ", 1)[0] + " " + new_unit1
                value2 = sub(
                    r"(-?\d+\.?\d* )(.*?)(-?\d+\.?\d* )",
                    f"{float(numbers[0]) * conversion2}" + r" \g<2>" + f"{float(numbers[1]) * conversion2} ",
                    value)
                value2 = value2.rsplit(" ", 1)[0] + " " + new_unit2
                value = f"{value1, value2}"
                return value
            else:
                # Operations for standard case (tuple)
                numbers = findall(r"-?\d+\.?\d* ", value)
                value1 = sub(r"(.*?)(-?\d+\.?\d* )(.*)", r"\g<1>" + f"{float(numbers[0]) * conversion1}" + ' ' + new_unit1, value)
                value2 = sub(r"(.*?)(-?\d+\.?\d* )(.*)", r"\g<1>" + f"{float(numbers[0]) * conversion2}" + ' ' + new_unit2, value)
                value = f"{value1, value2}"
                return value

    def replace_symbols(self, number: str) -> str:
        """Replaces the unicode number for the degree symbol with
        the actual symbol (°) if it can be found in the number.
        Args:
            number (str): string containing the value or conversion number
        Returns:
            str: the modified string
        """
        number = number.replace("&#176;", "°").replace(
            "&#177;", "±").replace("&#xb1;", "±").replace("&#x3bc;", "µ")
        return number

    def round_number(self, number: float, digits_of_value: int) -> str:
        """Round the given `number` to the specified number of `digits_of_value` and return it as a string.

        Args:
            number (float): The number to round.
            digits_of_value (int): The number of digits to which the `number` should be rounded.

        Returns:
            str: The rounded number as a string.
        """
        # print(f"Unrounded Number: {number}, Sign. Numbers: {digits_of_value}")
        if floor(number) == 0:
            number_of_zeros = len(str(number)[2:]) - len(str(number)[2:].lstrip('0'))
            digits_of_value += number_of_zeros
        rounded = round(
            number,
            digits_of_value - len(str(floor(number))) if digits_of_value - len(str(floor(number))) >= 0 else 0
        )
        rounded = int(rounded) if str(rounded)[-2:] == '.0' else rounded
        # print("Rounded Values: ", str(rounded))
        return str(rounded)

    # def round_number(self, number: float, digits_of_value: int) -> str:
    #     """Round the given `number` to the specified number of `digits_of_value` and return it as a string.
    #     Args:
    #         number (float): The number to round.
    #         digits_of_value (int): The number of significant digits to which the `number` should be rounded.
    #     Returns:
    #         str: The rounded number as a string.
    #     """
    #     negative = False
    #     if number < 0:
    #         number = abs(number)
    #         negative = True

    #     if number == 0:
    #         return '0'

    #     # Determine the magnitude of the number (number of digits before the decimal point)
    #     magnitude = int(log10(abs(number))) + 1 if number >= 1 else 0

    #     # Calculate the number of digits after the decimal point that should be included in the rounded result
    #     fractional_digits = max(digits_of_value - magnitude, 0)

    #     # Use string formatting to round the number to the appropriate number of fractional digits
    #     format_str = '{:.' + str(fractional_digits) + 'f}'
    #     rounded = format_str.format(round(number, fractional_digits))

    #     if negative:
    #         rounded = "-" + rounded
    #     return rounded

    def round_number_down(self, number: float, digits_of_value: int) -> str:
        """
        Round down the given `number` to the specified number of `digits_of_value` and return it as a string.
        Args:
            number (float): The number to round down.
            digits_of_value (int): The number of digits to which the `number` should be rounded down.
        Returns:
            str: The rounded down number as a string.
        """
        number = str(number)
        count = 0
        non_zero_seen = False
        for c in number:
            if c == '.':
                continue
            if c == '0' and not non_zero_seen:
                continue
            if c.isdigit():
                non_zero_seen = True
                count += 1
        count -= digits_of_value
        for _ in range(count):  # _ was i
            if number[-1] == '.':
                number = number[:-1]
                break
            else:
                number = number[:-1]
        if number[-1] == '.':
            number = number.replace('.', '')
        return number

    def round_number_up(self, number: float, digits_of_value: int) -> str:
        """
        Round up the given `number` to the specified number of `digits_of_value` and return it as a string.
        Args:
            number (float): The number to round up.
            digits_of_value (int): The number of digits to which the `number` should be rounded up.
        Returns:
            str: The rounded up number as a string.
        """
        if number == 0:
            return str(int(number))
        number = str(number)
        count = 0
        non_zero_seen = False
        for c in number:
            if c == '.':
                continue
            if c == '0' and not non_zero_seen:
                continue
            if c.isdigit():
                non_zero_seen = True
                count += 1
        count -= digits_of_value
        last_removed_char = None
        for _ in range(count):  # _ was i
            if number[-1] == '.':
                number = number[:-1]
                break
            else:
                last_removed_char = number[-1]
                number = number[:-1]

        if number[-1] == '.':
            number = number[:-1]

        if last_removed_char != '0':
            temp = int(number[-1]) + 1
            number = number[:-1]
            if temp == 10:
                try:
                    temp2 = int(number[-1]) + 1
                    number = number[:-1]
                    number += str(temp2) + '0'
                except ValueError:
                    pass
            else:
                number += str(temp)
        return number

    def round_values(self, value: str, conversion: str) -> str:
        """function rounds conversion values to match the number of decimal places of the value,
        and returns the updated string with rounded conversion values. It takes two arguments:
            * value, which is a string containing a value and its unit
            * conversion, which is a string containing a conversion formula.
        The function extracts the number of decimal places of the value
        and applies that same rounding to the conversion values in the conversion string.
        The updated conversion string is then returned.
        Args:
            value (str): string containing value and unit (e.g. "20.5 Nm")
            conversion (str): string containing a conversion formula
        Returns:
            str: string with rounded conversion values
        """
        my_value = value.split()[0]
        if len(findall(r"(\d+\.?\d*) +", value)) == 2:
            digits_of_value = len(findall(
                r"(\d+\.?\d*) +", value)[0].replace(".", "").replace("-", "").lstrip('0').strip())
            digits_of_value2 = len(findall(
                r"(\d+\.?\d*) +", value)[1].replace(".", "").replace("-", "").lstrip('0').strip())
        else:
            digits_of_value = len(findall(
                r"(\d+\.?\d*)", my_value)[0].replace(".", "").replace("-", "").lstrip('0'))
            digits_of_value2 = digits_of_value
        unit = findall(r"(\d+\.?\d* +)([a-zA-Z]+\.*[a-zA-Z]*\.?\d?)",
                       conversion.replace("to", "").replace("and", ""))
        unit = [match[1] for match in unit]
        # Cosmin: 15.11.2023 - commented out the following lines. They add an extra digit to the conversion
        # if len(unit) > 0:
        #     if unit[0] == "in.":
        #         digits_of_value += 1
        #         digits_of_value2 += 1
        my_conversion = findall(r"(-?\d+\.?\d*) +", conversion)
        my_conversion = [float(elem) for elem in my_conversion]
        if len(my_conversion) == 2:
            if "and" in conversion or "to" in conversion:
                rounded1 = self.round_number_up(
                    my_conversion[0], digits_of_value)
                rounded2 = self.round_number_down(
                    my_conversion[1], digits_of_value2)
            else:
                rounded1 = self.round_number(my_conversion[0], digits_of_value)
                if "+" in conversion or "-" in conversion or "U+00B1" in conversion or "±" in conversion:
                    rounded2 = self.round_number_down(
                        my_conversion[1], digits_of_value2)
                else:
                    rounded2 = self.round_number(
                        my_conversion[1], digits_of_value2)
            conversion = sub(r"(.*?)(-?\d+\.?\d*)(.*?)(-?\d+\.?\d*)",
                             r"\g<1>" + rounded1 + r"\g<3>" + rounded2, conversion)
            return conversion
        else:
            sign = ""
            if not conversion[0].isdigit() and not conversion[0] == "-":
                sign = conversion[0]
            my_conversion = self.round_number(
                my_conversion[0], digits_of_value)
            try:
                conversion = my_conversion + " " + unit[0]
            except IndexError:
                unit = conversion[-1]
                conversion = my_conversion + " °" + unit
            return sign + conversion

    def select_conversion(self, value: str, conversion: str) -> str:
        """Function follows the rule for when to choose which conversion if a unit
        can be converted to more units.
        Nm -> if value <= 50: lbf.in. else: lbf.ft
        psi -> if value <= 50: kPa else: MPa
        Args:
            value (str): _description_
            conversion (str): _description_
        Returns:
            str: The string containing the correct conversion selected from tuple
        """
        if search(r"\('.*?', '.*?'\)", conversion):
            conversion_tuple = literal_eval(conversion)
            if float(search(r"(\d+.?\d*)(.*?)", value).group(1)) <= 50:
                return conversion_tuple[0]
            return conversion_tuple[1]
        return conversion

    def create_excel(
            self,
            check_conversions=False,
            debug: bool = False,
            qt_window: QMainWindow = None,
            progress: Signal = Signal(0),
            console: Signal = Signal("")) -> int:
        """Function creates an xlsx file and fills it with the information for
        1. Value
        2. Conversion
        3. Rounded Conversion
        4. Pageblock

        Args:
            check_conversions: If set to True, the conversions will be validated.
            path (str, optional): Path where the xlsx file is saved. If not specified, the file is saved to the desktop.
        """
        try:
            values = self.individualize_values()

            workbook = Workbook()
            sheet = workbook.active
            sheet["A1"] = "Value"
            sheet["B1"] = "Conversion"
            sheet["C1"] = "Rounded Conversion"
            sheet["D1"] = "Pageblock"
            for cell in ["A1", "B1", "C1", "D1"]:
                sheet[cell].fill = PatternFill(
                    start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")

            max_progress = len(values) + 1
            for ind, value in enumerate(values, start=2):
                sheet[f"A{ind}"] = value[0]
                sheet[f"B{ind}"] = self.calculate_conversion(value[0])
                sheet[f"D{ind}"] = value[1]
                sheet[f"A{ind}"] = self.replace_symbols(sheet[f"A{ind}"].value)
                sheet[f"B{ind}"] = self.replace_symbols(sheet[f"B{ind}"].value)
                sheet[f"B{ind}"] = self.select_conversion(
                    sheet[f"A{ind}"].value, sheet[f"B{ind}"].value)
                sheet[f"C{ind}"] = self.round_values(
                    sheet[f"A{ind}"].value, sheet[f"B{ind}"].value)
                if qt_window is not None:
                    progress.emit(ind / max_progress * 100)

            filename = f'conversions_{basename(self.xml_path)}.xlsx'
            filepath = join(self.export_path, filename)
            workbook.save(filepath)

            if check_conversions:
                self.check_conversions(join(
                    self.export_path,
                    f'conversions_{basename(self.xml_path)}.xlsx')
                )

            if qt_window is not None:
                console.emit(
                    "Units table created successfully. See: " + join(
                        self.export_path,
                        f'conversions_{basename(self.xml_path)}.xlsx') + "\n")
        except Exception as err:
            if qt_window is not None and debug:
                progress.emit(100)
                console.emit("Error: " + str(err) + "\n" + format_exc() + "\n")
                return 1
        return 0

    def check_conversions(self, filepath: str):
        workbook = load_workbook(filepath)
        sheet = workbook.active

        validation_sheet = workbook.create_sheet("Validation")

        validation_sheet["A1"] = "In XML"
        validation_sheet["B1"] = "Result"
        validation_sheet["C1"] = "Corrected Version"
        header_fill = PatternFill(
            start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
        validation_sheet['A1'].fill = header_fill
        validation_sheet['B1'].fill = header_fill
        validation_sheet['C1'].fill = header_fill

        pgblk_list = self.extract_pageblocks()

        for ind, line in enumerate(pgblk_list):
            pgblk_list[ind] = pgblk_list[ind].split("\n")
        # Regular expression pattern to match values with units
        values = []
        for ind, pgblk in enumerate(pgblk_list):
            pgblk_nbr = search(r'(pgblknbr=")(\d+)(")', pgblk[0]).group(2)
            for ind, line in enumerate(pgblk):
                line_matches = []
                # Case 1: Between 791 and 1154 Nm
                if search(r"-?[0-9]+\.?[0-9]+ and -?[0-9]+\.?[0-9]+", line):
                    line = self.remove_x_and_super(line)
                    for elem in unit_list:
                        pattern = r"-?[0-9]+\.?[0-9]+ and -?[0-9]+\.?[0-9]+ " + \
                            elem + r" \(.*?\)"
                        matches = findall(pattern, line)
                        line_matches, values = self.check_for_substrings(
                            matches, line_matches, values, pgblk_nbr)
                # Case 2: Torque the screws (1-20) to 2 +0.5 Nm
                # elif "orque the" in line:
                elif search(r"-?\d+\.?\d* (?:&#177;|&#xb1;|\+|-){1}\d+\.?\d*", line):
                    line = self.remove_x_and_super(line)
                    for elem in unit_list:
                        pattern = r"-?\d+\.?\d* (?:&#177;|&#xb1;|\+|-){1}\d+\.?\d* " + \
                            elem + r" \(.*?\)"
                        matches = findall(pattern, line)
                        line_matches, values = self.check_for_substrings(
                            matches, line_matches, values, pgblk_nbr)
                # Case 3: 3.73 to 3.85 kg
                elif search(r"-?\d+\.?\d* to -?\d+\.?\d*", line):
                    line = self.remove_x_and_super(line)
                    for elem in unit_list:
                        pattern = r"-?\d+\.?\d* to -?\d+\.?\d* " + \
                            elem + r" \(.*?\)"
                        matches = findall(pattern, line)
                        line_matches, values = self.check_for_substrings(
                            matches, line_matches, values, pgblk_nbr)
                # Case 4: Normal case
                else:
                    line = self.remove_x_and_super(line)
                    for elem in unit_list:
                        pattern = r"(&#177;?|-?)([0-9]+\.?[0-9]* )" + \
                            rf"({elem})" + r"( \(.*?\))"
                        matches = findall(pattern, line)
                        line_matches, values = self.check_for_substrings(
                            matches, line_matches, values, pgblk_nbr)
        for i, elem in enumerate(values):
            if type(values[i][0]) == tuple:
                values[i] = (values[i][0][0] + values[i][0][1] + values[i][0][2] + values[i][0][3], values[i][1])

        row = 2
        for elem in values:
            temp = elem[0].replace("&#8804;", "≤").replace(
                "&#176;", "°").replace("&#177;", "±")
            extr_value = search(r"(.*?)( \(.*?)", temp).group(1)
            extr_conv = search(r"(.*?\()(.*?)(\).*?)", temp).group(2)
            workbook["Validation"].cell(row=row, column=1).value = temp
            for ind, row2 in enumerate(sheet.iter_rows(min_row=2, min_col=1, max_col=1)):
                if row2[0].value == extr_value:
                    if sheet.cell(row=ind + 2, column=3).value == extr_conv:
                        workbook["Validation"].cell(
                            row=row, column=2).value = chr(0x2713)
                        workbook["Validation"].cell(
                            row=row, column=2).font = Font(color='65DA65')
                        workbook["Validation"].cell(
                            row=row, column=2).alignment = Alignment(horizontal='center')
                    else:
                        workbook["Validation"].cell(
                            row=row, column=2).value = chr(0x2717)
                        workbook["Validation"].cell(
                            row=row, column=2).font = Font(color='F47174')
                        workbook["Validation"].cell(
                            row=row, column=2).alignment = Alignment(horizontal='center')
                        workbook["Validation"].cell(row=row, column=3).value = sheet.cell(
                            row=ind + 2, column=3).value
                    break
            row += 1

        workbook.save(filepath)


if __name__ == "__main__":
    # pass
    instance = UnitTable()
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\CRM-D9893-CO91-32-32-32RM_000-01_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AB05-27-82-12_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AW39-32-13-01_005-01_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-BD50-32-21-03_007-00_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-GE01-32-51-25_009-01_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CRM-D9893-BD50-32-21-03RM_004-01_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CRM-D9893-BD50-32-21-03RM_004-01_EN_TestM.xml")
    instance.set_xml(
        r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AR21-32-10-03_EN.xml")
    instance.create_excel(True)
