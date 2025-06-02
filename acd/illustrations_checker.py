from re import findall
from re import sub
from re import search

from os import walk
from os.path import join
from os.path import expanduser
from os.path import basename
from os.path import dirname
from os.path import normpath

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

from math import floor
from math import ceil

from ast import literal_eval

from lxml import etree

from .xml_processing import delete_first_line
from .xml_processing import replace_special_characters
from .xml_processing import linearize_xml

FILEPATH = dirname(__file__)


class illustrationChecker():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_list_of_illustrations(self, excel_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.excel_path = excel_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
            xml_content = linearize_xml(xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def create_dict_from_excel(self):
        xml_content = self.replace_entities()

        workbook = load_workbook(filename=self.excel_path)
        sheet = workbook.active

        pb_cell = None
        for row in sheet.iter_rows(min_row=1, max_row=10, min_col=1, max_col=10):
            for cell in row:
                if cell.value == "PB":
                    pb_cell = cell
                    break
            if pb_cell is not None:
                break

        if pb_cell is None:
            raise ValueError(
                "Cell with value 'PB' not found in the specified range.")

        min_row = pb_cell.row + 1
        max_row = min_row
        while sheet.cell(row=max_row, column=4).value is not None:
            max_row += 1
        min_col = pb_cell.column
        max_col = pb_cell.column + 6

        table_dict = {}
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            counter = 1
            element_num = len(table_dict) + 1
            temp = {}
            for cell in row:
                if counter == 1 and cell.value is None:
                    break
                if counter == 1:
                    temp["Pageblock"] = str(cell.value)
                elif counter == 2:
                    temp["Fig."] = str(cell.value)
                elif counter == 3:
                    temp["Sh."] = str(cell.value)
                elif counter == 4:
                    temp["Title"] = str(cell.value)
                elif counter == 7:
                    temp["ICN ATA"] = str(cell.value)
                else:
                    pass
                counter += 1
            if len(temp) != 0:
                table_dict[element_num] = temp

        workbook.close()

        return xml_content, table_dict

    def validate_table(self):
        xml_content, table_dict = self.create_dict_from_excel()
        xml_content = xml_content.replace("<csn>", "(").replace(
            "</csn>", ")").replace("<csn >", "(").replace("</csn >", ")")
        results = []
        xml_graphic_content = []
        # print(table_dict)
        for key, value in table_dict.items():
            pgblknbr = table_dict[key]["Pageblock"]

            validation_result = ("Passed", "")
            tree1 = etree.fromstring(xml_content)
            try:
                pgblk_content = tree1.xpath(
                    f'.//pgblk[@pgblknbr="{pgblknbr}"]')[0]
            except IndexError:
                pgblk_content = tree1.xpath('.//ipl')[0]
            # pgblk_content = search(r'(<pgblk.{0,150}' + pgblknbr + r'".*?</pgblk>)', xml_content).group(0)
            try:
                graphic_content = pgblk_content.xpath(
                    f".//graphic[title/text() = '{table_dict[key]['Title']}']")[0]
                graphic_content = etree.tostring(
                    graphic_content).decode("utf-8").lower()
                xml_graphic_content.append(graphic_content)
                # check sheetnumber
                sheetnum_xml = findall(r'sheetnbr="(.*?)"', graphic_content)
                sheetnum_xml = [int(element) for element in sheetnum_xml]
                if int(table_dict[key]["Sh."]) not in sheetnum_xml:
                    # print(sheetnum_xml, int(table_dict[key]["Sh."]))
                    # print("Sheet number not found")
                    validation_result = ("Failed", "Sheet number incorrect")
                # check ICN ATA
                icn_ata = table_dict[key]["ICN ATA"].replace(
                    ".cgm", "").replace(".gcm", "").lower()
                if "rm" in icn_ata:
                    # Case 1: 35 characters
                    icn_ata = search(
                        r"[a-z0-9]{3}-[a-z0-9]{4}-[a-z0-9]{2}-[a-z0-9]{2}-[a-z0-9]{4}-[a-z0-9]{5}-[a-z0-9]{5}-[a-z0-9]{3}", icn_ata).group(0)
                elif "rm" not in icn_ata:
                    # Case 2: 33 Characters
                    icn_ata = search(
                        r"[a-z0-9]{3}-[a-z0-9]{4}-[a-z0-9]{2}-[a-z0-9]{2}-[a-z0-9]{2}-[a-z0-9]{5}-[a-z0-9]{5}-[a-z0-9]{3}", icn_ata).group(0)
                if '"' + icn_ata + '"' not in graphic_content:
                    # print(f"ICN ATA not found: {table_dict[key]['ICN ATA'].replace('.cgm', '')} | {graphic_content}")
                    validation_result = ("Failed", "ICN ATA incorrect")
                # print(icn_ata)
                # print(graphic_content)

            except IndexError:
                # print("Title Not Found")
                validation_result = ("Failed", "Title incorrect")
                xml_graphic_content.append("Error")
            # print(validation_result)
            results.append(validation_result)

        self.create_validation_excel(table_dict, results, xml_graphic_content)

    def create_validation_excel(self, table_dict: dict, results: list, xml_graphic_content: list):
        workbook = Workbook()
        sheet = workbook.active

        # Headers
        sheet.cell(row=1, column=1).value = "Elem"
        sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=1).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=2).value = "Pageblock (LOI)"
        sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=2).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=3).value = "Figure (LOI)"
        sheet.cell(row=1, column=3).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=3).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=3).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=4).value = "Sheet (LOI)"
        sheet.cell(row=1, column=4).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=4).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=4).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=5).value = "Title (LOI)"
        sheet.cell(row=1, column=5).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=5).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=5).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=6).value = "ICN ATA (LOI)"
        sheet.cell(row=1, column=6).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=6).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=6).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=7).value = "ICN ATA (XML)"
        sheet.cell(row=1, column=7).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=7).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=7).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=8).value = "Validation Result"
        sheet.cell(row=1, column=8).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=8).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=8).fill = PatternFill(
            start_color='C1C1C1', end_color='C1C1C1', fill_type='solid')

        sheet.cell(row=1, column=9).value = "Comment"
        sheet.cell(row=1, column=9).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=9).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=9).fill = PatternFill(
            start_color='C1C1C1', end_color='C1C1C1', fill_type='solid')

        for ind, elem in enumerate(results, start=0):
            sheet.cell(
                row=ind + 2, column=1).value = list(table_dict.keys())[ind]
            sheet.cell(
                row=ind + 2, column=2).value = table_dict[list(table_dict.keys())[ind]]["Pageblock"]
            sheet.cell(
                row=ind + 2, column=3).value = table_dict[list(table_dict.keys())[ind]]["Fig."]
            sheet.cell(
                row=ind + 2, column=4).value = table_dict[list(table_dict.keys())[ind]]["Sh."]
            sheet.cell(
                row=ind + 2, column=5).value = table_dict[list(table_dict.keys())[ind]]["Title"]
            sheet.cell(
                row=ind + 2, column=6).value = table_dict[list(table_dict.keys())[ind]]["ICN ATA"]
            if search(r'(gnbr=")(.*?)(")', xml_graphic_content[ind]):
                temp = findall(r'(gnbr=")(.*?)(")', xml_graphic_content[ind])
                temp = [tuple_elem[1] for tuple_elem in temp]
                for elem in temp:
                    icn_ata_loi = table_dict[list(table_dict.keys())[ind]]["ICN ATA"].replace(
                        ".cgm", "").replace(".gcm", "").lower()
                    if icn_ata_loi[-4] == "_":
                        icn_ata_loi = icn_ata_loi[:-4]
                    if elem == icn_ata_loi or elem == table_dict[list(table_dict.keys())[ind]]["ICN ATA"].replace(".cgm", "").replace(".gcm", "").lower():
                        sheet.cell(row=ind + 2, column=7).value = elem.upper()
                        break
            sheet.cell(row=ind + 2, column=8).value = results[ind][0]
            sheet.cell(row=ind + 2, column=9).value = results[ind][1]

        workbook.save(join(
            self.export_path, f"ipl_table_check_{basename(normpath(self.xml_path))}.xlsx"))


class baselineReportFilter():
    def __init__(self) -> None:
        self.dir_path = None
        self.baseline_path = None
        self.ipl_content = None
        self.export_path = expanduser("~/Desktop")

    def set_base_directory(self, dir_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.dir_path = dir_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def find_base_report_excel(self):
        for root, dirs, files in walk(self.dir_path):
            for file in files:
                if "baseline" in file.lower():
                    return join(root, file)
        return None

    def find_ipl_and_get_content(self):
        for root, dirs, files in walk(self.dir_path):
            for file in files:
                if "dplist" in file.lower() and file.lower().endswith(".xml"):
                    with open(join(root, file), "r", encoding="utf-8") as _:
                        self.ipl_content = _.read()

    def find_files(self, directory: str, doc_num: str):
        matching_files = []
        for root, dirs, files in walk(directory):
            for file in files:
                if doc_num in file:
                    return True
        return False

    def filter_report(self):
        base_report_excel = self.find_base_report_excel()
        self.find_ipl_and_get_content()
        if base_report_excel is None or self.ipl_content is None:
            return None

        workbook = load_workbook(base_report_excel)
        sheet = workbook.active

        # Find row and col with cell value "Dossier / Review"
        dr_cell = None
        for row in sheet.iter_rows(min_col=1, max_col=15, min_row=1, max_row=25):
            for cell in row:
                if cell.value:
                    if cell.value.lower().strip().replace(' ', '') == "dossier/review":
                        dr_cell = cell
                        break
            if dr_cell is not None:
                break

        if dr_cell is None:
            raise ValueError(
                "Cell with value 'Dossier / Review' not found in the specified range.")

        min_row = dr_cell.row + 1
        max_row = min_row
        while sheet.cell(row=max_row, column=1).value is not None:
            max_row += 1
        min_col = dr_cell.column
        max_col = dr_cell.column + 2

        workbook_new = Workbook()
        path_sheet = workbook_new.create_sheet(title="Not in Directory")
        ipl_sheet = workbook_new.create_sheet(title="Not in IPL")

        # Headers for the path_sheet
        path_sheet.cell(row=1, column=1).value = "Doc. Number"
        path_sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        path_sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')
        path_sheet.cell(row=1, column=1).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        path_sheet.cell(row=1, column=2).value = "Used In"
        path_sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        path_sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')
        path_sheet.cell(row=1, column=2).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        # Headers for the ipl_sheet
        ipl_sheet.cell(row=1, column=1).value = "Doc. Number"
        ipl_sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        ipl_sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')
        ipl_sheet.cell(row=1, column=1).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        ipl_sheet.cell(row=1, column=2).value = "Used In"
        ipl_sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        ipl_sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')
        ipl_sheet.cell(row=1, column=2).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        path_sheet_index = 2
        for ind, row in enumerate(sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row-1), start=min_row):
            # print(sheet.cell(row=ind, column=2).value, sheet.cell(row=ind, column=3).value, sheet.cell(row=ind, column=4).value)
            doc_number = sheet.cell(
                row=ind, column=4).value.split("|")[0].strip()
            if sheet.cell(row=ind, column=2).value == "N/A" or sheet.cell(row=ind, column=2).value == "n/a":
                path_sheet.cell(row=path_sheet_index,
                                column=1).value = doc_number
                path_sheet.cell(row=path_sheet_index,
                                column=2).value = "No work"
                path_sheet_index += 1
                continue  # Skips the remaining code in the loop and move to the next iteration
            if self.find_files(dirname(base_report_excel), doc_number):
                path_sheet.cell(row=path_sheet_index,
                                column=1).value = doc_number
            else:
                path_sheet.cell(row=path_sheet_index,
                                column=1).value = doc_number
                path_sheet.cell(row=path_sheet_index,
                                column=2).value = "No work"
            path_sheet_index += 1

        pnr_list = findall(r"(\<pnr\>)(.*?)(\</pnr\>)", linearize_xml(
            replace_special_characters(delete_first_line(self.ipl_content))))
        pnr_list = [items[1] for items in pnr_list]

        ipl_sheet_index = 2
        for ind, row in enumerate(sheet.iter_rows(min_col=min_col, max_col=max_col, min_row=min_row, max_row=max_row-1), start=min_row):
            if sheet.cell(row=ind, column=2).value == "N/A":
                doc_number = sheet.cell(
                    row=ind, column=4).value.split("|")[0].strip()
                if sheet.cell(row=ind, column=2).value == "N/A" or sheet.cell(row=ind, column=2).value == "n/a":
                    ipl_sheet.cell(row=ipl_sheet_index,
                                   column=1).value = doc_number
                    ipl_sheet.cell(row=ipl_sheet_index,
                                   column=2).value = "No work"
                    ipl_sheet_index += 1
                    continue
                if doc_number not in pnr_list:
                    ipl_sheet.cell(row=ipl_sheet_index,
                                   column=1).value = doc_number
                    ipl_sheet.cell(row=ipl_sheet_index,
                                   column=2).value = "No work"
                else:
                    ipl_sheet.cell(row=ipl_sheet_index,
                                   column=1).value = doc_number
                ipl_sheet_index += 1

        # Remove default "Sheet" from workbook_new
        default_sheet = workbook_new["Sheet"]
        workbook_new.remove(default_sheet)

        workbook_new.save(join(
            self.export_path, f"baseline_report_filter_{basename(normpath(base_report_excel))}.xlsx"))


if __name__ == "__main__":
    # instance = illustrationChecker()
    # # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CRM-D9893-CO91-32-31-41RM_000-01_EN.xml")
    # # instance.set_list_of_illustrations(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CRM-D9893-CO91-32-31-41- List of illustrations.xlsm")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\32-51-21\CMM\CMM-D9893-GE01-32-51-21_EN.xml")
    # instance.set_list_of_illustrations(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\32-51-21\CMM\CMM-D9893-GE01-32-51-21_EN- List of illustrations.xlsm")
    # instance.set_export_path(r"C:\Users\bakalarz\Desktop\Export")
    # instance.validate_table()

    instance = baselineReportFilter()
    # instance.set_base_directory(r"C:\Users\bakalarz\Desktop\01_XML_Samples\WP_CO91_32-31-21\WP")
    instance.set_base_directory(r"C:\Users\bakalarz\Downloads\WP")
    instance.set_export_path(r"C:\Users\bakalarz\Desktop\Export")
    instance.filter_report()
