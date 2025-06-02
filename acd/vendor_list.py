from re import findall
from re import sub
from re import search

from os import remove
from os.path import join
from os.path import expanduser
from os.path import basename
from os.path import dirname
from os.path import normpath
from os.path import isdir
from os.path import getmtime
from os.path import abspath

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from selenium import webdriver

from json import loads

import sys

if sys.version_info >= (3, 12):
    from PySide6.QtWidgets import QMainWindow
    from PySide6.QtCore import Signal
else:
    from PySide2.QtWidgets import QMainWindow
    from PySide2.QtCore import Signal

from traceback import format_exc

import subprocess

import requests

from .xml_processing import delete_first_line
from .xml_processing import replace_special_characters
from .xml_processing import linearize_xml

from .archive import unarchive_file

import zipfile

FILEPATH = dirname(__file__)

VENDOR_EXCEPTIONS = ["0U583", "80020", "80205", "81343", "81348", "81349", "81350", "88006", "88041", "88044", "88827", "96906",
                     "99237", "99238", "C3544", "D7564", "D8286", "D9893", "F0110", "F0111", "F0112", "F0114", "F0115", "F1688",
                     "F1958", "I9001", "I9002", "I9003", "I9005", "I9006", "I9007", "I9008", "I9009", "K7766", "SC201", "U1653",
                     "NP"]

COUNTRY_IDS = {12: "ALBANIA", 71: "ALGERIA", 13: "ANDORRA", 9: "ANGOLA", 20: "ANTIGUA AND BARBUDA", 15: "ARGENTINA", 16: "ARMENIA", 7: "ARUBA",
               21: "AUSTRALIA", 22: "AUSTRIA", 23: "AZERBAIJAN", 32: "BAHAMAS", 31: "BAHRAIN", 29: "BANGLADESH", 40: "BARBADOS",
               35: "BELARUS", 25: "BELGIUM", 36: "BELIZE", 26: "BENIN", 42: "BHUTAN", 38: "BOLIVIA", 27: "SINT EUSTATIUS AND SABA BONAIRE",
               33: " AND HERZEGOVINA", 44: "BOTSWANA", 39: "BRAZIL", 41: "BRUNEI DARUSSALAM", 30: "BULGARIA", 28: "BURKINA FASO", 24: "BURUNDI",
               58: "CABO VERDE", 127: "CAMBODIA", 52: "CAMEROON", 46: "CANADA", 45: "CENTRAL AFRICAN REPUBLIC", 223: "CHAD", 49: "CHILE",
               50: "CHINA", 56: "COLOMBIA", 57: "COMOROS", 54: "CONGO", 53: "DEMOCRATIC REPUBLIC OF THE CONGO", 59: "COSTA RICA",
               51: "COTE D'IVOIRE", 107: "CROATIA", 60: "CUBA", 61: "CURACAO", 64: "CYPRUS", 65: "CZECH REPUBLIC", 69: "DENMARK",
               67: "DJIBOUTI", 68: "DOMINICA", 70: "DOMINICAN REPUBLIC", 72: "ECUADOR", 73: "EGYPT", 207: "EL SALVADOR", 96: "EQUATORIAL GUINEA",
               74: "ERITREA", 77: "ESTONIA", 218: "ESWATINI", 78: "ETHIOPIA", 83: "FAROE ISLANDS", 80: "FIJI", 79: "FINLAND", 82: "FRANCE",
               86: "GABON", 94: "GAMBIA", 252: "GAZA STRIP", 88: "GEORGIA", 66: "GERMANY", 90: "GHANA", 97: "GREECE", 99: "GREENLAND",
               98: "GRENADA", 100: "GUATEMALA", 92: "GUINEA", 95: "GUINEA-BISSAU", 103: "GUYANA", 108: "HAITI", 243: "HOLY SEE",
               106: "HONDURAS", 104: "HONG KONG", 109: "HUNGARY", 117: "ICELAND", 112: "INDIA", 110: "INDONESIA", 115: "ISLAMIC REPUBLIC OF IRAN",
               116: "IRAQ", 114: "IRELAND", 118: "ISRAEL", 119: "ITALY", 120: "JAMAICA", 123: "JAPAN", 122: "JORDAN", 124: "KAZAKHSTAN", 125: "KENYA",
               128: "KIRIBATI", 188: "DEMOCRATIC PEOPLE'S REPUBLIC OF KOREA", 130: "REPUBLIC OF KOREA", 131: "KUWAIT", 126: "KYRGYZSTAN",
               132: "LAO PEOPLE'S DEMOCRATIC REPUBLIC", 142: "LATVIA", 133: "LEBANON", 139: "LESOTHO", 134: "LIBERIA", 135: "LIBYA",
               137: "LIECHTENSTEIN", 140: "LITHUANIA", 141: "LUXEMBOURG", 148: "MADAGASCAR", 163: "MALAWI", 164: "MALAYSIA", 149: "MALDIVES",
               152: "MALI", 153: "MALTA", 151: "MARSHALL ISLANDS", 159: "MAURITANIA", 162: "MAURITIUS", 150: "MEXICO",
               84: "FEDERATED STATES OF MICRONESIA", 147: "REPUBLIC OF MOLDOVA", 146: "MONACO", 156: "MONGOLIA", 155: "MONTENEGRO", 145: "MOROCCO",
               158: "MOZAMBIQUE", 154: "MYANMAR", 263: "NACOMS", 166: "NAMIBIA", 1: "NATO & INTERNATIONAL ORG.", 176: "NAURU", 175: "NEPAL",
               173: "NETHERLANDS", 177: "NEW ZEALAND", 171: "NICARAGUA", 168: "NIGER", 170: "NIGERIA", 172: "NIUE", 2: "NON-NATO NATIONS",
               85: "NORTH MACEDONIA", 174: "NORWAY", 6: "NSPA", 4: "NSPA (NADB)", 5: "NSPA (NMCRL)", 260: "NSPA XML", 178: "OMAN", 179: "PAKISTAN",
               184: "PALAU", 191: "STATE OF PALESTINE", 180: "PANAMA", 185: "PAPUA NEW GUINEA", 190: "PARAGUAY", 182: "PERU", 183: "PHILIPPINES",
               186: "POLAND", 189: "PORTUGAL", 193: "QATAR", 195: "ROMANIA", 196: "RUSSIAN FEDERATION", 197: "RWANDA", 129: "SAINT KITTS AND NEVIS",
               136: "SAINT LUCIA", 244: "SAINT VINCENT AND THE GRENADINES", 251: "SAMOA", 208: "SAN MARINO", 213: "SAO TOME AND PRINCIPE",
               198: "SAUDI ARABIA", 200: "SENEGAL", 211: "SERBIA", 220: "SEYCHELLES", 206: "SIERRA LEONE", 201: "SINGAPORE",
               219: "SINT MAARTEN, DUTCH PART", 215: "SLOVAKIA", 216: "SLOVENIA", 205: "SOLOMON ISLANDS", 209: "SOMALIA", 256: "SOUTH AFRICA",
               212: "SOUTH SUDAN", 76: "SPAIN", 138: "SRI LANKA", 199: "SUDAN", 214: "SURINAME", 217: "SWEDEN", 48: "SWITZERLAND",
               221: "SYRIAN ARAB REPUBLIC", 235: "TAIWAN, PROVINCE OF CHINA", 226: "TAJIKISTAN", 236: "UNITED REPUBLIC OF TANZANIA",
               225: "THAILAND", 229: "TIMOR-LESTE", 224: "TOGO", 230: "TONGA", 231: "TRINIDAD AND TOBAGO", 232: "TUNISIA", 233: "TÜRKIYE",
               228: "TURKMENISTAN", 234: "TUVALU", 237: "UGANDA", 238: "UKRAINE", 14: "UNITED ARAB EMIRATES", 87: "UNITED KINGDOM",
               3: "UNITED NATIONS", 241: "UNITED STATES", 240: "URUGUAY", 242: "UZBEKISTAN", 249: "VANUATU", 245: "BOLIVARIAN REPUBLIC OF VENEZUELA",
               248: "VIET NAM", 253: "WEST BANK", 254: "YEMEN", 255: "YUGOSLAVIA", 257: "ZAMBIA", 258: "ZIMBABWE"}

# specifying the path to chrome driver is no longer needed
# https://stackoverflow.com/questions/76461596/unable-to-use-selenium-webdriver-getting-two-exceptions

def get_chrome_driver_version(
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")):
    driver_path = join(dirname(dirname(dirname(dirname(FILEPATH)))),
                       "seleniumdriver", "chromedriver.exe")
    if qt_window is not None and debug:
        console.emit(f"FILEPATH 1: {FILEPATH}")
        console.emit(f"driver_path 1: {driver_path}")
    with open(driver_path, 'rb') as _:
        license_content = _.read()
    installed_chrome_driver_version = search(
        r'("Surface Duo".*?Chrome\/)(.*?)( )', license_content.decode("latin-1")).group(2)
    return installed_chrome_driver_version


def update_chrome_driver(
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")):
    # Check if driver needs to be updated
    latest_chrome_version = requests.get(
        "https://chromedriver.storage.googleapis.com/LATEST_RELEASE").text
    installed_chrome_driver_version = get_chrome_driver_version(
        debug=debug, qt_window=qt_window, progress=progress, console=console)
    if latest_chrome_version.split('.')[0] != installed_chrome_driver_version.split('.')[0]:
        # Install newest driver
        installation_link = f"https://chromedriver.storage.googleapis.com/{latest_chrome_version}/chromedriver_win32.zip"
        headers = {'Content-Disposition': 'attachment'}
        response = requests.get(installation_link, headers=headers)
        driver_path = join(dirname(dirname(dirname(dirname(FILEPATH)))),
                           "seleniumdriver", "chromedriver.exe")
        if qt_window is not None and debug:
            console.emit(f"FILEPATH 2: {FILEPATH}")
            console.emit(f"driver_path 2: {driver_path}")
        with open(join(dirname(dirname(dirname(dirname(FILEPATH)))), "chromedriver_win32.zip"), "wb") as _:
            _.write(response.content)
        # unarchive_file(join(abspath("chromedriver_win32"), "chromedriver_win32.zip"))
        with zipfile.ZipFile(join(dirname(dirname(dirname(dirname(FILEPATH)))), "chromedriver_win32.zip"), 'r') as _:
            _.extractall(join(dirname(dirname(dirname(dirname(FILEPATH)))), "seleniumdriver"))
        remove(join(dirname(dirname(dirname(dirname(FILEPATH)))), "chromedriver_win32.zip"))


class VendorList():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
            xml_content = linearize_xml(xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def get_vendor_list(self) -> str:
        """Function to extract the vendor list from the xml.

        Returns:
            list: list of vendors.
        """
        xml_content = self.replace_entities()
        if "<title>Vendor List</title>" not in xml_content and "<title>VENDOR LIST</title>" not in xml_content:
            return xml_content, []
        vendlist = search(
            r"(\<vendlist.*?\>)(.*?)(\</vendlist\>)", xml_content).group(2)
        vendlist = [match for match in findall(
            r"(?:\<vendata.*?\>)(.*?)(?:\</vendata\>)", vendlist)]

        # Capture vendor code and vendor info in tuples
        vendor_info = []
        for elem in vendlist:
            temp = ()
            if search(r"(\<mfr\>)(.*?)(\</mfr\>)", elem):
                mfr = search(r"(\<mfr\>)(.*?)(\</mfr\>)", elem).group(2)
            else:
                mfr = ""
            if search(r"(\<mad\>)(.*?)(\</mad\>)", elem):
                mad = search(r"(\<mad\>)(.*?)(\</mad\>)", elem).group(2)
            else:
                mad = ""
            temp = (mfr, mad)
            if mfr not in VENDOR_EXCEPTIONS:
                vendor_info.append(temp)
        # remove duplicates
        vendor_info = list(dict.fromkeys(vendor_info))
        return xml_content, vendor_info

    def get_vendor_codes_from_itemdata(self):
        """Function to extract the vendor codes from the itemdata elements.
        """
        xml_content, vendor_info = self.get_vendor_list()

        # Get all itemdata elements
        itemdata = []
        for match in findall(r"(?:\<itemdata.*?\>)(.*?)(?:\</itemdata\>)", xml_content):
            itemdata.append(match)

        mfr_list = []
        # for elem in itemdata:
        #     if search(r"(\<mfr\>)(.*?)(\</mfr\>)", elem):
        #         print(search(r"(\<mfr\>)(.*?)(\</mfr\>)", elem).group(2))
        #         if search(r"(\<mfr\>)(.*?)(\</mfr\>)", elem).group(2) not in VENDOR_EXCEPTIONS:
        #             mfr_list.append(
        #                 search(r"(\<mfr\>)(.*?)(\</mfr\>)", elem).group(2))
        for elem in itemdata:
            matches = findall(
                r"(\<mfr\>)(.*?)(\</mfr\>)|(\<optmfr\>.*?\<mfr\>)(.*?)(\</mfr\>)", elem)
            for groups in matches:
                if groups[1] and groups[1] not in VENDOR_EXCEPTIONS:
                    mfr_list.append(groups[1])
                if groups[4] and groups[4] not in VENDOR_EXCEPTIONS:
                    mfr_list.append(groups[4])

        # remove duplicates
        mfr_list = list(dict.fromkeys(mfr_list))
        print(mfr_list)
        return vendor_info, mfr_list

    def lookup_vendor(self, vendor_code: str, list_to_check: list) -> bool:
        """Function to check if a vendor code is in the given list."""
        if list_to_check == []:
            return False

        if isinstance(list_to_check[0], tuple):
            for item in list_to_check:
                if vendor_code == item[0]:
                    return True
            return False
        else:
            return vendor_code in list_to_check

    def check_vendor_codes(self):
        """Function to create an excel file with two sheets"""
        vendor_info, mfr_list = self.get_vendor_codes_from_itemdata()

        workbook = Workbook()
        vendlist_sheet = workbook.active
        vendlist_sheet.title = "Vendor List"
        iplnom_sheet = workbook.create_sheet(title="Itemdata")

        vendlist_sheet.cell(row=1, column=1).value = "Vendor Code"
        vendlist_sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        vendlist_sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')

        vendlist_sheet.cell(row=1, column=2).value = "Found in itemdata"
        vendlist_sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        vendlist_sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')

        for ind, item in enumerate(vendor_info):
            vendlist_sheet.cell(row=ind + 2, column=1).value = item[0]
            if self.lookup_vendor(item[0], mfr_list):
                vendlist_sheet.cell(row=ind + 2, column=2).value = "True"
            else:
                vendlist_sheet.cell(row=ind + 2, column=2).value = "False"

        iplnom_sheet.cell(row=1, column=1).value = "Vendor Code"
        iplnom_sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        iplnom_sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')

        iplnom_sheet.cell(row=1, column=2).value = "Found in Vendor List"
        iplnom_sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        iplnom_sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')

        for ind, item in enumerate(mfr_list):
            iplnom_sheet.cell(row=ind + 2, column=1).value = item
            if self.lookup_vendor(item, vendor_info):
                iplnom_sheet.cell(row=ind + 2, column=2).value = "True"
            else:
                iplnom_sheet.cell(row=ind + 2, column=2).value = "False"

        workbook.save(
            join(self.export_path, f"vendor_check_{basename(normpath(self.xml_path))}.xlsx"))

        return mfr_list

    def get_addresses(
            self,
            debug: bool = False,
            qt_window: QMainWindow = None,
            progress: Signal = Signal(0),
            console: Signal = Signal("")):
        mfr_list = self.check_vendor_codes()
        web_driver_path = join(dirname(
            dirname(dirname(dirname(FILEPATH)))), "seleniumdriver", "chromedriver.exe")
        # if qt_window is not None and debug:
        #     console.emit(f"FILEPATH 3: {FILEPATH}")
        #     console.emit(f"web_driver_path 3: {web_driver_path}")
        mad_dict_list = []
        max_progress = len(mfr_list)
        for ind, code in enumerate(mfr_list):
            if qt_window is not None:
                # max progress is 80% of the total progress	
                progress.emit(int(ind / max_progress * 80))
            #driver = webdriver.Chrome(web_driver_path)
            driver = webdriver.Chrome()
            driver.get(
                f"https://eportal.nspa.nato.int/Codification/CageTool/CageTool/GetCageDetails?cagecode={code}")
            html_content = driver.page_source

            info_dict = search(r"(\<pre.*?\>)(.*?)(\</pre\>)",
                               html_content).group(2)
            info_dict = loads(info_dict)
            mad_dict_list.append(info_dict)

            driver.quit()
        return mfr_list, mad_dict_list

    def create_vendor_list(
            self,
            debug: bool = False,
            qt_window: QMainWindow = None,
            progress: Signal = Signal(0),
            console: Signal = Signal("")) -> int:

        try:
            # update_chrome_driver(debug, qt_window, progress, console)
            mfr_list, mad_dict_list = self.get_addresses(
                debug, qt_window, progress, console)

            vendlist = "<vendlist><title>VENDOR LIST</title>content1</vendlist>"

            vendata = ""
            max_progress = len(mfr_list)
            for ind, elem in enumerate(mfr_list):
                if qt_window is not None:
                    # from 80% to 100% of the total progress
                    progress.emit(int(80 + ind / max_progress * 20))
                name = mad_dict_list[ind]["name"]
                if "I. ALLGAEU" in name:
                    name = name.replace("I. ALLGAEU", "").strip()
                street1 = mad_dict_list[ind]["streetLine1"]
                street2 = mad_dict_list[ind]["streetLine2"]
                postal_and_city = mad_dict_list[ind]["geoAddressPostalZone"] + \
                    " " + mad_dict_list[ind]["geoAddressCity"]
                country = mad_dict_list[ind]["countryId"]
                try:
                    country = COUNTRY_IDS[country]
                except KeyError:
                    pass
                # concatenate strings by comma and white space
                if street2 is not None:
                    mad_info = ", ".join(
                        [name, street1, street2, postal_and_city, str(country)])
                else:
                    mad_info = ", ".join(
                        [name, street1, postal_and_city, str(country)])
                vendata += f"<vendata><mfr>{elem}</mfr><mad>{mad_info}</mad></vendata>"

            vendlist = vendlist.replace("content1", vendata)

            with open(join(self.export_path, f"vendor_list_{basename(normpath(self.xml_path))}.xml"), "w", encoding="utf-8") as _:
                _.write(vendlist)
            if qt_window is not None:
                progress.emit(100)
                console.emit(
                    f"Vendor list successfully created at {self.export_path}")
        except Exception as err:
            if qt_window is not None and debug:
                console.emit(f"Error: {err}. Full traceback: {format_exc()}")
            return 1
        return 0


if __name__ == "__main__":
    instance = VendorList()
    instance.set_xml(
        r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-BD50-32-21-03_007-00_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\322121DPLIST_UK-RM.XML")
    instance.set_export_path(r"C:\Users\bakalarz\Desktop\Export")
    # instance.check_vendor_codes()
    instance.create_vendor_list()
