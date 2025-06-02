from re import findall
from re import sub
from re import search

from os.path import join
from os.path import expanduser
from os.path import basename
from os.path import dirname
from os.path import normpath

from traceback import format_exc

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

from math import floor
from math import ceil

from ast import literal_eval

from .xml_processing import delete_first_line
from .xml_processing import replace_special_characters
from .xml_processing import linearize_xml

FILEPATH = dirname(__file__)


class FCChecker():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")
        self.intermediate_row = "1"

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
            xml_content = linearize_xml(xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def get_rows(self) -> str:
        xml_content = self.replace_entities()
        if not search(r"(\<table.*?\>)(\<title\>Fits.*?)(\</table\>)", xml_content):
            return None

        table = search(
            r"(\<table.*?\>)(\<title\>Fits.*?)(\</table\>)", xml_content).group(2)
        table = search(r"(\<tbody\>)(.*?)(\</tbody\>)", table).group(2)
        rows = findall(r"(\<row\>)(.*?)(\</row\>)", table)
        rows = [row[1] for row in rows]
        return rows

    def calculate_conversion(self, value: str, rounding_direction: str) -> str:
        value = literal_eval(value)
        if value == 0:
            return "0.0000"
        conversion = value * 0.03937
        # shorten the conversion to 5 decimal places
        # while len(sub(r"(.*?)(\.)(.*?)", r"\3", str(conversion))) > 5:
        #     conversion = str(conversion)[:-1]
        # conversion = str(conversion)
        # if literal_eval(conversion) > 0:
        #     if conversion[-2] != "9":
        #         conversion = conversion[:-1]
        #         last_digit = conversion[-1]
        #         conversion = conversion[:-1] + str(int(last_digit) + 1)
        #     elif conversion[-2] == "9" and conversion[-3] != "9":
        #         conversion = conversion[:-2]
        #         last_digit = conversion[-1]
        #         conversion = conversion[:-1] + str(int(last_digit) + 1) + "0"
        #     elif conversion[-2] == "9" and conversion[-3] == "9" and conversion[-4] != "9":
        #         conversion = conversion[:-3]
        #         last_digit = conversion[-1]
        #         conversion = conversion[:-1] + str(int(last_digit) + 1) + "00"
        #     elif conversion[-2] == "9" and conversion[-3] == "9" and conversion[-4] == "9" and conversion[-5] != "9":
        #         conversion = conversion[:-4]
        #         last_digit = conversion[-1]
        #         conversion = conversion[:-1] + str(int(last_digit) + 1) + "000"
        #     else:
        #         conversion = conversion[:-6]
        #         last_digit = conversion[-1]
        #         conversion = conversion[:-1] + \
        #             str(int(last_digit) + 1) + ".0000"
        # else:
        #     conversion = conversion[:-1]

        if rounding_direction == "min":
            conversion = ceil(conversion * 10000) / 10000
        if rounding_direction == "max":
            conversion = floor(conversion * 10000) / 10000

        conversion = str(conversion)
        while len(conversion.split('.')[1]) < 4:
            conversion += '0'
        return conversion

    def check_conversion(self, info: list) -> list:
        """Function to check if the given conversion is correct.
        args: info (list): list of strings containing the information to be checked.
        returns: result (list): list of lists containing the information to be exported."""
        result = []
        count_first_if = 0
        for elem in info:
            inner_result = []
            if search(r"(-*\d+.*\d*  \(-*\d+.*\d*\))", elem):
                count_first_if += 1
                if info[0].isdigit():
                    self.intermediate_row = info[0]
                inner_result.append(self.intermediate_row)

                value = elem.split(" ")[0]
                inner_result.append(value)

                if count_first_if % 2 == 0 or count_first_if == 7:  # even
                    expected_conversion = self.calculate_conversion(
                        value, "max")
                else:                                               # odd
                    expected_conversion = self.calculate_conversion(
                        value, "min")

                inner_result.append(expected_conversion)

                given_conversion = elem.split(" ")[2]
                inner_result.append(given_conversion)

                if expected_conversion == given_conversion.replace("(", "").replace(")", ""):
                    inner_result.append("True")
                else:
                    inner_result.append("False")
            if inner_result:
                result.append(inner_result)
        return result

    def check_rows(self) -> str:
        rows = self.get_rows()
        if rows is None:
            return

        workbook = Workbook()
        conv_sheet = workbook.active
        conv_sheet.title = "Conversion Check"
        assy_sheet = workbook.create_sheet(title="Assembly Clearance Check")

        # Headers for the conversion sheet
        conv_sheet.cell(row=1, column=1).value = "Row"
        conv_sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        conv_sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')

        conv_sheet.cell(row=1, column=2).value = "Value"
        conv_sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        conv_sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')

        conv_sheet.cell(row=1, column=3).value = "Expected Conversion"
        conv_sheet.cell(row=1, column=3).font = Font(
            bold=True, italic=True, size=16)
        conv_sheet.cell(row=1, column=3).alignment = Alignment(
            horizontal='center')

        conv_sheet.cell(row=1, column=4).value = "Given Conversion"
        conv_sheet.cell(row=1, column=4).font = Font(
            bold=True, italic=True, size=16)
        conv_sheet.cell(row=1, column=4).alignment = Alignment(
            horizontal='center')

        conv_sheet.cell(row=1, column=5).value = "Validation Result"
        conv_sheet.cell(row=1, column=5).font = Font(
            bold=True, italic=True, size=16)
        conv_sheet.cell(row=1, column=5).alignment = Alignment(
            horizontal='center')

        # Headers for the assembly clearance sheet
        assy_sheet.merge_cells(
            start_row=1, start_column=2, end_row=1, end_column=7)
        assy_sheet.cell(row=1, column=2).value = "Original Manufacturer Limits"
        assy_sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')
        assy_sheet.cell(row=1, column=2).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        assy_sheet.merge_cells(
            start_row=1, start_column=8, end_row=1, end_column=10)
        assy_sheet.cell(row=1, column=8).value = "In-Service Wear Limits"
        assy_sheet.cell(row=1, column=8).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=1, column=8).alignment = Alignment(
            horizontal='center')
        assy_sheet.cell(row=1, column=8).fill = PatternFill(
            start_color='E6E6E6', end_color='E6E6E6', fill_type='solid')

        assy_sheet.cell(row=2, column=1).value = "Row"
        assy_sheet.cell(row=2, column=1).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=1).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=2).value = "Expected Min."
        assy_sheet.cell(row=2, column=2).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=2).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=3).value = "Given Min."
        assy_sheet.cell(row=2, column=3).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=3).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=4).value = "Validation Result"
        assy_sheet.cell(row=2, column=4).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=4).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=5).value = "Expected Max."
        assy_sheet.cell(row=2, column=5).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=5).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=6).value = "Given Max."
        assy_sheet.cell(row=2, column=6).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=6).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=7).value = "Validation Result"
        assy_sheet.cell(row=2, column=7).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=7).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=8).value = "Expected Max."
        assy_sheet.cell(row=2, column=8).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=8).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=9).value = "Given Max."
        assy_sheet.cell(row=2, column=9).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=9).alignment = Alignment(
            horizontal='center')

        assy_sheet.cell(row=2, column=10).value = "Validation Result"
        assy_sheet.cell(row=2, column=10).font = Font(
            bold=True, italic=True, size=16)
        assy_sheet.cell(row=2, column=10).alignment = Alignment(
            horizontal='center')

        next_empty_row_conv_sheet = 2
        next_empty_row_assy_sheet = 3
        for ind, row in enumerate(rows):
            try:
                elem = search(r"(\<entry.*?\>)", row).group(1)
                entry = findall(r"(\<entry.*?\>)(.*?)(\</entry\>)", row)
                paras = [para[1].replace("<para>", " ").replace(
                    "</para>", " ").replace("   ", "  ").strip() for para in entry]
                paras = [sub(r"\<para.*?\>", "", para) for para in paras]
                # print(paras)
                if 'morerows="1"' in elem and 'matched' not in paras[2].lower():
                    # print("Do conversion and calculation")
                    result = self.check_conversion(paras)
                    # print(f"Result: {result} | from: morerows case\n")
                    for elem in result:
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=1).value = elem[0]
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=2).value = elem[1]
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=3).value = elem[2]
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=4).value = elem[3]
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=5).value = elem[4]
                        next_empty_row_conv_sheet += 1

                    #  get paras for the next row in rows
                    entry_next = findall(
                        r"(\<entry.*?\>)(.*?)(\</entry\>)", rows[ind + 1])
                    paras_next = [para[1].replace("<para>", " ").replace(
                        "</para>", " ").strip() for para in entry_next]
                    paras_next = [sub(r"\<para.*?\>", "", para)
                                  for para in paras_next]
                    # print(paras)
                    # print(paras_next)
                    # print("\n")
                    # Insert Row Number
                    assy_sheet.cell(row=next_empty_row_assy_sheet,
                                    column=1).value = paras[0]

                    # Get id min and max
                    # id_value_1 = literal_eval(paras_next[1].split(" ")[0])
                    # id_value_2 = literal_eval(paras_next[2].split(" ")[0])
                    id_value_1 = float(paras_next[1].split(" ")[0])
                    id_value_2 = float(paras_next[2].split(" ")[0])
                    id_min = min(id_value_1, id_value_2)
                    id_max = max(id_value_1, id_value_2)
                    # Get od min and max
                    # od_value_1 = literal_eval(paras[2].split(" ")[0])
                    # od_value_2 = literal_eval(paras[3].split(" ")[0])
                    od_value_1 = float(paras[2].split(" ")[0])
                    od_value_2 = float(paras[3].split(" ")[0])
                    od_min = min(od_value_1, od_value_2)
                    od_max = max(od_value_1, od_value_2)

                    assy_min = str(round(id_min - od_max, 3))
                    while len(sub(r"(.*?)(\.)(.*?)", r"\3", assy_min)) > 3:
                        assy_min = assy_min[:-1]
                    decimal_places = len(assy_min.split(".")[-1])
                    while decimal_places < 3:
                        assy_min += "0"
                        decimal_places += 1

                    assy_max = str(round(id_max - od_min, 3))
                    while len(sub(r"(.*?)(\.)(.*?)", r"\3", assy_max)) > 3:
                        assy_max = assy_max[:-1]
                    decimal_places = len(assy_max.split(".")[-1])
                    while decimal_places < 3:
                        assy_max += "0"
                        decimal_places += 1

                    assy_sheet.cell(row=next_empty_row_assy_sheet,
                                    column=2).value = assy_min
                    assy_sheet.cell(row=next_empty_row_assy_sheet,
                                    column=3).value = paras[4].split(" ")[0]
                    if assy_min == paras[4].split(" ")[0]:
                        assy_sheet.cell(row=next_empty_row_assy_sheet,
                                        column=4).value = "True"
                    else:
                        assy_sheet.cell(row=next_empty_row_assy_sheet,
                                        column=4).value = "False"

                    assy_sheet.cell(row=next_empty_row_assy_sheet,
                                    column=5).value = assy_max
                    assy_sheet.cell(row=next_empty_row_assy_sheet,
                                    column=6).value = paras[5].split(" ")[0]
                    if assy_max == paras[5].split(" ")[0]:
                        assy_sheet.cell(row=next_empty_row_assy_sheet,
                                        column=7).value = "True"
                    else:
                        assy_sheet.cell(row=next_empty_row_assy_sheet,
                                        column=7).value = "False"

                    try:
                        if paras_next[3] != "-":
                            # Get id min and max from permitted clearence
                            p_id_value_1 = None
                            p_id_value_2 = None
                            if paras_next[3]:
                                p_id_value_1 = literal_eval(paras_next[3].split(" ")[0])
                            if paras_next[4]:
                                p_id_value_2 = literal_eval(paras_next[4].split(" ")[0])
                            if p_id_value_1 and p_id_value_2:
                                p_id_min = min(p_id_value_1, p_id_value_2)
                                p_id_max = max(p_id_value_1, p_id_value_2)
                            # Get od min and max from permitted clearence
                            p_od_value_1 = None
                            p_od_value_2 = None
                            if paras[6]:
                                p_od_value_1 = literal_eval(paras[6].split(" ")[0])
                            if paras[7]:
                                p_od_value_2 = literal_eval(paras[7].split(" ")[0])
                            if p_od_value_1 and p_od_value_2:
                                p_od_min = min(p_od_value_1, p_od_value_2)
                                p_od_max = max(p_od_value_1, p_od_value_2)

                            # print(p_id_max, p_od_min)
                            if "p_id_max" in locals() and "p_od_min" in locals():
                                perm_max = str(round(p_id_max - p_od_min, 3))
                                while len(sub(r"(.*?)(\.)(.*?)", r"\3", perm_max)) > 3:
                                    perm_max = perm_max[:-1]
                                decimal_places = len(perm_max.split(".")[-1])
                                while decimal_places < 3:
                                    perm_max += "0"
                                    decimal_places += 1

                                assy_sheet.cell(
                                    row=next_empty_row_assy_sheet, column=8).value = perm_max
                                assy_sheet.cell(row=next_empty_row_assy_sheet,
                                                column=9).value = paras[8].split(" ")[0]
                                if perm_max == paras[8].split(" ")[0]:
                                    assy_sheet.cell(
                                        row=next_empty_row_assy_sheet, column=10).value = "True"
                                else:
                                    assy_sheet.cell(
                                        row=next_empty_row_assy_sheet, column=10).value = "False"
                    except IndexError:
                        pass

                    next_empty_row_assy_sheet += 1

                else:
                    # print("Do only conversion")
                    result = self.check_conversion(paras)
                    # print(f"Result: {result} | from: standard case\n")
                    for elem in result:
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=1).value = elem[0]
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=2).value = elem[1]
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=3).value = elem[2]
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=4).value = elem[3]
                        conv_sheet.cell(row=next_empty_row_conv_sheet,
                                        column=5).value = elem[4]
                        next_empty_row_conv_sheet += 1
            except AttributeError or ValueError:
                print(f"paras[2]: {paras[2]}")
                print(f"The Error traceback for {row}\n{format_exc()}")
                continue
        workbook.save(join(
            self.export_path, f"fits_and_clearances_check_{basename(normpath(self.xml_path))}.xlsx"))

        return rows


if __name__ == "__main__":
    instance = FCChecker()
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AW39-32-13-01_005-01_EN.xml")
    instance.set_xml(
        r"C:\Users\bakalarz\Downloads\CMM-D9893-CO91-32-21-12_000-01_EN.xml")
    instance.set_export_path(r"C:\Users\bakalarz\Desktop\Export")
    instance.check_rows()
