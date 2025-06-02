from re import findall
from re import sub
from re import search
from re import escape

from os.path import join
from os.path import expanduser
from os.path import basename
from os.path import dirname
from os.path import normpath

from lxml import etree

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment

from inflect import engine

from .xml_processing import delete_first_line
from .xml_processing import replace_special_characters
from .xml_processing import linearize_xml

FILEPATH = dirname(__file__)


class IPLChecker():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def get_pgblks(self):
        xml_content = self.replace_entities()

        tree = etree.fromstring(xml_content)

        pgblk3000 = tree.xpath("//pgblk[@pgblknbr='3000']")[0]
        pgblk3000_content = linearize_xml(
            etree.tostring(pgblk3000, encoding="unicode"))

        pgblk6000 = tree.xpath("//pgblk[@pgblknbr='6000']")[0]
        pgblk6000_content = linearize_xml(
            etree.tostring(pgblk6000, encoding="unicode"))

        pgblk7000 = tree.xpath("//pgblk[@pgblknbr='7000']")[0]
        pgblk7000_content = linearize_xml(
            etree.tostring(pgblk7000, encoding="unicode"))

        return pgblk3000_content, pgblk6000_content, pgblk7000_content

    # def create_dict(self, pgblk_content, determiners):
    #     pgblk_content = etree.fromstring(pgblk_content)
    #     my_dict = {}
    #     build_dict = []
    #     for para in pgblk_content.xpath('//para'):
    #         para_text = ''.join(para.itertext())
    #         sentences = para_text.split('. ')
    #         if sentences is None:
    #             sentences = [para_text]
    #         for sentence in sentences:
    #             item_numbers = findall(
    #                 r"\(\d{1,3}[a-zA-Z]{0,2}\-\d{1,4}[a-zA-Z]{0,2}\)", sentence)
    #             temp = []
    #             for item_number in item_numbers:
    #                 match = search(r"(" + "|".join(determiners) + r")(.*)(" + escape(item_number) + ")", sentence)
    #                 if match:
    #                     nomenclature = match.group(2).strip()
    #                     while search(r"(" + "|".join(determiners) + r")(.*)", nomenclature):
    #                         nomenclature = search(
    #                             r"(" + "|".join(determiners) + r")(.*)", nomenclature).group(2)
    #                     if (item_number, nomenclature.strip()):
    #                         temp.append((item_number, nomenclature))
    #                 build_dict.append(temp)
    #     for i, elem in enumerate(build_dict):
    #         for j, subelem in enumerate(elem):
    #             elem[j] = list(elem[j])
    #             if len(elem) >= 2:
    #                 elem[j][1] = elem[j][1].replace(elem[j - 1][1], "")
    #             if len(elem) >= 3:
    #                 elem[j][1] = elem[j][1].replace(elem[j - 2][1], "")
    #             if len(elem) >= 4:
    #                 elem[j][1] = elem[j][1].replace(elem[j - 3][1], "")
    #             elem[j][1] = sub(
    #                 r"\(\d{1,3}[a-zA-Z]{0,2}\-\d{1,4}[a-zA-Z]{0,2}\)", "", elem[j][1]).strip()
    #             elem[j] = tuple(elem[j])
    #     print(build_dict)
    #     build_dict = dict([item for sublist in build_dict for item in sublist])

    #     last_valid_value = None
    #     processed_data = {}
    #     for key, value in build_dict.items():
    #         if value == '' or value == ',' or value == 'and':
    #             if last_valid_value:
    #                 processed_data[key] = last_valid_value
    #         else:
    #             processed_data[key] = value
    #             last_valid_value = value

    #     for key, value in processed_data.items():
    #         processed_data[key] = value.replace(", ", "")
    #     return processed_data

    def create_dict(self, pgblk_content):
        pgblk_content = etree.fromstring(pgblk_content)
        my_dict = {}
        pattern = r'([a-zA-Z\s]+)\s*\((\d+-\d+)\)'
        prev_item_name = None
        for para in pgblk_content.xpath('//para'):
            para_text = ''.join(para.itertext())
            matches = findall(pattern, para_text)
            for match in matches:
                nomenclature = match[0].strip().replace("assemblies", "assy")
                item_number = match[1]
                if nomenclature in ['', 'and', ',', 'or', ' ']:
                    nomenclature = prev_item_name
                my_dict[item_number] = nomenclature
                prev_item_name = nomenclature
        words_to_remove = ['from ', 'the ', 'and ', 'an ', 'with ', 'other ', 'one ', 'second ', 'remove ', 'to ', 'install ', 'applied ', 'of ', 'surfaces ', 'step ', 'is ', 'correctly ', 'necessary ', 'between ', 'for ']
        my_new_dict = {}
        for key, value in my_dict.items():
            value = value.lower().strip()
            key = '(' + key + ')'
            while len(value.split(' ')) > 3:
                value = ' '.join(value.split(' ')[1:])
            for word in words_to_remove:
                value = value.replace(word, '')

            my_new_dict[key] = value
        # return dict(sorted(my_new_dict.items(), key=lambda x: [int(d) for d in findall(r'\d+', x[0])]))
        return my_new_dict

    def lowercase_dict_values(self, input_dict):
        output_dict = {}
        for key, value in input_dict.items():
            output_dict[key] = value.lower()
        return output_dict

    def create_proc_dicts(self):
        pgblk3000_content, pgblk6000_content, pgblk7000_content = self.get_pgblks()
        pgblk3000_content = pgblk3000_content.replace(
            "<csn>", "(").replace("</csn>", ")")
        pgblk6000_content = pgblk6000_content.replace(
            "<csn>", "(").replace("</csn>", ")")
        pgblk7000_content = pgblk7000_content.replace(
            "<csn>", "(").replace("</csn>", ")")

        determiners = [" the ", " each ", " an "]
        dict_disassembly = self.lowercase_dict_values(
            self.create_dict(pgblk3000_content))    # 3000
        dict_repair = self.lowercase_dict_values(
            self.create_dict(pgblk6000_content))         # 6000
        dict_assembly = self.lowercase_dict_values(
            self.create_dict(pgblk7000_content))       # 7000

        # print(dict_repair)
        return dict_disassembly, dict_repair, dict_assembly

    def get_dplist(self):
        xml_content = self.replace_entities()

        dict_ipl_table = {}
        tree = etree.fromstring(xml_content)

        dplist = tree.xpath("//dplist")[0]
        dplist_content = linearize_xml(
            etree.tostring(dplist, encoding="unicode"))

        tree = etree.fromstring(dplist_content)
        for figure in tree.findall(".//figure"):
            fignbr = figure.get('fignbr')
            for itemdata in figure.findall('prtlist/itemdata'):
                itemnbr = itemdata.get('itemnbr')
                while itemnbr.startswith("0"):
                    itemnbr = itemnbr[1:]
                itemnbr = f"({fignbr}-{itemnbr})"
                kwd, adt = "", ""
                kwd = itemdata.xpath('iplnom/nom/kwd')[0].text
                try:
                    adt = itemdata.xpath('iplnom/nom/adt')[0].text
                except IndexError:
                    pass
                # No 'adt' needed for screw and nut
                if "screw" in kwd.lower() or "nut" in kwd.lower() or "bolt" in kwd.lower():
                    adt = ""
                else:
                    if ',' in adt:
                        temp = adt.split(',')
                        adt = ' '.join(temp[::-1]).replace("  ", " ")
                    if ',' in kwd:
                        temp = kwd.split(',')
                        kwd = ' '.join(temp[::-1]).replace("  ", " ")
                    nom = f"{adt.strip()} {kwd.strip()}".strip()
                    if "assy" in adt.lower() or "assembly" in adt.lower() or "component" in adt.lower():
                        temp = adt.split(',')
                        last = temp[-1]
                        temp.pop()
                        adt = ' '.join(temp).replace("  ", " ").strip() + " "
                        nom = f"{adt}{kwd} {last}".strip()
                dict_ipl_table[itemnbr] = nom
        return self.lowercase_dict_values(dict_ipl_table)

    def _lookup_item_number(self, key: str, dict_to_check: dict, without_variants: bool, remove_letters: bool = False) -> bool:
        """Checks if the given key matches one of the keys in the passed dictionary.
        Returns True if a match is found, False otherwise.

        If the parameter "without_variants" is True, we remove the variants from the keys of the dictionary.
        If the parameter "remove_letters" is True, we remove the variants from the passed key.

        Args:
            key (str): Key to be checked
            dict_to_check (dict): Either "Disassemly", "Repair", "Assembly" or "IPL" -dictionary
            without_variants (bool): whether variants should be ignored or not
            remove_letters (bool, optional): whether to remove the letters from the key. Defaults to False.

        Returns:
            bool: Value holds information if a match was found or not
        """
        key_copy = key
        dict_to_check_copy = dict_to_check.copy()
        if remove_letters:
            key_copy = ''.join(filter(lambda x: not x.isalpha(), key_copy))
        if without_variants:
            for k in list(dict_to_check_copy.keys()):
                new_key = ''.join(filter(lambda x: not x.isalpha(), k))
                if new_key != k:
                    dict_to_check_copy[new_key] = dict_to_check_copy.pop(k)
        found = False
        for key2, value in dict_to_check_copy.items():
            if key_copy == key2:
                found = True
                break
        return found

    def _lookup_nomenclature(self, key: str, value: str, dict_to_check: dict, without_variants: bool) -> bool:
        """_summary_

        Args:
            key (str): _description_
            value (str): _description_
            dict_to_check (dict): _description_

        Returns:
            bool: _description_
        """
        if "screw" in value.lower() or "nut" in value.lower() or "bolt" in value.lower() or "washer" in value.lower():
            found = True
        else:
            dict_to_check_copy = dict_to_check.copy()
            original_key = key
            if without_variants:
                original_key = sub("[a-zA-Z]", "", original_key)
                dict_to_check_copy = {}
                for k, v in dict_to_check.items():
                    new_key = sub("[a-zA-Z]", "", k)
                    dict_to_check_copy[new_key] = v
            found = False
            try:
                if value in dict_to_check_copy[original_key] or value in engine().plural(dict_to_check_copy[original_key]) or engine().plural(value) in dict_to_check_copy[original_key]:
                    found = True
            except:
                pass
            if found is False:
                try:
                    if value.replace("assy", "assembly") in dict_to_check_copy[original_key] or value.replace("assy", "assembly") in engine().plural(dict_to_check_copy[original_key]) or engine().plural(value.replace("assy", "assembly")) in dict_to_check_copy[original_key]:
                        found = True
                except:
                    pass
            if found is False:
                try:
                    if value.replace("assembly", "assy") in dict_to_check_copy[original_key] or value.replace("assembly", "assy") in engine().plural(dict_to_check_copy[original_key]) or engine().plural(value.replace("assembly", "assy")) in dict_to_check_copy[original_key]:
                        found = True
                except:
                    pass
        return found

    def remove_digit_words(self, nomenclature: str) -> str:
        """Removes all digits written out as a word from a string

        Args:
            nomenclature (str): _description_

        Returns:
            str: _description_
        """
        digit_words = [' one', 'one ', ' two', 'two ', ' three', 'three ', ' four', 'four ', ' five', 'five ', ' six', 'six ', ' seven', 'seven ', ' eight', 'eight ', ' nine', 'nine ', ' ten', 'ten ', ' eleven', 'eleven ', ' twelve', 'twelve ', ' thirteen',
                       'thirteen ', ' fourteen', 'fourteen ', ' fifteen', 'fifteen ', ' sixteen', 'sixteen ', ' seventeen', 'seventeen ', ' eighteen', 'eighteen ', ' nineteen', 'nineteen ', ' twenty', 'twenty ', ' thirty', 'thirty ', ' forty', 'forty ', ' fifty', 'fifty ']
        for word in digit_words:
            if word in nomenclature:
                nomenclature = nomenclature.replace(word, "")
        return nomenclature

    def check(self):
        dict_disassembly, dict_repair, dict_assembly = self.create_proc_dicts()
        dict_ipl_table = self.get_dplist()

        workbook = Workbook()
        sheet_disassembly = workbook.active
        sheet_disassembly.title = "Disassembly"

        sheet_repair = workbook.create_sheet(title="Repair")

        sheet_assembly = workbook.create_sheet(title="Assembly")

        sheet_ipl = workbook.create_sheet(title="IPL")

        sheet_columns = {
            "Disassembly": {
                1: ("Item Number", True, 16),
                2: ("Nomenclature", True, 16),
                3: ("IPL (with variant)", False, 12),
                4: ("Assembly (with variant)", False, 12),
                5: ("IPL (without variant)", False, 12),
                6: ("Assembly (without variant)", False, 12)
            },
            "Repair": {
                1: ("Item Number", True, 16),
                2: ("Nomenclature", True, 16),
                3: ("IPL (with variant)", False, 12),
                4: ("IPL (without variant)", False, 12)
            },
            "Assembly": {
                1: ("Item Number", True, 16),
                2: ("Nomenclature", True, 16),
                3: ("IPL (with variant)", False, 12),
                4: ("Disassembly (with variant)", False, 12),
                5: ("IPL (without variant)", False, 12),
                6: ("Disassembly (without variant)", False, 12)
            },
            "IPL": {
                1: ("Item Number", True, 16),
                2: ("Nomenclature", True, 16),
                3: ("Disassembly (with variant)", False, 12),
                4: ("Repair (with variant)", False, 12),
                5: ("Assembly (with variant)", False, 12),
                6: ("Disassembly (without variant)", False, 12),
                7: ("Repair (without variant)", False, 12),
                8: ("Assembly (without variant)", False, 12)
            }
        }

        for sheet_name, columns in sheet_columns.items():
            for column, (header, is_bold, size) in columns.items():
                cell = workbook[sheet_name].cell(row=1, column=column)
                cell.value = header
                cell.font = Font(bold=is_bold, italic=True, size=size)
        # Checks for sheet "IPL"
        row = 2
        for key, value in dict_ipl_table.items():
            workbook["IPL"].cell(row=row, column=1).value = key
            workbook["IPL"].cell(row=row, column=2).value = self.remove_digit_words(
                value.lower()).strip()
            # Disassembly with variant check
            if self._lookup_item_number(key, dict_disassembly, False):
                if self._lookup_nomenclature(key, value, dict_disassembly, False):
                    workbook["IPL"].cell(row=row, column=3).value = chr(
                        0x2713)  # Check Symbol
                    workbook["IPL"].cell(
                        row=row, column=3).font = Font(color='65DA65')
                    workbook["IPL"].cell(row=row, column=3).alignment = Alignment(
                        horizontal='center')
                else:
                    workbook["IPL"].cell(row=row, column=3).value = "!"
                    workbook["IPL"].cell(
                        row=row, column=3).font = Font(color='FF9100')
                    workbook["IPL"].cell(row=row, column=3).alignment = Alignment(
                        horizontal='center')
            else:
                workbook["IPL"].cell(row=row, column=3).value = chr(
                    0x2717)  # Cross Symbol
                workbook["IPL"].cell(
                    row=row, column=3).font = Font(color='F47174')
                workbook["IPL"].cell(row=row, column=3).alignment = Alignment(
                    horizontal='center')
            # Disassembly without variant check
            if self._lookup_item_number(key, dict_disassembly, True, True):
                if self._lookup_nomenclature(key, value, dict_disassembly, True):
                    workbook["IPL"].cell(row=row, column=6).value = chr(0x2713)
                    workbook["IPL"].cell(
                        row=row, column=6).font = Font(color='65DA65')
                    workbook["IPL"].cell(row=row, column=6).alignment = Alignment(
                        horizontal='center')
                else:
                    workbook["IPL"].cell(row=row, column=6).value = "!"
                    workbook["IPL"].cell(
                        row=row, column=6).font = Font(color='FF9100')
                    workbook["IPL"].cell(row=row, column=6).alignment = Alignment(
                        horizontal='center')
            else:
                workbook["IPL"].cell(row=row, column=6).value = chr(0x2717)
                workbook["IPL"].cell(
                    row=row, column=6).font = Font(color='F47174')
                workbook["IPL"].cell(row=row, column=6).alignment = Alignment(
                    horizontal='center')
            # Repair with variant check
            if self._lookup_item_number(key, dict_repair, False):
                if self._lookup_nomenclature(key, value, dict_repair, False):
                    workbook["IPL"].cell(row=row, column=4).value = chr(0x2713)
                    workbook["IPL"].cell(
                        row=row, column=4).font = Font(color='65DA65')
                    workbook["IPL"].cell(row=row, column=4).alignment = Alignment(
                        horizontal='center')
                else:
                    workbook["IPL"].cell(row=row, column=4).value = "!"
                    workbook["IPL"].cell(
                        row=row, column=4).font = Font(color='FF9100')
                    workbook["IPL"].cell(row=row, column=4).alignment = Alignment(
                        horizontal='center')
            else:
                workbook["IPL"].cell(row=row, column=4).value = chr(0x2717)
                workbook["IPL"].cell(
                    row=row, column=4).font = Font(color='F47174')
                workbook["IPL"].cell(row=row, column=4).alignment = Alignment(
                    horizontal='center')
            # Repair without variant check
            if self._lookup_item_number(key, dict_repair, True, True):
                if self._lookup_nomenclature(key, value, dict_repair, True):
                    workbook["IPL"].cell(row=row, column=7).value = chr(0x2713)
                    workbook["IPL"].cell(
                        row=row, column=7).font = Font(color='65DA65')
                    workbook["IPL"].cell(row=row, column=7).alignment = Alignment(
                        horizontal='center')
                else:
                    workbook["IPL"].cell(row=row, column=7).value = "!"
                    workbook["IPL"].cell(
                        row=row, column=7).font = Font(color='FF9100')
                    workbook["IPL"].cell(row=row, column=7).alignment = Alignment(
                        horizontal='center')
            else:
                workbook["IPL"].cell(row=row, column=7).value = chr(0x2717)
                workbook["IPL"].cell(
                    row=row, column=7).font = Font(color='F47174')
                workbook["IPL"].cell(row=row, column=7).alignment = Alignment(
                    horizontal='center')
            # Assembly with variant check
            if self._lookup_item_number(key, dict_assembly, False):
                if self._lookup_nomenclature(key, value, dict_assembly, False):
                    workbook["IPL"].cell(row=row, column=5).value = chr(0x2713)
                    workbook["IPL"].cell(
                        row=row, column=5).font = Font(color='65DA65')
                    workbook["IPL"].cell(row=row, column=5).alignment = Alignment(
                        horizontal='center')
                else:
                    workbook["IPL"].cell(row=row, column=5).value = "!"
                    workbook["IPL"].cell(
                        row=row, column=5).font = Font(color='FF9100')
                    workbook["IPL"].cell(row=row, column=5).alignment = Alignment(
                        horizontal='center')
            else:
                workbook["IPL"].cell(row=row, column=5).value = chr(0x2717)
                workbook["IPL"].cell(
                    row=row, column=5).font = Font(color='F47174')
                workbook["IPL"].cell(row=row, column=5).alignment = Alignment(
                    horizontal='center')
            # Assembly without variant check
            if self._lookup_item_number(key, dict_assembly, True, True):
                if self._lookup_nomenclature(key, value, dict_assembly, True):
                    workbook["IPL"].cell(row=row, column=8).value = chr(0x2713)
                    workbook["IPL"].cell(
                        row=row, column=8).font = Font(color='65DA65')
                    workbook["IPL"].cell(row=row, column=8).alignment = Alignment(
                        horizontal='center')
                else:
                    workbook["IPL"].cell(row=row, column=8).value = "!"
                    workbook["IPL"].cell(
                        row=row, column=8).font = Font(color='FF9100')
                    workbook["IPL"].cell(row=row, column=8).alignment = Alignment(
                        horizontal='center')
            else:
                workbook["IPL"].cell(row=row, column=8).value = chr(0x2717)
                workbook["IPL"].cell(
                    row=row, column=8).font = Font(color='F47174')
                workbook["IPL"].cell(row=row, column=8).alignment = Alignment(
                    horizontal='center')
            row += 1
        # Checks for sheet "Disassembly"
        row = 2
        for key, value in dict_disassembly.items():
            workbook["Disassembly"].cell(row=row, column=1).value = key
            workbook["Disassembly"].cell(
                row=row, column=2).value = self.remove_digit_words(value.lower()).strip()
            # IPL with variant check
            if self._lookup_item_number(key, dict_ipl_table, False):
                if self._lookup_nomenclature(key, value, dict_ipl_table, False):
                    workbook["Disassembly"].cell(
                        row=row, column=3).value = chr(0x2713)
                    workbook["Disassembly"].cell(
                        row=row, column=3).font = Font(color='65DA65')
                    workbook["Disassembly"].cell(
                        row=row, column=3).alignment = Alignment(horizontal='center')
                else:
                    workbook["Disassembly"].cell(row=row, column=3).value = "!"
                    workbook["Disassembly"].cell(
                        row=row, column=3).font = Font(color='FF9100')
                    workbook["Disassembly"].cell(
                        row=row, column=3).alignment = Alignment(horizontal='center')
            else:
                workbook["Disassembly"].cell(
                    row=row, column=3).value = chr(0x2717)
                workbook["Disassembly"].cell(
                    row=row, column=3).font = Font(color='F47174')
                workbook["Disassembly"].cell(
                    row=row, column=3).alignment = Alignment(horizontal='center')
            # IPL without variant check
            if self._lookup_item_number(key, dict_ipl_table, True):
                if self._lookup_nomenclature(key, value, dict_ipl_table, True):
                    workbook["Disassembly"].cell(
                        row=row, column=5).value = chr(0x2713)
                    workbook["Disassembly"].cell(
                        row=row, column=5).font = Font(color='65DA65')
                    workbook["Disassembly"].cell(
                        row=row, column=5).alignment = Alignment(horizontal='center')
                else:
                    workbook["Disassembly"].cell(row=row, column=5).value = "!"
                    workbook["Disassembly"].cell(
                        row=row, column=5).font = Font(color='FF9100')
                    workbook["Disassembly"].cell(
                        row=row, column=5).alignment = Alignment(horizontal='center')
            else:
                workbook["Disassembly"].cell(
                    row=row, column=5).value = chr(0x2717)
                workbook["Disassembly"].cell(
                    row=row, column=5).font = Font(color='F47174')
                workbook["Disassembly"].cell(
                    row=row, column=5).alignment = Alignment(horizontal='center')
            # Assembly with variant
            if self._lookup_item_number(key, dict_assembly, False):
                if self._lookup_nomenclature(key, value, dict_assembly, False):
                    workbook["Disassembly"].cell(
                        row=row, column=4).value = chr(0x2713)
                    workbook["Disassembly"].cell(
                        row=row, column=4).font = Font(color='65DA65')
                    workbook["Disassembly"].cell(
                        row=row, column=4).alignment = Alignment(horizontal='center')
                else:
                    workbook["Disassembly"].cell(row=row, column=4).value = "!"
                    workbook["Disassembly"].cell(
                        row=row, column=4).font = Font(color='FF9100')
                    workbook["Disassembly"].cell(
                        row=row, column=4).alignment = Alignment(horizontal='center')
            else:
                workbook["Disassembly"].cell(
                    row=row, column=4).value = chr(0x2717)
                workbook["Disassembly"].cell(
                    row=row, column=4).font = Font(color='F47174')
                workbook["Disassembly"].cell(
                    row=row, column=4).alignment = Alignment(horizontal='center')
            # Assembly without variant
            if self._lookup_item_number(key, dict_assembly, True):
                if self._lookup_nomenclature(key, value, dict_assembly, True):
                    workbook["Disassembly"].cell(
                        row=row, column=6).value = chr(0x2713)
                    workbook["Disassembly"].cell(
                        row=row, column=6).font = Font(color='65DA65')
                    workbook["Disassembly"].cell(
                        row=row, column=6).alignment = Alignment(horizontal='center')
                else:
                    workbook["Disassembly"].cell(row=row, column=6).value = "!"
                    workbook["Disassembly"].cell(
                        row=row, column=6).font = Font(color='FF9100')
                    workbook["Disassembly"].cell(
                        row=row, column=6).alignment = Alignment(horizontal='center')
            else:
                workbook["Disassembly"].cell(
                    row=row, column=6).value = chr(0x2717)
                workbook["Disassembly"].cell(
                    row=row, column=6).font = Font(color='F47174')
                workbook["Disassembly"].cell(
                    row=row, column=6).alignment = Alignment(horizontal='center')
            row += 1
        # Checks for sheet "Repair"
        row = 2
        for key, value in dict_repair.items():
            workbook["Repair"].cell(row=row, column=1).value = key
            workbook["Repair"].cell(row=row, column=2).value = self.remove_digit_words(
                value.lower()).strip()
            if self._lookup_item_number(key, dict_ipl_table, False):
                if self._lookup_nomenclature(key, value, dict_ipl_table, False):
                    workbook["Repair"].cell(
                        row=row, column=3).value = chr(0x2713)
                    workbook["Repair"].cell(
                        row=row, column=3).font = Font(color='65DA65')
                    workbook["Repair"].cell(
                        row=row, column=3).alignment = Alignment(horizontal='center')
                else:
                    workbook["Repair"].cell(row=row, column=3).value = "!"
                    workbook["Repair"].cell(
                        row=row, column=3).font = Font(color='FF9100')
                    workbook["Repair"].cell(
                        row=row, column=3).alignment = Alignment(horizontal='center')
            else:
                workbook["Repair"].cell(row=row, column=3).value = chr(0x2717)
                workbook["Repair"].cell(
                    row=row, column=3).font = Font(color='F47174')
                workbook["Repair"].cell(
                    row=row, column=3).alignment = Alignment(horizontal='center')
            if self._lookup_item_number(key, dict_ipl_table, True):
                if self._lookup_nomenclature(key, value, dict_ipl_table, True):
                    workbook["Repair"].cell(
                        row=row, column=4).value = chr(0x2713)
                    workbook["Repair"].cell(
                        row=row, column=4).font = Font(color='65DA65')
                    workbook["Repair"].cell(
                        row=row, column=4).alignment = Alignment(horizontal='center')
                else:
                    workbook["Repair"].cell(row=row, column=4).value = "!"
                    workbook["Repair"].cell(
                        row=row, column=4).font = Font(color='FF9100')
                    workbook["Repair"].cell(
                        row=row, column=4).alignment = Alignment(horizontal='center')
            else:
                workbook["Repair"].cell(row=row, column=4).value = chr(0x2717)
                workbook["Repair"].cell(
                    row=row, column=4).font = Font(color='F47174')
                workbook["Repair"].cell(
                    row=row, column=4).alignment = Alignment(horizontal='center')
            row += 1
        # Checks for sheet "Assembly"
        row = 2
        for key, value in dict_assembly.items():
            workbook["Assembly"].cell(row=row, column=1).value = key
            workbook["Assembly"].cell(
                row=row, column=2).value = self.remove_digit_words(value.lower()).strip()
            # IPL with variant
            if self._lookup_item_number(key, dict_ipl_table, False):
                if self._lookup_nomenclature(key, value, dict_ipl_table, False):
                    workbook["Assembly"].cell(
                        row=row, column=3).value = chr(0x2713)
                    workbook["Assembly"].cell(
                        row=row, column=3).font = Font(color='65DA65')
                    workbook["Assembly"].cell(
                        row=row, column=3).alignment = Alignment(horizontal='center')
                else:
                    workbook["Assembly"].cell(row=row, column=3).value = "!"
                    workbook["Assembly"].cell(
                        row=row, column=3).font = Font(color='FF9100')
                    workbook["Assembly"].cell(
                        row=row, column=3).alignment = Alignment(horizontal='center')
            else:
                workbook["Assembly"].cell(
                    row=row, column=3).value = chr(0x2717)
                workbook["Assembly"].cell(
                    row=row, column=3).font = Font(color='F47174')
                workbook["Assembly"].cell(
                    row=row, column=3).alignment = Alignment(horizontal='center')
            # IPL without variant
            if self._lookup_item_number(key, dict_ipl_table, True):
                if self._lookup_nomenclature(key, value, dict_ipl_table, True):
                    workbook["Assembly"].cell(
                        row=row, column=5).value = chr(0x2713)
                    workbook["Assembly"].cell(
                        row=row, column=5).font = Font(color='65DA65')
                    workbook["Assembly"].cell(
                        row=row, column=5).alignment = Alignment(horizontal='center')
                else:
                    workbook["Assembly"].cell(row=row, column=5).value = "!"
                    workbook["Assembly"].cell(
                        row=row, column=5).font = Font(color='FF9100')
                    workbook["Assembly"].cell(
                        row=row, column=5).alignment = Alignment(horizontal='center')
            else:
                workbook["Assembly"].cell(
                    row=row, column=5).value = chr(0x2717)
                workbook["Assembly"].cell(
                    row=row, column=5).font = Font(color='F47174')
                workbook["Assembly"].cell(
                    row=row, column=5).alignment = Alignment(horizontal='center')
            # Disassembly with variant
            if self._lookup_item_number(key, dict_disassembly, False):
                if self._lookup_nomenclature(key, value, dict_disassembly, False):
                    workbook["Assembly"].cell(
                        row=row, column=4).value = chr(0x2713)
                    workbook["Assembly"].cell(
                        row=row, column=4).font = Font(color='65DA65')
                    workbook["Assembly"].cell(
                        row=row, column=4).alignment = Alignment(horizontal='center')
                else:
                    workbook["Assembly"].cell(row=row, column=4).value = "!"
                    workbook["Assembly"].cell(
                        row=row, column=4).font = Font(color='FF9100')
                    workbook["Assembly"].cell(
                        row=row, column=4).alignment = Alignment(horizontal='center')
            else:
                workbook["Assembly"].cell(
                    row=row, column=4).value = chr(0x2717)
                workbook["Assembly"].cell(
                    row=row, column=4).font = Font(color='F47174')
                workbook["Assembly"].cell(
                    row=row, column=4).alignment = Alignment(horizontal='center')
            # Disassembly without variant
            if self._lookup_item_number(key, dict_disassembly, False):
                if self._lookup_nomenclature(key, value, dict_disassembly, False):
                    workbook["Assembly"].cell(
                        row=row, column=6).value = chr(0x2713)
                    workbook["Assembly"].cell(
                        row=row, column=6).font = Font(color='65DA65')
                    workbook["Assembly"].cell(
                        row=row, column=6).alignment = Alignment(horizontal='center')
                else:
                    workbook["Assembly"].cell(row=row, column=6).value = "!"
                    workbook["Assembly"].cell(
                        row=row, column=6).font = Font(color='FF9100')
                    workbook["Assembly"].cell(
                        row=row, column=6).alignment = Alignment(horizontal='center')
            else:
                workbook["Assembly"].cell(
                    row=row, column=6).value = chr(0x2717)
                workbook["Assembly"].cell(
                    row=row, column=6).font = Font(color='F47174')
                workbook["Assembly"].cell(
                    row=row, column=6).alignment = Alignment(horizontal='center')
            row += 1

        workbook.save(
            join(self.export_path, f"Checker_ipl_{basename(normpath(self.xml_path))}.xlsx"))

        self.update_excel_file(join(
            self.export_path, f"Checker_ipl_{basename(normpath(self.xml_path))}.xlsx"), "Disassembly")
        self.update_excel_file(join(
            self.export_path, f"Checker_ipl_{basename(normpath(self.xml_path))}.xlsx"), "Repair")
        self.update_excel_file(join(
            self.export_path, f"Checker_ipl_{basename(normpath(self.xml_path))}.xlsx"), "Assembly")

    def update_excel_file(self, filename, sheet_name):
        wb = load_workbook(filename)
        ws = wb[sheet_name]

        for row in ws.iter_rows(min_row=2, min_col=2):
            if row[0].value == "and":
                row[0].value = row[0].offset(row=-1).value
            if row[0].value is not None:
                if "and" in row[0].value:
                    row[0].value = row[0].value.replace("and", "").strip()

        wb.save(filename)

