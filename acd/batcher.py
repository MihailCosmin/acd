"""This module provides function to batch different functions.
"""
from os import rename
from os import mkdir
from os.path import join
from os.path import isdir
from os.path import isfile

from tqdm import tqdm

from json import dump

from traceback import format_exc

from openpyxl import utils
from openpyxl import load_workbook

import sys

if sys.version_info >= (3, 12):
    from PySide6.QtWidgets import QMainWindow
    from PySide6.QtCore import Signal
else:
    from PySide2.QtWidgets import QMainWindow
    from PySide2.QtCore import Signal

from .filelist import *  # noqa # pylint: disable=unused-import, import-error, ungrouped-imports, wrong-import-position, wildcard-import, unused-wildcard-import
from .graphics import *  # noqa # pylint: disable=unused-import, import-error, ungrouped-imports, wrong-import-position, wildcard-import, unused-wildcard-import
from .file_info import *  # noqa # pylint: disable=unused-import, import-error, ungrouped-imports, wrong-import-position, wildcard-import, unused-wildcard-import
from .constants import FUNC_DICT  # noqa # pylint: disable=unused-import, import-error, ungrouped-imports, wrong-import-position, wildcard-import, unused-wildcard-import


# Batcher function
def batch(img_dir: str, ext_list: list, funcs: tuple) -> list:
    """
    This function takes a directory and returns a negative version of all images in the directory.

    img_dir: str, the directory to batch
    ext_list: list, the list of extensions to batch
    func: tuple with the functions to use on the images
    NOTE: If you give only one func, the tuple needs to end with a comma. Ex: (negative,)

    Example:
    batcher('C:\\Users\\username\\Desktop\\IMG', ("to_256", "compress_img"))

    returns: None
    """
    return_list = []

    assert isinstance(funcs, tuple), "Third argument (funcs) must be a tuple."
    for file_ in tqdm(list_files(img_dir, True)):
        if any(file_.endswith(ext) for ext in ext_list):
            for func in funcs:
                if "=" in func:
                    param = float(func.split("=")[1])
                    func = func.split("=")[0]
                try:
                    if not isfile(file_.replace(file_.split('\\')[-1], FUNC_DICT[func] + file_.split('\\')[-1])) \
                            and FUNC_DICT[func] not in file_:
                        return_list += [globals()[func](file_, param)]
                except KeyError:
                    return_list += [globals()[func](file_)]
    return return_list

def rename_illustrations(
        folder: str,
        loi_excel: str,
        undo: bool = False,
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:

    try:
        idr_list = []
        workbook = load_workbook(loi_excel, data_only=True)
        if "COVER" in workbook.sheetnames:
            sheet = workbook["COVER"]
        else:
            sheet = workbook[workbook.sheetnames[0]]

        if "cmp" in loi_excel.lower():
            start_cell = 'D6'
            end_cell = 'J2000'
        else:
            start_cell = 'B10'
            end_cell = 'H2000'

        loi_do = {}
        loi_undo = {}
        if qt_window is not None:
            progress.emit(0)
            console.emit("Reading LOI...")

        for row in sheet.iter_rows(min_row=sheet[start_cell].row, max_row=sheet[end_cell].row,
                                min_col=utils.column_index_from_string(start_cell[0]),
                                max_col=utils.column_index_from_string(end_cell[0])):
            if row[6].value is None and row[5].value is None:
                continue
            row_data = [cell.value for cell in row]

            icn = ""
            if "cmp" in loi_excel.lower():
                if row_data[5] is not None and row_data[6] is not None:
                    icn = row_data[5].split(".")[0] + "§Fig. " + row_data[6].split("Fig. ")[1]
            else:
                if row_data[6] is None:
                    icn = row_data[5].split(".")[0]
                else:
                    icn = row_data[6].split(".")[0]

            loi_do[str(row_data[0]) + "_" + str(row_data[1]) + "_" + str(row_data[2])] = icn
            if "cmp" in loi_excel.lower():
                icn = icn.split("§")[0]
            loi_undo[icn] = str(row_data[0]) + "_" + str(row_data[1]) + "_" + str(row_data[2])

        if qt_window is not None:
            progress.emit(10)
            console.emit("Finished reading LOI...\nRenaming illustrations...\n")

        scope = loi_undo if undo else loi_do       
        if debug and qt_window is not None:
            if not isdir(join(qt_window.exe_path, "debug")):
                mkdir(join(qt_window.exe_path, "debug"))
            with open(join(qt_window.exe_path, "debug", "scope.json"), "w") as file_:
                dump(scope, file_, indent=4)
        if not scope:
            if qt_window is not None:
                progress.emit(100)
                console.emit("Could not read LOI corerectly.")
            return 1

        max_progress = len(scope)
        for source, target in scope.items():
            title = None
            if "cmp" in loi_excel.lower():
                try:
                    title = target.split("§")[1]
                    title = title.replace("Fig. ", "Fig.")
                    title = title.replace("Sheet ", "Sh")
                except IndexError:
                    pass
                target = target.split("§")[0]
            if isfile(join(folder, source + ".cgm")):
                rename(join(folder, source + ".cgm"), join(folder, target + ".cgm"))
            if isfile(join(folder, source + ".idr")):
                rename(join(folder, source + ".idr"), join(folder, target + ".idr"))
                if title is None:
                    idr_list.append(join(folder, target + ".idr"))
                else:
                    idr_list.append(join(folder, target + ".idr§" + title))
            if isfile(join(folder, source + ".svg")):
                rename(join(folder, source + ".svg"), join(folder, target + ".svg"))
            if qt_window is not None:
                progress.emit(int((list(scope.keys()).index(source) + 1) / max_progress * 90 + 10))
                console.emit(f"Renamed {source} to {target}.")
        with open(join(folder, "idr_list.txt"), "w", encoding="utf-8") as file_:
            file_.write("\n".join(idr_list))
    except Exception as err:
        if debug:
            print(format_exc())
        if qt_window is not None:
            progress.emit(100)
            console.emit(f"Finished with error: {err}")
        return 1
    if qt_window is not None:
        progress.emit(100)
        console.emit("Done!")
    return 0

if __name__ == "__main__":
    rename_illustrations(
        r"C:\Users\munteanu\Downloads\TEST",
        r"C:\Users\munteanu\Downloads\TEST\CMM-D9893-C091-32-11-21_000-01_EN - List of illustrations.xlsm",
        False
    )
