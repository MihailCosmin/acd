from os import getcwd
from os import rename
from os import remove
from os.path import join

from time import sleep
from pandas import ExcelWriter
from pandas import read_excel as pd_read_excel

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment

from requests import get

import sys

if sys.version_info >= (3, 12):
    from PySide6.QtWidgets import QMainWindow
    from PySide6.QtCore import Signal
else:
    from PySide2.QtWidgets import QMainWindow
    from PySide2.QtCore import Signal

def download_excel(excel: str, path: str = None) -> None:
    """download_excel downloads an Excel file from a onedrive.live.com URL.

    Args:
        excel (str): URL to the Excel file [onedrive.live.com].
        path (str, optional): Path to save the Excel file. Defaults to None.
    """

    path = join(getcwd(), "temp.xlsx") if path is None else path
    excel = excel.replace("view", "download") if "view" in excel else excel
    with open(path, 'wb') as _:
        _.write(get(excel).content)

def get_excel_sheet_names(excel: str) -> list:
    """
    Get the names of the sheets in an Excel file

    Args:
        excel (str): Excel file path and name

    Returns:
        list: List of sheet names

    """
    workbook = load_workbook(filename=excel)
    return workbook.sheetnames

def colum_number_to_letter(num: int) -> str:
    """Converts column number to column letter.

    Args:
        num (int): Column number

    Returns:
        str: Column letter
    """

    string = ""
    while num > 0:
        module = (num - 1) % 26
        string = chr(65 + module) + string
        num = int((num - module) / 26)
    return string

def format_excel(
        excel_file: str,
        sheets: list = None,
        inplace: bool = False,
        header_rows: int = 1,
        header_format: dict = None,
        body_format: dict = None,
        remove_columns: list = None,
        auto_filter_columns: bool = False,
        freeze_panes: tuple = None,
        column_widths: dict = None,
        column_alignments: dict = None,
        debug: bool = False,
        progress_start: int = None,
        progress_end: int = None,
        qt_window: QMainWindow = None,
        progress: Signal = None,
        console: Signal = None):
    """Format an Excel file

    Args:
        excel_file (str): Excel file path and name
        sheets (list, optional): List of sheets to format. Defaults to None.
        inplace (bool, optional): If the excel file should be saved inplace or not. Defaults to True.
        header_rows (int, optional): Number of header rows. Defaults to 1.
        header_format (dict, optional): Header format. Defaults to None.
        body_format (dict, optional): Body format. Defaults to None.
        remove_columns (list, optional): List of columns to remove. Defaults to None.
        auto_filter_columns (bool, optional): If the columns should be filtered. Defaults to False.
        freeze_panes (tuple, optional): Tuple with the row and column to freeze. Defaults to None.
    """

    sheet_dfs = ()

    # if specific sheets are not specified, get all sheets
    if not isinstance(sheets, list):
        sheets = get_excel_sheet_names(excel_file)

    # write to a new file - Will replace original if inplace is True
    writer = ExcelWriter(excel_file.replace('.xlsx', '(2).xlsx'), engine='xlsxwriter')  # pylint: disable=abstract-class-instantiated
    workbook = writer.book  # pylint: disable=E1101, no-member

    # Create body and header formats
    body_format = workbook.add_format(body_format)
    header_format = workbook.add_format(header_format)

    # process scope sheets
    for ind, sheet in enumerate(sheets):
        if qt_window is not None and debug:
            console.emit(f'Formatting sheet: {sheet} no: {ind + 1} of {len(sheets)}')
        if sheet in get_excel_sheet_names(excel_file):
            sheet_d_frame = pd_read_excel(excel_file, sheet_name=sheet, dtype=str)
            if remove_columns is not None:
                for column in remove_columns:
                    # remove column by index
                    if isinstance(column, int):
                        sheet_d_frame = sheet_d_frame.drop(sheet_d_frame.columns[column], axis=1)
            sheet_dfs = sheet_dfs + (sheet_d_frame, )

            # add sheet to workbook
            workbook.add_worksheet(sheet)
            # print(workbook.sheetnames)
            # worksheet = writer.sheets[sheet]  # this maybe works with newer versions of pandas (2.0.0) and xlsxwriter
            worksheet = workbook.get_worksheet_by_name(sheet)

        # Write the dataframe to the worksheet
        merge_cells = 0
        for idx, col in enumerate(sheet_d_frame.columns):
            # if cell contains "unnamed" or "Unnamed", merge to previous cell(s)
            # get content from last valid cell
            if ("Unnamed" not in col and "unnamed" not in col) and merge_cells == 0:
                last_valid_cell = col
                worksheet.write(0, idx, col, header_format)
            elif not ("Unnamed" in col or "unnamed" in col) and merge_cells > 0:
                worksheet.merge_range(0, idx - merge_cells - 1, 0, idx - 1, last_valid_cell, header_format)
                worksheet.write(0, idx, col, header_format)
                merge_cells = 0
            else:
                merge_cells += 1

        # keep track of skipped rows (empty rows)
        skipped = 0
        for r_idx, (_, row) in enumerate(sheet_d_frame.iterrows()):
            # format additional header rows
            if header_rows > 1:
                header_rows -= 1
                cell_format = header_format
            else:
                cell_format = body_format

            # get row content in order to skip empty rows
            row_content = ""
            for c_idx, value in enumerate(row):
                row_content += str(value) if str(value) != 'nan' else ''
                row_content = row_content.replace('\n', '')
                row_content = row_content.strip()

            # skip empty rows
            if len(row_content) == 0 or (len(row_content) == 1 and str(row_content) == '0'):
                skipped += 1
            elif len(row_content) != 0:
                for c_idx, value in enumerate(row):
                    worksheet.write(r_idx + 1 - skipped, c_idx, str(value) if str(value) != 'nan' else '', cell_format)

        for i, column in enumerate(sheet_d_frame.columns):
            column_letter = get_column_letter(i + 1)
            max_length = max(sheet_d_frame[column].astype(str).map(len).max(), len(column))
            worksheet.set_column(f'{column_letter}:{column_letter}', max_length + 5)
        if auto_filter_columns is not None:
            worksheet.autofilter(0, 0, len(sheet_d_frame), len(sheet_d_frame.columns) - 1)
        if freeze_panes is not None:
            worksheet.freeze_panes(freeze_panes[0], freeze_panes[1])
        if column_widths is not None:
            for column, width in column_widths.items():
                worksheet.set_column(column, column, width)
        if qt_window is not None:
            progress.emit(int(progress_start + (ind / len(sheets)) * (progress_end - progress_start)))

    writer.close()

    if column_alignments is not None:
        workbook = load_workbook(excel_file.replace('.xlsx', '(2).xlsx'))
        worksheet = workbook.active
        for column, alignment in column_alignments.items():
            column = column + 1
            for cell in worksheet[colum_number_to_letter(column)][1:]:  # skip header
                cell.alignment = Alignment(horizontal=alignment)
        workbook.save(excel_file.replace('.xlsx', '(2).xlsx'))
        workbook.close()

    # replace original file
    if inplace:
        removed = False
        while not removed:
            try:
                remove(excel_file)
                removed = True
            except PermissionError:
                writer.close()
                sleep(1)
        renamed = False
        while not renamed:
            try:
                rename(excel_file.replace('.xlsx', '(2).xlsx'), excel_file)
                renamed = True
            except PermissionError:
                writer.close()
                sleep(1)
