from datetime import datetime

from copy import deepcopy

from dataclasses import dataclass
from dataclasses import asdict

from decimal import Decimal

import sys

from typing import Callable
from typing import Optional

from os import listdir
from os.path import join
from os.path import expanduser
from os.path import dirname
from os.path import basename
from os.path import isfile
from os.path import isdir
from os.path import abspath

from .filepath import clean_path

# from re import search  # To be replaced by regex.search, see below

from html import escape as html_escape

from io import StringIO
from json import dump
from json import dumps

import elementpath

from regex import fullmatch
from regex import escape as regex_escape
from regex import V1
from regex import compile as regex_compile
from regex import error as RegexError

from lxml import etree

from os import environ
from os import system

from .xml_processing import get_schema_from_xml
from .xml_processing import delete_first_line
from .xml_processing import translate_xsd_regex_to_python
from .xml_processing import is_in_set
from .s1000d import get_brex_ref
from .s1000d import ref_dict_to_str
from .s1000d import find_document_by_reference
from .s1000d import collect_csdb_schemas
from .default_brex import default_brex_dmc
from .default_brex import default_brex_path
from .default_brex import find_default_brex_fallback


NS_DICT = {'rdf': r'http://www.w3.org/1999/02/22-rdf-syntax-ns#',
            'xsi': r'http://www.w3.org/2001/XMLSchema-instance'}

SNS_MODES = ("normal", "strict", "unstrict")

XPATH_VERSIONS = ("1.0", "2.0")

# S1000D issues whose BREX rules are only guaranteed safe under XPath 1.0
# semantics (e.g. `=` comparison on node-sets) -- port of the version table
# in s1kd-brexcheck's `brex_requires_xpath2` (s1kd-brexcheck.c:1296-1322).
# Everything else (4.0+, or a BREX with no/unrecognised
# xsi:noNamespaceSchemaLocation) defaults to XPath 2.0.
_XPATH1_ONLY_ISSUE_PREFIXES = (
    "http://www.s1000d.org/S1000D_2-0",
    "http://www.s1000d.org/S1000D_2-1",
    "http://www.s1000d.org/S1000D_2-2",
    "http://www.s1000d.org/S1000D_2-3",
    "http://www.s1000d.org/S1000D_3-0",
)

# Shared look of the two formatted reports (category D6): `to_excel_report`
# writes these as openpyxl `PatternFill`/`Font` colours (ARGB-style hex with no
# leading '#'), `to_html_report` mirrors them in `_REPORT_HTML_CSS`, so a
# workbook and an HTML report of the same run read as the same document.
_REPORT_PALETTE = {
    "header": "1F4E79",   # table header band / titles
    "grid": "D9D9D9",     # cell borders
    "error": "FCE4E4",    # failing violation rows
    "warning": "FFF2CC",  # non-failing (severity fail="no") rows
    "ok": "E2F0D9",       # passing documents
    "band": "F5F7FA",     # zebra banding
    "muted": "808080",    # secondary text / informational tabs
}

# `allowedObjectFlag` spelled out for a human reader of the Excel/HTML reports;
# the raw '0'/'1'/'2' stays the value in the JSON and XML ones.
_REPORT_FLAG_LABELS = {
    '0': "0 - not allowed",
    '1': "1 - required",
    '2': "2 - value",
}

# Column order of the Excel report's Violations sheet, and the keys
# `_report_violation_rows` builds each row with (the HTML report shows a
# reader-friendly subset of the same rows).
_REPORT_VIOLATION_COLUMNS = (
    "Document", "BREX", "Line", "Flag", "Severity", "Status", "Rule ID",
    "BR decision", "Context", "Object path", "Object use", "Finding",
    "Allowed (single)", "Allowed (pattern)", "Allowed (range)",
    "Node xpath", "Node",
)

# Inline stylesheet of `to_html_report`. Deliberately self-contained -- no
# webfont, image or CDN reference -- so the report opens identically off a
# network share, an e-mail attachment or a `file://` path. The full light
# palette is defined on bare `:root`; the dark one is redefined twice, once
# under `prefers-color-scheme` (guarded so an explicit light choice still wins)
# and once under `[data-theme="dark"]` (so the toggle wins in both directions).
_REPORT_HTML_CSS = """
:root {
  color-scheme: light dark;
  --bg: #f4f6f8; --panel: #ffffff; --ink: #171a1f; --muted: #5c6672;
  --line: #e3e7ec; --accent: #1f4e79; --accent-ink: #ffffff;
  --error-bg: #fdecec; --error-ink: #9d2933;
  --warn-bg: #fff5df; --warn-ink: #8a6100;
  --ok-bg: #e7f4ea; --ok-ink: #1d6b34;
  --code-bg: #f2f4f7;
}
@media (prefers-color-scheme: dark) {
  :root:not([data-theme="light"]) {
    --bg: #101319; --panel: #181d25; --ink: #e6eaf0; --muted: #98a2b3;
    --line: #2a313c; --accent: #7fb3e3; --accent-ink: #0d1117;
    --error-bg: #3a1e22; --error-ink: #ff9ba2;
    --warn-bg: #3a2f16; --warn-ink: #f2c66b;
    --ok-bg: #16301f; --ok-ink: #86d99b;
    --code-bg: #11151c;
  }
}
:root[data-theme="dark"] {
  --bg: #101319; --panel: #181d25; --ink: #e6eaf0; --muted: #98a2b3;
  --line: #2a313c; --accent: #7fb3e3; --accent-ink: #0d1117;
  --error-bg: #3a1e22; --error-ink: #ff9ba2;
  --warn-bg: #3a2f16; --warn-ink: #f2c66b;
  --ok-bg: #16301f; --ok-ink: #86d99b;
  --code-bg: #11151c;
}
* { box-sizing: border-box; }
body {
  margin: 0; padding: 24px; background: var(--bg); color: var(--ink);
  font: 14px/1.5 "Segoe UI", system-ui, -apple-system, Roboto, Arial, sans-serif;
}
h1 { margin: 0 0 4px; font-size: 22px; letter-spacing: -0.01em; }
p { margin: 0; }
.muted { color: var(--muted); font-size: 12px; }
.mono, code, pre {
  font-family: "Cascadia Mono", Consolas, "SF Mono", Menlo, monospace;
  font-size: 12px;
}
.page-head {
  display: flex; align-items: flex-start; justify-content: space-between;
  gap: 16px; flex-wrap: wrap; margin-bottom: 18px;
  border-bottom: 3px solid var(--accent); padding-bottom: 12px;
}
#theme-toggle {
  display: inline-flex; align-items: center; gap: 8px; cursor: pointer;
  background: var(--panel); color: var(--ink); border: 1px solid var(--line);
  border-radius: 999px; padding: 7px 14px; font: inherit; font-size: 13px;
}
#theme-toggle:hover { border-color: var(--accent); }
.theme-icon { width: 12px; height: 12px; border-radius: 50%;
  background: linear-gradient(135deg, var(--accent) 50%, transparent 50%);
  border: 1px solid var(--accent); }
.cards { display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 16px; }
.card {
  flex: 1 1 132px; background: var(--panel); border: 1px solid var(--line);
  border-left: 4px solid var(--muted); border-radius: 8px; padding: 12px 14px;
  display: flex; flex-direction: column; gap: 2px;
}
.card-value { font-size: 26px; font-weight: 650; line-height: 1.1; }
.card-label { color: var(--muted); font-size: 12px; text-transform: uppercase;
  letter-spacing: 0.04em; }
.card.error { border-left-color: var(--error-ink); background: var(--error-bg); }
.card.error .card-value { color: var(--error-ink); }
.card.warning { border-left-color: var(--warn-ink); background: var(--warn-bg); }
.card.warning .card-value { color: var(--warn-ink); }
.card.ok { border-left-color: var(--ok-ink); background: var(--ok-bg); }
.card.ok .card-value { color: var(--ok-ink); }
.card.neutral { border-left-color: var(--accent); }
.chips { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 18px; }
.chip {
  display: inline-flex; align-items: center; gap: 8px; background: var(--panel);
  border: 1px solid var(--line); border-radius: 999px; padding: 4px 12px;
  font-size: 12px;
}
.chip-key { color: var(--muted); }
.chip-value { font-weight: 650; }
.section {
  background: var(--panel); border: 1px solid var(--line); border-radius: 8px;
  margin-bottom: 16px; overflow: hidden;
}
.section > summary {
  cursor: pointer; padding: 12px 16px; display: flex; align-items: center;
  gap: 10px; font-weight: 650; list-style: none;
}
.section > summary::-webkit-details-marker { display: none; }
.section > summary::before {
  content: "\\25B8"; color: var(--muted); transition: transform .15s ease;
}
.section[open] > summary::before { transform: rotate(90deg); }
.section-title { font-size: 15px; }
.pill {
  background: var(--accent); color: var(--accent-ink); border-radius: 999px;
  padding: 1px 9px; font-size: 12px; font-weight: 650;
}
.controls {
  display: flex; flex-wrap: wrap; gap: 12px; align-items: center;
  padding: 0 16px 12px;
}
.controls input[type="search"] {
  flex: 1 1 260px; padding: 7px 11px; border-radius: 6px; font: inherit;
  border: 1px solid var(--line); background: var(--bg); color: var(--ink);
}
.switch { display: inline-flex; align-items: center; gap: 6px; font-size: 13px;
  color: var(--muted); cursor: pointer; }
.table-wrap { overflow-x: auto; border-top: 1px solid var(--line); }
table.grid { border-collapse: collapse; width: 100%; font-size: 13px; }
table.grid th {
  position: sticky; top: 0; z-index: 1; text-align: left; white-space: nowrap;
  background: var(--accent); color: var(--accent-ink); font-weight: 650;
  padding: 9px 12px;
}
table.grid td {
  padding: 8px 12px; border-bottom: 1px solid var(--line);
  vertical-align: top; word-break: break-word;
}
table.grid tbody tr:nth-child(even) { background: color-mix(in srgb, var(--bg) 55%, transparent); }
table.grid td.num { text-align: right; white-space: nowrap; }
table.grid td.details { min-width: 240px; }
table.grid td.details p { margin: 0 0 4px; }
.finding { color: var(--error-ink); }
.status {
  display: inline-block; border-radius: 4px; padding: 1px 8px; font-size: 12px;
  font-weight: 650; white-space: nowrap;
}
.status.error, .status.failed { background: var(--error-bg); color: var(--error-ink); }
.status.warning { background: var(--warn-bg); color: var(--warn-ink); }
.status.passed { background: var(--ok-bg); color: var(--ok-ink); }
.status.skipped { background: var(--code-bg); color: var(--muted); }
code { background: var(--code-bg); border-radius: 4px; padding: 1px 5px; }
pre {
  background: var(--code-bg); border: 1px solid var(--line); border-radius: 6px;
  padding: 10px; overflow-x: auto; margin: 6px 0 0; white-space: pre-wrap;
}
.details details > summary { cursor: pointer; color: var(--muted); font-size: 12px; }
.empty {
  background: var(--panel); border: 1px solid var(--line); border-radius: 8px;
  padding: 32px; text-align: center; color: var(--ok-ink);
}
.empty-mark { font-size: 32px; display: block; margin-bottom: 8px; }
"""

# Inline behaviour of `to_html_report`: the dark/light override on top of the
# reader's `prefers-color-scheme` (remembered per browser, every storage access
# guarded so a `file://` document with site data blocked still renders), and a
# live filter over the violations table.
_REPORT_HTML_JS = """
(function () {
  var root = document.documentElement;
  var toggle = document.getElementById('theme-toggle');
  function stored(key, value) {
    try {
      if (value === undefined) { return window.localStorage.getItem(key); }
      window.localStorage.setItem(key, value);
    } catch (e) { /* private window, file:// with site data blocked, ... */ }
    return null;
  }
  function currentlyDark() {
    if (root.dataset.theme) { return root.dataset.theme === 'dark'; }
    return window.matchMedia('(prefers-color-scheme: dark)').matches;
  }
  var saved = stored('brex-report-theme');
  if (saved === 'dark' || saved === 'light') { root.dataset.theme = saved; }
  if (toggle) {
    toggle.addEventListener('click', function () {
      var next = currentlyDark() ? 'light' : 'dark';
      root.dataset.theme = next;
      stored('brex-report-theme', next);
    });
  }

  var table = document.getElementById('violations');
  if (!table) { return; }
  var filter = document.getElementById('violation-filter');
  var errorsOnly = document.getElementById('errors-only');
  var counter = document.getElementById('violation-count');
  var rows = Array.prototype.slice.call(table.tBodies[0].rows);
  function apply() {
    var needle = (filter && filter.value || '').toLowerCase();
    var only = errorsOnly && errorsOnly.checked;
    var shown = 0;
    rows.forEach(function (row) {
      var hide = (only && row.dataset.status !== 'error') ||
        (needle && row.textContent.toLowerCase().indexOf(needle) === -1);
      row.hidden = hide;
      if (!hide) { shown++; }
    });
    if (counter) {
      counter.textContent = shown === rows.length
        ? rows.length + ' violations'
        : shown + ' of ' + rows.length + ' violations';
    }
  }
  if (filter) { filter.addEventListener('input', apply); }
  if (errorsOnly) { errorsOnly.addEventListener('change', apply); }
  apply();
})();
"""


class BrexNotFound(Exception):
    pass

class NoBrexDefined(Exception):
    pass

class NoSchemaDeclared(Exception):
    """Raised by `_check_rules` (opt-in via `set_require_schema`) when the
    checked object has no `xsi:noNamespaceSchemaLocation` on its root
    element. Not raised by default: an S1000D <= 3.0 / DTD-based object
    legitimately has no such attribute (it is schema-qualified, if at all,
    through a DOCTYPE-driven mechanism outside `xsi:noNamespaceSchemaLocation`
    entirely), and `_check_rules` already handles that case correctly --
    every schema-qualified `contextRules`/`rulesContext` group is skipped and
    every unqualified one still applies (see `test_legacy_brex_spellings.py`,
    which depends on exactly this). `set_require_schema` is for a caller who
    wants to be told loudly, instead of silently under-checking, when an
    object that was supposed to carry a schema declaration does not."""
    pass

def clean_xpath(xpath):
    """Clean the xpath extra tabs, spaces and new lines"""
    xpath = xpath.strip().replace("\n", "").replace("\t", "")
    while "  " in xpath:
        xpath = xpath.replace("  ", " ")
    return xpath


@dataclass
class BrexViolation:
    """Structured record of one BREX content-rule violation (`allowedObjectFlag`
    `0`/`1`/`2`), the canonical shape violations are reported in once they
    leave the internal per-BREX `'0'/'1'/'2'` violation lists `_check_rules`
    still returns (kept for backward compatibility -- existing callers and
    the test suite key off `result[brex_path]['0']` etc. directly). Built by
    `BrexChecker.violations()` from a `validate()`/`_check_rules()` result;
    `to_json_report`/`to_xml_report` both derive their output from this list
    rather than walking the raw nested dicts a second time.

    Attributes:
        document (str): path of the checked object this violation was found in
        brex (str): path of the BREX file the violated rule came from
        rule_id (Optional[str]): `structureObjectRule`/`objrule` `@id`, when the
            rule declares one
        br_decision_ident_number (Optional[str]): `brDecisionRef/@brDecisionIdentNumber`,
            the business-rule decision customers quote (category A5)
        flag (str): `allowedObjectFlag`/`@objappl` this violation was recorded
            under -- `'0'` (must not be present), `'1'` (must be present) or
            `'2'` (value constrained)
        rules_context (str): the rule's `rulesContext`/`@context` qualifier,
            or `''` for an unqualified rule
        severity (Optional[str]): resolved `brSeverityLevel` (own or
            `defaultBrSeverityLevel`), `None` when neither is declared
        fail (bool): whether this violation counts as an error (`True`) or,
            per a `.brseveritylevels` file marking this severity `fail="no"`,
            as a non-failing warning (see `_is_severity_failure`)
        object_path (str): the rule's `objectPath`/`objpath` XPath expression
        object_use (Optional[str]): the rule's `objectUse`/`objuse` description
        allowed_values (dict): `{'single': [...], 'pattern': [...], 'range': [...],
            'tailoring': [...]}` -- empty lists for a flag `0`/`1` violation
            with no value constraint
        node_xpath (Optional[str]): canonical XPath of the violating node
            (category D2), `None` when no node backs the violation (e.g. a
            boolean flag-0 result or a flag-1 "required but missing" result)
        line (any): source line number of the violating node, `'x'` when
            unknown, or a descriptive placeholder for a multi-line-origin
            boolean result
        node_snippet (Optional[str]): serialised copy of the violating node
            (category D3), `None` when no node backs the violation
        duplicate (bool): `True` when an identical violation (same `flag`,
            `rules_context`, `object_path` and `node_xpath`) was already
            reported against an earlier, more specific layer of the same
            layered-BREX chain -- see `_deduplicate_violations`. Excluded
            from `_count_violations`/`_append_summary`/`run_summary` and
            omitted from `to_xml_report`/`to_json_report`, but kept in the
            raw `result[brex_path]['0'/'1'/'2']` lists for transparency.
    """
    document: str
    brex: str
    rule_id: Optional[str]
    br_decision_ident_number: Optional[str]
    flag: str
    rules_context: str
    severity: Optional[str]
    fail: bool
    object_path: str
    object_use: Optional[str]
    allowed_values: dict
    node_xpath: Optional[str]
    line: any
    node_snippet: Optional[str]
    duplicate: bool = False

    def to_dict(self) -> dict:
        """Plain-dict form, e.g. for JSON serialisation."""
        return asdict(self)

class BrexChecker():
    def __init__(self):
        self._xml_path = None
        self._xml_content = None
        self._xml_dir = None

        self._brex_list = (None, None)
        self._brex_dir_path = (None, None)
        self._brex_search_paths = []
        self._brex_recursive_search = True
        self._use_default_brex = False
        self._brex_fallbacks = []

        self._severity_levels_path = None
        self._severity_levels_search = True
        self._severity_levels = (None, False)

        self._xinclude = False
        self._resolve_entities = True
        self._load_external_dtd = False
        self._allow_network = False
        self._ignore_empty = False
        self._xpath_version = None
        self._require_schema_declaration = False

        self._rule_stats = {}
        # Cache of `_show_rules` output keyed by (brex path, schema), so a
        # BREX is parsed and its rule list (including compiled selectors)
        # extracted once per run rather than once per data module checked.
        # Ref §3.12.
        self._rule_cache = {}

    def set_xml_dir(self, dir_path: str) -> None:
        """_summary_

        Args:
            dir_path (str): _description_
        """
        self._xml_dir = dir_path

    def set_xml(self, xml: str):
        """Function with which the user can set the xml to be checked
        Args:
            xml (str): xml file path
        """
        with open(clean_path(xml), "r", encoding="utf-8") as f:
            self._xml_content = f.read()
        self._xml_path = xml
        if self._brex_dir_path[0] is None and self._brex_dir_path[1] is not True:
            self._brex_dir_path = (dirname(xml), False)

    def _init_brex_list(self):
        if self._brex_list[0] is not None:
            return

        if self._use_default_brex:
            self._brex_fallbacks = []
            schema = get_schema_from_xml(self._xml_content)
            self._brex_list = ([default_brex_path(default_brex_dmc(schema))], True)
            return

        chain, fallbacks = self._walk_brex_chain(self._xml_path, self._brex_dir_path[0])
        self._brex_list = (chain, True)
        self._brex_fallbacks = fallbacks

        if len(self._brex_list[0]) == 0:
            raise NoBrexDefined(f"Brex files couldn't be found\n\
                    Please use set_brex_path method to input the directory containing ALL brex data modules or \
                    use override_brex_list if the brex data modules are in different directories.\
                    expected brex: {ref_dict_to_str(get_brex_ref(self._xml_path))}".replace("                ", ""))
        else:
            for brex in self._brex_list[0]:
                if not isfile(brex):
                    raise BrexNotFound(f"Referenced Brex: {brex} is not in {self._brex_dir_path[0]}.\n\
                    Please use set_brex_path method to input the directory containing ALL brex data modules or \
                    use override_brex_list if the brex data modules are in different directories.".replace("                ", ""))

    def _walk_brex_chain(self, xml_start: str, search_dir: str) -> tuple:
        """Walk the `brexDmRef`/`brexref` layering chain starting from
        `xml_start`, resolving each reference to a file on disk: `search_dir`
        first, then every path added via `add_brex_search_path`, then falling
        back to a built-in default BREX when the reference names one
        (`find_default_brex_fallback`) -- same resolution order
        `_init_brex_list` has always used. A visited-set cycle guard stops
        the walk if a reference resolves back to a file already seen
        (including a BREX that self-references to terminate the chain).

        Extracted from `_init_brex_list` so `lint_brex_layers` can resolve
        the same layer chain starting from a BREX file directly, independent
        of any checked object.

        Args:
            xml_start (str): path of the document to read the first
                `brexDmRef`/`brexref` from (the checked object for
                `_init_brex_list`; a BREX file itself for `lint_brex_layers`)
            search_dir (str): primary directory to resolve references
                against

        Returns:
            tuple: `(brex_list, fallbacks)` -- `brex_list` is the resolved
                chain in walk order (nearest reference first); `fallbacks`
                lists any built-in BREX substitutions, same shape as
                `self._brex_fallbacks`
        """
        brex_list = []
        fallbacks = []
        xml = xml_start
        visited = {xml}
        while True:
            brex_ref_dict = get_brex_ref(xml)
            if brex_ref_dict is None:
                break
            brex_ref = ref_dict_to_str(brex_ref_dict)
            if brex_ref in xml:
                # A BREX that references itself terminates the layering walk.
                # When this is the *first* reference in the walk -- i.e.
                # `xml_start` itself is the self-referencing BREX, as when a
                # master/default BREX shipped in a CSDB is checked directly
                # (Ref §3.11: BREX data modules are checked like any other
                # object) -- it is its own sole applicable BREX rather than
                # an empty, unresolved chain. Mirrors s1kd's `main()`, where
                # `strcmp(brex_fnames[0], dmod_fnames[i]) == 0` for exactly
                # this case because `find_brex_fname_from_doc` already
                # resolved the self-reference to the file itself. A
                # self-reference reached on a *later* iteration (walking a
                # referenced BREX's own chain) needs no such fixup: that
                # BREX is already in `brex_list` from the previous
                # iteration's `.append(resolved)`.
                if xml == xml_start:
                    brex_list.append(xml)
                break
            resolved = find_document_by_reference(brex_ref, search_dir,
                                                   recursive=self._brex_recursive_search)
            if resolved is None:
                for search_path in self._brex_search_paths:
                    resolved = find_document_by_reference(brex_ref, search_path,
                                                           recursive=self._brex_recursive_search)
                    if resolved is not None:
                        break
            if resolved is None:
                # The referenced BREX isn't on disk anywhere we looked: fall
                # back to the built-in default BREX if the reference names
                # one of them (search_brex_fname_from_default_brex).
                fallback_dmc = find_default_brex_fallback(brex_ref_dict)
                if fallback_dmc is None:
                    break
                resolved = default_brex_path(fallback_dmc)
                fallbacks.append({
                    'Reference': brex_ref,
                    'UsedBuiltinBrex': fallback_dmc,
                    'BuiltinBrexPath': resolved
                })
            if resolved in visited:
                break
            visited.add(resolved)
            brex_list.append(resolved)
            xml = resolved
        return brex_list, fallbacks

    def set_brex_path(self, brex_path: str):
        """Function with which the user can set a path where the brex files are
        located in case they are located in another directory than the xml.
        Function call can be omitted when the Brex has the same directory, the xml has.
        Args:
            brex (str): brex file path
        """
        if isdir(brex_path):
            self._brex_dir_path = (brex_path, True)
        else:
            raise BrexNotFound(f"The given path {brex_path} seems to be leading to a file. \
                Please make sure to input the path of the directory containing ALL brex data modules or \
                use override_brex_list if the brex data modules are in different directories.".replace("                ", ""))

    def add_brex_search_path(self, brex_path: str):
        """Add an additional directory to search for referenced BREX data
        modules, equivalent to `s1kd-brexcheck`'s repeatable `-I`/`--include`
        option. Can be called multiple times to register several search
        paths.

        Each path added here is only searched if the referenced BREX was not
        found in the primary path (`set_brex_path`, or the checked XML's own
        directory when `set_brex_path` was not called); paths are then tried
        in the order they were added, stopping at the first match.

        Args:
            brex_path (str): directory to add to the BREX search path list
        """
        if isdir(brex_path):
            self._brex_search_paths.append(brex_path)
        else:
            raise BrexNotFound(f"The given search path {brex_path} seems to be leading to a file. \
                Please make sure to input the path of a directory containing brex data modules.".replace("                ", ""))

    def clear_brex_search_paths(self):
        """Remove all additional BREX search paths added via
        `add_brex_search_path`.
        """
        self._brex_search_paths = []

    def set_brex_recursive_search(self, enabled: bool):
        """Enable or disable recursive search (equivalent to
        `s1kd-brexcheck`'s `-r`/`--recursive`) of the primary BREX directory
        (`set_brex_path`, or the checked XML's own directory) and of every
        path added via `add_brex_search_path`. Enabled by default; disable it
        to only look directly inside each search directory, ignoring
        subdirectories.

        Args:
            enabled (bool): whether to search subdirectories recursively
        """
        self._brex_recursive_search = enabled

    def override_brex_list(self, _brex_list: list):
        """The user can specify a list with specific paths and brex files

        Args:
            _brex_list (list): list of strings containing paths of different brex paths
        """
        for brex_elem in _brex_list:
            if isfile(brex_elem) is False:
                raise BrexNotFound(f"Brex could not be found in given directory {brex_elem}. \
                                     Please specify the absolute path.")
        self._brex_list = (_brex_list, True)
        self._brex_fallbacks = []

    def use_default_brex(self, enabled: bool = True):
        """Equivalent to `s1kd-brexcheck -B`/`--default-brex`: ignore any
        `brexDmRef`/`brexref` the object carries and check only against the
        built-in default BREX matching the object's declared S1000D schema
        version (see `default_brex.default_brex_dmc`), with no further
        `brexDmRef` layering. Overrides `set_brex_path` / `override_brex_list`
        while enabled.

        Args:
            enabled (bool, optional): if True, switch to default-BREX-only
                mode; if False, return to normal `brexDmRef` resolution.
                Defaults to True.
        """
        self._use_default_brex = enabled
        self._brex_list = (None, None)
        self._brex_fallbacks = []

    def set_severity_levels_path(self, path: str):
        """Explicitly set the path to the `.brseveritylevels` file used to resolve
        `brSeverityLevel` values, e.g.:
        `<brSeverityLevels><brSeverityLevel value="brsl01" fail="yes">Error</brSeverityLevel>...`

        This overrides the default behaviour of searching the checked XML's
        directory and its parents for a file named `.brseveritylevels`
        (see `_find_severity_levels_file`); once called, that search is skipped
        and this exact path is used instead.

        Args:
            path (str): path to the severity levels XML file
        """
        self._severity_levels_path = path
        self._severity_levels = (None, False)

    def set_severity_levels_search(self, enabled: bool):
        """Enable or disable the default parent-directory search for a
        `.brseveritylevels` file (see `_find_severity_levels_file`). Enabled by
        default; this is the override for callers who need to turn it off, e.g.
        to ignore a `.brseveritylevels` file that exists but should not apply to
        this run. Has no effect once `set_severity_levels_path` has been called.

        Args:
            enabled (bool): if False, no severity-levels file is auto-discovered
                and every violation fails, as if no severity levels applied at all
        """
        self._severity_levels_search = enabled
        self._severity_levels = (None, False)

    def _find_severity_levels_file(self) -> str:
        """Search the checked XML's directory, then its parent directories in
        turn, for a file named `.brseveritylevels`. Adapts the generic
        `find_config` helper (`s1kd_tools.c:30-56`) that s1kd tools use to
        auto-discover CSDB-wide configuration files, searching from the checked
        file's directory instead of the process's current directory.

        Returns:
            str: full path to the first `.brseveritylevels` file found; None if
                `self._xml_path` is not set yet, or no file is found before the
                filesystem root is reached
        """
        if not self._xml_path:
            return None
        current = abspath(dirname(self._xml_path))
        while True:
            candidate = join(current, ".brseveritylevels")
            if isfile(candidate):
                return candidate
            parent = dirname(current)
            if parent == current:
                return None
            current = parent

    def _get_severity_levels(self) -> dict:
        """Parse the `.brseveritylevels` file into `{value: {'fail': bool, 'type': str}}`,
        caching the result. Port of the `brsl` lookup table in `s1kd-brexcheck.c`.

        Resolves the file from `set_severity_levels_path` if one was set explicitly;
        otherwise, unless disabled via `set_severity_levels_search`, auto-discovers
        it via `_find_severity_levels_file`.

        Returns:
            dict: severity-level lookup table; empty if no file was set or found
        """
        if self._severity_levels[1]:
            return self._severity_levels[0]
        levels = {}
        path = self._severity_levels_path
        if path is None and self._severity_levels_search:
            path = self._find_severity_levels_file()
        if path is not None:
            with open(clean_path(path), "r", encoding="utf-8") as _:
                content = _.read()
            tree = etree.parse(StringIO(content))
            for level in tree.findall('.//brSeverityLevel'):
                value = level.get('value')
                if value is None:
                    continue
                levels[value] = {
                    'fail': level.get('fail') != 'no',
                    'type': "".join(level.itertext()) or None
                }
        self._severity_levels = (levels, True)
        return levels

    def _is_severity_failure(self, severity: str) -> bool:
        """Decide whether a violation at the given business-rule severity level
        counts as a failure.

        Port of `is_failure` (`s1kd-brexcheck.c:569-605`): a violation always fails
        unless a `.brseveritylevels` file is set or auto-discovered, defines this
        exact severity value, and marks it `fail="no"`.

        Args:
            severity (str): resolved `brSeverityLevel` value, or None

        Returns:
            bool: True if the violation should count as a failing error
        """
        if severity is None:
            return True
        level = self._get_severity_levels().get(severity)
        if level is None:
            return True
        return level['fail']

    def set_xinclude(self, enabled: bool = True):
        """Equivalent to `s1kd-brexcheck`'s `--xinclude`: resolve `xi:include`
        elements in the checked object and every BREX file before checking,
        via lxml's `ElementTree.xinclude()`. Mirrors `read_xml_doc`'s
        `xmlXIncludeProcessFlags` call (`s1kd_tools.c:538-539`), which is
        applied uniformly to every CSDB object it reads. Disabled by default,
        matching libxml2's `XML_PARSE_XINCLUDE` default.

        Args:
            enabled (bool): whether to process XInclude directives
        """
        self._xinclude = enabled

    def set_resolve_entities(self, enabled: bool = True):
        """Control whether entity references are substituted with their
        declared content (lxml's `resolve_entities` parser option),
        equivalent to `--noent`. Entities declared in the internal DTD
        subset are read the same way regardless (see `_check_notation_rules`,
        which reads them directly off `docinfo.internalDTD`); this only
        affects whether an entity *reference* elsewhere in the content
        survives in the parsed tree as a placeholder or is replaced in place
        before rules are checked. Enabled by default -- lxml's own default,
        already stricter than `s1kd-brexcheck`'s default of leaving
        references unresolved unless `--noent` is given.

        Args:
            enabled (bool): whether to substitute entity references with content
        """
        self._resolve_entities = enabled

    def set_entity_resolution(self, load_external_dtd: bool = True, allow_network: bool = False):
        """Enable resolving entities declared in an *external* DTD subset
        (`SYSTEM`/`PUBLIC` entities), equivalent to `s1kd-brexcheck`'s
        `--dtdload` (and, if `allow_network` is also set, `--net`). Both are
        off by default: an external DTD is not fetched, and network access is
        never allowed unless explicitly requested here, so that parsing a
        checked object cannot trigger unexpected file or network access on
        its own.

        Args:
            load_external_dtd (bool): fetch and parse the object's external DTD subset
            allow_network (bool): allow the parser to resolve `http(s)://`
                DTD/entity references over the network; only takes effect
                when `load_external_dtd` is also True. Defaults to False.
        """
        self._load_external_dtd = load_external_dtd
        self._allow_network = allow_network and load_external_dtd

    def set_xml_catalog(self, catalog_path: str):
        """Register an XML catalog file for resolving external DTD/entity/
        schema references, equivalent to `--xml-catalog <file>`
        (`xmlLoadCatalog`). lxml has no direct catalog-loading binding, so
        this appends the path to the `XML_CATALOG_FILES` environment
        variable, which libxml2 reads the first time it needs to consult the
        global catalog -- the standard way to drive libxml2's catalog
        resolution from Python. Can be called multiple times to register
        several catalogs, same as repeating the command-line flag.

        Note: libxml2 only reads `XML_CATALOG_FILES` once per process, on its
        first catalog lookup, so a catalog registered after that point may
        not take effect within the same process.

        Args:
            catalog_path (str): path to an XML (or SGML) catalog file
        """
        if not isfile(catalog_path):
            raise BrexNotFound(f"The given catalog path {catalog_path} does not point to a file.")
        entries = environ.get("XML_CATALOG_FILES", "").split()
        if catalog_path not in entries:
            entries.append(catalog_path)
            environ["XML_CATALOG_FILES"] = " ".join(entries)

    def set_ignore_empty(self, enabled: bool = True):
        """Equivalent to `-e`/`--ignore-empty`: silently skip a checked object
        that is empty or not well-formed XML, instead of raising. In
        directory mode (`set_xml_dir`) the file is left out of the results
        entirely, matching `s1kd-brexcheck`'s `continue`; for a single object
        (`set_xml`/`validate()`), the skip is reported as
        `{"Skipped": True, "Summary": "..."}` instead of raising.

        Args:
            enabled (bool): whether to skip empty/non-XML input instead of raising
        """
        self._ignore_empty = enabled

    def set_xpath_version(self, version: str = None) -> None:
        """Force the XPath version used to compile every BREX content rule's
        `objectPath`, equivalent to `s1kd-brexcheck`'s `-X`/`--xpath-version`.

        By default (`version=None`) the version is chosen per BREX file the
        way s1kd's `DYNAMIC` mode does -- see `_brex_requires_xpath2`: XPath
        1.0 for a BREX declaring S1000D issue 2.0-3.0 (matching what those
        issues were written/validated against, including reliance on XPath
        1.0's `=`-on-node-set semantics), XPath 2.0 for 4.0+ or a BREX with
        no/unrecognised declared schema. Pass `"1.0"` or `"2.0"` to force
        that version for every BREX regardless of its declared issue.

        Args:
            version (str): `"1.0"`, `"2.0"`, or `None` to restore dynamic
                per-BREX selection

        Raises:
            ValueError: if `version` is not one of `XPATH_VERSIONS` or `None`
        """
        if version is not None and version not in XPATH_VERSIONS:
            raise ValueError(
                f"xpath_version must be one of {XPATH_VERSIONS} or None, got {version!r}"
            )
        self._xpath_version = version

    def _brex_requires_xpath2(self, brex_schema: str) -> bool:
        """Decide whether a BREX's content rules should be compiled with
        `elementpath.XPath2Parser` (True) or `elementpath.XPath1Parser`
        (False). Port of `brex_requires_xpath2` (s1kd-brexcheck.c:1296-1322):
        an explicit `set_xpath_version` override always wins; otherwise a
        BREX declaring S1000D issue 2.0-3.0 gets XPath 1.0 and everything
        else -- 4.0+, or no/unrecognised `xsi:noNamespaceSchemaLocation` --
        gets XPath 2.0. Decided from the *BREX's own* declared schema, not
        the checked object's, matching the C original
        (`xmlDocGetRootElement(brex)` rather than the dmod being checked).

        Args:
            brex_schema (str): the BREX's own declared schema URI (its root
                element's `xsi:noNamespaceSchemaLocation`), or `None` if absent

        Returns:
            bool: `True` to use XPath 2.0, `False` to use XPath 1.0
        """
        if self._xpath_version == "1.0":
            return False
        if self._xpath_version == "2.0":
            return True
        if brex_schema is None:
            return True
        return not brex_schema.startswith(_XPATH1_ONLY_ISSUE_PREFIXES)

    def set_require_schema(self, enabled: bool = True) -> None:
        """Whether `_check_rules` should raise `NoSchemaDeclared` instead of
        silently proceeding when the checked object's root element has no
        `xsi:noNamespaceSchemaLocation` attribute.

        Off by default (`enabled=False`), matching `s1kd-brexcheck`, which
        never errors on a missing schema declaration either: without it,
        `_check_rules` simply cannot resolve which `contextRules`/`rulesContext`
        groups apply, so only unqualified rules are checked -- correct,
        expected behaviour for a genuine S1000D <= 3.0 / DTD-based object,
        which never carries `xsi:noNamespaceSchemaLocation` in the first
        place (see the legacy-spelling support in `_show_rules`/
        `_get_object_rule_nodes`, and `test_legacy_brex_spellings.py`, which
        depends on this object shape checking cleanly with no error).

        Turn this on when checking objects that are all expected to be
        XSD/4.0+-based and should always declare a schema: a missing
        declaration on one of them is then far more likely an authoring
        mistake (or a caller that forgot to pass the resolved XML at all)
        than a legitimate DTD-based object, and silently under-checking it
        against unqualified rules only would hide the mistake instead of
        surfacing it.

        Args:
            enabled (bool): raise `NoSchemaDeclared` instead of silently
                degrading to unqualified-rules-only checking
        """
        self._require_schema_declaration = enabled

    def _build_xml_parser(self) -> etree.XMLParser:
        """Build the lxml parser used for both the checked object and every
        BREX file, honouring the parser options set via `set_resolve_entities`
        / `set_entity_resolution`. Mirrors `DEFAULT_PARSE_OPTS`
        (`s1kd_tools.c:14`), which `read_xml_doc` applies uniformly to every
        CSDB object it reads.

        Returns:
            etree.XMLParser: configured parser
        """
        return etree.XMLParser(
            resolve_entities=self._resolve_entities,
            load_dtd=self._load_external_dtd,
            no_network=not self._allow_network,
            huge_tree=True,
        )

    def _finish_parse(self, tree: any) -> any:
        """Apply XInclude processing to a freshly parsed tree, if enabled via
        `set_xinclude`. Equivalent to `read_xml_doc`'s
        `xmlXIncludeProcessFlags` call (`s1kd_tools.c:538-539`).

        Args:
            tree (any): parsed `ElementTree`

        Returns:
            any: the same tree, with XInclude directives resolved in place if enabled
        """
        if self._xinclude:
            tree.xinclude()
        return tree

    def _parse_xml_file(self, path: str) -> any:
        """Parse an XML file from disk with the configured parser options.

        Args:
            path (str): path to the XML file

        Returns:
            any: parsed `ElementTree`
        """
        return self._finish_parse(etree.parse(path, parser=self._build_xml_parser()))

    def _parse_xml_text(self, content: str) -> any:
        """Parse XML held in a string with the configured parser options.

        Args:
            content (str): XML content

        Returns:
            any: parsed `ElementTree`
        """
        return self._finish_parse(etree.parse(StringIO(content), parser=self._build_xml_parser()))

    def _is_valid_xml_file(self, path: str) -> bool:
        """Return whether `path` parses as well-formed XML with the
        configured parser options. Used by `set_ignore_empty` to decide
        whether a checked object should be silently skipped, equivalent to
        the `read_xml_doc(...) == NULL` check in `s1kd-brexcheck.c:2151-2160`.

        Args:
            path (str): path to the object to test

        Returns:
            bool: True if the file parses as XML; False if it is missing,
                empty, or not well-formed
        """
        try:
            self._parse_xml_file(path)
            return True
        except (etree.XMLSyntaxError, OSError):
            return False

    def _parse_brex_root(self, brex_path: str) -> any:
        """Parse a BREX file into its root element, same read-strip-parse
        steps `_get_object_rule_nodes`/`_get_sns_rules_group`/
        `_get_notation_rules_group` each perform inline. Used by the lint
        checks that need the whole BREX tree (`snsRules`, `contextRules`
        groups) rather than just the `objectPath` nodes
        `_get_object_rule_nodes` returns -- which can be empty for a BREX
        that carries no content rules at all, e.g. an SNS-table-only BREX.

        Args:
            brex_path (str): path to the BREX file

        Returns:
            any: root `lxml.etree._Element`
        """
        with open(clean_path(brex_path), "r", encoding="utf-8") as _:
            brex_content = _.read()
        brex_content = delete_first_line(brex_content)
        return self._parse_xml_text(brex_content).getroot()

    def _get_object_rule_nodes(self, brex: str, schema: str = None) -> any:
        """Return all `objectPath` nodes whose enclosing `contextRules` is
        unqualified or targets the given schema, selected with the descendant
        axis so nested/grouped rules at any depth are found. Uses real XPath
        (`Element.xpath`) rather than the restricted ElementPath `findall`,
        which also avoids lxml's "This search incorrectly ignores the root
        element" FutureWarning on a leading `//`.

        Args:
            brex (str): path of the brex
            schema (str): the object's declared schema; rules whose
                `rulesContext`/`context` names a different schema are
                excluded at selection time. `None` returns every rule.

        Returns:
            any: Set of nodes
        """
        with open(clean_path(brex), "r", encoding="utf-8") as _:
            brex_content = _.read()
        brex_content = delete_first_line(brex_content)
        root = self._parse_xml_text(brex_content).getroot()
        if schema is None:
            nodes = root.xpath('//contextRules//structureObjectRule/objectPath')
            # S1000D <= 3.0 spelling
            nodes += root.xpath('//contextrules//objrule/objpath')
        else:
            nodes = root.xpath(
                '//contextRules[not(@rulesContext) or @rulesContext=$schema]'
                '//structureObjectRule/objectPath',
                schema=schema,
            )
            # S1000D <= 3.0 spelling
            nodes += root.xpath(
                '//contextrules[not(@context) or @context=$schema]//objrule/objpath',
                schema=schema,
            )
        return nodes

    def _show_rules(self, brex: str, schema: str = None, debug: bool = False) -> any:
        """Creates a, in nested dictionaries structured, JSON file containing all necessary information about the brex rules i.e.
        xpath, objectflag, objectUse, objectValues et Al.

        Args:
            brex (str): brex_path
            schema (str): the object's declared schema, passed through to
                `_get_object_rule_nodes` to filter rules at selection time

        Returns:
            any: Nested Dictionary
        """
        nodes_to_check = self._get_object_rule_nodes(brex, schema)
        default_br_severity_level = None
        # XPath version (category: XPath selection) is a property of this
        # BREX file's own declared S1000D issue, decided once per file --
        # see `_brex_requires_xpath2`. Irrelevant when there are no rules to
        # compile a selector for.
        xpath_parser = elementpath.XPath2Parser
        if len(nodes_to_check) > 0:
            brex_root = nodes_to_check[0].getroottree().getroot()
            default_br_severity_level = brex_root.get('defaultBrSeverityLevel')
            brex_schema = brex_root.get(f'{{{NS_DICT["xsi"]}}}noNamespaceSchemaLocation')
            if not self._brex_requires_xpath2(brex_schema):
                xpath_parser = elementpath.XPath1Parser
        allowed_object_flag_dict = []
        for counter, x in enumerate(nodes_to_check):
            values_allowed = []
            regex_allowed = []
            ranges_allowed = []
            value_tailoring = []
            for objectValue in x.getparent().xpath('objectValue|objval'):
                # S1000D <= 3.0 spells these @valtype and @val1[~@val2] instead
                # of @valueForm and @valueAllowed (a range is written as two
                # attributes rather than one "first~last" string).
                value_form = objectValue.get('valueForm', objectValue.get('valtype'))
                value_allowed = objectValue.get('valueAllowed')
                if value_allowed is None and objectValue.get('val1') is not None:
                    value_allowed = objectValue.get('val1')
                    val2 = objectValue.get('val2')
                    if val2 is not None:
                        value_allowed = f"{value_allowed}~{val2}"
                if value_form == "pattern":
                    regex_allowed.append(translate_xsd_regex_to_python(value_allowed))
                elif value_form == "range":
                    ranges_allowed.append(value_allowed)
                else:
                    # "single" -- and anything without a form at all, since
                    # @valueForm/@valtype is optional in the schema. Exact
                    # string equality is the fallback branch in the C
                    # reference too (`check_node_values`,
                    # `s1kd-brexcheck.c:219-248`, reached whenever `form` is
                    # NULL). Dispatching only on the three known forms would
                    # silently drop a form-less valueAllowed, turning a rule
                    # whose only objectValue omits the form into a no-op and
                    # making a rule that mixes a form-less objectValue with a
                    # formed one report false positives. Ref §2.C.
                    values_allowed.append(value_allowed)
                # Category C4: @valueTailoring distinguishes "lexical" (a project
                # may extend this allowed-value set) from "restrictable" (a
                # project may only narrow it). Absent on S1000D <= 3.0's @val1/
                # @val2 spelling and on plenty of 4.x rules too (193 of 2462 in
                # the evidence base), so only recorded per objectValue when the
                # BREX actually declares it -- nothing to distinguish otherwise.
                tailoring = objectValue.get('valueTailoring')
                if tailoring is not None:
                    value_tailoring.append({
                        'valueForm': value_form,
                        'valueAllowed': value_allowed,
                        'valueTailoring': tailoring,
                    })
            context_group = next(x.iterancestors('contextRules', 'contextrules'), None)
            context_rules = (
                context_group.get('rulesContext', context_group.get('context', ''))
                if context_group is not None else ''
            )
            br_decision_ref = x.getparent().find('brDecisionRef')
            br_decision_ident_number = br_decision_ref.get('brDecisionIdentNumber') if br_decision_ref is not None else None
            br_severity_level = x.getparent().get('brSeverityLevel')
            if br_severity_level is None:
                br_severity_level = default_br_severity_level
            # Register every namespace in scope at this objectPath node (lxml's
            # nsmap includes prefixes declared on ancestors), rather than relying
            # on a hard-coded rdf+xsi dictionary. The default namespace (lxml key
            # None) is remapped to '' as elementpath expects. NS_DICT is kept as
            # a base so rdf/xsi stay resolvable even if a rule's local scope
            # happens not to declare them. Ref §3.12.
            namespaces = dict(NS_DICT)
            namespaces.update(
                {(prefix or ''): uri for prefix, uri in x.nsmap.items()}
            )
            xpath_text = str(nodes_to_check[counter].text)
            # Read objectUse as its full text content -- a rule description
            # can carry child markup (e.g. inline formatting), and `.text`
            # alone silently drops everything after the first child element.
            # A rule with no objectUse at all (schema allows it; `lint_brex`'s
            # MissingObjectUse finding flags it separately) is tolerated as
            # None instead of the old `[0].text` raising IndexError. Ref
            # §3.12, category D4.
            object_use_nodes = x.getparent().xpath('objectUse|objuse')
            object_use = ''.join(object_use_nodes[0].itertext()) if object_use_nodes else None
            # Compile the objectPath selector once here, when the rule dict
            # is built, instead of once per (rule, document) pair inside
            # `_check_object_flag_0/1/2` -- reused unchanged across every
            # document the checker evaluates it against, since a compiled
            # `elementpath.Selector` is stateless and only ever combined with
            # a fresh `XPathContext` per document (see `_select_with_nodes`).
            # A compile failure is captured here instead of raised, and
            # reported as an `xpathError` by the flag helpers the first time
            # the rule is used, matching prior per-document behaviour. Ref
            # §3.12.
            try:
                selector = elementpath.Selector(xpath_text, namespaces=namespaces, parser=xpath_parser)
                selector_error = None
            except elementpath.ElementPathError as e:
                selector = None
                selector_error = str(e)
            allowed_object_flag_dict.append({
                    'xpath': xpath_text,
                    'Brex': str(brex),
                    'ruleId': x.getparent().get('id'),
                    'ObjectFlag': x.get('allowedObjectFlag', x.get('objappl')),
                    'objectUse': object_use,
                    'contextRules': context_rules,
                    'values_allowed': values_allowed,
                    'regex_allowed': regex_allowed,
                    'ranges_allowed': ranges_allowed,
                    'value_tailoring': value_tailoring,
                    'brDecisionIdentNumber': br_decision_ident_number,
                    'brSeverityLevel': br_severity_level,
                    'namespaces': namespaces,
                    'selector': selector,
                    'selector_error': selector_error,
                }
            )
        if debug:
            with open(clean_path(join(expanduser("~/Desktop"), f'brex_{basename(brex)}.json')), 'w', encoding="utf-8") as _:
                for elem in allowed_object_flag_dict:
                    dumpable = {k: v for k, v in elem.items() if k != 'selector'}
                    _.write(dumps(dumpable, indent=4, ensure_ascii=False))
        return allowed_object_flag_dict

    def _get_content_rules(self, brex: str, schema: str = None, debug: bool = False) -> list:
        """Cached wrapper around `_show_rules`: parses a BREX and extracts its
        rule list (including compiled selectors, see `_show_rules`) once per
        `(brex, schema)` pair for the lifetime of this checker instance,
        instead of once per data module checked -- `_check_rules` is called
        once per document, and a directory-mode `validate()` run checks many
        documents against the same BREX chain. Ref §3.12.

        Args:
            brex (str): brex_path
            schema (str): the object's declared schema
            debug (bool): forwarded to `_show_rules` on a cache miss

        Returns:
            list: same shape as `_show_rules`
        """
        cache_key = (brex, schema)
        cached = self._rule_cache.get(cache_key)
        if cached is None:
            cached = self._show_rules(brex, schema=schema, debug=debug)
            self._rule_cache[cache_key] = cached
        return cached

    def regex_builder(self, attribute_name: str, attribute_value: str, xpath):
        """Build a raw-text search regex for one attribute name/value pair.
        Both are escaped with `regex.escape` before being embedded (Ref
        §3.12) -- an attribute name or value containing regex metacharacters
        (e.g. `.`, `(`, `[`) previously produced a wrong or invalid pattern.

        Args:
            attribute_name (str): attribute name to search for
            attribute_value (str): attribute value to search for, or None to
                match any value
        Returns:
            str: regex pattern
        """
        escaped_name = regex_escape(str(attribute_name))
        if attribute_value is not None:
            build_regex = f'({escaped_name})(.*?)("{regex_escape(str(attribute_value))}")'
        else:
            build_regex = f'({escaped_name})(.*?)(")(.*?)(")'
        return build_regex

    def _select_with_nodes(self, selector: any, root: any) -> tuple:
        """Evaluate a compiled `elementpath.Selector` the same way `Selector.select`
        does, while also returning the raw XPath node backing each item of a
        node-set result. `Selector.select` (via `XPathToken.get_results`) reduces
        every node to its plain value (an lxml element, or a bare string for an
        attribute/text result), discarding the node's parent/position -- exactly
        the information needed to compute a violating node's canonical XPath and
        a copy of its owning element (categories D2/D3). This re-implements that
        reduction from the lower-level, un-formatted `root_token.select()` so the
        formatted half of the return value stays identical to plain `.select()`.

        Args:
            selector (any): compiled rule selector (`elementpath.Selector`)
            root (any): document root to evaluate the selector against

        Returns:
            tuple: `(result, nodes)`. `result` is exactly what `selector.select(root)`
                would return. `nodes` is `None` when `result` is a bare scalar (no
                node backs a computed boolean/number), otherwise a list of raw
                `elementpath` XPath node objects (or `None` per position for a
                non-node item) aligned with `result`.
        """
        context = elementpath.XPathContext(root, schema=selector.parser.schema)
        raw_items = list(selector.root_token.select(context))

        values = []
        nodes = []
        for item in raw_items:
            if isinstance(item, elementpath.xpath_nodes.XPathNode):
                values.append(item.value)
                nodes.append(item)
            else:
                values.append(item)
                nodes.append(None)

        if len(raw_items) == 1 and not isinstance(
                raw_items[0], (elementpath.xpath_nodes.ElementNode, elementpath.xpath_nodes.DocumentNode)):
            if isinstance(raw_items[0], (bool, int, float, Decimal)):
                return raw_items[0], None
            elif selector.root_token.label in ('function', 'literal'):
                return values[0], None

        return values, nodes

    def _resolve_owning_element(self, node: any) -> any:
        """Resolve a raw XPath node (from `_select_with_nodes`) to the lxml
        element that backs it, walking up to the parent for an attribute/text
        result (which has no element of its own). Shared by
        `_node_xpath_and_copy` (categories D2/D3) and `_node_line_number`
        (category D1).

        Args:
            node (any): raw XPath node from `_select_with_nodes`'s `nodes`
                list, or `None`

        Returns:
            any: the backing `lxml.etree._Element`, or `None` when `node` is
                `None` or does not resolve to one
        """
        if node is None:
            return None
        element = getattr(node, 'obj', None)
        if not isinstance(element, etree._Element):
            parent = getattr(node, 'parent', None)
            element = getattr(parent, 'obj', None) if parent is not None else None
        if not isinstance(element, etree._Element):
            return None
        return element

    def _node_line_number(self, node: any) -> any:
        """Real line number of a violating node, read from the parsed tree's
        `sourceline` (lxml's binding to libxml2's `xmlGetLineNo`) instead of
        scanning the raw XML text for the attribute name. An attribute or
        text result has no `sourceline` of its own, so it is reported against
        its owning element's line, same as `_node_xpath_and_copy` does for
        the node's XPath/copy. Ref §3.12, category D1.

        Args:
            node (any): raw XPath node from `_select_with_nodes`'s `nodes`
                list, or `None` when no such node is available

        Returns:
            any: the 1-based line number (`int`), or `None` when it cannot be
                resolved (no backing node, or the node carries no line info)
        """
        element = self._resolve_owning_element(node)
        if element is None:
            return None
        return element.sourceline

    def _node_xpath_and_copy(self, node: any, deep_copy_nodes: bool = False) -> tuple:
        """Resolve a raw XPath node (from `_select_with_nodes`) into the two
        fields `s1kd-brexcheck`'s `dump_nodes_xml` attaches to every violation:
        the node's canonical XPath (port of `xpath_of`, `s1kd_tools.c:59-144`,
        using `elementpath`'s own equivalent node-path computation instead of
        re-walking the tree) and a copy of its owning element, serialised to an
        XML string. An attribute or text result is reported against its owning
        element (`if (node->type == XML_ATTRIBUTE_NODE) node = node->parent;` in
        the C original), since a bare attribute/text value has no subtree of its
        own to copy.

        Args:
            node (any): raw XPath node from `_select_with_nodes`'s `nodes` list,
                or `None` when the violation has no backing node (e.g. a flag-1
                "required but missing" violation, or a boolean-valued rule)
            deep_copy_nodes (bool): copy the full subtree (all descendants),
                equivalent to `-8`/`--deep-copy-nodes`. Defaults to a shallow
                copy of just the element's own tag and attributes, matching
                `xmlCopyNode(node, 2)` (properties only, no children).

        Returns:
            tuple: `(canonical_xpath, xml_snippet)`, both `None` when `node` is
                `None` or does not resolve to an lxml element
        """
        if node is None:
            return None, None

        try:
            canonical_xpath = node.extended_path
        except AttributeError:
            canonical_xpath = None

        element = self._resolve_owning_element(node)
        if element is None:
            return canonical_xpath, None

        try:
            if deep_copy_nodes:
                copy_elem = deepcopy(element)
            else:
                copy_elem = etree.Element(element.tag, nsmap=element.nsmap)
                for key, val in element.attrib.items():
                    copy_elem.set(key, val)
            xml_snippet = etree.tostring(copy_elem, encoding="unicode")
        except (TypeError, ValueError):
            xml_snippet = None

        return canonical_xpath, xml_snippet

    def _resolve_selector(self, brex_violations: dict, value: dict) -> any:
        """Return a rule's precompiled selector (built once in `_show_rules`
        and reused unchanged for every document, Ref §3.12), or record an
        `xpathError` and return `None` when it failed to compile.

        Args:
            brex_violations (dict): violations accumulator to record a
                compile failure against
            value (dict): rule dict from `_show_rules`, carrying `selector`/
                `selector_error`

        Returns:
            any: compiled `elementpath.Selector`, or `None` on failure
        """
        selector = value.get('selector')
        if selector is None:
            brex_violations[value["Brex"]]['xpathError'].append({
                'Description': value["objectUse"],
                'Xpath': value['xpath'],
                'Error': value.get('selector_error') or "Failed to compile objectPath",
                'BrDecisionIdentNumber': value.get('brDecisionIdentNumber')}
            )
        return selector

    def _check_object_flag_0(self, schema: str, brex_violations: dict, root: any, value: any,
                              deep_copy_nodes: bool = False):
        if value['contextRules'] == schema or value['contextRules'] == "":
            selector = self._resolve_selector(brex_violations, value)
            if selector is None:
                return brex_violations
            try:
                result, nodes = self._select_with_nodes(selector, root)
            except elementpath.ElementPathError as e:
                brex_violations[value["Brex"]]['xpathError'].append({
                    'Description': value["objectUse"],
                    'Xpath': value['xpath'],
                    'Error': str(e),
                    'BrDecisionIdentNumber': value.get('brDecisionIdentNumber')}
                )
                return brex_violations
            # A scalar result (boolean, or a bare number/string function call
            # with no comparison, e.g. `count(...)`/`string(...)`) has no
            # node of its own to report against, and is evaluated for
            # truthiness the same way XPath's effective boolean value would
            # (non-empty string, non-zero number, or the boolean itself) --
            # mirroring how `_check_object_flag_1` already treats these same
            # three scalar types. Only `bool` was handled here before: a
            # bare number result crashed on `len(result)` (`count(...)`
            # returning `0` or `1`), and a bare string result was silently
            # miscounted as a node-set, iterating over its characters.
            is_scalar = isinstance(result, (bool, str, int, float, Decimal))
            is_hit = bool(result) if is_scalar else len(result) > 0
            self._record_rule_hit(value, matched=is_hit, violated=is_hit)
            if is_scalar:
                if result:
                    line_label = ("(Boolean condition -> Interpret XPath)" if isinstance(result, bool)
                                  else "(Scalar condition -> Interpret XPath)")
                    brex_violations[value["Brex"]]['0'].append({
                        'Line': line_label,
                        'Description': value["objectUse"],
                        'Xpath': value['xpath'],
                        'NodeXpath': None,
                        'Object': None,
                        'RuleId': value.get('ruleId'),
                        'RulesContext': value['contextRules'],
                        'BrDecisionIdentNumber': value.get('brDecisionIdentNumber'),
                        'BrSeverityLevel': value.get('brSeverityLevel'),
                        'Fail': self._is_severity_failure(value.get('brSeverityLevel'))}
                    )
            else:
                for idx, element in enumerate(result):
                    node = nodes[idx] if nodes else None
                    line_no = self._node_line_number(node)
                    if line_no is None:
                        line_no = "x"
                    node_xpath, node_copy = self._node_xpath_and_copy(node, deep_copy_nodes)
                    brex_violations[value["Brex"]]['0'].append({
                        'Line': line_no,
                        'Description': value["objectUse"],
                        'Xpath': value['xpath'],
                        'NodeXpath': node_xpath,
                        'Object': node_copy,
                        'RuleId': value.get('ruleId'),
                        'RulesContext': value['contextRules'],
                        'BrDecisionIdentNumber': value.get('brDecisionIdentNumber'),
                        'BrSeverityLevel': value.get('brSeverityLevel'),
                        'Fail': self._is_severity_failure(value.get('brSeverityLevel'))}
                    )
        return brex_violations

    def _check_object_flag_1(self, schema: str, brex_violations: dict, root: any, value: any,
                              deep_copy_nodes: bool = False):
        if value['contextRules'] == schema or value['contextRules'] == "":
            selector = self._resolve_selector(brex_violations, value)
            if selector is None:
                return brex_violations
            try:
                result, nodes = self._select_with_nodes(selector, root)
            except elementpath.ElementPathError as e:
                brex_violations[value["Brex"]]['xpathError'].append({
                    'Description': value["objectUse"],
                    'Xpath': value['xpath'],
                    'Error': str(e),
                    'BrDecisionIdentNumber': value.get('brDecisionIdentNumber')}
                )
                return brex_violations
            # A scalar result (boolean, or a bare number/string function
            # call with no comparison, e.g. `count(...)`/`string(...)`) has
            # no node-set to measure, and is evaluated for truthiness the
            # same way XPath's effective boolean value would be. `Decimal`
            # belongs in the same tuple as `_check_object_flag_0`/`_2`:
            # `elementpath` returns `decimal.Decimal` -- not `float` -- for
            # any XPath arithmetic or decimal literal (`count(//b) div 4`,
            # `1.5`), which otherwise fell through to `len(result)` and
            # crashed with `TypeError: object of type 'decimal.Decimal' has
            # no len()`.
            is_scalar = isinstance(result, (bool, str, int, float, Decimal))
            violation = not result if is_scalar else len(result) == 0
            value_violations = []
            if violation:
                brex_violations[value["Brex"]]['1'].append({
                            'Description': value["objectUse"],
                            'Xpath': value['xpath'],
                            'NodeXpath': None,
                            'Object': None,
                            'RuleId': value.get('ruleId'),
                            'RulesContext': value['contextRules'],
                            'BrDecisionIdentNumber': value.get('brDecisionIdentNumber'),
                            'BrSeverityLevel': value.get('brSeverityLevel'),
                            'Fail': self._is_severity_failure(value.get('brSeverityLevel'))}
                            )
            elif not is_scalar and (
                    value["values_allowed"] or value["regex_allowed"] or value["ranges_allowed"]):
                value_violations = self._check_object_values(value, result, nodes, deep_copy_nodes)
                brex_violations[value["Brex"]]['2'].extend(value_violations)
            self._record_rule_hit(value, matched=not violation, violated=violation or bool(value_violations))
        return brex_violations

    def _check_object_values(self, value: any, elements: any, nodes: any = None,
                              deep_copy_nodes: bool = False) -> list:
        """Check a set of matched nodes against a rule's `objectValue` children.

        Shared by any flag whose matched nodes must additionally satisfy a value
        constraint (port of `check_objects_values`, `s1kd-brexcheck.c:275-304`,
        which applies value checking to a rule's matched node-set regardless of
        `allowedObjectFlag`). Ref §3.8.

        Args:
            value (any): rule dict from `_show_rules`, carrying `values_allowed`
                / `regex_allowed` / `ranges_allowed` / `value_tailoring`
            elements (any): node-set matched by `value['xpath']`
            nodes (any): raw XPath nodes aligned with `elements`, from
                `_select_with_nodes`, used to compute `NodeXpath`/`Object`
                (categories D2/D3); `None` when no such alignment is available
            deep_copy_nodes (bool): copy the full subtree instead of just the
                element's own tag and attributes, see `_node_xpath_and_copy`

        Returns:
            list: one violation dict per element whose value matches none of the
                allowed values, patterns or ranges
        """
        violations = []
        for idx, element in enumerate(elements):
            valid_elem = False
            if isinstance(element, etree._Element):
                # Full recursive text content, as `xmlNodeGetContent`
                # (`s1kd-brexcheck.c:232`) gives the C reference -- `.text`
                # alone stops at the first child element, so an element with
                # mixed content (`<title>Some <emphasis>bold</emphasis>
                # text</title>`) would be compared as just "Some " and
                # reported as violating a rule it satisfies. Same defect
                # §4.4 fixed for `objectUse` one level up.
                element_value = ''.join(element.itertext())
            else:
                element_value = element if isinstance(element, str) else str(element)
            if element_value not in value["values_allowed"]:
                if len(value["regex_allowed"]) > 0:
                    if any(bool(fullmatch(regex, element_value, V1)) for regex in value["regex_allowed"]):
                        valid_elem = True
                if not valid_elem and len(value["ranges_allowed"]) > 0:
                    if any(is_in_set(element_value, value_range) for value_range in value["ranges_allowed"]):
                        valid_elem = True
            else:
                valid_elem = True
            if not valid_elem:
                node = nodes[idx] if nodes else None
                line_no = self._node_line_number(node)
                if line_no is None:
                    line_no = "x"
                node_xpath, node_copy = self._node_xpath_and_copy(node, deep_copy_nodes)
                violations.append({
                    'Line': line_no,
                    'Description': f'Element/Attribute ({element_value}) did not match the object values.',
                    'Xpath': value['xpath'],
                    'NodeXpath': node_xpath,
                    'Object': node_copy,
                    'Single Values': [value["values_allowed"]],
                    'Pattern Values': [value["regex_allowed"]],
                    'Range Values': [value["ranges_allowed"]],
                    'ValueTailoring': value.get('value_tailoring', []),
                    'ObjectUse': value["objectUse"],
                    'RuleId': value.get('ruleId'),
                    'RulesContext': value['contextRules'],
                    'BrDecisionIdentNumber': value.get('brDecisionIdentNumber'),
                    'BrSeverityLevel': value.get('brSeverityLevel'),
                    'Fail': self._is_severity_failure(value.get('brSeverityLevel'))})
        return violations

    def _check_object_flag_2(self, schema: str, brex_violations: dict, root: any, value: any,
                              deep_copy_nodes: bool = False):
        if ('values_allowed' in value or 'regex_allowed' in value or 'ranges_allowed' in value) and (value['contextRules'] == schema or value['contextRules'] == ""):
            selector = self._resolve_selector(brex_violations, value)
            if selector is None:
                return brex_violations
            try:
                result, nodes = self._select_with_nodes(selector, root)
            except elementpath.ElementPathError as e:
                brex_violations[value["Brex"]]['xpathError'].append({
                    'Description': value["objectUse"],
                    'Xpath': value['xpath'],
                    'Error': str(e),
                    'BrDecisionIdentNumber': value.get('brDecisionIdentNumber')}
                )
                return brex_violations
            if type(result) is not bool:
                # A bare number/string function result (e.g. `count(...)`,
                # `string(...)` with no comparison) is itself the value to
                # check, not a node-set to iterate -- wrap it as the sole
                # item of a one-element list instead of passing a raw
                # scalar straight to `_check_object_values`, which
                # previously crashed on a number (not iterable at all) and
                # silently iterated a string character by character.
                is_scalar = isinstance(result, (str, int, float, Decimal))
                elements = [result] if is_scalar else result
                elements_nodes = [None] if is_scalar else nodes
                value_violations = self._check_object_values(value, elements, elements_nodes, deep_copy_nodes)
                brex_violations[value["Brex"]]['2'].extend(value_violations)
                self._record_rule_hit(value, matched=len(elements) > 0, violated=bool(value_violations))
            else:
                self._record_rule_hit(value, matched=bool(result), violated=False)
        return brex_violations

    def _rule_stat_key(self, value: dict) -> tuple:
        """Stable key identifying "the same rule" for `rule_hit_statistics`
        across every document checked: the BREX file it came from, its
        `rulesContext`/`context` qualifier, and its `objectPath`/`objpath`
        text. Two rules with identical text under the same qualifier in the
        same file are indistinguishable for statistics purposes, same as
        `lint_brex`'s duplicate-id/duplicate-decision-number checks already
        treat them as interchangeable.

        Args:
            value (dict): rule dict from `_show_rules`

        Returns:
            tuple: `(Brex, contextRules, xpath)`
        """
        return (value['Brex'], value['contextRules'], value['xpath'])

    def _record_rule_evaluated(self, value: dict) -> None:
        """Record that a content rule was in scope and considered for the
        document currently being checked (per-rule hit statistics, P2
        differentiator). Called once per rule per document from
        `_check_rules`, before dispatching to `_check_object_flag_0/1/2`,
        regardless of whether that dispatch goes on to find a match, a
        violation, or nothing at all -- including an `allowedObjectFlag="2"`
        rule with no `objectValue` children, which never reaches any check
        function (the same silent no-op `lint_brex`'s `EmptyValueFlag2`
        finding flags statically; here it shows up at runtime as a rule
        stuck at 0 matched / 0 violated across the whole data set).

        Args:
            value (dict): rule dict from `_show_rules`
        """
        key = self._rule_stat_key(value)
        stat = self._rule_stats.get(key)
        if stat is None:
            stat = {
                'Brex': value['Brex'],
                'ContextRules': value['contextRules'],
                'Xpath': value['xpath'],
                'ObjectFlag': value.get('ObjectFlag'),
                'ObjectUse': value.get('objectUse'),
                'BrDecisionIdentNumber': value.get('brDecisionIdentNumber'),
                'Evaluated': 0,
                'Matched': 0,
                'Violated': 0,
            }
            self._rule_stats[key] = stat
        stat['Evaluated'] += 1

    def _record_rule_hit(self, value: dict, matched: bool, violated: bool) -> None:
        """Record the outcome of actually checking a rule against the
        current document (per-rule hit statistics): whether its `objectPath`
        located something (`matched` -- a non-empty node-set or a true
        boolean result, independent of what `allowedObjectFlag` does with
        that outcome: a found forbidden node for flag `0`, a found required
        node for flag `1`, or a found value-constrained node for flag `2`)
        and whether that evaluation produced a violation (`violated`). Not
        called at all when the rule's selector raised (`xpathError`), since
        neither can be determined for that document; `_record_rule_evaluated`
        alone still counts the attempt.

        Args:
            value (dict): rule dict from `_show_rules`
            matched (bool): whether `value['xpath']` located something in
                this document
            violated (bool): whether this evaluation produced a violation
        """
        key = self._rule_stat_key(value)
        stat = self._rule_stats.get(key)
        if stat is None:
            self._record_rule_evaluated(value)
            stat = self._rule_stats[key]
        if matched:
            stat['Matched'] += 1
        if violated:
            stat['Violated'] += 1

    def reset_rule_statistics(self) -> None:
        """Clear the per-rule hit statistics accumulated so far (see
        `rule_hit_statistics`). Statistics accumulate across every call to
        `_check_rules` -- i.e. across an entire `validate()` directory-mode
        run, or across as many single-object `validate()` calls as the
        caller makes -- since the point of the feature is to see which rules
        never fire across a whole data set, not a single document. Call this
        to start a fresh data set.
        """
        self._rule_stats = {}

    def rule_hit_statistics(self) -> list:
        """Per-rule hit statistics accumulated across every check performed
        so far (P2 differentiator: "which BREX rules never fire against its
        data set"). Scoped to content rules (`structureObjectRule`/
        `objrule`); SNS and notation rules are checked as a single pass per
        document rather than rule-by-rule, so they have no per-rule
        breakdown here.

        Returns:
            list: one entry per distinct rule (see `_rule_stat_key`):
                `{"Brex", "ContextRules", "Xpath", "ObjectFlag", "ObjectUse",
                "BrDecisionIdentNumber", "Evaluated", "Matched", "Violated"}`.
                `Evaluated` counts every document the rule was in scope for;
                `Matched` counts documents where `Xpath` located something;
                `Violated` counts documents where that rule produced a
                violation. A rule with `Evaluated > 0` and `Matched == 0`
                never found its target anywhere in the data set checked so
                far; one with `Matched > 0` and `Violated == 0` always found
                a conforming target.
        """
        return [dict(stat) for stat in self._rule_stats.values()]

    def _get_sns_rules_group(self) -> any:
        """Merge the `snsRules` element from every active BREX into one root.

        Port of `check_brex_sns` (`s1kd-brexcheck.c:1144-1173`), which builds a
        combined `snsRulesGroup` document from the `snsRules` of every BREX
        passed to the tool so a code can be checked against rules defined in
        any of them.

        Returns:
            any: `snsRulesGroup` lxml element (childless if no BREX defines `snsRules`)
        """
        group = etree.Element("snsRulesGroup")
        for brex in self._brex_list[0]:
            with open(clean_path(brex), "r", encoding="utf-8") as _:
                brex_content = _.read()
            brex_content = delete_first_line(brex_content)
            brex_tree = self._parse_xml_text(brex_content)
            sns_rules = brex_tree.find(".//snsRules")
            if sns_rules is not None:
                group.append(deepcopy(sns_rules))
        return group

    def _sns_should_check(self, code: str, tag: str, ctx: any, sns_mode: str = "normal") -> bool:
        """Decide whether an SNS code level needs to be checked.

        Port of `should_check` (`s1kd-brexcheck.c:1038-1054`):

        - `strict`: always check every level; a placeholder code ("0" for
          sub/sub-sub-system, "00"/"0000" otherwise) is not treated as
          shorthand and must itself match a defined `snsCode`.
        - `unstrict`: check a level only if the current scope defines any
          rule for that level at all; if it defines none, any code (whether
          or not it looks like a placeholder) is accepted without checking.
        - `normal` (default): a non-placeholder code is always checked; a
          placeholder code is only checked if the current scope actually
          defines rules for that level.

        Args:
            code (str): the SNS code value from the data module's dmCode
            tag (str): the SNS rule element to look for (snsSystem/snsSubSystem/snsSubSubSystem/snsAssy)
            ctx (any): current scope to search within (an snsRulesGroup or a matched snsXxx node)
            sns_mode (str): one of `SNS_MODES` ("normal", "strict", "unstrict")

        Returns:
            bool: True if this level should be checked
        """
        if sns_mode == "strict":
            return True
        if sns_mode == "unstrict":
            return ctx.find(f".//{tag}") is not None
        if tag in ("snsSubSystem", "snsSubSubSystem"):
            non_placeholder = code != "0"
        else:
            non_placeholder = code not in ("00", "0000")
        return non_placeholder or ctx.find(f".//{tag}") is not None

    def _check_sns_rules(self, sns_rules_group: any, dmod_root: any, sns_mode: str = "normal") -> any:
        """Check a data module's SNS code against the merged SNS rules.

        Port of `check_brex_sns_rules` (`s1kd-brexcheck.c:1057-1144`): walks
        `systemCode` -> `subSystemCode` -> `subSubSystemCode` -> `assyCode` down
        `snsSystem` / `snsSubSystem` / `snsSubSubSystem` / `snsAssy`, stopping at
        the first failing level.

        Args:
            sns_rules_group (any): merged `snsRulesGroup` element from `_get_sns_rules_group`
            dmod_root (any): root element of the data module being checked
            sns_mode (str): one of `SNS_MODES` ("normal", "strict", "unstrict");
                see `_sns_should_check`

        Returns:
            any: dict describing the first failing level, or None if the SNS code is valid
                 (or the object being checked is not a data module, or has no `dmCode`)
        """
        if dmod_root.tag != "dmodule":
            return None

        dm_code = dmod_root.find(".//dmIdent/dmCode")
        if dm_code is None:
            return None

        system_code = dm_code.get("systemCode", "")
        sub_system_code = dm_code.get("subSystemCode", "")
        sub_sub_system_code = dm_code.get("subSubSystemCode", "")
        assy_code = dm_code.get("assyCode", "")

        levels = (
            ("systemCode", "snsSystem", system_code,
             system_code),
            ("subSystemCode", "snsSubSystem", sub_system_code,
             f"{system_code}-{sub_system_code}"),
            ("subSubSystemCode", "snsSubSubSystem", sub_sub_system_code,
             f"{system_code}-{sub_system_code}{sub_sub_system_code}"),
            ("assyCode", "snsAssy", assy_code,
             f"{system_code}-{sub_system_code}{sub_sub_system_code}-{assy_code}"),
        )

        ctx = sns_rules_group
        for code_name, tag, code, invalid_value in levels:
            if not self._sns_should_check(code, tag, ctx, sns_mode):
                continue
            match = ctx.xpath(f".//{tag}[snsCode=$code]", code=code)
            if not match:
                return {"code": code_name, "invalidValue": invalid_value}
            ctx = match[0]

        return None

    def _get_notation_rules_group(self) -> any:
        """Merge the `notationRuleList` element from every active BREX into one root.

        Port of the notation-rule loading step in `check_brex_notations`
        (`s1kd-brexcheck.c:1229-1256`), which builds a combined
        `notationRuleGroup` document from the `notationRuleList` of every
        BREX passed to the tool so an entity's notation can be checked
        against rules defined in any of them.

        Returns:
            any: `notationRuleGroup` lxml element (childless if no BREX defines `notationRuleList`)
        """
        group = etree.Element("notationRuleGroup")
        for brex in self._brex_list[0]:
            with open(clean_path(brex), "r", encoding="utf-8") as _:
                brex_content = _.read()
            brex_content = delete_first_line(brex_content)
            brex_tree = self._parse_xml_text(brex_content)
            notation_rule_list = brex_tree.find(".//notationRuleList")
            if notation_rule_list is not None:
                group.append(deepcopy(notation_rule_list))
        return group

    def _check_entity_notation(self, entity_name: str, notation_name: str, notation_rule_group: any) -> any:
        """Check a single unparsed (`NDATA`) entity's notation against the notation rules.

        Port of `check_entity` (`s1kd-brexcheck.c:1176-1201`). A notation is
        accepted if some `notationRule/notationName` names it with
        `@allowedNotationFlag != "0"`. Otherwise the entity is reported
        against the `objectUse` of the first `notationRule` in the merged
        rule group -- the C original's fallback XPath,
        `(//notationRule[notationName=X]|//notationRule)[1]`, unions a
        subset with its own superset, so it always resolves to the first
        `notationRule` in document order regardless of whether it actually
        names this notation.

        Args:
            entity_name (str): the `<!ENTITY>` name (for reporting only)
            notation_name (str): the NDATA notation the entity declares --
                for an unparsed entity, libxml2/lxml store the NDATA target
                name in the entity's content, which is what `entity->content`
                reads in the C original
            notation_rule_group (any): merged `notationRuleGroup` element
                from `_get_notation_rules_group`

        Returns:
            any: dict describing the violation, or None if the notation is allowed
        """
        allowed = notation_rule_group.xpath(
            ".//notationRule[notationName=$name and notationName/@allowedNotationFlag != '0']",
            name=notation_name,
        )
        if allowed:
            return None

        rules = notation_rule_group.xpath(".//notationRule")
        rule = rules[0] if rules else None
        object_use = None
        if rule is not None:
            use_node = rule.find("objectUse")
            if use_node is not None:
                object_use = "".join(use_node.itertext()) or None

        return {
            "Entity": entity_name,
            "Notation": notation_name,
            "Description": object_use or f"Notation '{notation_name}' is not allowed.",
        }

    def _check_notation_rules(self, notation_rule_group: any, dmod_tree: any) -> list:
        """Check every unparsed entity declared in the object's internal DTD subset.

        Port of `check_brex_notation_rules` (`s1kd-brexcheck.c:1204-1227`):
        walks the `ENTITY` declarations of the internal DTD subset, and for
        each external, unparsed (`NDATA`) entity checks its notation via
        `_check_entity_notation`. Objects with no internal DTD subset (the
        common case for XSD-validated S1000D 4.x+ content) are not checked,
        matching the original's `if (!(dtd = dmod_doc->intSubset)) return 0;`.

        Args:
            notation_rule_group (any): merged `notationRuleGroup` element
                from `_get_notation_rules_group`
            dmod_tree (any): parsed `ElementTree` of the object being checked

        Returns:
            list: violation records, one per entity naming a disallowed notation
        """
        internal_dtd = dmod_tree.docinfo.internalDTD
        if internal_dtd is None:
            return []

        violations = []
        for entity in internal_dtd.iterentities():
            # Unparsed (NDATA) entities are the only ones with both a system
            # identifier and content set (content holds the NDATA notation
            # name); internal entities have no system_url, external parsed
            # entities have no content -- this mirrors the C original's
            # etype == XML_EXTERNAL_GENERAL_UNPARSED_ENTITY check.
            if entity.system_url is None or entity.content is None:
                continue
            violation = self._check_entity_notation(entity.name, entity.content, notation_rule_group)
            if violation is not None:
                violations.append(violation)
        return violations

    def _get_non_context_rules(self) -> list:
        """Collect `nonContextRules/nonContextRule` entries (category A4) from
        every BREX in the active chain.

        These are the human-readable business rules a BREX carries with no
        machine-checkable `objectPath`/`objectValue` form -- S1000D's own escape
        hatch for BRs that can't be expressed as a rule (e.g. "deletion of data
        modules is treated as a special case of update"). Neither
        `s1kd-brexcheck` nor our own checker read them before; they are not
        violations and are never checked, only surfaced, so authors don't lose
        track of a whole rule class that was previously dropped silently.

        Returns:
            list: one entry per `nonContextRule`, in BREX-chain then document
                order: `{"Brex": path, "Text": str or None,
                "BrDecisionIdentNumber": str or None}`
        """
        entries = []
        for brex in self._brex_list[0]:
            with open(clean_path(brex), "r", encoding="utf-8") as _:
                brex_content = _.read()
            brex_content = delete_first_line(brex_content)
            brex_tree = self._parse_xml_text(brex_content)
            for rule in brex_tree.findall(".//nonContextRules/nonContextRule"):
                br_decision_ref = rule.find("brDecisionRef")
                br_decision_ident_number = (
                    br_decision_ref.get("brDecisionIdentNumber") if br_decision_ref is not None else None
                )
                text_source = rule.find("simplePara")
                if text_source is None:
                    text_source = rule
                text = "".join(text_source.itertext()).strip() or None
                entries.append({
                    "Brex": brex,
                    "Text": text,
                    "BrDecisionIdentNumber": br_decision_ident_number,
                })
        return entries

    def _remove_deleted_elements(self, node: any) -> None:
        """Recursively drop elements marked as deleted from a parsed tree.

        Port of `rem_delete_nodes`/`rem_delete_elems` (`s1kd_tools.c:1054-1088`),
        `s1kd-brexcheck`'s `-^`/`--remove-deleted` option: an element carrying
        `@changeType="delete"` (or the legacy `@change="delete"` spelling the C
        original also checks) is removed along with its whole subtree before any
        rule is checked, so content staged for deletion in a change-marked
        revision does not trigger BREX violations. Children are only visited when
        the element itself is kept, matching the C original.

        Args:
            node (any): element to inspect, e.g. the checked document's root
        """
        change = node.get('change', node.get('changeType'))
        if change == 'delete':
            parent = node.getparent()
            if parent is not None:
                parent.remove(node)
            return
        for child in list(node):
            self._remove_deleted_elements(child)

    def _check_rules(self, debug: bool = False, progress_callback: Callable[[int, int, str], None] = None,
                      sns_mode: str = "normal", remove_deleted: bool = False,
                      deep_copy_nodes: bool = False, check_sns: bool = False,
                      check_notations: bool = False) -> dict:
        """Traverses through every node of the brex and checks the rules through the given xpaths.
        For objectFlag 0 we also get the line of the error
        For objectFlag 1 we only get the Description of the rule that was violated
        For objectFlag 2 we get a list containing all 'single' values and a list containing all 'pattern' values
                         and we might get the line of the error

        Args:
            debug (bool): dump intermediate rule/error data for inspection
            progress_callback (Callable[[int, int, str], None]): optional, called as
                `progress_callback(current, total, "rules")` after each content rule
                is checked -- e.g. `lambda current, total, stage: bar.update(1)` if the
                caller wants a `tqdm` bar of their own. `None` (the default) reports no
                progress; the library itself no longer depends on `tqdm`. See `validate`,
                whose directory-mode file loop reports the same way with `stage="files"`.
            sns_mode (str): one of `SNS_MODES` ("normal", "strict", "unstrict");
                see `_sns_should_check`
            remove_deleted (bool): equivalent to `s1kd-brexcheck -^`/`--remove-deleted`;
                drop elements marked `@changeType="delete"` (see `_remove_deleted_elements`)
                before every check (content rules, SNS, notations)
            deep_copy_nodes (bool): equivalent to `-8`/`--deep-copy-nodes`; the `Object`
                field of every content-rule violation record holds a full recursive copy
                of the violating element instead of just its own tag and attributes
                (see `_node_xpath_and_copy`)
            check_sns (bool): equivalent to `s1kd-brexcheck -S`/`--sns`; opt in to
                checking the object's SNS against the BREX `snsRules` (see
                `_check_sns_rules`). Off by default, matching `opts->check_sns` in
                the C original. When off, the result has no `'sns'` key at all.
            check_notations (bool): equivalent to `s1kd-brexcheck -n`/`--notations`;
                opt in to checking the object's unparsed entity notations against the
                BREX `notationRuleList` (see `_check_notation_rules`). Off by default,
                matching `opts->check_notations` in the C original. When off, the
                result has no `'notations'` key at all.

        Raises:
            NoSchemaDeclared: if `set_require_schema(True)` is active and the checked
                object has no `xsi:noNamespaceSchemaLocation`. Off by default -- see
                `set_require_schema`.

        Returns:
            any: Dictionary with all errors
        """
        schema = get_schema_from_xml(self._xml_content)
        if schema is None and self._require_schema_declaration:
            raise NoSchemaDeclared(
                f"{self._xml_path} has no xsi:noNamespaceSchemaLocation on its root "
                "element, so schema-qualified BREX rules (rulesContext) cannot be "
                "selected for it -- only unqualified rules would be checked. If this "
                "is a genuine S1000D <= 3.0 / DTD-based object, call "
                "set_require_schema(False) (the default) to check it against "
                "unqualified rules only, as intended."
            )
        brex_violations_dict = {}
        for brex in self._brex_list[0]:
            brex_violations_dict[brex] = {
                '0': [],
                '1': [],
                '2': [],
                'xpathError': []
            }
        brex_violations_dict["brexFallback"] = list(self._brex_fallbacks)
        root = self._parse_xml_file(self._xml_path)

        if remove_deleted:
            self._remove_deleted_elements(root.getroot())

        dmod_root = root.getroot()
        # Both checks are opt-in, matching `opts->check_sns` / `opts->check_notations`
        # in `s1kd-brexcheck.c` (both default to `false` -- `:1994`, `:2084`, `:2091`
        # -- and the calls are gated at `:1376` and `:1383`).
        if check_sns and dmod_root.tag == "dmodule":
            sns_rules_group = self._get_sns_rules_group()
            sns_error = self._check_sns_rules(sns_rules_group, dmod_root, sns_mode)
            brex_violations_dict["sns"] = [] if sns_error is None else [{
                "code": sns_error["code"],
                "invalidValue": sns_error["invalidValue"],
                "Description": f"{sns_error['code']} is not valid according to the SNS rules.",
            }]

        if check_notations:
            notation_rule_group = self._get_notation_rules_group()
            brex_violations_dict["notations"] = self._check_notation_rules(notation_rule_group, root)

        brex_violations_dict["nonContextRules"] = self._get_non_context_rules()

        all_content_rules = []
        for brex in self._brex_list[0]:
            content_rules = self._get_content_rules(brex, schema=schema, debug=debug)
            all_content_rules += content_rules

        if debug:
            with open(clean_path(join(expanduser("~/Desktop"), "All_content_rules.txt")), 'w', encoding="utf-8") as _:
                for rule in all_content_rules:
                    _.write(str(rule) + "\n")
        total_content_rules = len(all_content_rules)
        for idx, value in enumerate(all_content_rules):
            self._record_rule_evaluated(value)
            if value["ObjectFlag"] == '0':
                brex_violations_dict |= self._check_object_flag_0(
                    schema, brex_violations_dict, root, value, deep_copy_nodes)
            if value["ObjectFlag"] == '1':
                brex_violations_dict |= self._check_object_flag_1(
                    schema, brex_violations_dict, root, value, deep_copy_nodes)
            has_values = value["values_allowed"] != [] or value["regex_allowed"] != [] or value["ranges_allowed"] != []
            # S1000D <= 3.0 rules commonly omit @objappl entirely for a
            # value-only constraint (no presence/absence semantics); s1kd's
            # is_invalid falls through to the value check in that case too.
            if has_values and value["ObjectFlag"] in ('2', None):
                brex_violations_dict |= self._check_object_flag_2(
                    schema, brex_violations_dict, root, value, deep_copy_nodes)
            if progress_callback is not None:
                progress_callback(idx + 1, total_content_rules, "rules")
        self._deduplicate_violations(brex_violations_dict)
        return brex_violations_dict

    def _deduplicate_violations(self, brex_violations_dict: dict) -> None:
        """Mark, but do not remove, a content-rule violation ('0'/'1'/'2') as a
        duplicate when an earlier, more specific layer of the same layered-BREX
        chain (`self._brex_list[0]`, walked project-specific-layer first -- see
        `_walk_brex_chain`/`lint_brex_layers`) already reported the identical
        violation.

        Layered BREX commonly re-states or inherits the same rule verbatim
        across layers (a project BREX re-declaring a rule the master BREX
        already defines, or a project BREX simply not overriding one it
        inherits): checking the same object against both layers then produces
        two violation records for what is really one real-world defect, one
        under each layer's own `Brex` key in the result. `lint_brex_layers`
        already treats a rule as "the same" across layers when it shares its
        `rulesContext`/`context` qualifier and exact `objectPath`/`objpath`
        text (a documented simplification); a violation additionally needs to
        land on the same node (or, for a flag-1 "required but missing"/boolean
        flag-0 result with no backing node, `NodeXpath` is `None` for both,
        which still correctly identifies them as the same violation) to count
        as the same violation rather than merely the same rule.

        Marks in place via a `'Duplicate'` key on the violation dict, kept in
        the raw per-BREX lists this method's caller (`_check_rules`) returns
        (`result[brex_path]['0'/'1'/'2']`) so nothing disappears from that
        structure or changes its length -- only `_count_violations` (and, in
        turn, `_append_summary`/`run_summary`) and `to_xml_report`/
        `to_json_report` skip a `Duplicate: True` entry.

        A violation identity that recurs within a single BREX file (rather
        than across two files in the chain) is deduplicated the same way:
        the identity key below intentionally does not include `Brex`, since
        two rules sharing `rulesContext`+`objectPath`+the same violating node
        are the same real-world defect regardless of whether they came from
        one file or two.

        Args:
            brex_violations_dict (dict): the in-progress result being built by
                `_check_rules`, mutated in place
        """
        seen = set()
        for brex in self._brex_list[0]:
            brex_result = brex_violations_dict.get(brex)
            if not isinstance(brex_result, dict):
                continue
            for flag in ('0', '1', '2'):
                for violation in brex_result.get(flag, []):
                    key = (flag, violation.get('RulesContext'), violation.get('Xpath'), violation.get('NodeXpath'))
                    if key in seen:
                        violation['Duplicate'] = True
                    else:
                        seen.add(key)
                        violation['Duplicate'] = False

    def _count_violations(self, object_flag_dict: dict) -> tuple:
        """Counts the number of actual Brex violations (flags 0, 1 and 2, plus SNS and
        notation rules) for a xml, and tallies content-rule violations by their
        resolved `brSeverityLevel`. `xpathError` entries are diagnostics about a rule
        that could not be evaluated, not violations, and are excluded from the count.
        A content-rule violation marked `'Duplicate': True` (the same violation
        already reported against an earlier, more specific layer of the same
        layered-BREX chain -- see `_deduplicate_violations`) is also excluded.

        A content-rule violation whose resolved `brSeverityLevel` is marked `fail="no"`
        in the `.brseveritylevels` file (see `_is_severity_failure`) is reported as a
        warning instead of an error and does not count towards the failing total. SNS
        and notation-rule violations have no associated severity level and always
        count as errors; they are tallied under the `None` severity key alongside
        content-rule violations that carry no `brSeverityLevel` at all, since neither
        case resolved to one.

        Shared by `_append_summary` (per-document "N Errors, M Warnings" string) and
        `run_summary` (per-run totals, category D7), so both report the same counts.
        Accepts `object_flag_dict` either before or after `validate()` adds its
        `Summary` key (and, for a skipped object, `Skipped`) -- both are ignored
        here alongside `brexFallback` and `nonContextRules`, since none of the
        three is a violation list.

        Args:
            object_flag_dict (dict): mapping of brex path to its '0'/'1'/'2'/'xpathError' violation
                lists, plus optional 'sns' / 'notations' keys holding SNS and notation violations,
                a 'brexFallback' key listing any built-in BREX substitutions, and a
                'nonContextRules' key listing informational category-A4 entries (see
                `_get_non_context_rules`) -- none of these three is itself a violation and
                none counts towards the total

        Returns:
            tuple: (error_count: int, warning_count: int, severity_counts: dict) where
                `severity_counts` maps a `BrSeverityLevel` value (or `None`) to the number
                of violations resolved at that level, regardless of `Fail`/`Warning` status
        """
        error_count = 0
        warning_count = 0
        severity_counts = {}
        for key, brex_result in object_flag_dict.items():
            if key in ("brexFallback", "Summary", "Skipped", "nonContextRules"):
                continue
            if key in ("sns", "notations"):
                error_count += len(brex_result)
                if brex_result:
                    severity_counts[None] = severity_counts.get(None, 0) + len(brex_result)
                continue
            for flag in ('0', '1', '2'):
                for violation in brex_result[flag]:
                    if violation.get('Duplicate'):
                        # Same violation already counted against an earlier,
                        # more specific layer of the same BREX chain -- see
                        # `_deduplicate_violations`.
                        continue
                    severity = violation.get('BrSeverityLevel')
                    severity_counts[severity] = severity_counts.get(severity, 0) + 1
                    if violation.get('Fail', True):
                        error_count += 1
                    else:
                        warning_count += 1
        return error_count, warning_count, severity_counts

    def _append_summary(self, object_flag_dict: dict) -> str:
        """Counts the number of actual Brex violations (flags 0, 1 and 2, plus SNS and
        notation rules) for a xml. `xpathError` entries are diagnostics about a rule
        that could not be evaluated, not violations, and are excluded from the count.

        A content-rule violation whose resolved `brSeverityLevel` is marked `fail="no"`
        in the `.brseveritylevels` file (see `_is_severity_failure`) is reported as a
        warning instead of an error and does not count towards the failing total. SNS
        and notation-rule violations have no associated severity level and always
        count as errors.

        Args:
            object_flag_dict (dict): mapping of brex path to its '0'/'1'/'2'/'xpathError' violation
                lists, plus optional 'sns' / 'notations' keys holding SNS and notation violations,
                a 'brexFallback' key listing any built-in BREX substitutions, and a
                'nonContextRules' key listing informational category-A4 entries (none of
                these three is itself a violation and none counts towards the total)

        Returns:
            str: human-readable violation count, e.g. "3 Errors" or "3 Errors, 1 Warnings"
        """
        error_count, warning_count, _ = self._count_violations(object_flag_dict)
        if warning_count:
            return f"{error_count} Errors, {warning_count} Warnings"
        return f"{error_count} Errors"

    def _is_single_object_result(self, result: dict) -> bool:
        """Distinguish a single-object `validate()` result from a directory-mode
        result (a mapping of `{filename: single-object result}`).

        A single-object result always carries a top-level `Summary` string
        (see `validate`/`_append_summary`); a directory-mode result does not,
        since `Summary` only appears nested inside each per-file result.

        Args:
            result (dict): a `validate()` return value

        Returns:
            bool: True if `result` is a single-object result
        """
        return isinstance(result.get("Summary"), str)

    def run_summary(self, result: dict) -> dict:
        """Build a per-run summary (category D7): how many checked documents passed
        or failed, and how many violations were recorded at each business-rule
        severity level, port of `s1kd-brexcheck -T`/`--totals`.

        Unlike `_append_summary` (a per-document string folded into `validate()`
        itself), this is a separate, opt-in conversion over a finished `validate()`
        result -- the same pattern `to_xml_report` uses -- so existing callers of
        `validate()` see no change in its return shape; `to_xml_report` also calls
        this to embed the same totals as a `<summary>` node.

        Accepts either a single-object result (`validate()` after `set_xml`) or a
        directory-mode result (`validate()` after `set_xml_dir`, a mapping of
        `{filename: single-object result}`), distinguished the same way
        `_is_single_object_result` does.

        A document "passes" when it was checked and its error count (as computed
        by `_count_violations`, i.e. severity `fail="no"` violations counted as
        warnings, not errors) is zero; warnings alone do not fail a document. A
        skipped document (`set_ignore_empty`, category "Skipped") is counted
        separately and excluded from `DocumentsChecked`/passed/failed, since it was
        never actually checked. Note directory mode cannot report every skipped
        document: `validate()` drops a skipped file from its results mapping
        entirely rather than recording a `Skipped` marker for it (see `validate`),
        so `DocumentsSkipped` there only ever reflects what a single-object result
        can carry -- 0 in practice, since directory mode is the only mode that
        actually skips files today.

        Args:
            result (dict): a `validate()` return value

        Returns:
            dict: {
                "DocumentsChecked": int,   # passed + failed, excludes skipped
                "DocumentsPassed": int,
                "DocumentsFailed": int,
                "DocumentsSkipped": int,
                "Errors": int,
                "Warnings": int,
                "ViolationsBySeverity": dict,  # BrSeverityLevel value (or None) -> count
            }
        """
        if self._is_single_object_result(result):
            documents = [result]
        else:
            documents = [doc for doc in result.values() if isinstance(doc, dict)]

        documents_checked = 0
        documents_passed = 0
        documents_failed = 0
        documents_skipped = 0
        total_errors = 0
        total_warnings = 0
        severity_counts = {}

        for doc_result in documents:
            if doc_result.get("Skipped"):
                documents_skipped += 1
                continue
            error_count, warning_count, doc_severity_counts = self._count_violations(doc_result)
            documents_checked += 1
            total_errors += error_count
            total_warnings += warning_count
            for severity, count in doc_severity_counts.items():
                severity_counts[severity] = severity_counts.get(severity, 0) + count
            if error_count:
                documents_failed += 1
            else:
                documents_passed += 1

        return {
            "DocumentsChecked": documents_checked,
            "DocumentsPassed": documents_passed,
            "DocumentsFailed": documents_failed,
            "DocumentsSkipped": documents_skipped,
            "Errors": total_errors,
            "Warnings": total_warnings,
            "ViolationsBySeverity": severity_counts,
        }

    def _violation_from_dict(self, document: str, brex: str, flag: str, violation: dict) -> BrexViolation:
        """Convert one raw violation record (a `result[brex_path][flag]` list
        entry, as built by `_check_object_flag_0/1/2`/`_check_object_values`)
        into a structured `BrexViolation`. The single conversion point both
        `violations()` and `to_xml_report` (via `_append_error_node`) go
        through, so the JSON and XML reports are both derived from the same
        structured shape rather than each walking the raw dict independently.

        Args:
            document (str): path of the checked object the violation is from
            brex (str): path of the BREX file the violated rule came from
            flag (str): `'0'`, `'1'` or `'2'` -- which violation list `violation`
                came from
            violation (dict): one raw violation record

        Returns:
            BrexViolation: the structured equivalent
        """
        return BrexViolation(
            document=document,
            brex=brex,
            rule_id=violation.get('RuleId'),
            br_decision_ident_number=violation.get('BrDecisionIdentNumber'),
            flag=flag,
            rules_context=violation.get('RulesContext', ''),
            severity=violation.get('BrSeverityLevel'),
            fail=violation.get('Fail', True),
            object_path=violation.get('Xpath'),
            # Flag 0/1 records carry the rule's objectUse only under
            # 'Description'; a value violation (flag 2, or a flag-1 rule's
            # follow-on value check, §3.8) carries the rule's objectUse
            # separately under 'ObjectUse' and repurposes 'Description' for
            # its own "did not match the object values" message -- prefer
            # 'ObjectUse' so this always ends up being the rule's own
            # description, matching `BrexViolation.object_use`.
            object_use=violation.get('ObjectUse', violation.get('Description')),
            allowed_values={
                'single': (violation.get('Single Values') or [[]])[0],
                'pattern': (violation.get('Pattern Values') or [[]])[0],
                'range': (violation.get('Range Values') or [[]])[0],
                'tailoring': violation.get('ValueTailoring', []),
            },
            node_xpath=violation.get('NodeXpath'),
            line=violation.get('Line'),
            node_snippet=violation.get('Object'),
            duplicate=bool(violation.get('Duplicate')),
        )

    def _result_documents(self, result: dict) -> dict:
        """Normalise either `validate()` result shape into a
        `{document path: single-object result}` mapping, so everything that
        walks a result per document (`violations`,
        `_json_document_sections`) names its documents the same way
        `to_xml_report` does when it builds one `document` node per object.

        Args:
            result (dict): a `validate()` return value, single-object
                (after `set_xml`) or directory-mode (after `set_xml_dir`),
                distinguished the same way `_is_single_object_result` does

        Returns:
            dict: one entry keyed by `self._xml_path` for a single-object
                result, one per filename key for a directory-mode result
        """
        if self._is_single_object_result(result):
            return {(self._xml_path or ""): result}
        return {name: doc for name, doc in result.items() if isinstance(doc, dict)}

    def violations(self, result: dict) -> list:
        """Flatten a `validate()`/`_check_rules()` result into a list of
        structured `BrexViolation` records -- the canonical, typed form
        `to_json_report` and `to_xml_report` both derive their output from
        (via `_violation_from_dict`), instead of each re-walking the raw
        nested `result[brex_path]['0'/'1'/'2']` dicts independently.

        Only content-rule violations (`allowedObjectFlag`/`@objappl` `'0'`,
        `'1'`, `'2'`) are represented: SNS (category A2) and notation-rule
        (A3) violations have no `objectPath`/`objectUse`/allowed-values shape
        to report through `BrexViolation`, and `nonContextRules`/
        `brexFallback` are informational, not violations -- all three remain
        available in `result` itself and in `to_xml_report`'s XML.

        Accepts either a single-object result (`validate()` after `set_xml`)
        or a directory-mode result (`validate()` after `set_xml_dir`, a
        mapping of `{filename: single-object result}`), distinguished the
        same way `_is_single_object_result` does. A skipped document
        (`set_ignore_empty`, category "Skipped") contributes nothing, since
        it was never actually checked.

        Args:
            result (dict): a `validate()` return value

        Returns:
            list: `BrexViolation` records, in document, then BREX, then flag
                ('0'/'1'/'2') order -- including a `duplicate=True` entry
                (see `_deduplicate_violations`); filter those out explicitly
                if only the reportable set is wanted, the way
                `to_json_report`/`to_xml_report` do
        """
        return [violation for violation, _raw in self._violation_records(result)]

    def _json_document_sections(self, result: dict) -> dict:
        """Flatten everything a report carries that `violations()` deliberately
        leaves out -- SNS (category A2) and notation (A3) violations, plus the
        informational `nonContextRules` (A4) and `brexFallback` entries --
        into flat lists for `to_json_report`, the JSON counterpart of the
        `sns`/`notations`/`nonContextRules` nodes `_append_sns_notation_nodes`
        and `_append_non_context_rules_node` put inside each XML `document`
        node.

        Without these, `to_json_report` contradicted itself on any object
        whose only findings were SNS or notation ones: its `summary` counted
        them (`_count_violations` does), while its `violations` list -- which
        by design only holds content-rule violations, the only kind
        `BrexViolation`'s `objectPath`/`objectUse`/allowed-values shape fits
        -- came back empty, with no way for a consumer to learn what had been
        counted.

        Flat lists, each entry tagged with its own `document` path, rather
        than a per-document nesting: the report's `violations` list is flat
        the same way, with `BrexViolation.document` identifying the object.
        Key names mirror the XML report's element names (`invalidValue`,
        `invalidNotation`, `objectUse`), so a consumer of either report reads
        the same vocabulary. A document that was checked and passed simply
        contributes no entry -- the JSON equivalent of the XML report's
        `<noErrors/>` child -- as does a skipped one (`set_ignore_empty`),
        which was never checked at all, matching `violations()`.

        `brexFallback` has no counterpart in `to_xml_report`'s
        `-x`-compatible shape; it is carried by the `validate()` result and,
        now, this report.

        Args:
            result (dict): a `validate()` return value

        Returns:
            dict: `{"sns": [...], "notations": [...], "nonContextRules": [...],
                "brexFallback": [...]}`, ready to merge into the report payload
        """
        sns_entries = []
        notation_entries = []
        non_context_entries = []
        fallback_entries = []
        for docname, doc_result in self._result_documents(result).items():
            if doc_result.get("Skipped"):
                continue
            for sns_error in doc_result.get("sns") or []:
                sns_entries.append({
                    "document": docname,
                    "code": sns_error.get("code"),
                    "invalidValue": sns_error.get("invalidValue"),
                    "objectUse": sns_error.get("Description"),
                })
            for notation_error in doc_result.get("notations") or []:
                notation_entries.append({
                    "document": docname,
                    "entity": notation_error.get("Entity"),
                    "invalidNotation": notation_error.get("Notation"),
                    "objectUse": notation_error.get("Description"),
                })
            for non_context_rule in doc_result.get("nonContextRules") or []:
                non_context_entries.append({
                    "document": docname,
                    "brex": non_context_rule.get("Brex"),
                    "brDecisionIdentNumber": non_context_rule.get("BrDecisionIdentNumber"),
                    "text": non_context_rule.get("Text"),
                })
            for fallback in doc_result.get("brexFallback") or []:
                fallback_entries.append({
                    "document": docname,
                    "reference": fallback.get("Reference"),
                    "usedBuiltinBrex": fallback.get("UsedBuiltinBrex"),
                    "builtinBrexPath": fallback.get("BuiltinBrexPath"),
                })
        return {
            "sns": sns_entries,
            "notations": notation_entries,
            "nonContextRules": non_context_entries,
            "brexFallback": fallback_entries,
        }

    def to_json_report(self, result: dict, indent: int = 2) -> str:
        """Convert a `validate()` result into a JSON report derived from the
        structured `BrexViolation` list (see `violations`), rather than
        serialising the raw `result` dict verbatim -- the ad-hoc JSON shape
        category D6 originally flagged as the only report format available.
        `to_xml_report` is the equivalent for the `-x`-compatible XML shape;
        both are now derived from the same `_violation_from_dict` conversion.

        A `duplicate=True` violation (see `_deduplicate_violations`) is left
        out, matching `to_xml_report` and `_count_violations`/`run_summary`.

        Alongside `violations` (content-rule violations only, the only kind
        `BrexViolation` can represent) the report carries the same SNS,
        notation and `nonContextRules` content `to_xml_report` emits per
        `document` node, plus `brexFallback` -- see `_json_document_sections`.
        Every count in `summary` is therefore backed by something the report
        itself lists.

        Args:
            result (dict): a `validate()` return value
            indent (int): `json.dumps` indent; `None` for compact output

        Returns:
            str: `{"summary": run_summary(result), "violations": [...],
                "sns": [...], "notations": [...], "nonContextRules": [...],
                "brexFallback": [...]}` -- each `violations` entry the
                `dataclasses.asdict()` form of one `BrexViolation`, the rest
                flat `document`-tagged dicts
        """
        payload = {
            "summary": self.run_summary(result),
            "violations": [v.to_dict() for v in self.violations(result) if not v.duplicate],
        }
        payload.update(self._json_document_sections(result))
        return dumps(payload, indent=indent, ensure_ascii=False)

    def _append_sns_notation_nodes(self, document_node: any, result: dict) -> None:
        """Append the `sns` and `notations` nodes of one `document` node, port of
        the corresponding fragments of `check_brex_sns_rules`
        (`s1kd-brexcheck.c:1077-1134`) and `check_brex_notation_rules`
        (`s1kd-brexcheck.c:1213-1224`): an empty `<noErrors/>` child when the
        object was checked and passed, one `<error>` per violation otherwise,
        and no `sns`/`notations` node at all when that check did not run
        (`result` has no `'sns'`/`'notations'` key -- not a data module, or no
        BREX in the chain defines the relevant rules).

        Args:
            document_node (any): the `document` element to append to
            result (dict): single-object `validate()` result
        """
        sns_violations = result.get("sns")
        if sns_violations is not None:
            sns_node = etree.SubElement(document_node, "sns")
            if sns_violations:
                for sns_error in sns_violations:
                    error_node = etree.SubElement(sns_node, "error")
                    etree.SubElement(error_node, "code").text = sns_error.get("code")
                    etree.SubElement(error_node, "invalidValue").text = sns_error.get("invalidValue")
            else:
                etree.SubElement(sns_node, "noErrors")

        notation_violations = result.get("notations")
        if notation_violations is not None:
            notations_node = etree.SubElement(document_node, "notations")
            if notation_violations:
                for notation_error in notation_violations:
                    error_node = etree.SubElement(notations_node, "error")
                    etree.SubElement(error_node, "invalidNotation").text = notation_error.get("Notation")
                    etree.SubElement(error_node, "objectUse").text = notation_error.get("Description")
            else:
                etree.SubElement(notations_node, "noErrors")

    def _append_non_context_rules_node(self, document_node: any, result: dict) -> None:
        """Append the `nonContextRules` node (category A4): one `<nonContextRule>`
        per human-readable, non-machine-checkable business rule collected from
        every BREX in the chain (see `_get_non_context_rules`).

        Unlike `sns`/`notations`, these are informational only -- not a check
        that passed or failed -- so there is no `<noErrors/>` counterpart; the
        node is simply omitted when the chain defines no `nonContextRule` at all
        (or `result` has no `'nonContextRules'` key, e.g. a skipped object).

        Args:
            document_node (any): the `document` element to append to
            result (dict): single-object `validate()` result
        """
        entries = result.get("nonContextRules")
        if not entries:
            return
        container = etree.SubElement(document_node, "nonContextRules")
        for entry in entries:
            rule_node = etree.SubElement(container, "nonContextRule")
            br_decision_ident_number = entry.get("BrDecisionIdentNumber")
            if br_decision_ident_number is not None:
                etree.SubElement(rule_node, "brDecisionRef", brDecisionIdentNumber=br_decision_ident_number)
            etree.SubElement(rule_node, "text").text = entry.get("Text")

    def _append_error_node(self, brex_node: any, violation: BrexViolation) -> None:
        """Append one `error` node for a content-rule violation, port of the
        `<error>` construction in `check_brex_rules` (`s1kd-brexcheck.c:900-938`).

        Unlike the C original, where one `<error>` can hold several `<object>`
        children (every node matched by the same rule), each of our violation
        records already corresponds to a single matched node (see
        `_check_object_flag_0`/`_check_object_values`), so this emits at most
        one `<object>` child per `<error>`.

        Args:
            brex_node (any): the `brex` element to append to
            violation (BrexViolation): the structured violation to render,
                built by `_violation_from_dict`
        """
        error_node = etree.SubElement(brex_node, "error")

        if violation.severity:
            error_node.set("brSeverityLevel", violation.severity)
            if not violation.fail:
                error_node.set("fail", "no")
        else:
            error_node.set("fail", "yes")

        if violation.br_decision_ident_number is not None:
            etree.SubElement(error_node, "brDecisionRef",
                              brDecisionIdentNumber=violation.br_decision_ident_number)

        object_path_node = etree.SubElement(error_node, "objectPath", allowedObjectFlag=violation.flag)
        object_path_node.text = violation.object_path

        etree.SubElement(error_node, "objectUse").text = violation.object_use

        if violation.node_snippet is not None:
            object_node = etree.SubElement(error_node, "object")
            if violation.line is not None:
                object_node.set("line", str(violation.line))
            if violation.node_xpath is not None:
                object_node.set("xpath", violation.node_xpath)
            object_node.append(etree.fromstring(violation.node_snippet.encode("utf-8")))

    def _build_document_node(self, result: dict, docname: str) -> any:
        """Build one `document` node (a single checked object's report), port of
        the `documentNode` built in `check_brex`/`check_brex_rules`
        (`s1kd-brexcheck.c:1373-1374`, `827-965`).

        Args:
            result (dict): single-object `validate()` result
            docname (str): path of the checked object, reported as `document/@path`

        Returns:
            any: `document` lxml element
        """
        document_node = etree.Element("document")
        document_node.set("path", docname)

        self._append_sns_notation_nodes(document_node, result)
        self._append_non_context_rules_node(document_node, result)

        for brex_path, brex_result in result.items():
            if brex_path in ("sns", "notations", "brexFallback", "Summary", "Skipped", "nonContextRules"):
                continue
            if not isinstance(brex_result, dict):
                continue
            brex_node = etree.SubElement(document_node, "brex")
            brex_node.set("path", brex_path)
            for flag in ('0', '1', '2'):
                for violation in brex_result.get(flag, []):
                    structured = self._violation_from_dict(docname, brex_path, flag, violation)
                    if structured.duplicate:
                        # Already reported against an earlier, more specific
                        # layer of the same chain -- see `_deduplicate_violations`.
                        continue
                    self._append_error_node(brex_node, structured)
            for xpath_error in brex_result.get("xpathError", []):
                error_node = etree.SubElement(brex_node, "xpathError")
                error_node.text = xpath_error.get("Xpath")
                if xpath_error.get("Error") is not None:
                    error_node.set("error", xpath_error["Error"])

        return document_node

    def _append_run_summary_node(self, root: any, result: dict) -> None:
        """Append the `<summary>` node (category D7) as the first child of the
        `<brexCheck>` root, port of `s1kd-brexcheck -T`'s totals line into the
        `-x` XML report shape. Built from `run_summary`, so the XML report and
        the `run_summary()` dict always agree on the same counts.

        Args:
            root (any): the `brexCheck` root element to prepend the summary to
            result (dict): the same `validate()` result being converted
        """
        totals = self.run_summary(result)
        summary_node = etree.Element("summary")
        summary_node.set("documentsChecked", str(totals["DocumentsChecked"]))
        summary_node.set("documentsPassed", str(totals["DocumentsPassed"]))
        summary_node.set("documentsFailed", str(totals["DocumentsFailed"]))
        summary_node.set("documentsSkipped", str(totals["DocumentsSkipped"]))
        summary_node.set("errors", str(totals["Errors"]))
        summary_node.set("warnings", str(totals["Warnings"]))
        for severity, count in totals["ViolationsBySeverity"].items():
            severity_node = etree.SubElement(summary_node, "severity")
            if severity is not None:
                severity_node.set("value", severity)
            severity_node.set("count", str(count))
        root.insert(0, summary_node)

    def to_xml_report(self, result: dict) -> str:
        """Convert a `validate()` result into an XML report compatible with the
        `s1kd-brexcheck -x` shape:
        `brexCheck/{summary/severity,document/{sns,notations,brex/{error/{brDecisionRef,
        objectPath,objectUse,object},xpathError}}}`. `summary` (category D7) is our
        own addition -- not part of `-x`'s shape -- carrying the same per-run totals
        as `run_summary()`.

        Accepts either a single-object result (`validate()` after `set_xml`) or
        a directory-mode result (`validate()` after `set_xml_dir`, a mapping of
        `{filename: single-object result}`), distinguished the same way
        `_is_single_object_result` does.

        Args:
            result (dict): a `validate()` return value

        Returns:
            str: serialised `<brexCheck>` XML document
        """
        root = etree.Element("brexCheck")
        if self._is_single_object_result(result):
            root.append(self._build_document_node(result, self._xml_path or ""))
        else:
            for filename, file_result in result.items():
                if not isinstance(file_result, dict):
                    continue
                root.append(self._build_document_node(file_result, filename))
        self._append_run_summary_node(root, result)
        return etree.tostring(root, encoding="unicode", pretty_print=True)

    def _violation_records(self, result: dict) -> list:
        """Walk a `validate()` result once and pair every content-rule
        violation with the raw record it was built from.

        `violations()` exposes only the structured half; the formatted
        reports (`to_excel_report`/`to_html_report`) additionally want the
        raw record's `'Description'`, which is the only place the *found*
        value of a flag-2 violation is carried ("Element/Attribute (X) did
        not match the object values.") and which `BrexViolation` deliberately
        drops in favour of the rule's own `objectUse` (see
        `_violation_from_dict`). Keeping the walk here means all four report
        formats share one traversal and one set of skip rules.

        Args:
            result (dict): a `validate()` return value, single-object or
                directory-mode

        Returns:
            list: `(BrexViolation, raw record dict)` tuples, in document,
                then BREX, then flag ('0'/'1'/'2') order, including
                `duplicate=True` entries
        """
        records = []
        for docname, doc_result in self._result_documents(result).items():
            if doc_result.get("Skipped"):
                continue
            for brex_path, brex_result in doc_result.items():
                if brex_path in ("sns", "notations", "brexFallback", "Summary", "Skipped", "nonContextRules"):
                    continue
                if not isinstance(brex_result, dict):
                    continue
                for flag in ('0', '1', '2'):
                    for violation in brex_result.get(flag, []):
                        records.append(
                            (self._violation_from_dict(docname, brex_path, flag, violation), violation)
                        )
        return records

    def _document_stats(self, result: dict) -> list:
        """Per-document error/warning tallies, the batch-mode counterpart of
        `run_summary`'s run-wide totals: one row per checked object, so a
        directory run's report can show which files failed instead of only a
        grand total. Shared by `to_excel_report` and `to_html_report`.

        A document "passes" when its error count is zero, the same rule
        `run_summary` applies -- warnings alone do not fail it, they are
        reported in their own column.

        Args:
            result (dict): a `validate()` return value

        Returns:
            list: `{"document", "errors", "warnings", "severities", "status"}`
                dicts, one per document, in result order; `status` is one of
                `"Passed"`, `"Failed"` or `"Skipped"`
        """
        rows = []
        for docname, doc_result in self._result_documents(result).items():
            if doc_result.get("Skipped"):
                rows.append({
                    "document": docname,
                    "errors": 0,
                    "warnings": 0,
                    "severities": {},
                    "status": "Skipped",
                })
                continue
            errors, warnings, severities = self._count_violations(doc_result)
            rows.append({
                "document": docname,
                "errors": errors,
                "warnings": warnings,
                "severities": severities,
                "status": "Failed" if errors else "Passed",
            })
        return rows

    def _xpath_error_rows(self, result: dict) -> list:
        """Flatten the `xpathError` diagnostics (a rule whose `objectPath`
        could not be compiled or evaluated, see `_resolve_selector`) of every
        document into `document`-tagged rows.

        These are not violations -- `_count_violations` excludes them and
        `violations()` does not represent them -- but they mean a rule was
        silently *not* checked, which a report reader needs to see. The XML
        report already emits them as `xpathError` nodes; this is the same
        content for the Excel and HTML reports.

        Args:
            result (dict): a `validate()` return value

        Returns:
            list: `{"document", "brex", "objectPath", "objectUse", "error",
                "brDecisionIdentNumber"}` dicts
        """
        rows = []
        for docname, doc_result in self._result_documents(result).items():
            if doc_result.get("Skipped"):
                continue
            for brex_path, brex_result in doc_result.items():
                if not isinstance(brex_result, dict):
                    continue
                for xpath_error in brex_result.get('xpathError', []):
                    rows.append({
                        "document": docname,
                        "brex": brex_path,
                        "objectPath": xpath_error.get('Xpath'),
                        "objectUse": xpath_error.get('Description'),
                        "error": xpath_error.get('Error'),
                        "brDecisionIdentNumber": xpath_error.get('BrDecisionIdentNumber'),
                    })
        return rows

    def _report_source(self) -> str:
        """What was checked, for a report header: the directory in batch mode
        (`set_xml_dir`), otherwise the single object's path."""
        return self._xml_dir or self._xml_path or ""

    def _report_violation_rows(self, result: dict) -> list:
        """Render-ready violation rows shared by `to_excel_report` and
        `to_html_report`, so both reports carry the same columns in the same
        order and stay consistent with the JSON/XML ones.

        `duplicate=True` violations are left out, matching every other report
        and `run_summary`'s counts.

        Args:
            result (dict): a `validate()` return value

        Returns:
            list: one dict per reportable violation, keyed by the column
                labels in `_REPORT_VIOLATION_COLUMNS`
        """
        rows = []
        for violation, raw in self._violation_records(result):
            if violation.duplicate:
                continue
            allowed = violation.allowed_values or {}
            # 'Description' is the rule's objectUse for a flag-0/1 record and
            # the "Element/Attribute (X) did not match the object values."
            # message for a value one -- only worth a column in the latter
            # case, where it is the only carrier of the value actually found.
            message = raw.get('Description')
            if message == violation.object_use:
                message = None
            rows.append({
                "Document": basename(str(violation.document)) or str(violation.document),
                "BREX": basename(str(violation.brex)) or str(violation.brex),
                "Line": violation.line,
                "Flag": _REPORT_FLAG_LABELS.get(violation.flag, violation.flag),
                "Severity": violation.severity,
                "Status": "Error" if violation.fail else "Warning",
                "Rule ID": violation.rule_id,
                "BR decision": violation.br_decision_ident_number,
                "Context": violation.rules_context,
                "Object path": violation.object_path,
                "Object use": violation.object_use,
                "Finding": message,
                "Allowed (single)": ", ".join(str(_) for _ in allowed.get('single') or []),
                "Allowed (pattern)": ", ".join(str(_) for _ in allowed.get('pattern') or []),
                "Allowed (range)": ", ".join(str(_) for _ in allowed.get('range') or []),
                "Node xpath": violation.node_xpath,
                "Node": violation.node_snippet,
            })
        return rows

    def to_excel_report(self, result: dict, path: str) -> str:
        """Convert a `validate()` result into a formatted Excel workbook --
        the fourth report format alongside `to_json_report` (machine-readable),
        `to_xml_report` (`s1kd-brexcheck -x` compatible) and the raw result
        dict, and the one meant to be read by a human reviewer or handed to a
        customer.

        Accepts either a single-object result (`validate()` after `set_xml`)
        or a batch/directory-mode one (`validate()` after `set_xml_dir`, a
        mapping of `{filename: single-object result}`), distinguished the same
        way `_is_single_object_result` does. In batch mode the *Summary* sheet
        gains a per-document pass/fail table (`_document_stats`) and every
        other sheet's `Document` column identifies the object a row came from,
        so one workbook covers a whole folder run.

        Sheets, in order (a sheet whose content is empty is omitted, except
        the always-present first two):

        - **Summary** -- run totals (`run_summary`), violations by
          `brSeverityLevel`, and the per-document table
        - **Violations** -- one row per reportable content-rule violation
          (`_report_violation_rows`); `duplicate=True` entries are excluded,
          matching the other reports
        - **SNS**, **Notations** -- category A2/A3 violations
          (`_json_document_sections`)
        - **XPath errors** -- rules that could not be evaluated
          (`_xpath_error_rows`); not violations, but rules that were silently
          not checked
        - **Non-context rules**, **BREX fallback** -- informational

        Formatting: a coloured, bold, frozen header row with autofilter,
        borders on every cell, tuned column widths, wrapped text, and rows
        tinted by outcome (red for an error, amber for a warning, green for a
        passing document).

        Args:
            result (dict): a `validate()` return value
            path (str): destination `.xlsx` path; parent directory must exist

        Raises:
            ImportError: if `openpyxl` is not installed

        Returns:
            str: `path`, for convenience
        """
        try:
            # Imported here rather than at module scope: openpyxl is only
            # needed by this one report format, and importing brex_checker
            # itself must stay cheap (see the lazy `acd/__init__.py`).
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - depends on the environment
            raise ImportError(
                "to_excel_report needs openpyxl (pip install openpyxl); "
                "use to_json_report/to_xml_report/to_html_report otherwise."
            ) from exc

        def excel_value(value):
            """Excel-safe cell value: control characters openpyxl refuses are
            stripped, and an over-long node snippet is truncated well inside
            the 32767-character cell limit."""
            if value is None or isinstance(value, (int, float, bool)):
                return value
            text = str(value)
            text = ''.join(
                char for char in text
                if char in '\t\n\r' or ord(char) >= 32
            )
            if len(text) > 2000:
                text = text[:2000] + ' [...]'
            return text

        palette = _REPORT_PALETTE
        thin = Side(style="thin", color=palette["grid"])
        cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill("solid", fgColor=palette["header"])
        title_font = Font(bold=True, size=16, color=palette["header"])
        muted_font = Font(size=9, color=palette["muted"])
        label_font = Font(bold=True, size=11, color=palette["header"])
        top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        fills = {
            key: PatternFill("solid", fgColor=palette[key])
            for key in ("error", "warning", "ok", "band")
        }

        def write_table(worksheet, headers, rows, widths, tint=None, start_row=1,
                        autofilter=True, freeze=True):
            """Header + body of one table, styled; returns the next free row."""
            for column, header in enumerate(headers, start=1):
                cell = worksheet.cell(row=start_row, column=column, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = cell_border
                cell.alignment = header_align
            worksheet.row_dimensions[start_row].height = 26
            for offset, row in enumerate(rows):
                row_number = start_row + 1 + offset
                fill = fills.get(tint(row)) if tint else None
                if fill is None and offset % 2:
                    fill = fills["band"]
                for column, header in enumerate(headers, start=1):
                    cell = worksheet.cell(
                        row=row_number, column=column, value=excel_value(row.get(header))
                    )
                    cell.border = cell_border
                    cell.alignment = top_left
                    if fill is not None:
                        cell.fill = fill
            for column, width in enumerate(widths, start=1):
                worksheet.column_dimensions[get_column_letter(column)].width = width
            last_row = start_row + len(rows)
            if autofilter:
                worksheet.auto_filter.ref = (
                    f"A{start_row}:{get_column_letter(len(headers))}{max(last_row, start_row)}"
                )
            if freeze:
                worksheet.freeze_panes = worksheet.cell(row=start_row + 1, column=1)
            return last_row + 2

        def add_sheet(name, tab_color):
            worksheet = workbook.create_sheet(name)
            worksheet.sheet_properties.tabColor = tab_color
            return worksheet

        workbook = Workbook()
        summary = workbook.active
        summary.title = "Summary"
        summary.sheet_properties.tabColor = palette["header"]
        summary.sheet_view.showGridLines = False

        totals = self.run_summary(result)
        document_rows = self._document_stats(result)
        violation_rows = self._report_violation_rows(result)
        sections = self._json_document_sections(result)
        xpath_error_rows = self._xpath_error_rows(result)

        summary.merge_cells("A1:D1")
        summary["A1"] = "BREX check report"
        summary["A1"].font = title_font
        summary.merge_cells("A2:D2")
        summary["A2"] = (
            f"{self._report_source()}  --  generated "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        summary["A2"].font = muted_font
        summary.column_dimensions["A"].width = 34
        summary.column_dimensions["B"].width = 16

        row = 4
        summary.cell(row=row, column=1, value="Run totals").font = label_font
        row += 1
        totals_start = row
        for label, key in (
            ("Documents checked", "DocumentsChecked"),
            ("Documents passed", "DocumentsPassed"),
            ("Documents failed", "DocumentsFailed"),
            ("Documents skipped", "DocumentsSkipped"),
            ("Errors", "Errors"),
            ("Warnings", "Warnings"),
        ):
            label_cell = summary.cell(row=row, column=1, value=label)
            value_cell = summary.cell(row=row, column=2, value=totals[key])
            for cell in (label_cell, value_cell):
                cell.border = cell_border
            value_cell.font = Font(bold=True)
            if totals[key]:
                if key in ("DocumentsFailed", "Errors"):
                    for cell in (label_cell, value_cell):
                        cell.fill = fills["error"]
                elif key == "Warnings":
                    for cell in (label_cell, value_cell):
                        cell.fill = fills["warning"]
                elif key == "DocumentsPassed":
                    for cell in (label_cell, value_cell):
                        cell.fill = fills["ok"]
            row += 1
        summary.row_dimensions[totals_start].height = 18

        row += 1
        summary.cell(row=row, column=1, value="Violations by severity").font = label_font
        row += 1
        severity_rows = [
            {"Severity": severity if severity is not None else "(none declared)", "Count": count}
            for severity, count in sorted(
                totals["ViolationsBySeverity"].items(), key=lambda item: str(item[0])
            )
        ] or [{"Severity": "(none)", "Count": 0}]
        # No freeze pane on this sheet: its tables start well down the page,
        # and freezing there would pin the whole block above them. The
        # documents table below carries the autofilter instead -- it is the
        # one worth filtering on a folder run.
        row = write_table(summary, ["Severity", "Count"], severity_rows, [34, 16],
                          start_row=row, autofilter=False, freeze=False)

        summary.cell(row=row, column=1, value="Documents").font = label_font
        row += 1
        write_table(
            summary,
            ["Document", "Status", "Errors", "Warnings"],
            [
                {
                    "Document": basename(str(entry["document"])) or str(entry["document"]),
                    "Status": entry["status"],
                    "Errors": entry["errors"],
                    "Warnings": entry["warnings"],
                }
                for entry in document_rows
            ],
            [52, 14, 12, 12],
            tint=lambda entry: (
                "error" if entry["Status"] == "Failed"
                else "warning" if entry["Warnings"] else "ok" if entry["Status"] == "Passed" else None
            ),
            start_row=row,
            freeze=False,
        )

        violations_sheet = add_sheet("Violations", palette["error"])
        write_table(
            violations_sheet,
            list(_REPORT_VIOLATION_COLUMNS),
            violation_rows,
            [26, 22, 8, 22, 14, 10, 16, 16, 16, 46, 46, 34, 26, 26, 26, 34, 40],
            tint=lambda entry: "error" if entry["Status"] == "Error" else "warning",
        )

        if sections["sns"]:
            write_table(
                add_sheet("SNS", palette["error"]),
                ["Document", "Code", "Invalid value", "Object use"],
                [
                    {
                        "Document": basename(str(entry["document"])) or str(entry["document"]),
                        "Code": entry["code"],
                        "Invalid value": entry["invalidValue"],
                        "Object use": entry["objectUse"],
                    }
                    for entry in sections["sns"]
                ],
                [26, 22, 22, 60],
                tint=lambda entry: "error",
            )
        if sections["notations"]:
            write_table(
                add_sheet("Notations", palette["error"]),
                ["Document", "Entity", "Invalid notation", "Object use"],
                [
                    {
                        "Document": basename(str(entry["document"])) or str(entry["document"]),
                        "Entity": entry["entity"],
                        "Invalid notation": entry["invalidNotation"],
                        "Object use": entry["objectUse"],
                    }
                    for entry in sections["notations"]
                ],
                [26, 22, 22, 60],
                tint=lambda entry: "error",
            )
        if xpath_error_rows:
            write_table(
                add_sheet("XPath errors", palette["warning"]),
                ["Document", "BREX", "Object path", "Object use", "Error"],
                [
                    {
                        "Document": basename(str(entry["document"])) or str(entry["document"]),
                        "BREX": basename(str(entry["brex"])) or str(entry["brex"]),
                        "Object path": entry["objectPath"],
                        "Object use": entry["objectUse"],
                        "Error": entry["error"],
                    }
                    for entry in xpath_error_rows
                ],
                [26, 22, 46, 46, 46],
                tint=lambda entry: "warning",
            )
        if sections["nonContextRules"]:
            write_table(
                add_sheet("Non-context rules", palette["muted"]),
                ["Document", "BREX", "BR decision", "Text"],
                [
                    {
                        "Document": basename(str(entry["document"])) or str(entry["document"]),
                        "BREX": basename(str(entry["brex"] or "")) or str(entry["brex"]),
                        "BR decision": entry["brDecisionIdentNumber"],
                        "Text": entry["text"],
                    }
                    for entry in sections["nonContextRules"]
                ],
                [26, 22, 20, 80],
            )
        if sections["brexFallback"]:
            write_table(
                add_sheet("BREX fallback", palette["muted"]),
                ["Document", "Reference", "Used built-in BREX", "Built-in BREX path"],
                [
                    {
                        "Document": basename(str(entry["document"])) or str(entry["document"]),
                        "Reference": entry["reference"],
                        "Used built-in BREX": entry["usedBuiltinBrex"],
                        "Built-in BREX path": entry["builtinBrexPath"],
                    }
                    for entry in sections["brexFallback"]
                ],
                [26, 40, 20, 60],
            )

        workbook.save(clean_path(path))
        return path

    def to_html_report(self, result: dict, path: str = None,
                       title: str = "BREX check report") -> str:
        """Convert a `validate()` result into a self-contained, formatted HTML
        report -- the browser-readable counterpart of `to_excel_report`,
        carrying the same content and built from the same shared row helpers.

        Accepts either a single-object result or a batch/directory-mode one,
        distinguished the same way `_is_single_object_result` does; in batch
        mode the report gains a per-document pass/fail table
        (`_document_stats`) and every table's `Document` column names the
        object a row came from.

        The output is one HTML document with no external requests at all --
        CSS and JS inline, no fonts, images or CDN links -- so it can be
        e-mailed or opened straight off a network share. It follows the
        reader's `prefers-color-scheme` for dark/light and carries a toggle
        that overrides it (remembered in `localStorage`, guarded so a
        `file://` document with site data blocked still renders), plus a live
        text filter and an errors-only switch over the violations table.

        Args:
            result (dict): a `validate()` return value
            path (str): optional destination; the report is written there
                (UTF-8) as well as returned
            title (str): heading shown at the top of the report

        Returns:
            str: the complete HTML document
        """
        totals = self.run_summary(result)
        document_rows = self._document_stats(result)
        violation_rows = self._report_violation_rows(result)
        sections = self._json_document_sections(result)
        xpath_error_rows = self._xpath_error_rows(result)

        def cell(value, css_class=None, mono=False):
            text = "" if value is None else str(value)
            classes = " ".join(_ for _ in (css_class, "mono" if mono else None) if _)
            attr = f' class="{classes}"' if classes else ""
            return f"<td{attr}>{html_escape(text)}</td>"

        def table(headers, body_rows, css_class="grid"):
            head = "".join(f"<th>{html_escape(_)}</th>" for _ in headers)
            body = "".join(f"<tr>{row}</tr>" for row in body_rows)
            return (
                f'<div class="table-wrap"><table class="{css_class}">'
                f"<thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>"
            )

        def section(heading, count, content, open_by_default=True):
            return (
                f'<details class="section"{" open" if open_by_default else ""}>'
                f'<summary><span class="section-title">{html_escape(heading)}</span>'
                f'<span class="pill">{count}</span></summary>{content}</details>'
            )

        parts = [
            "<!doctype html>",
            '<html lang="en"><head><meta charset="utf-8">',
            '<meta name="viewport" content="width=device-width, initial-scale=1">',
            f"<title>{html_escape(title)}</title>",
            f"<style>{_REPORT_HTML_CSS}</style>",
            "</head><body>",
            '<header class="page-head"><div>',
            f"<h1>{html_escape(title)}</h1>",
            f'<p class="muted">{html_escape(self._report_source())} &middot; generated '
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
            "</div>",
            '<button id="theme-toggle" type="button" aria-label="Toggle dark mode">'
            '<span class="theme-icon"></span><span class="theme-label">Theme</span></button>',
            "</header>",
        ]

        cards = [
            ("Documents checked", totals["DocumentsChecked"], "neutral"),
            ("Passed", totals["DocumentsPassed"], "ok" if totals["DocumentsPassed"] else "neutral"),
            ("Failed", totals["DocumentsFailed"], "error" if totals["DocumentsFailed"] else "neutral"),
            ("Errors", totals["Errors"], "error" if totals["Errors"] else "ok"),
            ("Warnings", totals["Warnings"], "warning" if totals["Warnings"] else "neutral"),
        ]
        if totals["DocumentsSkipped"]:
            cards.append(("Skipped", totals["DocumentsSkipped"], "neutral"))
        parts.append('<section class="cards">')
        for label, value, tone in cards:
            parts.append(
                f'<div class="card {tone}"><span class="card-value">{value}</span>'
                f'<span class="card-label">{html_escape(label)}</span></div>'
            )
        parts.append("</section>")

        if totals["ViolationsBySeverity"]:
            chips = "".join(
                f'<span class="chip"><span class="chip-key">'
                f'{html_escape(str(severity) if severity is not None else "no severity")}'
                f'</span><span class="chip-value">{count}</span></span>'
                for severity, count in sorted(
                    totals["ViolationsBySeverity"].items(), key=lambda item: str(item[0])
                )
            )
            parts.append(f'<section class="chips">{chips}</section>')

        if len(document_rows) > 1:
            parts.append(section(
                "Documents",
                len(document_rows),
                table(
                    ["Document", "Status", "Errors", "Warnings"],
                    [
                        cell(basename(str(entry["document"])) or entry["document"], mono=True)
                        + f'<td><span class="status {entry["status"].lower()}">'
                        f"{html_escape(entry['status'])}</span></td>"
                        + cell(entry["errors"], "num")
                        + cell(entry["warnings"], "num")
                        for entry in document_rows
                    ],
                ),
            ))

        if violation_rows:
            controls = (
                '<div class="controls">'
                '<input id="violation-filter" type="search" '
                'placeholder="Filter violations (document, rule, xpath, text...)">'
                '<label class="switch"><input id="errors-only" type="checkbox">'
                "<span>Errors only</span></label>"
                '<span id="violation-count" class="muted"></span></div>'
            )
            body_rows = []
            for entry in violation_rows:
                details = []
                if entry["Finding"]:
                    details.append(f'<p class="finding">{html_escape(entry["Finding"])}</p>')
                for label in ("Allowed (single)", "Allowed (pattern)", "Allowed (range)"):
                    if entry[label]:
                        details.append(
                            f'<p><span class="muted">{html_escape(label)}:</span> '
                            f'<code>{html_escape(entry[label])}</code></p>'
                        )
                if entry["Node xpath"]:
                    details.append(
                        f'<p><span class="muted">node:</span> '
                        f'<code>{html_escape(str(entry["Node xpath"]))}</code></p>'
                    )
                if entry["Node"]:
                    details.append(
                        "<details><summary>node source</summary>"
                        f'<pre>{html_escape(str(entry["Node"]))}</pre></details>'
                    )
                status = entry["Status"].lower()
                body_rows.append(
                    f'<tr data-status="{status}">'
                    + cell(entry["Document"], mono=True)
                    + cell(entry["Line"], "num")
                    + f'<td><span class="status {status}">{html_escape(entry["Status"])}</span></td>'
                    + cell(entry["Severity"])
                    + cell(entry["Flag"])
                    + cell(entry["Rule ID"] or entry["BR decision"], mono=True)
                    + cell(entry["Object path"], mono=True)
                    + cell(entry["Object use"])
                    + f'<td class="details">{"".join(details)}</td>'
                    + "</tr>"
                )
            head = "".join(
                f"<th>{_}</th>"
                for _ in ("Document", "Line", "Status", "Severity", "Flag", "Rule",
                          "Object path", "Object use", "Details")
            )
            parts.append(section(
                "Violations",
                len(violation_rows),
                controls
                + '<div class="table-wrap"><table class="grid" id="violations">'
                f"<thead><tr>{head}</tr></thead><tbody>{''.join(body_rows)}</tbody></table></div>",
            ))
        else:
            parts.append(
                '<section class="empty"><span class="empty-mark">&#10003;</span>'
                "<p>No BREX violations found.</p></section>"
            )

        if sections["sns"]:
            parts.append(section("SNS violations", len(sections["sns"]), table(
                ["Document", "Code", "Invalid value", "Object use"],
                [
                    cell(basename(str(entry["document"])) or entry["document"], mono=True)
                    + cell(entry["code"], mono=True)
                    + cell(entry["invalidValue"], mono=True)
                    + cell(entry["objectUse"])
                    for entry in sections["sns"]
                ],
            )))
        if sections["notations"]:
            parts.append(section("Notation violations", len(sections["notations"]), table(
                ["Document", "Entity", "Invalid notation", "Object use"],
                [
                    cell(basename(str(entry["document"])) or entry["document"], mono=True)
                    + cell(entry["entity"], mono=True)
                    + cell(entry["invalidNotation"], mono=True)
                    + cell(entry["objectUse"])
                    for entry in sections["notations"]
                ],
            )))
        if xpath_error_rows:
            parts.append(section("XPath errors", len(xpath_error_rows), table(
                ["Document", "BREX", "Object path", "Object use", "Error"],
                [
                    cell(basename(str(entry["document"])) or entry["document"], mono=True)
                    + cell(basename(str(entry["brex"])) or entry["brex"], mono=True)
                    + cell(entry["objectPath"], mono=True)
                    + cell(entry["objectUse"])
                    + cell(entry["error"])
                    for entry in xpath_error_rows
                ],
            )))
        if sections["nonContextRules"]:
            parts.append(section("Non-context rules", len(sections["nonContextRules"]), table(
                ["Document", "BREX", "BR decision", "Text"],
                [
                    cell(basename(str(entry["document"])) or entry["document"], mono=True)
                    + cell(basename(str(entry["brex"] or "")) or entry["brex"], mono=True)
                    + cell(entry["brDecisionIdentNumber"], mono=True)
                    + cell(entry["text"])
                    for entry in sections["nonContextRules"]
                ],
            ), open_by_default=False))
        if sections["brexFallback"]:
            parts.append(section("BREX fallback", len(sections["brexFallback"]), table(
                ["Document", "Reference", "Used built-in BREX", "Built-in BREX path"],
                [
                    cell(basename(str(entry["document"])) or entry["document"], mono=True)
                    + cell(entry["reference"], mono=True)
                    + cell(entry["usedBuiltinBrex"])
                    + cell(entry["builtinBrexPath"], mono=True)
                    for entry in sections["brexFallback"]
                ],
            ), open_by_default=False))

        parts.append(f"<script>{_REPORT_HTML_JS}</script>")
        parts.append("</body></html>")
        html = "\n".join(parts)
        if path:
            with open(clean_path(path), "w", encoding="utf-8") as report_file:
                report_file.write(html)
        return html

    def lint_brex(self, brex_path: str, csdb_schemas: any = None) -> list:
        """Self-consistency lint pass over a single BREX file, meant to be run
        before that BREX is used to check anything (e.g. before passing it to
        `override_brex_list`/`set_brex_path`), so authoring defects surface on
        their own instead of silently under-checking every object validated
        against it.

        Unlike `_show_rules`/`_check_rules`, this is not tied to any checked
        object: it inspects every `structureObjectRule` in the file
        regardless of `rulesContext`/`allowedObjectFlag`, and has no
        equivalent in `s1kd-brexcheck` -- it is our own addition, reusing the
        same rule-selection (`_get_object_rule_nodes`) and value-parsing
        (`translate_xsd_regex_to_python`, the `~`/`|` range-and-set grammar
        `is_in_set` relies on) building blocks the checker itself uses.

        Checks performed, one finding per issue found:

        - every `objectPath`/`objpath` compiles as an XPath expression;
        - every `structureObjectRule`/`objrule` has an `objectUse`/`objuse`
          child (a missing one is what `_show_rules`'s `...[0].text` would
          raise `IndexError` on);
        - an `allowedObjectFlag="2"`/`objappl="2"` rule carries at least one
          `objectValue`/`objval` child -- otherwise it is a silent no-op,
          since `_check_object_flag_2` only ever checks a rule that declares
          `values_allowed`/`regex_allowed`/`ranges_allowed`;
        - every `valueForm="pattern"`/`valtype="pattern"` value is a valid
          XSD regular expression once translated
          (`translate_xsd_regex_to_python`) and compiled (`regex` `V1`);
        - every `valueForm="range"`/`valtype="range"` value parses as a
          `~`-range or `|`-set, i.e. no empty member and no empty range
          bound (`is_in_set`/`is_in_range` do not raise on these, they just
          silently mis-compare, so a bad range is otherwise invisible);
        - a rule's resolved `brSeverityLevel` (its own, or the BREX root's
          `defaultBrSeverityLevel`) exists in the currently configured
          `.brseveritylevels` file (see `set_severity_levels_path`/
          `set_severity_levels_search`); skipped entirely when no severity
          levels are configured/discoverable, since there is nothing to
          check against;
        - duplicate `structureObjectRule/@id` values;
        - duplicate `brDecisionIdentNumber` values (via `brDecisionRef`);
        - when `csdb_schemas` is given, a `contextRules`/`contextrules` group
          whose `rulesContext`/`context` names a schema absent from it (see
          `_lint_unreachable_rules_context`);
        - the `snsRules` table itself, if the BREX declares one (see
          `_lint_sns_rules_table`): duplicate `snsCode` among sibling
          elements at the same level, a level element with no `snsTitle`,
          and an `snsCode` that does not match a `valueForm="pattern"`
          content rule declared elsewhere in the same file for the
          corresponding `dmCode` attribute (`systemCode`/`subSystemCode`/
          `subSubSystemCode`/`assyCode`).

        Args:
            brex_path (str): path to the BREX file to lint
            csdb_schemas (any): optional iterable of schema URI strings
                actually used by objects in the target CSDB (see
                `s1000d.collect_csdb_schemas`). When given, enables the
                unreachable-`rulesContext` check; `None` (the default) skips
                it, since without a CSDB to compare against there is no way
                to tell a legitimately-unused-yet schema from a typo.

        Returns:
            list: finding dicts, each carrying at least `Category` and
                `Description`, plus whichever of `Xpath`/`Id`/
                `BrDecisionIdentNumber`/`Line`/`ValueAllowed`/`RulesContext`/
                `Level`/`SnsCode` apply to that finding; empty if the BREX
                has no self-consistency issues
        """
        findings = []
        nodes_to_check = self._get_object_rule_nodes(brex_path, schema=None)

        default_br_severity_level = None
        # Compile with the same XPath version `_show_rules` would actually
        # use for this BREX (see `_brex_requires_xpath2`), so InvalidXPath
        # findings reflect real checking behaviour rather than always
        # assuming XPath 2.0.
        xpath_parser = elementpath.XPath2Parser
        if len(nodes_to_check) > 0:
            brex_root = nodes_to_check[0].getroottree().getroot()
            default_br_severity_level = brex_root.get('defaultBrSeverityLevel')
            brex_schema = brex_root.get(f'{{{NS_DICT["xsi"]}}}noNamespaceSchemaLocation')
            if not self._brex_requires_xpath2(brex_schema):
                xpath_parser = elementpath.XPath1Parser

        severity_levels = self._get_severity_levels()

        seen_ids = {}
        seen_br_decision_numbers = {}

        for object_path in nodes_to_check:
            rule = object_path.getparent()
            xpath_text = str(object_path.text) if object_path.text is not None else ""
            line = rule.sourceline

            rule_id = rule.get('id')
            if rule_id:
                seen_ids.setdefault(rule_id, []).append(line)

            br_decision_ref = rule.find('brDecisionRef')
            br_decision_ident_number = (
                br_decision_ref.get('brDecisionIdentNumber') if br_decision_ref is not None else None
            )
            if br_decision_ident_number:
                seen_br_decision_numbers.setdefault(br_decision_ident_number, []).append(line)

            base_fields = {
                'Xpath': xpath_text,
                'Id': rule_id,
                'BrDecisionIdentNumber': br_decision_ident_number,
                'Line': line,
            }

            namespaces = dict(NS_DICT)
            namespaces.update({(prefix or ''): uri for prefix, uri in object_path.nsmap.items()})
            try:
                elementpath.Selector(xpath_text, namespaces=namespaces, parser=xpath_parser)
            except elementpath.ElementPathError as e:
                findings.append({
                    'Category': 'InvalidXPath',
                    'Description': f"objectPath does not compile: {e}",
                    **base_fields,
                })

            if not rule.xpath('objectUse|objuse'):
                findings.append({
                    'Category': 'MissingObjectUse',
                    'Description': "structureObjectRule has no objectUse.",
                    **base_fields,
                })

            allowed_object_flag = object_path.get('allowedObjectFlag', object_path.get('objappl'))
            object_value_nodes = rule.xpath('objectValue|objval')

            if allowed_object_flag == '2' and not object_value_nodes:
                findings.append({
                    'Category': 'EmptyValueFlag2',
                    'Description': (
                        'allowedObjectFlag="2" rule has no objectValue children '
                        'and is a silent no-op.'
                    ),
                    **base_fields,
                })

            for object_value in object_value_nodes:
                value_form = object_value.get('valueForm', object_value.get('valtype'))
                value_allowed = object_value.get('valueAllowed')
                if value_allowed is None and object_value.get('val1') is not None:
                    value_allowed = object_value.get('val1')
                    val2 = object_value.get('val2')
                    if val2 is not None:
                        value_allowed = f"{value_allowed}~{val2}"

                if value_form == 'pattern':
                    if value_allowed is None:
                        findings.append({
                            'Category': 'InvalidPattern',
                            'Description': 'valueForm="pattern" objectValue has no valueAllowed.',
                            'ValueAllowed': value_allowed,
                            **base_fields,
                        })
                        continue
                    try:
                        regex_compile(translate_xsd_regex_to_python(value_allowed), V1)
                    except RegexError as e:
                        findings.append({
                            'Category': 'InvalidPattern',
                            'Description': f"pattern is not a valid XSD regular expression: {e}",
                            'ValueAllowed': value_allowed,
                            **base_fields,
                        })
                elif value_form == 'range':
                    if value_allowed is None:
                        findings.append({
                            'Category': 'InvalidRange',
                            'Description': 'valueForm="range" objectValue has no valueAllowed.',
                            'ValueAllowed': value_allowed,
                            **base_fields,
                        })
                        continue
                    for member in value_allowed.split('|'):
                        bounds = member.split('~')
                        if member == '' or any(bound == '' for bound in bounds):
                            findings.append({
                                'Category': 'InvalidRange',
                                'Description': (
                                    f"range/set member {member!r} of valueAllowed "
                                    f"{value_allowed!r} does not parse as a range or set."
                                ),
                                'ValueAllowed': value_allowed,
                                **base_fields,
                            })

            rule_severity = rule.get('brSeverityLevel')
            effective_severity = rule_severity if rule_severity is not None else default_br_severity_level
            if effective_severity is not None and severity_levels and effective_severity not in severity_levels:
                findings.append({
                    'Category': 'UnknownSeverityLevel',
                    'Description': (
                        f"brSeverityLevel {effective_severity!r} is not defined in the "
                        "configured .brseveritylevels file."
                    ),
                    'BrSeverityLevel': effective_severity,
                    **base_fields,
                })

        for rule_id, lines in seen_ids.items():
            if len(lines) > 1:
                findings.append({
                    'Category': 'DuplicateId',
                    'Description': f"structureObjectRule/@id {rule_id!r} is used by {len(lines)} rules.",
                    'Id': rule_id,
                    'Lines': lines,
                })

        for number, lines in seen_br_decision_numbers.items():
            if len(lines) > 1:
                findings.append({
                    'Category': 'DuplicateBrDecisionIdentNumber',
                    'Description': f"brDecisionIdentNumber {number!r} is used by {len(lines)} rules.",
                    'BrDecisionIdentNumber': number,
                    'Lines': lines,
                })

        if csdb_schemas is not None:
            findings.extend(self._lint_unreachable_rules_context(brex_path, csdb_schemas))

        findings.extend(self._lint_sns_rules_table(brex_path, nodes_to_check))

        return findings

    def _lint_unreachable_rules_context(self, brex_path: str, csdb_schemas: any) -> list:
        """Find `contextRules`/`contextrules` groups whose `rulesContext`/
        `context` names a schema that no object in the given CSDB actually
        uses -- including a typo'd schema URI, which is indistinguishable
        from "legitimately unused" without a CSDB to compare against. Every
        rule nested under such a group can never match any object, since
        `_get_object_rule_nodes` only ever selects a qualified group when its
        schema string equals the checked object's schema exactly.

        An unqualified group (no `rulesContext`/`context` at all) always
        applies to every object and is never flagged.

        Args:
            brex_path (str): path to the BREX file to lint
            csdb_schemas (any): iterable of schema URI strings actually used
                by objects in the target CSDB, e.g. from
                `s1000d.collect_csdb_schemas`. An empty set means "nothing to
                compare against" (no valid XML found), not "every qualified
                group is unreachable" -- the check is skipped entirely.

        Returns:
            list: one `UnreachableRulesContext` finding per distinct
                unreachable schema string, carrying `RulesContext` and every
                `Lines` (the `contextRules`/`contextrules` element's line)
                it was declared on
        """
        schemas = set(csdb_schemas)
        if not schemas:
            return []

        root = self._parse_brex_root(brex_path)
        groups = root.xpath('//contextRules[@rulesContext]|//contextrules[@context]')

        lines_by_schema = {}
        for group in groups:
            schema = group.get('rulesContext', group.get('context'))
            if schema in schemas:
                continue
            lines_by_schema.setdefault(schema, []).append(group.sourceline)

        return [
            {
                'Category': 'UnreachableRulesContext',
                'Description': (
                    f"rulesContext {schema!r} names a schema no object in the given "
                    "CSDB uses; every rule in this group can never fire."
                ),
                'RulesContext': schema,
                'Lines': lines,
            }
            for schema, lines in lines_by_schema.items()
        ]

    def _lint_sns_rules_table(self, brex_path: str, content_rule_nodes: any) -> list:
        """Validate the `snsRules` table itself, if the BREX declares one:
        duplicate `snsCode` among sibling `snsSystem`/`snsSubSystem`/
        `snsSubSubSystem`/`snsAssy` elements at the same level, a level
        element with no `snsTitle`, and a `snsCode` that fails every
        `valueForm="pattern"` content rule declared elsewhere in the same
        file for the corresponding `dmCode` attribute (a rule whose
        `objectPath`/`objpath` ends in `@systemCode`, `@subSystemCode`,
        `@subSubSystemCode` or `@assyCode` -- a heuristic name match, since
        nothing formally links an `snsRules` level to a content rule).

        Neither `s1kd-brexcheck` nor our own checker validate the SNS table's
        own structure today -- only a data module's SNS *code* is checked
        against it (`_check_sns_rules`); a malformed table (a duplicate
        entry, a missing title, a code that contradicts the BREX's own
        pattern rule) is otherwise invisible.

        Args:
            brex_path (str): path to the BREX file to lint
            content_rule_nodes (any): `objectPath`/`objpath` nodes already
                selected by `lint_brex` (via `_get_object_rule_nodes`),
                reused here to build the pattern cross-reference without a
                second selection pass

        Returns:
            list: `DuplicateSnsCode` / `MissingSnsTitle` /
                `SnsCodeOutsidePattern` findings; empty if the BREX declares
                no `snsRules`, or its table has no such issues
        """
        attr_by_tag = {
            'snsSystem': 'systemCode',
            'snsSubSystem': 'subSystemCode',
            'snsSubSubSystem': 'subSubSystemCode',
            'snsAssy': 'assyCode',
        }

        patterns_by_attr = {}
        for object_path in content_rule_nodes:
            xpath_text = str(object_path.text) if object_path.text is not None else ""
            attr = xpath_text.rsplit('@', 1)[-1] if '@' in xpath_text else None
            if attr not in attr_by_tag.values():
                continue
            for object_value in object_path.getparent().xpath('objectValue|objval'):
                if object_value.get('valueForm', object_value.get('valtype')) != 'pattern':
                    continue
                value_allowed = object_value.get('valueAllowed', object_value.get('val1'))
                if value_allowed is None:
                    continue
                try:
                    translated = translate_xsd_regex_to_python(value_allowed)
                    regex_compile(translated, V1)
                except RegexError:
                    continue
                patterns_by_attr.setdefault(attr, []).append(translated)

        findings = []
        root = self._parse_brex_root(brex_path)
        for tag, attr in attr_by_tag.items():
            codes_by_parent = {}
            for element in root.xpath(f'.//{tag}'):
                code_node = element.find('snsCode')
                code = code_node.text if code_node is not None else None
                line = element.sourceline

                if element.find('snsTitle') is None:
                    findings.append({
                        'Category': 'MissingSnsTitle',
                        'Description': f"{tag} (snsCode={code!r}) has no snsTitle.",
                        'Level': tag,
                        'SnsCode': code,
                        'Line': line,
                    })

                if code is None:
                    continue

                codes_by_parent.setdefault((element.getparent(), code), []).append(line)

                patterns = patterns_by_attr.get(attr)
                if patterns and not any(bool(fullmatch(pattern, code, V1)) for pattern in patterns):
                    findings.append({
                        'Category': 'SnsCodeOutsidePattern',
                        'Description': (
                            f"{tag} snsCode {code!r} does not match any valueForm=\"pattern\" "
                            f"content rule declared for @{attr}."
                        ),
                        'Level': tag,
                        'SnsCode': code,
                        'Line': line,
                    })

            for (_parent, code), lines in codes_by_parent.items():
                if len(lines) > 1:
                    findings.append({
                        'Category': 'DuplicateSnsCode',
                        'Description': f"{tag} snsCode {code!r} is used by {len(lines)} sibling elements.",
                        'Level': tag,
                        'SnsCode': code,
                        'Lines': lines,
                    })

        return findings

    def lint_brex_layers(self, entry_brex_path: str) -> list:
        """Cross-BREX conflict detection across a layered BREX chain
        (category A8), independent of any checked object: resolves the same
        `brexDmRef`/`brexref` chain `_init_brex_list` would follow, but
        starting from `entry_brex_path` itself rather than from an object
        that references it (see `_walk_brex_chain`), then compares every
        content rule that appears in more than one layer.

        Layer order follows official S1000D `valueTailoring` terminology: a
        project's own BREX (closer to the object -- `entry_brex_path` is
        layer 0) is the "lower" layer that may only restrict what a "higher",
        more general BREX it references (via `brexDmRef`, at a larger layer
        index -- ultimately a master/default BREX) declares. Two rules are
        treated as "the same rule" when they share both their `rulesContext`/
        `context` qualifier and their exact `objectPath`/`objpath` text; this
        is a simplification (it does not, for example, catch an unqualified
        rule in one layer conflicting with a schema-qualified rule for the
        same path in another), but keeps the comparison unambiguous.

        Findings:

        - `ConflictingAllowedObjectFlag`: the same rule is required absent
          (`"0"`) by one layer and present (`"1"`) by another -- the only
          `allowedObjectFlag` pair no object could ever satisfy
          simultaneously. `"2"` (value constrained) is not compared this way:
          it is compatible with either on its own -- an absent object
          trivially satisfies a `"2"` rule too (nothing to check), and paired
          with `"1"` it only adds a value constraint on top of the presence
          requirement -- so a general layer that merely constrains an
          attribute's value while a more specific layer forbids it outright
          is a legitimate narrowing, not a conflict.
        - `RestrictableValueSetWidened`: the same rule declares a `single`-
          form `objectValue` with `valueTailoring="restrictable"` in two or
          more layers, and a lower (more specific) layer's allowed-value set
          is not a subset of a higher (more general) layer's set for that
          same rule -- i.e. the specific layer added values the general
          layer never allowed, which `restrictable` permits narrowing but
          not widening. `pattern`/`range` value forms are not compared: a
          general subset relationship between two regular expressions or
          ranges is not attempted here.

        Args:
            entry_brex_path (str): path to the BREX file to start the layer
                walk from (the same file that would be passed to
                `set_brex_path`/`override_brex_list`, or referenced by an
                object's own `brexDmRef`)

        Returns:
            list: finding dicts; empty if the chain resolves to a single
                BREX (nothing to compare), or every shared rule agrees
                across all layers that define it
        """
        search_dir = self._brex_dir_path[0] if self._brex_dir_path[0] else dirname(entry_brex_path)
        chain, _ = self._walk_brex_chain(entry_brex_path, search_dir)
        layers = [entry_brex_path] + chain
        if len(layers) < 2:
            return []

        rules_by_key = {}
        for index, layer in enumerate(layers):
            for rule in self._show_rules(layer, schema=None):
                key = (rule['contextRules'], rule['xpath'])
                rules_by_key.setdefault(key, []).append((index, rule))

        findings = []
        for (context_rules, xpath), entries in rules_by_key.items():
            if len(entries) < 2:
                continue

            flag_entries = [
                (index, rule['Brex'], rule.get('ObjectFlag'))
                for index, rule in entries
                if rule.get('ObjectFlag') is not None
            ]
            distinct_flags = {flag for _, _, flag in flag_entries}
            # "0" (must not be present) and "1" (must be present) are the only
            # truly irreconcilable pair: no object can satisfy both. "2" (value
            # constrained) is compatible with either on its own -- an object
            # that is simply absent trivially satisfies a "2" rule too (nothing
            # to check), and combined with "1" it just adds a value constraint
            # on top of the presence requirement. Verified against the real
            # ATABREX 01A/00A/S1000D-default chain: without this restriction,
            # a project layer forbidding an attribute ("0") that a more general
            # layer only value-constrains ("2") -- a legitimate narrowing, not
            # a conflict -- produced two false positives.
            if {'0', '1'} <= distinct_flags:
                findings.append({
                    'Category': 'ConflictingAllowedObjectFlag',
                    'Description': (
                        f"objectPath {xpath!r} is given contradictory allowedObjectFlag "
                        "values across the layered BREX chain."
                    ),
                    'Xpath': xpath,
                    'ContextRules': context_rules,
                    'Layers': [
                        {'LayerIndex': index, 'Brex': brex, 'ObjectFlag': flag}
                        for index, brex, flag in flag_entries
                    ],
                })

            restrictable_by_layer = [
                (index, rule['Brex'], {
                    entry['valueAllowed'] for entry in rule.get('value_tailoring', [])
                    if entry.get('valueForm') == 'single' and entry.get('valueTailoring') == 'restrictable'
                })
                for index, rule in entries
            ]
            restrictable_by_layer = [item for item in restrictable_by_layer if item[2]]

            for pos_lower in range(len(restrictable_by_layer)):
                for pos_higher in range(pos_lower + 1, len(restrictable_by_layer)):
                    lower_index, lower_brex, lower_set = restrictable_by_layer[pos_lower]
                    higher_index, higher_brex, higher_set = restrictable_by_layer[pos_higher]
                    extra = lower_set - higher_set
                    if extra:
                        findings.append({
                            'Category': 'RestrictableValueSetWidened',
                            'Description': (
                                f"objectPath {xpath!r} restrictable value set in "
                                f"{lower_brex!r} (layer {lower_index}) allows values not "
                                f"present in the more general layer {higher_brex!r} "
                                f"(layer {higher_index}): {sorted(extra)!r}."
                            ),
                            'Xpath': xpath,
                            'ContextRules': context_rules,
                            'LowerBrex': lower_brex,
                            'LowerLayerIndex': lower_index,
                            'HigherBrex': higher_brex,
                            'HigherLayerIndex': higher_index,
                            'ExtraValues': sorted(extra),
                        })

        return findings

    def validate(self, debug: bool = False, progress_callback: Callable[[int, int, str], None] = None,
                 sns_mode: str = "normal", remove_deleted: bool = False, deep_copy_nodes: bool = False,
                 check_sns: bool = False, check_notations: bool = False) -> dict:
        """Check xml against all brexes and dump the results into a JSon file

        Args:
            debug (bool): dump intermediate rule/error data for inspection
            progress_callback (Callable[[int, int, str], None]): optional progress
                reporter, called as `progress_callback(current, total, stage)` --
                `stage="files"` once per file in directory mode (`set_xml_dir`), and
                `stage="rules"` once per content rule checked within each document
                (forwarded to `_check_rules`). `None` (the default) reports no
                progress. Replaces the old hard `tqdm` dependency (`include_tqdm`):
                the library no longer imports `tqdm` at all, and a caller who wants
                a `tqdm` bar can drive one from the callback themselves, e.g.
                `lambda current, total, stage: bars[stage].update(1)`.
            sns_mode (str): SNS shorthand mode, one of `SNS_MODES`. Only has an
                effect when `check_sns=True`. Port of `should_check`
                (`s1kd-brexcheck.c:1038`):

                - `"normal"` (default): optional levels default to `0` / `00` /
                  `0000`, i.e. a placeholder code is only checked if the BREX
                  actually defines rules for that level.
                - `"strict"`: no shorthand — every level's code must match a
                  `snsCode` defined by the BREX, including placeholders.
                - `"unstrict"`: any code is valid at a level the BREX defines
                  no rules for, whether or not it looks like a placeholder.
            remove_deleted (bool): equivalent to `s1kd-brexcheck -^`/`--remove-deleted`;
                drop elements marked `@changeType="delete"` before checking. See
                `_remove_deleted_elements`.
            deep_copy_nodes (bool): equivalent to `-8`/`--deep-copy-nodes`; every
                content-rule violation's `Object` field holds a full recursive
                copy of the violating element (all descendants) instead of just
                its own tag and attributes. See `_node_xpath_and_copy`.
            check_sns (bool): equivalent to `s1kd-brexcheck -S`/`--sns`; check the
                object's SNS against the BREX `snsRules`. **Off by default**, matching
                `opts->check_sns` in the C original (`s1kd-brexcheck.c:1994`, gated at
                `:1376`), which reports no SNS violations for a plain invocation. When
                off, the result has no `'sns'` key at all, so `Summary`/`run_summary`
                counts and the XML report's `<sns>` node are unaffected.
            check_notations (bool): equivalent to `s1kd-brexcheck -n`/`--notations`;
                check the object's unparsed entity notations against the BREX
                `notationRuleList`. **Off by default**, matching `opts->check_notations`
                in the C original (`s1kd-brexcheck.c:2084`, `:2091`, gated at `:1383`).
                Note that when the resolved BREX chain declares no `notationRule` at
                all, opting in reports *every* unparsed entity as disallowed -- that is
                the C tool's own behaviour, which is only ever reached under `-n`.
                When off, the result has no `'notations'` key at all.

        Raises:
            ValueError: if `sns_mode` is not one of `SNS_MODES`
        """
        if sns_mode not in SNS_MODES:
            raise ValueError(f"sns_mode must be one of {SNS_MODES}, got {sns_mode!r}")
        if self._xml_dir:
            # Real extension test (was `.xml" in name`, which also matched
            # e.g. "foo.xml.bak") and no exclusion of BREX data modules --
            # s1kd-brexcheck validates BREX objects like any other, including
            # against themselves via their own `brexDmRef`/`brexref` chain.
            # Ref §3.11.
            files = [
                _ for _ in listdir(self._xml_dir)
                if _.lower().endswith(".xml") and isfile(join(self._xml_dir, _))
            ]
            had_explicit_brex_list = self._brex_list[0] is not None
            had_explicit_brex_dir_path = self._brex_dir_path[1] is True
            initial_brex_list = self._brex_list
            initial_brex_dir_path = self._brex_dir_path
            results = {}
            total_files = len(files)
            for idx, _xml in enumerate(files):
                xml_path = join(self._xml_dir, _xml)
                if self._ignore_empty and not self._is_valid_xml_file(xml_path):
                    if progress_callback is not None:
                        progress_callback(idx + 1, total_files, "files")
                    continue
                self.set_xml(xml_path)
                self._init_brex_list()
                result = self._check_rules(debug=debug, progress_callback=progress_callback, sns_mode=sns_mode,
                                            remove_deleted=remove_deleted, deep_copy_nodes=deep_copy_nodes,
                                            check_sns=check_sns, check_notations=check_notations)
                result["Summary"] = self._append_summary(result)
                results[_xml] = result
                self._brex_list = initial_brex_list if had_explicit_brex_list else (None, None)
                self._brex_dir_path = initial_brex_dir_path if had_explicit_brex_dir_path else (None, None)
                if progress_callback is not None:
                    progress_callback(idx + 1, total_files, "files")
            if debug:
                with open(clean_path(join(expanduser("~/Desktop"), f'Errors_{basename(self._xml_dir)}.json')), 'w', encoding="utf-8") as _:
                    dump(results, _, indent=4)
            return results
        else:
            if self._ignore_empty and not self._is_valid_xml_file(self._xml_path):
                result = {"Skipped": True, "Summary": "Skipped (empty or non-XML file)"}
                if debug:
                    with open(clean_path(join(expanduser("~/Desktop"), f'Errors_{basename(self._xml_path)}.json')), 'w', encoding="utf-8") as _:
                        dump(result, _, indent=4)
                return result
            self._init_brex_list()
            result = self._check_rules(debug=debug, progress_callback=progress_callback, sns_mode=sns_mode,
                                        remove_deleted=remove_deleted, deep_copy_nodes=deep_copy_nodes,
                                        check_sns=check_sns, check_notations=check_notations)
            summary = self._append_summary(result)
            result["Summary"] = summary
            if debug:
                with open(clean_path(join(expanduser("~/Desktop"), f'Errors_{basename(self._xml_path)}.json')), 'w', encoding="utf-8") as _:
                    dump(result, _, indent=4)
            return result

