from re import findall
from re import sub
from re import search
from re import escape

from os.path import join
from os.path import expanduser
from os.path import basename
from os.path import dirname
from os.path import normpath

from lxml import etree

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

from .xml_processing import delete_first_line
from .xml_processing import replace_special_characters
from .xml_processing import linearize_xml

FILEPATH = dirname(__file__)


class RefChecker():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
            xml_content = linearize_xml(xml_content)
            xml_content = sub(r'\<!ENTITY M\d{3}.*?"\>', "", xml_content)
            xml_content = sub(r"(&)(M\d{3})(;)",
                              r"~" + r"\g<2>" + r"§", xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def get_entities(self):
        xml_content = self.replace_entities()
        xml_content = etree.fromstring(xml_content)

        consumables_list = []
        spm_list = []

        for para in xml_content.xpath('//para'):
            para_text = ''.join(para.itertext())

            item_numbers = findall(r"M\d{3}", para_text)
            consumables_list.extend([(item_number, para_text.replace("~", "&").replace("§", ";")) for item_number in item_numbers])

            spm_numbers = findall(r"(SPM_\d{2}-\d{2}-\d{2}(P\d{2})?)", para_text)
            spm_list.extend([(''.join(spm_number), para_text.replace("~", "&").replace("§", ";")) for spm_number in spm_numbers])

        return consumables_list, spm_list

    def check_entities(self):
        consumables_list, spm_list = self.get_entities()

        missing_cons_ref_list = [elem for elem in consumables_list if not search("&" + elem[0] + ";", elem[1])]
        missing_spm_ref_list = [elem for elem in spm_list if not search("&" + elem[0] + ";", elem[1])]

        return missing_cons_ref_list, missing_spm_ref_list

    def write_to_excel(self):
        missing_cons_ref_list, missing_spm_ref_list = self.check_entities()

        workbook = Workbook()
        cons_sheet = workbook.active
        cons_sheet.title = "Consumables"
        spm_sheet = workbook.create_sheet(title="SPMs")

        cons_sheet.cell(row=1, column=1).value = "Entity"
        cons_sheet.cell(row=1, column=1).font = Font(bold=True, italic=True, size=16)
        cons_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        cons_sheet.cell(row=1, column=2).value = "Text"
        cons_sheet.cell(row=1, column=2).font = Font(bold=True, italic=True, size=16)
        cons_sheet.cell(row=1, column=2).alignment = Alignment(horizontal='center')

        for ind, item in enumerate(missing_cons_ref_list):
            cons_sheet.cell(row=ind + 2, column=1).value = item[0]
            cons_sheet.cell(row=ind + 2, column=2).value = item[1]

        spm_sheet.cell(row=1, column=1).value = "Entity"
        spm_sheet.cell(row=1, column=1).font = Font(bold=True, italic=True, size=16)
        spm_sheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        spm_sheet.cell(row=1, column=2).value = "Text"
        spm_sheet.cell(row=1, column=2).font = Font(bold=True, italic=True, size=16)
        spm_sheet.cell(row=1, column=2).alignment = Alignment(horizontal='center')

        for ind, item in enumerate(missing_spm_ref_list):
            spm_sheet.cell(row=ind + 2, column=1).value = item[0]
            spm_sheet.cell(row=ind + 2, column=2).value = item[1]

        workbook.save(join(self.export_path, f"reference_check_{basename(normpath(self.xml_path))}.xlsx"))

class CsnChecker():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
            xml_content = linearize_xml(xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def get_item_numbers(self):
        xml_content = self.replace_entities()

        xml_content = etree.fromstring(xml_content)
        item_numbers_list = []
        for para in xml_content.xpath('//para'):
            para_text = ''.join(para.itertext())
            sentences = para_text.split('. ')
            if sentences is None:
                sentences = [para_text]
            for sentence in sentences:
                item_numbers = findall(
                    r"\(\d{1,3}[a-zA-Z]{0,2}\-\d{1,4}[a-zA-Z]{0,2}\)", sentence)
                for item_number in item_numbers:
                    item_numbers_list.append((item_number, sentence))
        return item_numbers_list

    def write_to_excel(self):
        item_numbers_list = self.get_item_numbers()

        workbook = Workbook()
        worksheet = workbook.active

        worksheet.cell(row=1, column=1).value = "Item-Number"
        worksheet.cell(row=1, column=1).font = Font(bold=True, italic=True, size=16)
        worksheet.cell(row=1, column=1).alignment = Alignment(horizontal='center')

        worksheet.cell(row=1, column=2).value = "Text"
        worksheet.cell(row=1, column=2).font = Font(bold=True, italic=True, size=16)
        worksheet.cell(row=1, column=2).alignment = Alignment(horizontal='center')

        for ind, item in enumerate(item_numbers_list):
            worksheet.cell(row=ind + 2, column=1).value = item[0]
            worksheet.cell(row=ind + 2, column=2).value = item[1]

        workbook.save(
            join(self.export_path, f"CSN_check_{basename(normpath(self.xml_path))}.xlsx"))


class GraphicRefChecker:
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
            xml_content = linearize_xml(xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def check_graphics(self):
        xml_content = self.replace_entities()
        xml_content_bckp = xml_content
        xml_content = sub(r"\<graphic.*?\</graphic\>", "", xml_content)
        task_refids = findall(r'\<grphcref refid="graphic.*?/\>', xml_content)
        task_refids = [search(r'refid=".*?"', refid).group(0) for refid in task_refids]

        ref_ids_info = []
        keep_track_of_first_reference = []
        tree = etree.fromstring(xml_content_bckp)
        for ind, refid in enumerate(task_refids):
            if refid in keep_track_of_first_reference:
                pass
            else:
                keep_track_of_first_reference.append(refid)
                target_node = tree.xpath(f'.//*[@{task_refids[ind]}]/parent::*')[0]
                while target_node.tag != "task" and target_node.tag != "subtask":
                    target_node = target_node.getparent()
                id = search(r'(")(.*?)(")', refid).group(2)
                task_info = []
                for key, value in target_node.attrib.items():
                    pair = f"{key}: {value}"
                    task_info.append(pair)
                task_info = ', '.join(task_info)
                end_of_task = target_node.xpath(".//prclist1")[-1]
                if end_of_task.xpath(f'.//*[@key="{id}"]'):  # Checks if in the current task a node with the attribute "refid" exists
                    status = "Found at end of subtask"
                else:
                    status = "Not found at end of subtask"
                ref_ids_info.append((status, task_info, refid))
        return ref_ids_info

    def write_to_excel(self):
        ref_ids_info = self.check_graphics()

        workbook = Workbook()
        sheet = workbook.active

        sheet.cell(row=1, column=1).value = "Graphic Ref"
        sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=1).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=2).value = "First Reference"
        sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=2).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=3).value = "Status"
        sheet.cell(row=1, column=3).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=3).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=3).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        for ind, elem in enumerate(ref_ids_info, start=2):
            sheet.cell(row=ind, column=1).value = ref_ids_info[ind - 2][2]
            sheet.cell(row=ind, column=2).value = ref_ids_info[ind - 2][1]
            sheet.cell(row=ind, column=3).value = ref_ids_info[ind - 2][0]
            if ref_ids_info[ind - 2][0] == "Found at end of subtask":
                sheet.cell(row=ind, column=3).font = Font(color="65DA65")
            else:
                sheet.cell(row=ind, column=3).font = Font(color="F47174")

        workbook.save(join(
            self.export_path, f"graphic_ref_check_{basename(normpath(self.xml_path))}.xlsx"))


