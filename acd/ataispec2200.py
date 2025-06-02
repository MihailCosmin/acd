import sys

from io import StringIO

from re import findall
from re import sub

from os.path import join
from os.path import expanduser
from os.path import basename
from os.path import dirname
from os.path import normpath

from traceback import format_exc

from json import dump

from regex import search

from lxml import etree
from lxml.etree import tostring

from openpyxl import load_workbook

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill

import sys

if sys.version_info >= (3, 12):
    from PySide6.QtWidgets import QMainWindow
    from PySide6.QtCore import Signal
else:
    from PySide2.QtWidgets import QMainWindow
    from PySide2.QtCore import Signal

from .xml_processing import linearize_xml
from .xml_processing import delete_first_line
from .constants import TORQUE_VALUES_REGEX
from .constants import ITEMNUMBER_VALUES_REGEX
from .constants import NO_ITEMNUMBER_VALUES_REGEX
from .constants import OR_ITEMNUMBER_VALUES_REGEX
from .constants import ITEMDATA_REGEX
from .constants import ITEMNBR_REGEX
from .constants import PNR_REGEX
from .constants import KWD_REGEX
from .constants import ADT_REGEX
from .constants import MFR_REGEX
from .constants import IPLNOM_REGEX
from .xml_processing import replace_special_characters


FILEPATH = dirname(__file__)


class NoXmlSet(Exception):
    pass


class ConsumablesValidator():
    def __init__(self) -> None:
        self.xml_path = None
        self.xml_list = None
        self.main_dict = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function with huch the user can specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def _print_result(self):
        """Function checks the tables in the created dictionary for discrepancies and prints the results
        of the check to an xml file.
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = f"XML consumables validation for: {basename(normpath(self.xml_path))}\n"

        error_count = 0
        for key, value in self.main_dict.items():
            for key2, value2 in value.items():
                # Find amount of tasks per pageblock
                temp_task_list = []
                for task in value:
                    temp_task_list.append(task)

                # Check if con of table is in procedure
                for elem in value2['Table-con']:
                    if elem in value2['Procedure-con']:
                        pass
                    else:
                        error_count += 1
                        next_empty_row = sheet.max_row + 2
                        sheet[f"A{next_empty_row}"] = "    Consumable"
                        sheet[f"B{next_empty_row}"] = f"'{elem}'"
                        sheet[f"C{next_empty_row}"] = "from pageblock"
                        sheet[f"D{next_empty_row}"] = f"'{key}'"
                        sheet[f"E{next_empty_row}"] = "not found in the procedure of task"
                        sheet[f"F{next_empty_row}"] = f"'{key2 + 1}' of '{temp_task_list[-1] + 1}'"

                # Check if con of procedure is in table
                for elem in value2['Procedure-con']:
                    if elem in value2['Table-con']:
                        pass
                    else:
                        error_count += 1
                        next_empty_row = sheet.max_row + 2
                        sheet[f"A{next_empty_row}"] = "    Consumable"
                        sheet[f"B{next_empty_row}"] = f"'{elem}'"
                        sheet[f"C{next_empty_row}"] = "from pageblock"
                        sheet[f"D{next_empty_row}"] = f"'{key}'"
                        sheet[f"E{next_empty_row}"] = "not found in consumables table of task"
                        sheet[f"F{next_empty_row}"] = f"'{key2 + 1}' of '{temp_task_list[-1] + 1}'"

                # Check if ted of table is in procedure
                for elem in value2['Table-ted']:
                    if elem in value2['Procedure-ted']:
                        pass
                    else:
                        error_count += 1
                        next_empty_row = sheet.max_row + 2
                        sheet[f"A{next_empty_row}"] = "    Tool"
                        sheet[f"B{next_empty_row}"] = f"'{elem}'"
                        sheet[f"C{next_empty_row}"] = "from pageblock"
                        sheet[f"D{next_empty_row}"] = f"'{key}'"
                        sheet[f"E{next_empty_row}"] = "not found in consumables table of task"
                        sheet[f"F{next_empty_row}"] = f"'{key2 + 1}' of '{temp_task_list[-1] + 1}'"

                # Check if con of procedure is in table
                for elem in value2['Procedure-ted']:
                    if elem in value2['Table-ted']:
                        pass
                    else:
                        error_count += 1
                        next_empty_row = sheet.max_row + 2
                        sheet[f"A{next_empty_row}"] = "    Tool"
                        sheet[f"B{next_empty_row}"] = f"'{elem}'"
                        sheet[f"C{next_empty_row}"] = "from pageblock"
                        sheet[f"D{next_empty_row}"] = f"'{key}'"
                        sheet[f"E{next_empty_row}"] = "not found in the procedure of task"
                        sheet[f"F{next_empty_row}"] = f"'{key2 + 1}' of '{temp_task_list[-1] + 1}'"

        if error_count == 0:
            next_empty_row = sheet.max_row + 1
            sheet[f"A{next_empty_row}"] = "    No Errors detected"

        workbook.save(join(
            self.export_path, f"Validated_cons_{basename(normpath(self.xml_path))}.xlsx"))

    def validate_consumables(self, export: bool = False):
        """Traverses through every pageblock node in a xml file, counts the con and ted nodes for
        each pageblock and saves the text of the child nodes of each con and ted node.

        With optional parameters the user can decide if he wants to only get the result of
        the validation value or if he also wants a json file to be created displaying the
        gathered information containing all cons and teds from the xml, in a structured form.
        Args:
            export (bool, optional): Parameter to decide whether additional files should be
            created or not. Defaults to False.
        """
        if self.xml_path is None:
            raise NoXmlSet("No xml found to validate. Use the set_xml function to set the path of your xml \
                        before you use the validate_consumables function.".replace("                        ", ""))

        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
        xml_content = linearize_xml(xml_content)
        xml_content = delete_first_line(xml_content)
        xml_content = replace_special_characters(xml_content)
        # Insert replacements for all entitys
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        entities = linearize_xml(entities)
        xml_content = sub(r"(]>.*?<cmm)", entities + r'\1', xml_content)
        xml_content = etree.parse(StringIO(xml_content))
        all_pageblocks = xml_content.xpath(r"//pgblk")

        main_dict = {}
        table_con_list = []
        pro_con_list = []
        table_ted_list = []
        pro_ted_list = []
        for pgblk in all_pageblocks:
            main_dict[pgblk.attrib['pgblknbr']] = {}
            table_con_list.append(pgblk.attrib['pgblknbr'])

            check_task = pgblk.xpath(r".//task")
            for num, task in enumerate(check_task):
                main_dict[task.attrib['pgblknbr']][num] = {
                    'Table-con': [],
                    'Procedure-con': [],
                    'Table-ted': [],
                    'Procedure-ted': []
                }
                list_to_check_uniqueness = []

                table_con = task.xpath(
                    r"./topic/prclist1/prcitem1/prclist2[preceding-sibling::prcitem/title = 'Consumables']")
                if table_con:
                    for con in table_con[0].xpath(r".//con"):
                        connbr_text = ""
                        try:
                            connbr_text = str(con.xpath("connbr")[0].text)
                        except IndexError:
                            pass
                        conname_text = ""
                        try:
                            conname_text = str(
                                con.xpath("conname")[0].text).strip()
                        except IndexError:
                            pass
                        table_con_tuple = (connbr_text, conname_text)
                        if table_con_tuple in list_to_check_uniqueness:
                            pass
                        else:
                            list_to_check_uniqueness.append(table_con_tuple)
                            main_dict[pgblk.attrib['pgblknbr']
                                      ][num]['Table-con'].append(table_con_tuple)
                            table_con_list.append(f"{num}:{table_con_tuple}")
                    list_to_check_uniqueness.clear()

                procedure_con = task.xpath(r"./topic[title = 'Procedure']")
                procedure_con += task.xpath(r"./topic[title = 'Job Set-up']")
                if procedure_con:
                    for con in procedure_con[0].xpath(r".//con") + procedure_con[1].xpath(r".//con"):
                        connbr_text = ""
                        try:
                            connbr_text = str(con.xpath("connbr")[0].text)
                        except IndexError:
                            pass
                        try:
                            conname_text = str(
                                con.xpath("conname")[0].text).strip()
                        except IndexError:
                            pass

                        procedure_con_tuple = (connbr_text, conname_text)
                        if procedure_con_tuple in list_to_check_uniqueness:
                            pass
                        else:
                            list_to_check_uniqueness.append(
                                procedure_con_tuple)
                            main_dict[pgblk.attrib['pgblknbr']
                                      ][num]['Procedure-con'].append(procedure_con_tuple)
                            pro_con_list.append(f"{num}:{procedure_con_tuple}")
                    list_to_check_uniqueness.clear()

                table_ted = task.xpath(
                    r"./topic/prclist1/prcitem1/prclist2[preceding-sibling::prcitem/title = 'Special Tools']")
                if table_ted:
                    for ted in table_ted[0].xpath(r".//ted"):
                        toolnbr_text = ""
                        try:
                            toolnbr_text = str(ted.xpath("toolnbr")[0].text)
                        except IndexError:
                            pass
                        toolname_text = ""
                        try:
                            toolname_text = str(
                                ted.xpath("toolname")[0].text).strip()
                        except IndexError:
                            pass
                        table_ted_tuple = (toolnbr_text, toolname_text)
                        if table_ted_tuple in list_to_check_uniqueness:
                            pass
                        else:
                            list_to_check_uniqueness.append(table_ted_tuple)
                            main_dict[pgblk.attrib['pgblknbr']
                                      ][num]['Table-ted'].append(table_ted_tuple)
                            table_ted_list.append(f"{num}:{table_ted_tuple}")
                    list_to_check_uniqueness.clear()

                procedure_ted = task.xpath(r"./topic[title = 'Procedure']")
                procedure_ted += task.xpath(r"./topic[title = 'Job Set-up']")
                if procedure_ted:
                    for ted in procedure_ted[0].xpath(r".//ted") + procedure_ted[1].xpath(r".//ted"):
                        toolnbr_text = ""
                        try:
                            toolnbr_text = str(ted.xpath("toolnbr")[0].text)
                        except IndexError:
                            pass
                        toolname_text = ""
                        try:
                            toolname_text = str(
                                ted.xpath("toolname")[0].text).strip()
                        except IndexError:
                            pass
                        procedure_ted_tuple = (toolnbr_text, toolname_text)
                        if procedure_ted_tuple in list_to_check_uniqueness:
                            pass
                        else:
                            list_to_check_uniqueness.append(
                                procedure_ted_tuple)
                            main_dict[pgblk.attrib['pgblknbr']
                                      ][num]['Procedure-ted'].append(procedure_ted_tuple)
                            pro_ted_list.append(f"{num}:{procedure_ted_tuple}")
                    list_to_check_uniqueness.clear()

        if export:
            with open(join(self.export_path, f"main_dict_cons_{basename(normpath(self.xml_path))}.json"), "w", encoding="utf-8") as _:
                dump(main_dict, _, indent=4)

        self.main_dict = main_dict

        self._print_result()


class TorqueValuesValidator():
    def __init__(self) -> None:
        self.xml_path = None
        self.main_dict = None
        self.procedure_torques = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def _print_result(self):
        """Function checks the tables in the created dictionary for discrepancies and prints the results
        of the check to an xml file.
        """
        workbook = Workbook()
        sheet = workbook.active
        sheet["A1"] = f"XML torque values validation for: {basename(normpath(self.xml_path))}\n"

        error_count = 0
        # Check if Torque from table is in Procedure
        for entry in self.main_dict['Table']:
            found = False
            for entry2 in self.main_dict['Procedure']:
                if entry[0] != "-":
                    if entry[0] == entry2[0]:
                        if self.main_dict['Table'][entry] == self.main_dict['Procedure'][entry2]:
                            found = True
                            break
                else:
                    if entry[1] == entry2[1]:
                        if self.main_dict['Table'][entry] == self.main_dict['Procedure'][entry2]:
                            found = True
                            break
            if found is False:
                error_count += 1
                next_empty_row = sheet.max_row + 2
                sheet[f"A{next_empty_row}"] = "    Torque Value"
                sheet[f"B{next_empty_row}"] = f"'{entry}': '{self.main_dict['Table'][entry]}'"
                sheet[f"C{next_empty_row}"] = "from Table not found in Procedure"

        # Check if Torque from Procedure is in Table
        for entry in self.main_dict['Procedure']:
            found = False
            for entry2 in self.main_dict['Table']:
                if entry[0] != "-":
                    if entry[0] == entry2[0]:
                        if self.main_dict['Procedure'][entry] == self.main_dict['Table'][entry2]:
                            found = True
                            break
                else:
                    if entry[1] == entry2[1]:
                        if self.main_dict['Procedure'][entry] == self.main_dict['Table'][entry2]:
                            found = True
                            break
            if found is False:
                error_count += 1
                next_empty_row = sheet.max_row + 2
                sheet[f"A{next_empty_row}"] = "    Torque Value"
                sheet[f"B{next_empty_row}"] = f"'{entry}': '{self.main_dict['Procedure'][entry]}"
                sheet[f"C{next_empty_row}"] = "from Procedure not found in Table"

        if error_count == 0:
            next_empty_row = sheet.max_row + 2
            sheet[f"A{next_empty_row}"] = "    No Errors detected"

        workbook.save(join(
            self.export_path, f"Validated_tor_{basename(normpath(self.xml_path))}.xlsx"))

    def set_export_path(self, export_path: str):
        """Function with which the user can specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def validate_torque_values(
            self,
            export: bool = False,
            debug: bool = False,
            qt_window: QMainWindow = None,
            progress: Signal = Signal(0),
            console: Signal = Signal("")) -> int:
        """1. Gets the pageblock 8000 which contains the torque values from the table.
        2. Gets all strings that match the RegEx "torque the .*?.</para>". to get the
        torque values from the procedure.
        Through multiple manipulation steps we obtain the wanted information for the
        torque values and safe them into a dictionary.

        With optional parameters the user can decide if he wants to only get the result of
        the validation value or if he also wants a json file to be created displaying the
        gathered information containing all cons and teds from the xml, in a structured form.
        Args:
            export (bool, optional): _description_. Defaults to False.
        """
        try:
            with open(self.xml_path, "r", encoding="utf-8") as _:
                xml_content = _.read()
            xml_content = linearize_xml(xml_content)
            xml_content = delete_first_line(xml_content)
            xml_content_backup = xml_content
            xml_content = xml_content.replace("&nbsp;", " ")
            with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
                entities = _.read()
            entities = linearize_xml(entities)
            xml_content = sub(r"(]>.*?<cmm)", entities + r'\1', xml_content)
            xml_content = replace_special_characters(xml_content)
            xml_content = etree.parse(StringIO(xml_content))

            pageblock_8000 = xml_content.xpath(
                r"//pgblk[@pgblknbr = '8000']")  # Table Torques
            main_dict = {
                'Table': {

                },
                'Procedure': {

                }
            }
            temp = []
            para_list = []
            measuring_instruments = pageblock_8000[0].xpath(
                r"./task/topic/prclist1[preceding-sibling::title = 'Torque Values Table']")
            if measuring_instruments == []:
                measuring_instruments = pageblock_8000[0].xpath(
                    r"./task[title = 'Torque Values']")
            if measuring_instruments:
                for entry in measuring_instruments[0].xpath(r".//entry"):
                    if len(entry.xpath(r".//para")) > 1:
                        temp.clear()
                        if entry.xpath(r".//para")[0].text is not None:
                            for ind, elem in enumerate(entry.xpath(r".//para")):
                                temp.append(entry.xpath(r".//para")[ind].text)
                            multiple = temp[1]
                            para_list.append(multiple)
                        else:
                            for elem in entry.xpath(r".//csn"):
                                temp.append(elem.text)
                            multiple = ' '.join(temp)
                            para_list.append(multiple)
                    elif entry.xpath(r".//para")[0].text is None:
                        if len(entry.xpath(r".//csn")) > 1:
                            for ind, elem in enumerate(entry.xpath(r".//csn")):
                                temp.append(entry.xpath(r".//para/csn")[ind].text)
                            multiple = ' '.join(temp)
                            para_list.append(multiple)
                        else:
                            try:
                                para_list.append(entry.xpath(r".//csn")[0].text)
                            except IndexError:
                                temp_item_no = ''.join(
                                    entry.xpath(r".//para")[0].itertext())
                                temp_item_no = sub(
                                    r"( \()([0-9]+)(\))", r'\2', temp_item_no)
                                para_list.append(temp_item_no)
                    else:
                        para_list.append(entry.xpath(r".//para")[0].text)
                para_list = [
                    elem.lower() if elem is not None else elem for elem in para_list]
                para_list = para_list[3:]
                while len(para_list) != 0:
                    if para_list[0] == "":
                        para_list[0] = "-"
                    if para_list[1].endswith(' '):
                        para_list[1] = para_list[1][:-1]
                    main_dict['Table'][(
                        para_list[0], para_list[1].lower())] = para_list[2]
                    para_list = para_list[3:]

            # lxml to get all prcitem nodes and then use regex to extract the wanted information from the procedure
            xml_content_backup = xml_content_backup.replace("</csn >", "</csn>").replace(
                "<csn >", "<csn>").replace("<csn>", "(").replace("</csn>", ")")
            xml_content_backup = replace_special_characters(xml_content_backup)
            with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
                entities = _.read()
            entities = linearize_xml(entities)
            xml_content_backup = sub(
                r"(]>.*?<cmm)", entities + r'\1', xml_content_backup)
            tree = etree.parse(StringIO(xml_content_backup))
            for match in tree.xpath("./pgblk"):
                if "'pgblknbr': '15000'" in str(match.attrib):
                    match.getparent().remove(match)
            prcitems = []
            prcitems_content = ""
            prcitems += tree.findall(".//prcitem1")
            for elem in prcitems:
                prcitems_content += tostring(elem).decode("utf-8") + "\n"
            prcitems_content = replace_special_characters(prcitems_content.lower())
            prcitems_content = prcitems_content.split("\n")

            for elem in prcitems_content:
                elem = sub(r"\<.*?\>", "", elem)
                if "for pn" in elem and search(ITEMNUMBER_VALUES_REGEX, elem):
                    for match in findall(TORQUE_VALUES_REGEX, elem):
                        if search(ITEMNUMBER_VALUES_REGEX, elem):
                            item_num = search(ITEMNUMBER_VALUES_REGEX, elem).group(
                                3).replace("(", "").replace(")", "")
                            name = search(ITEMNUMBER_VALUES_REGEX, elem).group(
                                2).replace("(", "").replace(")", "")
                            torque = ''.join(match)
                            main_dict['Procedure'][(item_num, name)] = torque
                elif search(ITEMNUMBER_VALUES_REGEX, elem) and 'or' not in search(ITEMNUMBER_VALUES_REGEX, elem).group(4):
                    item_num = search(ITEMNUMBER_VALUES_REGEX, elem).group(
                        3).replace("(", "").replace(")", "")
                    name = search(ITEMNUMBER_VALUES_REGEX, elem).group(2)
                    if name.endswith(" "):
                        name = name[:-1]
                    torque = search(ITEMNUMBER_VALUES_REGEX, elem).group(5)
                    torque += search(ITEMNUMBER_VALUES_REGEX, elem).group(6)
                    torque += search(ITEMNUMBER_VALUES_REGEX, elem).group(7)
                    torque += search(ITEMNUMBER_VALUES_REGEX, elem).group(8)
                    torque += search(ITEMNUMBER_VALUES_REGEX, elem).group(9)
                    torque += search(ITEMNUMBER_VALUES_REGEX, elem).group(10)
                    torque += search(ITEMNUMBER_VALUES_REGEX, elem).group(11)
                    main_dict['Procedure'][(item_num, name)] = torque
                elif search(NO_ITEMNUMBER_VALUES_REGEX, elem):
                    item_num = search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(
                        4).replace("(", "").replace(")", "")
                    if item_num == "":
                        item_num = "-"
                    name = search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(2)
                    if name.endswith(" "):
                        name = name[:-1]
                    torque = search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(5)
                    torque += search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(6)
                    torque += search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(7)
                    torque += search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(8)
                    torque += search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(9)
                    torque += search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(10)
                    torque += search(NO_ITEMNUMBER_VALUES_REGEX, elem).group(11)
                    main_dict['Procedure'][(item_num, name)] = torque
                elif search(OR_ITEMNUMBER_VALUES_REGEX, elem):
                    item_num = search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(3).replace("(", "").replace(
                        ")", "") + " " + search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(5).replace("(", "").replace(")", "")
                    name = search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(2)
                    if name.endswith(" "):
                        name = name[:-1]
                    torque = search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(7)
                    torque += search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(8)
                    torque += search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(9)
                    torque += search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(10)
                    torque += search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(11)
                    torque += search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(12)
                    torque += search(OR_ITEMNUMBER_VALUES_REGEX, elem).group(13)
                    main_dict['Procedure'][(item_num, name)] = torque

            if export:
                with open(join(self.export_path, f"main_dict_torque_values_{basename(normpath(self.xml_path))}.json"), "w", encoding="utf-8") as _:
                    temp_dict = {key: {str(key2): value2 for key2, value2 in value.items(
                    )} for key, value in main_dict.items()}
                    dump(temp_dict, _, indent=4)

            self.main_dict = main_dict
            self.procedure_torques = main_dict['Procedure']
            self._print_result()

            self.create_torque_values_table()
            if qt_window is not None:
                console.emit(
                    "Torque Values table created successfully. See: " + join(
                        self.export_path,
                        f"Torque_Table_{basename(normpath(self.xml_path))}.xml") + "\n")
        except Exception as err:
            if qt_window is not None and debug:
                progress.emit(100)
                console.emit("Error: " + str(err) + "\n" + format_exc() + "\n")
                return 1
        return 0

    def create_torque_values_table(self):
        row = ""
        self.procedure_torques = dict(sorted(self.procedure_torques.items()))

        for ind, (key, value) in enumerate(self.procedure_torques.items(), start=2):
            number = key[0]
            for match in findall(r"\d+[a-z]?-\d+[a-z]?", number):
                number = number.replace(match, f"<csn>{match}</csn>")
            name = key[1].title()
            row += f'<row><?validrow {ind}?><entry rotate="0" valign="middle"><para chg="N" mark="1">{number}</para></entry><entry align="left" rotate="0" valign="middle"><para>{name}</para></entry><entry align="left" valign="middle"><para>{value}</para></entry></row>'

        table_content = f"<tbody>{row}</tbody>"
        with open(join(self.export_path, f"Torque_Table_{basename(normpath(self.xml_path))}.xml"), "w", encoding="utf-8") as _:
            _.write(table_content)


class cons_and_teds_checker():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")
        self.pgblk_contents = None

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def get_all_pgblks(self):
        """_summary_

        Returns:
            _type_: _description_
        """
        xml_content = self.replace_entities()
        tree = etree.fromstring(xml_content)
        pgblks = tree.xpath("//pgblk")
        pgblk_contents = [linearize_xml(etree.tostring(
            pgblk, encoding="unicode")) for pgblk in pgblks]
        # Remove Pageblocks 0, 1, 8000, 9000, 10000 from the list
        new_pgblk_contents = []
        for pgblk in pgblk_contents:
            if 'pgblknbr="0"' not in pgblk and 'pgblknbr="1"' not in pgblk and 'pgblknbr="8000"' not in pgblk and 'pgblknbr="9000"' not in pgblk and 'pgblknbr="10000"' not in pgblk and "isempty" not in pgblk:
                new_pgblk_contents.append(pgblk)
        self.pgblk_contents = new_pgblk_contents
        return new_pgblk_contents

    def collect_cons_and_teds(self):
        """_summary_

        Returns:
            _type_: _description_
        """
        pgblk_contents = self.get_all_pgblks()

        dict_cons = {}
        dict_tools = {}

        for pgblk in pgblk_contents:
            if search(r'(pgblknbr=")(\d+)(")', pgblk):
                pgblknbr = search(r'(pgblknbr=")(\d+)(")', pgblk).group(2)
            else:
                pgblknbr = None
            # Create list of tasks per pageblock
            tree = etree.fromstring(pgblk)
            tasks = tree.xpath("//task")
            tasks_content = [etree.tostring(task, encoding="unicode") for task in tasks]

            cons = []
            tools = []
            for ind, task in enumerate(tasks_content, start=1):
                task_num = ind
                # Get connbr and conname
                matches = findall(
                    r"(\<connbr\>)(.*?)(\</connbr\>\<conname\>)(.*?)(\</conname\>)", task)
                for match in matches:
                    if (match[1], match[3], task_num) not in cons:
                        cons.append((match[1], match[3], task_num))
                # Get toolnbr and toolname
                matches = findall(
                    r"(\<toolnbr\>)(.*?)(\</toolnbr\>\<toolname\>)(.*?)(\</toolname\>)", task)
                for match in matches:
                    if (match[1], match[3], task_num) not in tools:
                        tools.append((match[1], match[3], task_num))
            dict_cons[pgblknbr] = cons
            dict_tools[pgblknbr] = tools
        return dict_cons, dict_tools

    def _lookup_in_procedure(self, pgblknbr: str, nbr: str, name: str):
        pgblk_contents = self.pgblk_contents
        pgblk_to_check = ""
        for elem in pgblk_contents:
            if f'pgblknbr="{pgblknbr}"' in elem:
                pgblk_to_check = elem
        root = etree.fromstring(pgblk_to_check)
        procedures = root.xpath(
            "//topic[title[text()='Procedure'] or para[text()='Procedure']]")
        if procedures:
            for ind, procedure in enumerate(procedures):
                procedures[ind] = etree.tostring(procedure, encoding="unicode")
        for procedure in procedures:
            if nbr in procedure and name in procedure:
                return True
            elif nbr in procedure and name not in procedure:
                return (True, False)
        return False

    def _lookup_in_table(self, pgblknbr: str, nbr: str, name: str, item: str):
        pgblk_contents = self.pgblk_contents
        pgblk_to_check = ""
        for elem in pgblk_contents:
            if f'pgblknbr="{pgblknbr}"' in elem:
                pgblk_to_check = elem
        root = etree.fromstring(pgblk_to_check)
        tables = root.xpath(
            f"//prcitem1[prcitem/title[text()='{item}'] or prcitem/para[text()='{item}']]")
        if tables:
            for ind, table in enumerate(tables):
                tables[ind] = etree.tostring(table, encoding="unicode")
        for table in tables:
            if nbr in table and name in table:
                return True
            elif nbr in table and name not in table:
                return (True, False)
        return False

    def write_to_excel(self):
        """_summary_
        """
        dict_cons, dict_tools = self.collect_cons_and_teds()

        workbook = Workbook()
        sheet_cons = workbook.active
        sheet_cons.title = "Consumables"

        sheet_tools = workbook.create_sheet(title="Tools")

        sheet_columns = {
            "Consumables": {
                1: ("Pageblock", True, 16),
                2: ("Task", True, 16),
                3: ("Connbr", True, 16),
                4: ("Conname", True, 16),
                5: ("Used in Table", False, 16),
                6: ("Used in Procedure", False, 16)
            },
            "Tools": {
                1: ("Pageblock", True, 16),
                2: ("Task", True, 16),
                3: ("Toolnbr", True, 16),
                4: ("Toolname", True, 16),
                5: ("Used in Table", False, 16),
                6: ("Used in Procedure", False, 16)
            }
        }

        for sheet_name, columns in sheet_columns.items():
            for column, (header, is_bold, size) in columns.items():
                cell = workbook[sheet_name].cell(row=1, column=column)
                cell.value = header
                cell.font = Font(bold=is_bold, italic=True, size=size)

        row = 2
        for key, value in dict_cons.items():
            for elem in dict_cons[key]:
                workbook["Consumables"].cell(row=row, column=1).value = key
                workbook["Consumables"].cell(row=row, column=2).value = elem[2]
                workbook["Consumables"].cell(row=row, column=3).value = elem[0]
                workbook["Consumables"].cell(row=row, column=4).value = elem[1]
                if self._lookup_in_table(key, elem[0], elem[1], "Consumables"):
                    workbook["Consumables"].cell(
                        row=row, column=5).value = chr(0x2713)  # Check Symbol
                    workbook["Consumables"].cell(
                        row=row, column=5).font = Font(color='65DA65')
                    workbook["Consumables"].cell(
                        row=row, column=5).alignment = Alignment(horizontal='center')
                elif self._lookup_in_table(key, elem[0], elem[1], "Consumables") == (True, False):
                    workbook["Consumables"].cell(row=row, column=5).value = "!"
                    workbook["Consumables"].cell(
                        row=row, column=5).font = Font(color='FF9100')
                    workbook["Consumables"].cell(
                        row=row, column=5).alignment = Alignment(horizontal='center')
                else:
                    workbook["Consumables"].cell(
                        row=row, column=5).value = chr(0x2717)  # Cross Symbol
                    workbook["Consumables"].cell(
                        row=row, column=5).font = Font(color='F47174')
                    workbook["Consumables"].cell(
                        row=row, column=5).alignment = Alignment(horizontal='center')
                if self._lookup_in_procedure(key, elem[0], elem[1]):
                    workbook["Consumables"].cell(
                        row=row, column=6).value = chr(0x2713)  # Check Symbol
                    workbook["Consumables"].cell(
                        row=row, column=6).font = Font(color='65DA65')
                    workbook["Consumables"].cell(
                        row=row, column=6).alignment = Alignment(horizontal='center')
                elif self._lookup_in_procedure(key, elem[0], elem[1]) == (True, False):
                    workbook["Consumables"].cell(row=row, column=6).value = "!"
                    workbook["Consumables"].cell(
                        row=row, column=6).font = Font(color='FF9100')
                    workbook["Consumables"].cell(
                        row=row, column=6).alignment = Alignment(horizontal='center')
                else:
                    workbook["Consumables"].cell(
                        row=row, column=6).value = chr(0x2717)  # Cross Symbol
                    workbook["Consumables"].cell(
                        row=row, column=6).font = Font(color='F47174')
                    workbook["Consumables"].cell(
                        row=row, column=6).alignment = Alignment(horizontal='center')
                row += 1
        row = 2
        for key, value in dict_tools.items():
            for elem in dict_tools[key]:
                workbook["Tools"].cell(row=row, column=1).value = key
                workbook["Tools"].cell(row=row, column=2).value = elem[2]
                workbook["Tools"].cell(row=row, column=3).value = elem[0]
                workbook["Tools"].cell(row=row, column=4).value = elem[1]
                if self._lookup_in_table(key, elem[0], elem[1], "Special Tools"):
                    workbook["Tools"].cell(row=row, column=5).value = chr(
                        0x2713)  # Check Symbol
                    workbook["Tools"].cell(
                        row=row, column=5).font = Font(color='65DA65')
                    workbook["Tools"].cell(row=row, column=5).alignment = Alignment(
                        horizontal='center')
                elif self._lookup_in_table(key, elem[0], elem[1], "Special Tools") == (True, False):
                    workbook["Tools"].cell(row=row, column=5).value = "!"
                    workbook["Tools"].cell(
                        row=row, column=5).font = Font(color='FF9100')
                    workbook["Tools"].cell(row=row, column=5).alignment = Alignment(
                        horizontal='center')
                else:
                    workbook["Tools"].cell(row=row, column=5).value = chr(
                        0x2717)  # Cross Symbol
                    workbook["Tools"].cell(
                        row=row, column=5).font = Font(color='F47174')
                    workbook["Tools"].cell(row=row, column=5).alignment = Alignment(
                        horizontal='center')
                if self._lookup_in_procedure(key, elem[0], elem[1]):
                    workbook["Tools"].cell(row=row, column=6).value = chr(
                        0x2713)  # Check Symbol
                    workbook["Tools"].cell(
                        row=row, column=6).font = Font(color='65DA65')
                    workbook["Tools"].cell(row=row, column=6).alignment = Alignment(
                        horizontal='center')
                elif self._lookup_in_procedure(key, elem[0], elem[1]) == (True, False):
                    workbook["Tools"].cell(row=row, column=6).value = "!"
                    workbook["Tools"].cell(
                        row=row, column=6).font = Font(color='FF9100')
                    workbook["Tools"].cell(row=row, column=6).alignment = Alignment(
                        horizontal='center')
                else:
                    workbook["Tools"].cell(row=row, column=6).value = chr(
                        0x2717)  # Cross Symbol
                    workbook["Tools"].cell(
                        row=row, column=6).font = Font(color='F47174')
                    workbook["Tools"].cell(row=row, column=6).alignment = Alignment(
                        horizontal='center')
                row += 1

        workbook.save(
            join(self.export_path, f"Checker_ted_pbs_{basename(normpath(self.xml_path))}.xlsx"))

def ipl_to_dict(xml: str) -> dict:
    """Converts an IPL XML file to a dictionary.

    Args:
        xml (str): Path to the IPL XML file.

    Returns:
        dict: Dictionary containing the IPL data.

    """
    if xml.endswith(".xlsx"):
        return _ipl_to_dict_excel(xml)
    ipl_dict = {}

    with open(xml, "r", encoding="utf-8") as xml_in:
        content = xml_in.read().replace("\n", "")

    if search(ITEMDATA_REGEX, content):
        for itemdata in findall(ITEMDATA_REGEX, content):
            itemdata = "".join(itemdata)
            itemnbr = ""
            if search(ITEMNBR_REGEX, itemdata):
                itemnbr = search(ITEMNBR_REGEX, itemdata).group(1)
            if search(PNR_REGEX, itemdata):
                ipl_dict[search(PNR_REGEX, itemdata).group(1)] = {"Nomenclature": "", "MFR": ""}
                pnr = search(PNR_REGEX, itemdata).group(1)
                if search(IPLNOM_REGEX, itemdata):
                    if search(ADT_REGEX, itemdata):
                        if "ASS" in search(ADT_REGEX, itemdata).group(1):
                            ipl_dict[pnr]["Nomenclature"] = ipl_dict[pnr]["Nomenclature"] + search(KWD_REGEX, itemdata).group(1) + " "
                            ipl_dict[pnr]["Nomenclature"] = ipl_dict[pnr]["Nomenclature"] + search(ADT_REGEX, itemdata).group(1)
                        else:
                            ipl_dict[pnr]["Nomenclature"] = ipl_dict[pnr]["Nomenclature"] + search(ADT_REGEX, itemdata).group(1) + " "
                            ipl_dict[pnr]["Nomenclature"] = ipl_dict[pnr]["Nomenclature"] + search(KWD_REGEX, itemdata).group(1)

                        ipl_dict[pnr]["Nomenclature"] = ipl_dict[pnr]["Nomenclature"].strip()
                    else:
                        ipl_dict[pnr]["Nomenclature"] = ipl_dict[pnr]["Nomenclature"] + search(KWD_REGEX, itemdata).group(1)
                    if search(MFR_REGEX, itemdata):
                        ipl_dict[pnr]["MFR"] = ipl_dict[pnr]["MFR"] + search(MFR_REGEX, itemdata).group(1)
                ipl_dict[pnr]["Itemnbr"] = itemnbr

    return ipl_dict

def _ipl_to_dict_excel(excel: str) -> dict:
    """Converts an IPL Excel file to a dictionary.

    Args:
        excel (str): Path to the IPL Excel file.

    Returns:
        dict: Dictionary containing the IPL data.

    """
    ipl_dict = {}

    workbook = load_workbook(excel)
    sheet = workbook.active

    for row in sheet.iter_rows(min_row=2, min_col=2, max_col=5, values_only=True):
        if row[0] is not None and row[2] is not None:
            itemvar = row[1] if row[1] is not None else ""
            pn = row[2]
            ipl_dict[pn] = {"Nomenclature": row[3], "MFR": "", "Itemnbr": row[0] + itemvar}
        elif row[3] is not None:
            ipl_dict[pn]["Nomenclature"] += " " + row[3].strip()
    for value in ipl_dict.values():
        if "(" in value["Nomenclature"] and ")" in value["Nomenclature"]:
            value["MFR"] = value["Nomenclature"].split("(")[1].split(")")[0].strip()
        else:
            value["MFR"] = ""
        value["Nomenclature"] = value["Nomenclature"].split("(")[0].strip().replace(".", "")

    return ipl_dict

class pgblk_9000_ted_checker():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def prepare_xml(self) -> str:
        """The prepare_xml() function reads an XML file and performs
        a series of preprocessing steps on it.
        Returns:
            str: It returns the resulting XML content as a string.
        """
        with open(self.xml_path, "r", encoding="utf-8") as _:
            xml_content = _.read()
            xml_content = delete_first_line(xml_content)
            xml_content = replace_special_characters(xml_content)
        return xml_content

    def replace_entities(self) -> str:
        """Adds content of inmedISOEntities to the xml to handle special characters
        for further lxml processing.

        Returns:
            str: xml content with mentioned modifications.
        """
        xml_content = self.prepare_xml()
        with open(join(FILEPATH, "inmedISOEntities.ent"), "r", encoding="utf-8") as _:
            entities = _.read()
        xml_content = sub(r"(]>(.|\n)*?<cmm)", entities + r'\1', xml_content)
        return xml_content

    def get_all_pgblks(self):
        """_summary_

        Returns:
            _type_: _description_
        """
        xml_content = self.replace_entities()
        tree = etree.fromstring(xml_content)
        pgblks = tree.xpath("//pgblk")
        pgblk_contents = [linearize_xml(etree.tostring(
            pgblk, encoding="unicode")) for pgblk in pgblks]
        # Remove Pageblocks 0, 1, 8000, 9000, 10000 from the list
        new_pgblk_contents = []
        for pgblk in pgblk_contents:
            if 'pgblknbr="0"' not in pgblk[:500] and 'pgblknbr="1"' not in pgblk[:500] and 'pgblknbr="8000"' not in pgblk[:500] and 'pgblknbr="10000"' not in pgblk[:500] and "isempty" not in pgblk:
                new_pgblk_contents.append(pgblk)
        self.pgblk_contents = new_pgblk_contents
        return new_pgblk_contents

    def get_tools_from_pgblk_9000(self):
        pgblk_contents = self.get_all_pgblks()

        # Get toolnbr and toolname from table in pageblock 9000
        pgblk_9000_content = ""
        dict_tools = {}
        for pgblk in pgblk_contents:
            if 'pgblknbr="9000"' in pgblk:
                pgblk_9000_content = pgblk

        root = etree.fromstring(pgblk_9000_content)
        tool_table = root.xpath(
            "//topic[title[text()='List of Special Tools, Fixtures and Equipment'] or para[text()='List of Special Tools, Fixtures and Equipment']]")
        tool_table = etree.tostring(tool_table[0], encoding="unicode")

        matches = findall(
            r"(\<toolnbr\>)(.*?)(\</toolnbr\>\<toolname\>)(.*?)(\</toolname\>)", tool_table)
        for match in matches:
            dict_tools[match[1]] = match[3]

        # Get toolnbr and toolname from procedure of all pageblocks except 9000
        pgblk_contents_without_9000 = []
        dict_tools_procedure = {}

        for pgblk in pgblk_contents:
            if 'pgblknbr="9000"' not in pgblk:
                pgblk_contents_without_9000.append(pgblk)

        for pgblk in pgblk_contents_without_9000:
            root = etree.fromstring(pgblk)
            tool_proc = root.xpath(
                "//topic[title[text()='Procedure'] or para[text()='Procedure']]")
            for proc in tool_proc:
                tools = etree.tostring(proc, encoding="unicode")
                matches = findall(
                    r"(\<toolnbr\>)(.*?)(\</toolnbr\>\<toolname\>)(.*?)(\</toolname\>)", tools)
                for match in matches:
                    dict_tools_procedure[match[1]] = match[3]
        return dict_tools, dict_tools_procedure

    def write_to_excel(
            self,
            debug: bool = False,
            qt_window: QMainWindow = None,
            progress: Signal = Signal(0),
            console: Signal = Signal("")) -> int:

        dict_tools, dict_tools_procedure = self.get_tools_from_pgblk_9000()

        workbook = Workbook()
        sheet_cons = workbook.active
        sheet_cons.title = "Special Tools Table"

        sheet_tools = workbook.create_sheet(title="Procedure")

        sheet_columns = {
            "Special Tools Table": {
                1: ("Toolnbr", True, 16),
                2: ("Toolname", True, 16),
                3: ("Used in Procedure", False, 16)
            },
            "Procedure": {
                1: ("Toolnbr", True, 16),
                2: ("Toolname", True, 16),
                3: ("Used in Table", False, 16)
            }
        }

        for sheet_name, columns in sheet_columns.items():
            for column, (header, is_bold, size) in columns.items():
                cell = workbook[sheet_name].cell(row=1, column=column)
                cell.value = header
                cell.font = Font(bold=is_bold, italic=True, size=size)
        row = 2
        for key, value in dict_tools.items():
            workbook["Special Tools Table"].cell(row=row, column=1).value = key
            workbook["Special Tools Table"].cell(
                row=row, column=2).value = value
            if key in dict_tools_procedure and value in dict_tools_procedure.values():
                workbook["Special Tools Table"].cell(row=row, column=3).value = chr(
                    0x2713)  # Check Symbol
                workbook["Special Tools Table"].cell(
                    row=row, column=3).font = Font(color='65DA65')
                workbook["Special Tools Table"].cell(row=row, column=3).alignment = Alignment(
                    horizontal='center')
            else:
                workbook["Special Tools Table"].cell(row=row, column=3).value = chr(
                    0x2717)  # Cross Symbol
                workbook["Special Tools Table"].cell(
                    row=row, column=3).font = Font(color='F47174')
                workbook["Special Tools Table"].cell(row=row, column=3).alignment = Alignment(
                    horizontal='center')
            row += 1
        row = 2
        for key, value in dict_tools_procedure.items():
            workbook["Procedure"].cell(row=row, column=1).value = key
            workbook["Procedure"].cell(row=row, column=2).value = value
            if key in dict_tools and value in dict_tools.values():
                workbook["Procedure"].cell(row=row, column=3).value = chr(
                    0x2713)  # Check Symbol
                workbook["Procedure"].cell(
                    row=row, column=3).font = Font(color='65DA65')
                workbook["Procedure"].cell(row=row, column=3).alignment = Alignment(
                    horizontal='center')
            else:
                workbook["Procedure"].cell(row=row, column=3).value = chr(
                    0x2717)  # Cross Symbol
                workbook["Procedure"].cell(
                    row=row, column=3).font = Font(color='F47174')
                workbook["Procedure"].cell(row=row, column=3).alignment = Alignment(
                    horizontal='center')
            row += 1

        workbook.save(
            join(self.export_path, f"Checker_ted_pb9000_{basename(normpath(self.xml_path))}.xlsx"))

class AtaNumbering():
    def __init__(self) -> None:
        self.xml_path = None
        self.export_path = expanduser("~/Desktop")

    def set_xml(self, xml_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            xml_path (str): xml file path.
        """
        self.xml_path = xml_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def check_key_numbers(self):
        with open(self.xml_path, 'r', encoding="utf-8") as _:
            xml_content = _.read()

        # get key_number
        key_number = self.xml_path.split('-')
        key_number = '-'.join(key_number[3:6])
        key_number = key_number.lower()
        print(key_number)
        if "rm" in key_number:
            while len(key_number) > 10:
                key_number = key_number[:-1]
        else:
            while len(key_number) > 8:
                key_number = key_number[:-1]

        validations = []
        xml_content_lines = xml_content.split('\n')
        for ind, line in enumerate(xml_content_lines):
            if search(r'((key|id)=")(.*?)(")', line):
                key_number_of_line = search(r'((key|id)=")(.*?)(")', line).group(3).lower()
                if key_number not in key_number_of_line:
                    validations.append((ind + 1, key_number_of_line, "Failed"))
                else:
                    validations.append((ind + 1, key_number_of_line, "Passed"))
        return validations

    def write_to_excel(self):
        validations = self.check_key_numbers()

        workbook = Workbook()
        sheet = workbook.active

        sheet.cell(row=1, column=1).value = "Line"
        sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=1).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=2).value = "Found Key"
        sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=2).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        sheet.cell(row=1, column=3).value = "Status"
        sheet.cell(row=1, column=3).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=3).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=3).fill = PatternFill(
            start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

        for ind, elem in enumerate(validations, start=2):
            sheet.cell(row=ind, column=1).value = validations[ind - 2][0]
            sheet.cell(row=ind, column=2).value = validations[ind - 2][1]
            sheet.cell(row=ind, column=3).value = validations[ind - 2][2]
            if validations[ind - 2][2] == "Passed":
                sheet.cell(row=ind, column=3).font = Font(color="65DA65")
            else:
                sheet.cell(row=ind, column=3).font = Font(color="F47174")

        error_count = 0
        for cell in sheet['C']:
            if cell.value == "Failed":
                error_count += 1

        sheet.cell(row=1, column=5).value = f"Error Count: {error_count}"
        sheet.cell(row=1, column=5).font = Font(
            bold=True, italic=True, size=16)
        sheet.cell(row=1, column=5).alignment = Alignment(
            horizontal='center')
        sheet.cell(row=1, column=5).fill = PatternFill(
            start_color='F47174', end_color='F47174', fill_type='solid')

        workbook.save(join(
            self.export_path, f"keys_check_{basename(normpath(self.xml_path))}.xlsx"))


if __name__ == "__main__":
    # instance = ConsumablesValidator()
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-BD50-32-21-03_007-00_EN.xml")
    # instance.validate_consumables()

    # instance = TorqueValuesValidator()
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-BD50-32-21-03_007-00_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CRM-D9893-AR21-32-10-02RM_EN.xml")              # Doesnt Work
    # instance.set_xml(r"C:\Users\bakalarz\Downloads\CRM-D9893-AR21-32-10-02RM_EN.xml")                                 # Doesnt Work
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-GE01-32-51-25_009-01_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AR21-32-10-03_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AB06-27-82-23_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AB05-27-82-12_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AW39-32-13-01_005-01_EN.xml")
    # instance.set_xml(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CRM-D9893-BD50-32-21-03RM_004-01_EN.xml")
    # instance.validate_torque_values()

    instance = cons_and_teds_checker()
    instance.set_xml(
        r"D:\CMM Automation\REWORK\CMM\2.1 rework\CMM-D9893-C091-32-11-21_000-01_EN.xml")
    instance.set_export_path(r"D:\CMM Automation\REWORK\CMM\2.1 rework")
    instance.write_to_excel()

    # instance = pgblk_9000_ted_checker()
    # instance.set_xml(
    #     # r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AR21-32-10-03_EN.xml")
    #     r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-AB05-27-82-12_EN.xml")
    # instance.set_export_path(r"C:\Users\bakalarz\Desktop\Export")
    # instance.write_to_excel()

    # instance = AtaNumbering()
    # instance.set_xml(
    #     r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CRM-D9893-AR21-32-10-02RM_EN.xml")
    # # instance.set_xml(
    # #     r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-BD50-32-21-03_007-00_EN.xml")
    # instance.set_export_path(r"C:\Users\bakalarz\Desktop\Export")
    # instance.write_to_excel()
    # print(_ipl_to_dict_excel(r"D:\Automation\Illu Automation\CMM\322131DPLIST_UK.XML"))
    # print(ipl_to_dict(r"D:\CMM Automation\REWORK\ILLU\3D file and list for EFW\A320A321_IPC_GM_253101_098X_R12_AUG_15_23.xlsx"))
