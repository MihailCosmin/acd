"""Shared presentation layer for the formatted, human-readable check reports.

`BrexChecker` and `XmlChecker` each answer a different question -- does this
object obey its BREX, is this object even well-formed and schema-valid -- but
a reviewer reads both the same way and should not have to learn two layouts.
This module owns the parts that make the two look like one product:

- `REPORT_PALETTE` -- the colours, as openpyxl ARGB-style hex with no leading
  '#'. `REPORT_HTML_CSS` mirrors them, so a workbook and an HTML report of the
  same run read as the same document.
- `ExcelReport` -- a workbook with a styled Summary sheet, plus `write_table`
  for the header/body/tint/width/autofilter/freeze treatment every sheet uses.
- `HtmlReport` -- one self-contained HTML document: inline CSS and JS, no
  webfont, image or CDN reference, so it opens identically off a network
  share, an e-mail attachment or a `file://` path.

Nothing here knows what a violation, a BREX or a schema is. The callers build
render-ready rows -- lists of dicts keyed by column label -- and hand them
over; that split is what lets a second checker reuse the look without
inheriting the first one's vocabulary.
"""

from datetime import datetime
from html import escape as html_escape

from .filepath import clean_path

# Shared look of the two formatted reports: `ExcelReport` writes these as
# openpyxl `PatternFill`/`Font` colours (ARGB-style hex with no leading '#'),
# `REPORT_HTML_CSS` mirrors them as CSS custom properties.
REPORT_PALETTE = {
    "header": "1F4E79",   # table header band / titles
    "grid": "D9D9D9",     # cell borders
    "error": "FCE4E4",    # failing rows
    "warning": "FFF2CC",  # non-failing (warning) rows
    "ok": "E2F0D9",       # passing documents
    "band": "F5F7FA",     # zebra banding
    "muted": "808080",    # secondary text / informational tabs
}

# Inline stylesheet of every HTML report. Deliberately self-contained -- no
# webfont, image or CDN reference -- so the report opens identically off a
# network share, an e-mail attachment or a `file://` path. The full light
# palette is defined on bare `:root`; the dark one is redefined twice, once
# under `prefers-color-scheme` (guarded so an explicit light choice still wins)
# and once under `[data-theme="dark"]` (so the toggle wins in both directions).
REPORT_HTML_CSS = """
:root {
  color-scheme: light dark;
  --bg: #f4f6f8; --panel: #ffffff; --ink: #171a1f; --muted: #5c6672;
  --line: #e3e7ec; --accent: #1f4e79; --accent-ink: #ffffff;
  --error-bg: #fdecec; --error-ink: #9d2933;
  --warn-bg: #fff5df; --warn-ink: #8a6100;
  --ok-bg: #e7f4ea; --ok-ink: #1d6b34;
  --code-bg: #f2f4f7;
}
@media (prefers-color-scheme: dark) {
  :root:not([data-theme="light"]) {
    --bg: #101319; --panel: #181d25; --ink: #e6eaf0; --muted: #98a2b3;
    --line: #2a313c; --accent: #7fb3e3; --accent-ink: #0d1117;
    --error-bg: #3a1e22; --error-ink: #ff9ba2;
    --warn-bg: #3a2f16; --warn-ink: #f2c66b;
    --ok-bg: #16301f; --ok-ink: #86d99b;
    --code-bg: #11151c;
  }
}
:root[data-theme="dark"] {
  --bg: #101319; --panel: #181d25; --ink: #e6eaf0; --muted: #98a2b3;
  --line: #2a313c; --accent: #7fb3e3; --accent-ink: #0d1117;
  --error-bg: #3a1e22; --error-ink: #ff9ba2;
  --warn-bg: #3a2f16; --warn-ink: #f2c66b;
  --ok-bg: #16301f; --ok-ink: #86d99b;
  --code-bg: #11151c;
}
* { box-sizing: border-box; }
body {
  margin: 0; padding: 24px; background: var(--bg); color: var(--ink);
  font: 14px/1.5 "Segoe UI", system-ui, -apple-system, Roboto, Arial, sans-serif;
}
h1 { margin: 0 0 4px; font-size: 22px; letter-spacing: -0.01em; }
p { margin: 0; }
.muted { color: var(--muted); font-size: 12px; }
.mono, code, pre {
  font-family: "Cascadia Mono", Consolas, "SF Mono", Menlo, monospace;
  font-size: 12px;
}
.page-head {
  display: flex; align-items: flex-start; justify-content: space-between;
  gap: 16px; flex-wrap: wrap; margin-bottom: 18px;
  border-bottom: 3px solid var(--accent); padding-bottom: 12px;
}
#theme-toggle {
  display: inline-flex; align-items: center; gap: 8px; cursor: pointer;
  background: var(--panel); color: var(--ink); border: 1px solid var(--line);
  border-radius: 999px; padding: 7px 14px; font: inherit; font-size: 13px;
}
#theme-toggle:hover { border-color: var(--accent); }
.theme-icon { width: 12px; height: 12px; border-radius: 50%;
  background: linear-gradient(135deg, var(--accent) 50%, transparent 50%);
  border: 1px solid var(--accent); }
.cards { display: flex; flex-wrap: wrap; gap: 12px; margin-bottom: 16px; }
.card {
  flex: 1 1 132px; background: var(--panel); border: 1px solid var(--line);
  border-left: 4px solid var(--muted); border-radius: 8px; padding: 12px 14px;
  display: flex; flex-direction: column; gap: 2px;
}
.card-value { font-size: 26px; font-weight: 650; line-height: 1.1; }
.card-label { color: var(--muted); font-size: 12px; text-transform: uppercase;
  letter-spacing: 0.04em; }
.card.error { border-left-color: var(--error-ink); background: var(--error-bg); }
.card.error .card-value { color: var(--error-ink); }
.card.warning { border-left-color: var(--warn-ink); background: var(--warn-bg); }
.card.warning .card-value { color: var(--warn-ink); }
.card.ok { border-left-color: var(--ok-ink); background: var(--ok-bg); }
.card.ok .card-value { color: var(--ok-ink); }
.card.neutral { border-left-color: var(--accent); }
.chips { display: flex; flex-wrap: wrap; gap: 8px; margin-bottom: 18px; }
.chip {
  display: inline-flex; align-items: center; gap: 8px; background: var(--panel);
  border: 1px solid var(--line); border-radius: 999px; padding: 4px 12px;
  font-size: 12px;
}
.chip-key { color: var(--muted); }
.chip-value { font-weight: 650; }
.section {
  background: var(--panel); border: 1px solid var(--line); border-radius: 8px;
  margin-bottom: 16px; overflow: hidden;
}
.section > summary {
  cursor: pointer; padding: 12px 16px; display: flex; align-items: center;
  gap: 10px; font-weight: 650; list-style: none;
}
.section > summary::-webkit-details-marker { display: none; }
.section > summary::before {
  content: "\\25B8"; color: var(--muted); transition: transform .15s ease;
}
.section[open] > summary::before { transform: rotate(90deg); }
.section-title { font-size: 15px; }
.pill {
  background: var(--accent); color: var(--accent-ink); border-radius: 999px;
  padding: 1px 9px; font-size: 12px; font-weight: 650;
}
.controls {
  display: flex; flex-wrap: wrap; gap: 12px; align-items: center;
  padding: 0 16px 12px;
}
.controls input[type="search"] {
  flex: 1 1 260px; padding: 7px 11px; border-radius: 6px; font: inherit;
  border: 1px solid var(--line); background: var(--bg); color: var(--ink);
}
.switch { display: inline-flex; align-items: center; gap: 6px; font-size: 13px;
  color: var(--muted); cursor: pointer; }
.table-wrap { overflow-x: auto; border-top: 1px solid var(--line); }
table.grid { border-collapse: collapse; width: 100%; font-size: 13px; }
table.grid th {
  position: sticky; top: 0; z-index: 1; text-align: left; white-space: nowrap;
  background: var(--accent); color: var(--accent-ink); font-weight: 650;
  padding: 9px 12px;
}
table.grid td {
  padding: 8px 12px; border-bottom: 1px solid var(--line);
  vertical-align: top; word-break: break-word;
}
table.grid tbody tr:nth-child(even) { background: color-mix(in srgb, var(--bg) 55%, transparent); }
table.grid td.num { text-align: right; white-space: nowrap; }
table.grid td.details { min-width: 240px; }
table.grid td.details p { margin: 0 0 4px; }
.finding { color: var(--error-ink); }
.status {
  display: inline-block; border-radius: 4px; padding: 1px 8px; font-size: 12px;
  font-weight: 650; white-space: nowrap;
}
.status.error, .status.failed { background: var(--error-bg); color: var(--error-ink); }
.status.warning { background: var(--warn-bg); color: var(--warn-ink); }
.status.passed { background: var(--ok-bg); color: var(--ok-ink); }
.status.skipped { background: var(--code-bg); color: var(--muted); }
code { background: var(--code-bg); border-radius: 4px; padding: 1px 5px; }
pre {
  background: var(--code-bg); border: 1px solid var(--line); border-radius: 6px;
  padding: 10px; overflow-x: auto; margin: 6px 0 0; white-space: pre-wrap;
}
.details details > summary { cursor: pointer; color: var(--muted); font-size: 12px; }
.empty {
  background: var(--panel); border: 1px solid var(--line); border-radius: 8px;
  padding: 32px; text-align: center; color: var(--ok-ink);
}
.empty-mark { font-size: 32px; display: block; margin-bottom: 8px; }
"""

# Inline behaviour of every HTML report: the dark/light override on top of the
# reader's `prefers-color-scheme` (remembered per browser, every storage access
# guarded so a `file://` document with site data blocked still renders), and a
# live filter over the report's main table.
#
# The filter half is keyed off `data-` attributes rather than fixed ids, so a
# report whose main table is a list of findings rather than violations reuses
# it unchanged: mark the table `data-filterable` (optionally `data-noun`, for
# the "N of M <noun>" counter) and the controls `data-filter-input`,
# `data-filter-errors-only` and `data-filter-count`. Rows opt into the
# errors-only switch by carrying `data-status="error"`.
REPORT_HTML_JS = """
(function () {
  var root = document.documentElement;
  var toggle = document.getElementById('theme-toggle');
  function stored(key, value) {
    try {
      if (value === undefined) { return window.localStorage.getItem(key); }
      window.localStorage.setItem(key, value);
    } catch (e) { /* private window, file:// with site data blocked, ... */ }
    return null;
  }
  function currentlyDark() {
    if (root.dataset.theme) { return root.dataset.theme === 'dark'; }
    return window.matchMedia('(prefers-color-scheme: dark)').matches;
  }
  var saved = stored('acd-report-theme');
  if (saved === 'dark' || saved === 'light') { root.dataset.theme = saved; }
  if (toggle) {
    toggle.addEventListener('click', function () {
      var next = currentlyDark() ? 'light' : 'dark';
      root.dataset.theme = next;
      stored('acd-report-theme', next);
    });
  }

  var table = document.querySelector('table[data-filterable]');
  if (!table) { return; }
  var noun = table.dataset.noun || 'rows';
  var filter = document.querySelector('[data-filter-input]');
  var errorsOnly = document.querySelector('[data-filter-errors-only]');
  var counter = document.querySelector('[data-filter-count]');
  var rows = Array.prototype.slice.call(table.tBodies[0].rows);
  function apply() {
    var needle = (filter && filter.value || '').toLowerCase();
    var only = errorsOnly && errorsOnly.checked;
    var shown = 0;
    rows.forEach(function (row) {
      var hide = (only && row.dataset.status !== 'error') ||
        (needle && row.textContent.toLowerCase().indexOf(needle) === -1);
      row.hidden = hide;
      if (!hide) { shown++; }
    });
    if (counter) {
      counter.textContent = shown === rows.length
        ? rows.length + ' ' + noun
        : shown + ' of ' + rows.length + ' ' + noun;
    }
  }
  if (filter) { filter.addEventListener('input', apply); }
  if (errorsOnly) { errorsOnly.addEventListener('change', apply); }
  apply();
})();
"""


def excel_value(value):
    """Excel-safe cell value: control characters openpyxl refuses are stripped,
    and an over-long snippet is truncated well inside the 32767-character cell
    limit.

    Args:
        value: any cell value; numbers, booleans and `None` pass through

    Returns:
        the value as openpyxl will accept it
    """
    if value is None or isinstance(value, (int, float, bool)):
        return value
    text = str(value)
    text = ''.join(
        char for char in text
        if char in '\t\n\r' or ord(char) >= 32
    )
    if len(text) > 2000:
        text = text[:2000] + ' [...]'
    return text


class ExcelReport():
    """A styled openpyxl workbook: a Summary sheet with a title block, and
    `write_table` for the one table treatment every sheet shares (coloured
    bold frozen header, autofilter, borders, wrapped top-aligned cells, tuned
    widths, zebra banding, and rows tinted by outcome).

    Construction opens the workbook and writes the Summary sheet's heading;
    callers then fill Summary with their own blocks and add further sheets.

    Args:
        title (str): report heading, e.g. "BREX check report"
        source (str): what was checked -- a directory in batch mode, otherwise
            the single object's path -- shown under the heading with a
            generation timestamp

    Raises:
        ImportError: if `openpyxl` is not installed
    """

    def __init__(self, title: str, source: str):
        try:
            # Imported here rather than at module scope: openpyxl is only
            # needed by this one report format, and importing a checker must
            # stay cheap (see the lazy `acd/__init__.py`).
            from openpyxl import Workbook
            from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
            from openpyxl.utils import get_column_letter
        except ImportError as exc:  # pragma: no cover - depends on the environment
            raise ImportError(
                "the Excel report needs openpyxl (pip install openpyxl); "
                "use the JSON, XML or HTML report otherwise."
            ) from exc

        self._get_column_letter = get_column_letter
        self.palette = REPORT_PALETTE

        thin = Side(style="thin", color=self.palette["grid"])
        self.cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill("solid", fgColor=self.palette["header"])
        self.title_font = Font(bold=True, size=16, color=self.palette["header"])
        self.muted_font = Font(size=9, color=self.palette["muted"])
        self.label_font = Font(bold=True, size=11, color=self.palette["header"])
        self.bold = Font(bold=True)
        self.top_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
        self.header_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        self.fills = {
            key: PatternFill("solid", fgColor=self.palette[key])
            for key in ("error", "warning", "ok", "band")
        }

        self.workbook = Workbook()
        self.summary = self.workbook.active
        self.summary.title = "Summary"
        self.summary.sheet_properties.tabColor = self.palette["header"]
        self.summary.sheet_view.showGridLines = False

        self.summary.merge_cells("A1:D1")
        self.summary["A1"] = title
        self.summary["A1"].font = self.title_font
        self.summary.merge_cells("A2:D2")
        self.summary["A2"] = (
            f"{source}  --  generated "
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        )
        self.summary["A2"].font = self.muted_font
        self.summary.column_dimensions["A"].width = 34
        self.summary.column_dimensions["B"].width = 16

    def add_sheet(self, name: str, tab_color: str):
        """Append a worksheet with a coloured tab.

        Args:
            name (str): sheet name
            tab_color (str): a `REPORT_PALETTE` value

        Returns:
            the new worksheet
        """
        worksheet = self.workbook.create_sheet(name)
        worksheet.sheet_properties.tabColor = tab_color
        return worksheet

    def write_table(self, worksheet, headers: list, rows: list, widths: list,
                    tint=None, start_row: int = 1, autofilter: bool = True,
                    freeze: bool = True) -> int:
        """Write one styled table -- header row plus body -- and return the
        next free row.

        Args:
            worksheet: destination worksheet
            headers (list): column labels, also the keys read from each row
            rows (list): dicts keyed by column label
            widths (list): column widths, positionally matched to `headers`
            tint (callable): optional `row -> palette key or None`, tinting a
                whole row by outcome; rows it returns `None` for fall back to
                zebra banding
            start_row (int): 1-based row to write the header on
            autofilter (bool): put an autofilter over the table
            freeze (bool): freeze everything above the first body row

        Returns:
            int: the first free row after the table, plus one blank spacer row
        """
        for column, header in enumerate(headers, start=1):
            cell = worksheet.cell(row=start_row, column=column, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.cell_border
            cell.alignment = self.header_align
        worksheet.row_dimensions[start_row].height = 26
        for offset, row in enumerate(rows):
            row_number = start_row + 1 + offset
            fill = self.fills.get(tint(row)) if tint else None
            if fill is None and offset % 2:
                fill = self.fills["band"]
            for column, header in enumerate(headers, start=1):
                cell = worksheet.cell(
                    row=row_number, column=column, value=excel_value(row.get(header))
                )
                cell.border = self.cell_border
                cell.alignment = self.top_left
                if fill is not None:
                    cell.fill = fill
        for column, width in enumerate(widths, start=1):
            worksheet.column_dimensions[self._get_column_letter(column)].width = width
        last_row = start_row + len(rows)
        if autofilter:
            worksheet.auto_filter.ref = (
                f"A{start_row}:{self._get_column_letter(len(headers))}"
                f"{max(last_row, start_row)}"
            )
        if freeze:
            worksheet.freeze_panes = worksheet.cell(row=start_row + 1, column=1)
        return last_row + 2

    def write_totals(self, row: int, entries: list, heading: str = "Run totals") -> int:
        """Write a label/value block on the Summary sheet -- the run totals --
        tinting a row by the palette key its entry names when the value is
        non-zero.

        Args:
            row (int): 1-based row to write the heading on
            entries (list): `(label, value, palette key or None)` triples
            heading (str): block heading

        Returns:
            int: the first free row after the block
        """
        self.summary.cell(row=row, column=1, value=heading).font = self.label_font
        row += 1
        first = row
        for label, value, tone in entries:
            label_cell = self.summary.cell(row=row, column=1, value=label)
            value_cell = self.summary.cell(row=row, column=2, value=value)
            for cell in (label_cell, value_cell):
                cell.border = self.cell_border
            value_cell.font = self.bold
            if value and tone:
                for cell in (label_cell, value_cell):
                    cell.fill = self.fills[tone]
            row += 1
        self.summary.row_dimensions[first].height = 18
        return row

    def label(self, row: int, text: str) -> int:
        """Write a block heading on the Summary sheet.

        Args:
            row (int): 1-based row
            text (str): heading text

        Returns:
            int: the next row
        """
        self.summary.cell(row=row, column=1, value=text).font = self.label_font
        return row + 1

    def save(self, path: str) -> str:
        """Save the workbook.

        Args:
            path (str): destination `.xlsx` path; parent directory must exist

        Returns:
            str: `path`, for convenience
        """
        self.workbook.save(clean_path(path))
        return path


class HtmlReport():
    """One self-contained HTML document: inline CSS and JS, no webfont, image
    or CDN reference, so it can be e-mailed or opened straight off a network
    share. It follows the reader's `prefers-color-scheme` and carries a toggle
    that overrides it (remembered in `localStorage`, guarded so a `file://`
    document with site data blocked still renders).

    Construction emits everything up to and including the page header; callers
    then append cards, chips and sections, and `render` closes the document.

    Args:
        title (str): heading shown at the top of the report, and the `<title>`
        source (str): what was checked, shown under the heading with a
            generation timestamp
    """

    def __init__(self, title: str, source: str):
        self.parts = [
            "<!doctype html>",
            '<html lang="en"><head><meta charset="utf-8">',
            '<meta name="viewport" content="width=device-width, initial-scale=1">',
            f"<title>{html_escape(title)}</title>",
            f"<style>{REPORT_HTML_CSS}</style>",
            "</head><body>",
            '<header class="page-head"><div>',
            f"<h1>{html_escape(title)}</h1>",
            f'<p class="muted">{html_escape(source)} &middot; generated '
            f"{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>",
            "</div>",
            '<button id="theme-toggle" type="button" aria-label="Toggle dark mode">'
            '<span class="theme-icon"></span><span class="theme-label">Theme</span></button>',
            "</header>",
        ]

    @staticmethod
    def escape(text) -> str:
        """Escape a string for inclusion in the document body.

        Exposed so callers building their own markup fragments -- a details
        cell, say -- escape exactly the way the builder's own helpers do.

        Args:
            text: value to escape

        Returns:
            str: the escaped text
        """
        return html_escape("" if text is None else str(text))

    @staticmethod
    def cell(value, css_class: str = None, mono: bool = False) -> str:
        """One escaped `<td>`.

        Args:
            value: cell content; `None` renders empty
            css_class (str): optional extra class, e.g. `"num"`
            mono (bool): render in the monospace face

        Returns:
            str: the `<td>` markup
        """
        text = "" if value is None else str(value)
        classes = " ".join(_ for _ in (css_class, "mono" if mono else None) if _)
        attr = f' class="{classes}"' if classes else ""
        return f"<td{attr}>{html_escape(text)}</td>"

    @staticmethod
    def status_cell(value) -> str:
        """A `<td>` carrying a coloured status badge -- Error, Warning, Passed,
        Failed or Skipped.

        Args:
            value: the status text; its lower-case form selects the colour

        Returns:
            str: the `<td>` markup
        """
        text = "" if value is None else str(value)
        return f'<td><span class="status {text.lower()}">{html_escape(text)}</span></td>'

    @staticmethod
    def table(headers: list, body_rows: list, css_class: str = "grid",
              table_attrs: str = "") -> str:
        """A scrollable table. Wide content scrolls inside its own container
        rather than making the page scroll horizontally.

        Args:
            headers (list): column labels
            body_rows (list): pre-rendered `<td>` runs, one string per row
            css_class (str): table class
            table_attrs (str): extra attributes for the `<table>` element,
                e.g. an id or the `data-filterable` marker the filter JS keys off

        Returns:
            str: the table markup
        """
        head = "".join(f"<th>{html_escape(_)}</th>" for _ in headers)
        body = "".join(f"<tr>{row}</tr>" for row in body_rows)
        attrs = f" {table_attrs}" if table_attrs else ""
        return (
            f'<div class="table-wrap"><table class="{css_class}"{attrs}>'
            f"<thead><tr>{head}</tr></thead><tbody>{body}</tbody></table></div>"
        )

    @staticmethod
    def section(heading: str, count, content: str, open_by_default: bool = True) -> str:
        """A collapsible section with a count pill.

        Args:
            heading (str): section title
            count: number shown in the pill
            content (str): section body markup
            open_by_default (bool): render expanded

        Returns:
            str: the `<details>` markup
        """
        return (
            f'<details class="section"{" open" if open_by_default else ""}>'
            f'<summary><span class="section-title">{html_escape(heading)}</span>'
            f'<span class="pill">{count}</span></summary>{content}</details>'
        )

    @staticmethod
    def filter_controls(placeholder: str, input_id: str, checkbox_id: str,
                        count_id: str, switch_label: str = "Errors only") -> str:
        """The live text filter and errors-only switch that sit above a
        filterable table. The ids are the caller's to choose; the JS finds
        these controls by their `data-` attributes.

        Args:
            placeholder (str): search box placeholder
            input_id (str): id of the search input
            checkbox_id (str): id of the errors-only checkbox
            count_id (str): id of the "N of M" counter
            switch_label (str): label beside the checkbox

        Returns:
            str: the controls markup
        """
        return (
            '<div class="controls">'
            f'<input id="{input_id}" data-filter-input type="search" '
            f'placeholder="{html_escape(placeholder)}">'
            f'<label class="switch"><input id="{checkbox_id}" data-filter-errors-only '
            f'type="checkbox"><span>{html_escape(switch_label)}</span></label>'
            f'<span id="{count_id}" data-filter-count class="muted"></span></div>'
        )

    def add(self, markup: str) -> None:
        """Append raw markup to the document body.

        Args:
            markup (str): already-escaped HTML
        """
        self.parts.append(markup)

    def add_cards(self, cards: list) -> None:
        """The row of headline stat cards under the page header.

        Args:
            cards (list): `(label, value, tone)` triples, `tone` being one of
                `"error"`, `"warning"`, `"ok"` or `"neutral"`
        """
        self.parts.append('<section class="cards">')
        for label, value, tone in cards:
            self.parts.append(
                f'<div class="card {tone}"><span class="card-value">{value}</span>'
                f'<span class="card-label">{html_escape(label)}</span></div>'
            )
        self.parts.append("</section>")

    def add_chips(self, chips: list) -> None:
        """A row of key/value chips -- a breakdown too small to deserve a table.

        Args:
            chips (list): `(key, value)` pairs
        """
        if not chips:
            return
        markup = "".join(
            f'<span class="chip"><span class="chip-key">{html_escape(str(key))}</span>'
            f'<span class="chip-value">{value}</span></span>'
            for key, value in chips
        )
        self.parts.append(f'<section class="chips">{markup}</section>')

    def add_empty_state(self, message: str) -> None:
        """The all-clear panel shown instead of an empty table.

        Args:
            message (str): what was found to be clean
        """
        self.parts.append(
            '<section class="empty"><span class="empty-mark">&#10003;</span>'
            f"<p>{html_escape(message)}</p></section>"
        )

    def render(self, path: str = None) -> str:
        """Close the document, optionally writing it to disk.

        Args:
            path (str): optional destination; the report is written there
                (UTF-8) as well as returned

        Returns:
            str: the complete HTML document
        """
        parts = self.parts + [
            f"<script>{REPORT_HTML_JS}</script>",
            "</body></html>",
        ]
        html = "\n".join(parts)
        if path:
            with open(clean_path(path), "w", encoding="utf-8") as report_file:
                report_file.write(html)
        return html
