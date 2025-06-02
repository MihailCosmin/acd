from os.path import sep
from os.path import dirname
from os.path import basename

import sys

from re import sub
from re import search
from re import findall

from traceback import format_exc

import pandas as pd
import openpyxl

import sys

if sys.version_info >= (3, 12):
    from PySide6.QtWidgets import QMainWindow
    from PySide6.QtCore import Signal
else:
    from PySide2.QtWidgets import QMainWindow
    from PySide2.QtCore import Signal


from .filelist import list_files2
from .txt import get_textfile_content


def check_line_widths(
        svg_file: str,
        valid_widths: list,
        magnification: int = 5,
        highlight_color: str = r"#FF0000",
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:
    """Check the line widths of an svg file and highlight the lines that are not in the valid_widths list

    Args:
        svg_file (str): SVG file to check
        valid_widths (list): Valid line widths
        magnification (int, optional): Magnification factor for the line widths. Defaults to 5.
        highlight_color (str, optional): Highlight color for the invalid line widths. Defaults to r"#FF0000".
    """
    try:
        # with open(svg_file, "r", encoding="utf-8") as svg:
        # original_svg_content = new_svg_content = svg.read()
        original_svg_content = new_svg_content = get_textfile_content(svg_file)
        for line in findall(r'<line.*?/>', new_svg_content):
            for width in findall(r'(stroke-width=")(.*?)(")', ''.join(line)):
                if width[1] not in valid_widths:
                    print(f"width {width[1]} not in {valid_widths}")
                    # svg_content = sub(r'(stroke-width=")(.*?)(")', r'\1' + str(float(width[1]) * 10) + r'\3', svg_content)
                    width_value = width[1]
                    new_line = sub(r'(stroke-width=")(' + width_value + ')(")',
                                   lambda match, width_value=width_value: match.group(
                                       1) + str(float(width_value) * magnification) + match.group(3),
                                   line)

                    if search(r'stroke=".*?"', line) is not None:
                        new_svg_content = new_svg_content.replace(
                            line, sub(r'(stroke=")(.*?)(")', r'\1' + highlight_color + r'\3', new_line))
                    else:
                        new_svg_content = new_svg_content.replace(
                            line, sub(r'(<line )', r'\1' + rf'stroke="{highlight_color}" ', new_line))
        if original_svg_content != new_svg_content:
            if qt_window is not None:
                progress.emit(100)
                # console.emit(
                #     f"Non-conform line widths found in {svg_file}. Please check the file: {svg_file.replace('.svg', '_line_width_check.svg')}")
            with open(svg_file.replace('.svg', '_line_width_check.svg'), "w", encoding="utf-8") as svg:
                svg.write(new_svg_content)
            return 0
        elif qt_window is not None:
            progress.emit(100)
            # console.emit(f"No non-conform line widths found in {svg_file}.")
            return 1
    except Exception as _:
        if qt_window is not None and debug:
            progress.emit(100)
            console.emit(
                f"Error checking line widths in {svg_file}.\n{_}\n{format_exc()}")
        return 2
    return 1


def __check_line_widths(
        svg_file: str,
        valid_widths: list,
        magnification: int = 3,
        highlight_color: str = r"#FF0000",
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:
    """Check the line widths of an svg file and highlight the lines that are not in the valid_widths list

    Args:
        svg_file (str): SVG file to check
        valid_widths (list): Valid line widths
        magnification (int, optional): Magnification factor for the line widths. Defaults to 5.
        highlight_color (str, optional): Highlight color for the invalid line widths. Defaults to r"#FF0000".

    NOTE: This was an attempt to make the script work both for the svgs created from IsoDraw and from the convertion from cgm to svg.
    cgm to svg has a different style, the file cannot be open as image directly. We need a better convertor from cgm to svg.
    """
    try:
        # with open(svg_file, "r", encoding="utf-8") as svg:
        # original_svg_content = new_svg_content = svg.read()
        original_svg_content = new_svg_content = get_textfile_content(svg_file)
        for line in findall(r'<.*?>', new_svg_content):
            for width in findall(r'(stroke-width(?:=|:)"?)(.*?)("| |\/)', ''.join(line)):
                if width[1].rstrip("0") not in valid_widths:
                    print(
                        f"width {width[1].rstrip('0')} not in {valid_widths}")
                    # svg_content = sub(r'(stroke-width=")(.*?)(")', r'\1' + str(float(width[1]) * 10) + r'\3', svg_content)
                    width_value = width[1].rstrip("0")
                    new_line = sub(r'(stroke-width(?:=|:)"?)(.*?)("| |\/)',
                                   lambda match, width_value=width[1]: match.group(
                                       1) + str(float(width[1]) * magnification) + match.group(3),
                                   line)

                    if search(r'stroke(?:=|:)"?.*?("| |\/)', line) is not None:
                        print(f"11 line: {line}")
                        new_line = sub(
                            r'(stroke(?:=|:)\"?)(.*?)(\"| |\/)', r'\1' + highlight_color + r'\3', new_line)
                        print(f"11 new_line: {new_line}")
                        new_svg_content = new_svg_content.replace(
                            line, new_line)
                    else:
                        print(f"22 line: {line}")
                        new_line = sub(
                            r'(<[a-z]+ )', r'\1' + rf'stroke="{highlight_color}" ', new_line)
                        print(f"22 new_line: {new_line}")

                        new_svg_content = new_svg_content.replace(
                            line, new_line)
        if original_svg_content != new_svg_content:
            if qt_window is not None:
                progress.emit(100)
                # console.emit(
                #     f"Non-conform line widths found in {svg_file}. Please check the file: {svg_file.replace('.svg', '_line_width_check.svg')}")
            with open(svg_file.replace('.svg', '_line_width_check.svg'), "w", encoding="utf-8") as svg:
                svg.write(new_svg_content)
            return 0
        elif qt_window is not None:
            progress.emit(100)
            # console.emit(f"No non-conform line widths found in {svg_file}.")
            return 1
    except Exception as _:
        if qt_window is not None and debug:
            progress.emit(100)
            console.emit(
                f"Error checking line widths in {svg_file}.\n{_}\n{format_exc()}")
        return 2
    return 1


def batch_check_line_widths(
        svg_folder: str,
        valid_widths: list,
        magnification: int = 5,
        highlight_color: str = r"#FF0000",
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:
    """batch check the line widths of svg files in a folder and highlight the lines that are not in the valid_widths list

    Args:
        svg_folder (str): SVG folder to check
        valid_widths (list): List of valid line widths
        magnification (int, optional): Magnification factor for the line widths. Defaults to 5.
        highlight_color (str, optional): Highlight color for the invalid line widths. Defaults to r"#FF0000".
        qt_window (QMainWindow, optional): Qt window to show the progress. Defaults to None.
        progress (Signal, optional): Progress signal. Defaults to Signal(0).
        console (Signal, optional): Console signal. Defaults to Signal("").

    Returns:
        int: 0 if successful, 1 if not
    """
    result = {1: "Line Widths are OK",
              0: "Incorrect Line Widths found", 2: "Error"}
    results = {}
    try:
        scope = list_files2(svg_folder, True, ['svg'])
        for svg in scope:
            if qt_window is not None:
                progress.emit(int(scope.index(svg) / len(scope) * 100))
            results[svg.split(sep)[-1]] = result[check_line_widths(
                svg, valid_widths, magnification, highlight_color, debug, qt_window, progress, console)]

            if qt_window is not None and debug:
                console.emit(
                    f"Line Width check for {svg.split(sep)[-1]} finished. Result: {results[svg.split(sep)[-1]]}!")
    except Exception as _:
        return 1

    if qt_window is not None:
        # Save results to excel as follows: first columsn filename, second column result. If result is 0, highlight the cell
        df = pd.DataFrame.from_dict(
            results, orient='index', columns=['Result'])
        df.index.name = 'Filename'
        df.to_excel(svg_folder + sep + 'Line_Width_check.xlsx',
                    engine='xlsxwriter')

        workbook = openpyxl.load_workbook(
            svg_folder + sep + 'Line_Width_check.xlsx')
        worksheet = workbook.active

        # format header blue background white text bold
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color='00B0F0', end_color='00B0F0', fill_type='solid')
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                if cell.value == "Incorrect Line Widths found":
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
                    cell.font = openpyxl.styles.Font(color="9c0006")
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'),
                                                         right=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        top=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        bottom=openpyxl.styles.Side(border_style='thin', color='000000'))
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center')
                if cell.value == "Line Widths are OK":
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='c6efce', end_color='c6efce', fill_type='solid')
                    cell.font = openpyxl.styles.Font(color="006100")
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'),
                                                         right=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        top=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        bottom=openpyxl.styles.Side(border_style='thin', color='000000'))
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center')

        # set column width
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 30
        workbook.save(svg_folder + sep + 'Line_Width_check.xlsx')

        progress.emit(100)
        console.emit(
            f"\n\nLine Width check finished. Results saved to {svg_folder + sep + 'Line_Width_check.xlsx'}")

        return 0

    return results


def check_icn(
        svg_file: str,
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:
    """Check if the icn is in the svg file

    Args:
        svg_file (str): SVG file to check

    Returns:
        bool: True if the icn is in the svg file
    """
    icn = svg_file.split(sep)[-1].replace(".svg", ".CGM")
    # with open(svg_file, "r", encoding="utf-8") as svg:
    #     svg_content = svg.read()
    svg_content = get_textfile_content(svg_file)
    if search(rf'(<text.*?>)({icn})(</text>)', svg_content) is not None:
        return 1
    return 0


def check_icns(
        svg_folder: str,
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:
    """Check if the icn is in the svg files

    Args:
        svg_file (str): SVG file to check

    Returns:
        bool: True if the icn is in the svg file
    """
    result = {1: "ICN found", 0: "ICN not found"}
    results = {}
    try:
        scope = list_files2(svg_folder, True, ['svg'])
        for svg in scope:
            if qt_window is not None:
                progress.emit(int(scope.index(svg) / len(scope) * 100))
            results[svg.split(sep)[-1]] = result[check_icn(svg, debug=debug,
                                                           qt_window=qt_window, progress=progress, console=console)]
            if qt_window is not None and debug:
                console.emit(
                    f"ICN check for {svg.split(sep)[-1]} finished. Result: {results[svg.split(sep)[-1]]}!")
    except Exception as _:
        return 1

    if qt_window is not None:
        # Save results to excel as follows: first columsn filename, second column result. If result is 0, highlight the cell
        df = pd.DataFrame.from_dict(
            results, orient='index', columns=['Result'])
        df.index.name = 'Filename'
        df.to_excel(svg_folder + sep + 'ICN_check.xlsx', engine='xlsxwriter')

        workbook = openpyxl.load_workbook(svg_folder + sep + 'ICN_check.xlsx')
        worksheet = workbook.active

        # format header blue background white text bold
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color='00B0F0', end_color='00B0F0', fill_type='solid')
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                if cell.value == "ICN not found":
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
                    cell.font = openpyxl.styles.Font(color="9c0006")
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'),
                                                         right=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        top=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        bottom=openpyxl.styles.Side(border_style='thin', color='000000'))
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center')
                if cell.value == "ICN found":
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='c6efce', end_color='c6efce', fill_type='solid')
                    cell.font = openpyxl.styles.Font(color="006100")
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'),
                                                         right=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        top=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        bottom=openpyxl.styles.Side(border_style='thin', color='000000'))
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center')
        # set column width
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 20
        workbook.save(svg_folder + sep + 'ICN_check.xlsx')

        progress.emit(100)
        console.emit(
            f"ICN check finished. Results saved to {svg_folder + sep + 'ICN_check.xlsx'}")
        return 0
    return results


def check_text_format(
        svg_file: str,
        valid_font_family: list = None,
        valid_font_size: list = None,
        valid_fill: list = None,
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:
    """Check the text format of an svg file

    Args:
        svg_file (str): SVG file to check

    """
    if valid_font_family is None:
        valid_font_family = ["'Helvetica'"]

    if valid_font_size is None:
        valid_font_size = ['2.822', '3.174', '4.586', '5.291']

    if valid_fill is None:
        valid_fill = ['#000000']

    # with open(svg_file, "r", encoding="utf-8") as svg:
    #     original_svg_content = new_svg_content = svg.read()
    original_svg_content = new_svg_content = get_textfile_content(svg_file)
    for text in findall(r'(<text.*?>)(.*?)(</text>)', new_svg_content):
        original_full_text = new_full_text = "".join(text)
        font_family = search(r'(font-family=")(.*?)(")', text[0])
        font_size = search(r'(font-size=")(.*?)(")', text[0])
        fill = search(r'(fill=")(.*?)(")', text[0])
        error = False
        if font_family is not None:
            if font_family[2] not in valid_font_family:
                error = True
        if font_size is not None:
            if font_size[2] not in valid_font_size:
                error = True
        if fill is not None:
            if fill[2] not in valid_fill:
                error = True
        if error:
            new_full_text = new_full_text.replace(
                fill.group(0), 'fill="#FF0000"')
            new_svg_content = new_svg_content.replace(
                original_full_text, new_full_text)

    if original_svg_content != new_svg_content:
        with open(svg_file.replace('.svg', '_text_format_check.svg'), "w", encoding="utf-8") as svg:
            svg.write(new_svg_content)
        return 0
    else:
        return 1


def batch_check_text_format(
        svg_folder: str,
        valid_font_family: list = None,
        valid_font_size: list = None,
        valid_fill: list = None,
        debug: bool = False,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:
    """batch check the text format of svg files in a folder

    Args:
        svg_folder (str): SVG folder to check
        valid_font_family (list, optional): Valid font families. Defaults to None.
        valid_font_size (list, optional): Valid font sizes. Defaults to None.
        valid_fill (list, optional): Valid fill colors. Defaults to None.
        qt_window (QMainWindow, optional): Qt window to show the progress. Defaults to None.
        progress (Signal, optional): Progress signal. Defaults to Signal(0).
        console (Signal, optional): Console signal. Defaults to Signal("").

    Returns:
        int: 0 if successful, 1 if not
    """
    if valid_font_family is None:
        valid_font_family = ["'Helvetica'"]

    if valid_font_size is None:
        valid_font_size = ['2.822', '3.174', '4.586', '5.291']

    if valid_fill is None:
        valid_fill = ['#000000']
    result = {1: "Text format is OK",
              0: "Incorrect text format found", 2: "Error"}
    results = {}
    try:
        scope = list_files2(svg_folder, True, ['svg'])
        for svg in scope:
            if qt_window is not None:
                progress.emit(int(scope.index(svg) / len(scope) * 100))
            results[svg.split(sep)[-1]] = result[check_text_format(svg, valid_font_family, valid_font_size,
                                                                   valid_fill, debug, qt_window, progress, console)]

            if qt_window is not None and debug:
                console.emit(
                    f"Text Format check for {svg.split(sep)[-1]} finished. Result: {results[svg.split(sep)[-1]]}!")
    except Exception as _:
        if debug:
            print(format_exc())
        if console != Signal(""):
            console.emit(
                f"Error checking text format in {svg_folder}.\n{_}\n{format_exc()}")
        return 1

    if qt_window is not None:
        # Save results to excel as follows: first columsn filename, second column result. If result is 0, highlight the cell
        df = pd.DataFrame.from_dict(
            results, orient='index', columns=['Result'])
        df.index.name = 'Filename'
        df.to_excel(svg_folder + sep + 'Text_Format_check.xlsx',
                    engine='xlsxwriter')

        workbook = openpyxl.load_workbook(
            svg_folder + sep + 'Text_Format_check.xlsx')
        worksheet = workbook.active

        # format header blue background white text bold
        for row in worksheet.iter_rows(min_row=1, max_row=1):
            for cell in row:
                cell.fill = openpyxl.styles.PatternFill(
                    start_color='00B0F0', end_color='00B0F0', fill_type='solid')
                cell.font = openpyxl.styles.Font(color="FFFFFF", bold=True)
        for row in worksheet.iter_rows(min_row=2, min_col=2, max_col=2):
            for cell in row:
                if cell.value == "Incorrect text format found":
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='ffc7ce', end_color='ffc7ce', fill_type='solid')
                    cell.font = openpyxl.styles.Font(color="9c0006")
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'),
                                                         right=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        top=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        bottom=openpyxl.styles.Side(border_style='thin', color='000000'))
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center')
                if cell.value == "Text format is OK":
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color='c6efce', end_color='c6efce', fill_type='solid')
                    cell.font = openpyxl.styles.Font(color="006100")
                    cell.border = openpyxl.styles.Border(left=openpyxl.styles.Side(border_style='thin', color='000000'),
                                                         right=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        top=openpyxl.styles.Side(
                        border_style='thin', color='000000'),
                        bottom=openpyxl.styles.Side(border_style='thin', color='000000'))
                    cell.alignment = openpyxl.styles.Alignment(
                        horizontal='center')

        # set column width
        worksheet.column_dimensions['A'].width = 50
        worksheet.column_dimensions['B'].width = 30
        workbook.save(svg_folder + sep + 'Text_Format_check.xlsx')

        progress.emit(100)
        console.emit(
            f"\n\nText Format check finished. Results saved to {svg_folder + sep + 'Text_Format_check.xlsx'}")
        return 0
    return results


def check_illu_text(
        svg_file: str,
        qt_window: QMainWindow = None,
        progress: Signal = Signal(0),
        console: Signal = Signal("")) -> int:
    """Check the text format of an svg file

    Args:
        svg_file (str): SVG file to check

    """
    # with open(svg_file, "r", encoding="utf-8") as svg:
    #     svg_content = svg.read()
    svg_content = get_textfile_content(svg_file)
    for text in findall(r'(<text.*?>)(.*?)(</text>)', svg_content):
        print(text[1])


if __name__ == '__main__':
    # check_line_widths(r"D:\Automation\Illu Automation\ICN-CO91-32-32-33-D9893-00657-A01_000.svg",
    #                   ["0.12", "0.18", "0.35", "0.6", "0.7"])
    check_text_format(
        r"D:\Automation\Illu Automation\ICN-CO91-32-32-33-D9893-00657-A01_000.svg")
