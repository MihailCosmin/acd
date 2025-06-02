from os import remove
from os.path import isfile
from os.path import basename
from os.path import join
from os.path import expanduser
from os.path import dirname
from os.path import normpath

from re import sub
from re import search
from re import findall

from traceback import format_exc

from urllib import request
from io import StringIO
from io import BytesIO
from sys import exit as done

from lxml import etree
from xml.etree.ElementTree import parse
from xml.etree.ElementTree import Element
from xml.etree import ElementTree as ET

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.styles import Alignment
from openpyxl.styles import PatternFill


from .xml_processing import linearize_xml
from .xml_processing import delete_first_line
from .unit_list import unit_list
from .python_func import check_brackets


class XmlSchemaValidator():
    """
    XmlSchemaValidator: Class that validates xml files against a schema.

    """

    def __init__(self, debug: bool = False):
        self.debug = debug

    def _add_root_path_to_schema_location(
            self,
            xml_content: str,
            root_path: str) -> str:
        """
        Searches for all instances that match the regular expression and
        adds the root to the schema names

        Args:
            xml_content (str): Content of the schema (after the first line was deleted)
            root_path (str): String that contains the root that will be added to the schema names

        Returns:
            str: String containing the content of the schema,
            but with the changes mentioned in the description

        """

        return sub(r'(schemaLocation=")(.*?)("/>)', r'\1' + root_path + r'\2\3', xml_content)

    def _get_schema_url_and_root(self, linearized_file: str) -> tuple:
        """
        Searches if a schema url is given inside the xml document.
        If a url matches the regular expression,
        the whole url and only the root part of the url are saved inside a list.

        Args:
            linearized_file (str): Linearized content of the xml document,

        Returns:
            tuple: A tuple containing exactly two strings

        """

        schema_url_regex = r'(xsi:noNamespaceSchemaLocation=")(.*?)(")'
        if search(schema_url_regex, linearized_file):
            schema = str(search(schema_url_regex, linearized_file).group(2))
            if self.debug:
                print(f"Schema: {schema}")
            root_part = schema.replace(schema.split('/')[-1], "")
            return schema, root_part
        return None, None

    def _new_urlopen(self, schema_url: str) -> str:
        """
        Works like the urlopen from the request library.
        However, urlopen saves the content as a binarystring which we don't want to.
        Thus we call the function decode to receive a regular string.

        Args:
            schema_url (str): The url of the schema, included in the xml

        Returns:
            str: A string which contains the content of the schema

        """

        with request.urlopen(schema_url) as _:
            return _.read().decode('utf-8')

    def _lxml_parser(self, schema: str, xml_file: str) -> tuple:
        """
        Validate xml document against schema with lxml.
        If there are errors display the first error and in which line it appears.
        If there are no errors display that there are no errors.

        Args:
            schema (str): String containing the content of the schema to parse
            xml_file (str): The path to the xml file to be validated

        Returns:
            tuple: A String followed by a boolean value

        """

        schema = StringIO(schema)
        schema = etree.parse(schema)
        schema = etree.XMLSchema(schema)

        with open(xml_file, "r", encoding="utf-8") as _:
            xml_content = _.read()

        xml_content = etree.parse(StringIO(xml_content))
        try:
            schema.assertValid(xml_content)
            return None
        except (etree.DocumentInvalid):
            print("Validation error(s):")
        return schema.error_log[0].line, schema.error_log[0].message

    def _lxml_parser_all_errors(self, schema: str, xml_file: str) -> tuple:
        """
        Validate xml document against schema with lxml.
        If there are multiple errors, try to fix them and display the error(s) and
        in which line they appear.
        If the error can't be fixed, stop the program.
        If there are no errors display that there are no errors.

        Args:
            schema (str): String containing the content of the schema to parse
            xml_file (str): The path to the xml file to be validated

        Returns:
            tuple: A String followed by a boolean value

        """

        schema = StringIO(schema)
        schema = etree.parse(schema)
        schema = etree.XMLSchema(schema)

        with open(xml_file, "r", encoding="utf-8") as _:
            xml_content = _.read()

        try:
            xml_content = etree.parse(StringIO(xml_content))
        except ValueError:
            xml_content = xml_content.encode('utf-8')  # Cosmin 20.10.2023 added to deal with encoding issues
            xml_content = etree.parse(BytesIO(xml_content))
        try:
            schema.assertValid(xml_content)
            return None
        except (etree.DocumentInvalid):
            print("Validation error(s):")
            for ind, error in enumerate(list(set(schema.error_log))):
                print(f"{ind}  Line {error.line}: {error.message}")
            return schema.error_log[0].line, schema.error_log[0].message

    def test_delete_error(
            self,
            xml_file: str,
            line_error: int,
            error_message: str,
            schema_url: str) -> str:
        """
        Function that get's the error message and 'corrects' the error in order to display the next one.
        If the error message has olny one suggestion for a mandatory tag,
        the function adds the missing tag into the xml.
        If the error message has more than one suggestions we differentiate between two cases:
        1. Element causing the error is not listed in the schema:
        Replace parent and all childs of the element white new line
        2. Element causing the error is listed in the schema:
        This is a critical error we can't solve (yet), exit program

        Args:
            xml_file (str): Path to the xml file to validate
            line_error (int): line number where the error occured
            error_message (str): error message that is produced by lxml
            schema_url (str): The url of the schema, included in the xml

        Returns:
            str: string that contains the edited xml where the errors were fixed

        """

        with open(xml_file, "r", encoding="utf-8") as _:
            xml_content = _.read()

        schema_content = self._new_urlopen(schema_url)

        content_error = search(r"(Element ')(.*?)(')", error_message).group(2)
        fix_error = f"</{content_error}>"

        # Save every line as an element of a list
        my_xml_list = xml_content.split("\n")
        if "Expected is one of" in error_message:
            if search(content_error, schema_content) is None:
                line = line_error - 1
                if fix_error in my_xml_list[line_error - 1] or\
                        ("/>" in my_xml_list[line_error - 1] and
                         "<" not in my_xml_list[line_error - 1]):
                    my_xml_list[line_error - 1] = "\n"
                else:
                    my_xml_list[line_error - 1] = "\n"
                    line += 1
                    while fix_error not in my_xml_list[line]:
                        my_xml_list[line] = "\n"
                        line += 1
                    my_xml_list[line] = "\n"
                    line += 1
            else:
                print(
                    f"Critical error found in line {line_error} -> Can't resolve!")
                done()

        if "Expected is (" in error_message:
            expected_element = search(
                r'(Expected is \(\ )(.*?)( \)\.)', error_message).group(2)
            my_xml_list[line_error - 1] = sub(
                r'( *)(\<)', r'\1' +
                f'<{expected_element}></{expected_element}>' + r'\2',
                my_xml_list[line_error - 1]
            )

        # Add \n to the end of every element
        my_xml_list = [element + "\n" for element in my_xml_list]
        my_xml_list = ''.join(my_xml_list).replace("\n\n", "\n")

        with open("edited_xml.xml", "w", encoding="utf-8") as _:
            _.write(my_xml_list)

        return "edited_xml.xml"

    def validate_xml(
            self,
            xml_file: str,
            mode: bool = False,
            debug: bool = False) -> str:
        """
        Main file of the script. It calls all function in the right order to validate
        a XML document against a schema:
        1. We open the XML and save the content in a variable
        2. Linearize the XML document with the linearize_xml() function
        3. Obtain a list containing two strings ( [schema, root_part] ) with
        the _get_schema_url_and_root() fct. and safe them respectively into variables
        4. Open and save the content of the schema with the _new_urlopen() function
        5. Modify schema content with the functions:
        - linearize_xml(),
        - delete_first_line(),
        - _add_root_path_to_schema_location()
        6. Validate the schema with the _lxml_parser_all_errors() function

        Args:
            xml_file (str): path to the xml file to validate
            mode (bool, optional): If True, the script will try to fix the first error in order
            to find all subsequent errors.
            Defaults to False.
            debug (bool, optional): If True, the script will print the content of the schema and
            the xml file.
            Defaults to False.

        Returns:
            str: Result of the validation

        """

        if debug and not self.debug:
            self.debug = True

        with open(xml_file, "r", encoding="utf-8") as _:
            xml_content = _.read()

        xml_content = linearize_xml(xml_content)

        schema, root_part = self._get_schema_url_and_root(xml_content)

        schema_content = self._new_urlopen(schema)
        schema_content = delete_first_line(schema_content)
        schema_content = linearize_xml(schema_content)
        schema_content = self._add_root_path_to_schema_location(
            schema_content, root_part)

        # lxml part for each error
        if mode is True:
            error = self._lxml_parser_all_errors(schema_content, xml_file)
            original_error = error
            run_num = 1
            while True:
                if debug:
                    print(f"run_num: {run_num}")
                error = self._lxml_parser_all_errors(schema_content, xml_file)
                if error == original_error:
                    if error is None and original_error is None and str(format_exc()).strip() == 'NoneType: None':
                        print("2 No error found")
                    else:
                        print(f"2 Critical error found -> Can't resolve!\nerror: {error}\noriginal_error: {original_error}\nformat_exc: {format_exc()}")
                    break
                if debug:
                    print(f"error: {error}")
                if error is None:
                    break
                xml_file = self.test_delete_error(
                    xml_file, error[0], error[1], schema)
                run_num += 1
                
        else:
            error = self._lxml_parser(schema_content, xml_file)
            if debug:
                print(f"error: {error}")
            return 0

        if isfile("edited_xml.xml"):
            remove("edited_xml.xml")


class Punctuation():
    def __init__(self) -> None:
        self.file_path = None
        self.export_path = expanduser("~/Desktop")

    def set_text_file(self, text_file_path: str):
        """Function with which the user can set the xml to be checked.

        Args:
            text_file_path (str): text file path.
        """
        self.file_path = text_file_path

    def set_export_path(self, export_path: str):
        """Function to specify a path different to the default path (desktop),
        where to export the files the script produces.
        Args:
            export_path (str): path to where the generated files should be exported.
        """
        self.export_path = export_path

    def check_brackets(self, file_path: str, linearized=False):
        """Checks if all brackets in a file are matched or not

        Args:
            file_path (str): _description_
            linearized (bool): if set to "True" the position is given by it's col number instead of line number
        Returns:
            _type_: _description_
        """
        with open(file_path, "r", encoding="utf-8") as _:
            file_content = _.read()
        # file_content = linearize_xml(file_content)
        # for para in findall(r"(\<para\>)(.*?)(\</para\>)", file_content):
        #     para = para[1]
        #     if "refint" in para:
        #         print(f"Para: {para}")
        #     new_para = para
        #     for sub in findall(r"\<.*?/\>", para):
        #         new_para = new_para.replace(sub, "")
        #     if "refint" in para:
        #         print(f"NewPara: {new_para}")
        #     file_content = file_content.replace(para, new_para)
        # with open("testDeleteLater2.xml", "w", encoding="utf-8") as _:
        #     _.write(file_content)

        mismatch_list = []
        # stack = []
        # lines = file_content.split("\n")

        # # Check opening bracket mismatch
        # for ind, line in enumerate(lines, start=1):
        #     for ind2, char in enumerate(line, start=1):
        #         if stack:
        #             if char == ')':
        #                 stack.pop()
        #             if char == '(':
        #                 mismatch_list.append((f"Bracket '(' at line: {stack[0][1]}, column: {stack[0][2]}", line))
        #                 stack.pop()
        #                 stack.append(('(', ind, ind2))
        #         else:
        #             if char == '(':
        #                 stack.append(('(', ind, ind2))
        # # Check closing bracket mismatch
        # for ind, line in enumerate(lines, start=1):
        #     for ind2, char in enumerate(line, start=1):
        #         if stack:
        #             if char == '(':
        #                 stack.pop()
        #             if char == ')':
        #                 mismatch_list.append((f"Bracket ')' at line: {stack[0][1]}, column: {stack[0][2]}", line))
        #                 stack.pop()
        #                 stack.append((')', ind, ind2))
        #         else:
        #             if char == ')':
        #                 stack.append((')', ind, ind2))

        # tree = etree.fromstring(file_content.encode('utf-8'))
        tree = etree.parse(file_path)

        for para in tree.iter('para'):
            text = ''.join(para.itertext())
            if text:
                opening_bracket_list = findall(r"\(", text)
                closing_bracket_list = findall(r"\)", text)
                line_number = 0
                if len(opening_bracket_list) > len(closing_bracket_list):
                    line_number = para.sourceline
                    mismatch_list.append(
                        (f"Bracket '(' at para in line: {line_number}", text))
                if len(opening_bracket_list) < len(closing_bracket_list):
                    line_number = para.sourceline
                    mismatch_list.append(
                        (f"Bracket ')' at para in line: {line_number}", text))
        return mismatch_list

    def check_fullstops(self, file_path: str):
        with open(file_path, 'r', encoding="utf-8") as _:
            file_content = _.read()

        tree = parse(file_path)
        root = tree.getroot()

        occurences = []

        for para in root.iter('para'):
            text = para.text
            if text and '..' in text:
                char_position = text.index('..')

                # Calculate line_number
                line_number = "Not Found"
                if text in file_content:
                    index = file_content.index(text)
                    line_number = file_content.count('\n', 0, index) + 1

                occurence_info = {
                    'char_position': char_position,
                    'line_number': line_number,
                    'line_text': text.strip()
                }
                occurences.append(occurence_info)
        return occurences

    def check_hard_spaces(self):
        with open(self.file_path, 'r', encoding="utf-8") as _:
            content = _.read()
        if search(r"\<\?.*?]\>", linearize_xml(content)):
            part_before_cmm = search(
                r"\<\?.*?]\>", linearize_xml(content)).group(0)
        else:
            part_before_cmm = ""

        # modified_content = linearize_xml(content)
        # modified_content = etree.fromstring(modified_content)
        modified_content = etree.parse(self.file_path)
        paras = modified_content.xpath(".//para")
        modified_content = etree.tostring(modified_content).decode()
        for elem in paras:
            elem_string = etree.tostring(elem).decode().replace(
                ' xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" xmlns:rdf="http://www.w3.org/1999/02/22-rdf-syntax-ns#" xmlns:dc="http://www.purl.org/dc/elements/1.1/"', '')
            matches = findall(r"\d+\.*\d* [a-zA-Z]", elem_string)
            modified_elem = elem_string
            for match in matches:
                # We replace an empty space with a non-breaking space
                modified_match = match.replace(' ', ' ')
                modified_elem = modified_elem.replace(match, modified_match)
            modified_content = modified_content.replace(
                elem_string, modified_elem)

            # TODO: Add a check for the following cases:
            # matches_2 = findall()

        with open(join(self.export_path, f"added_hard_spaces_{basename(normpath(self.file_path))}.xml"), 'w', encoding="utf-8") as _:
            _.write(part_before_cmm + modified_content)

    def check_punctuation(self):
        workbook = Workbook()
        brackets_sheet = workbook.active
        brackets_sheet.title = "Brackets"
        fullstops_sheet = workbook.create_sheet(title="Full Stops")

        # Headers for the "Brackets" sheet
        brackets_sheet.cell(
            row=1, column=1).value = "Unmatched brackets found:"
        brackets_sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        brackets_sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')

        brackets_sheet.cell(row=1, column=2).value = "Text:"
        brackets_sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        brackets_sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')

        brackets_mismatches = self.check_brackets(self.file_path)
        for ind, elem in enumerate(brackets_mismatches, start=2):
            brackets_sheet.cell(
                row=ind, column=1).value = brackets_mismatches[ind - 2][0]
            brackets_sheet.cell(
                row=ind, column=2).value = brackets_mismatches[ind - 2][1]

        # Headers for the "Fullstops" sheet
        fullstops_sheet.cell(row=1, column=1).value = "Occurences found:"
        fullstops_sheet.cell(row=1, column=1).font = Font(
            bold=True, italic=True, size=16)
        fullstops_sheet.cell(row=1, column=1).alignment = Alignment(
            horizontal='center')

        fullstops_sheet.cell(row=1, column=2).value = "Text:"
        fullstops_sheet.cell(row=1, column=2).font = Font(
            bold=True, italic=True, size=16)
        fullstops_sheet.cell(row=1, column=2).alignment = Alignment(
            horizontal='center')

        fullstops_occurences = self.check_fullstops(self.file_path)
        for ind, elem in enumerate(fullstops_occurences, start=2):
            fullstops_sheet.cell(
                row=ind, column=1).value = f"Double fullstop '..' at para in line {fullstops_occurences[ind-2]['char_position']} line {fullstops_occurences[ind-2]['line_number']}"
            fullstops_sheet.cell(
                row=ind, column=2).value = fullstops_occurences[ind - 2]["line_text"]

        workbook.save(join(
            self.export_path, f"punctuation_check_{basename(normpath(self.file_path))}.xlsx"))

        # TODO: Continue here
        self.check_hard_spaces()


if __name__ == "__main__":
    from os import listdir
    for xml in listdir(r"C:\Users\munteanu\Downloads\Liebherr S1000D\CMP_27-20-10"):
        if xml.lower().endswith(".xml"):
            instance = XmlSchemaValidator()
            instance.validate_xml(join(
                r"C:\Users\munteanu\Downloads\Liebherr S1000D\CMP_27-20-10", xml), True, True)

    # instance = Punctuation()
    # # instance.set_text_file(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CMM-D9893-GE01-32-51-25_009-01_EN.xml")
    # instance.set_text_file(r"C:\Users\bakalarz\Desktop\01_XML_Samples\ATA\CRM-D9893-CO91-32-31-41RM_000-01_EN.xml")
    # instance.set_export_path(r"C:\Users\bakalarz\Desktop\Export")
    # instance.check_punctuation()
