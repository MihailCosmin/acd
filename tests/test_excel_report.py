"""`BrexChecker.to_excel_report` -- the formatted Excel report (category D6).

Covers both result shapes `validate()` can return (single object after
`set_xml`, batch/folder after `set_xml_dir`), the sheet set, the actual
formatting (header band, borders, freeze pane, autofilter, outcome tinting)
and the conditional sheets that only appear when the run produced that kind
of finding.
"""

import pytest

from openpyxl import load_workbook

from acd.brex_checker import BrexChecker

DMODULE_SCHEMA = "http://www.s1000d.org/S1000D_4-2/xml_schema_flat/dmodule.xsd"

SEVERITY_LEVELS_CONTENT = """<brSeverityLevels>
  <brSeverityLevel value="brsl01" fail="yes">Error</brSeverityLevel>
  <brSeverityLevel value="brsl02" fail="no">Warning</brSeverityLevel>
</brSeverityLevels>
"""

BREX_CONTENT = """<brex>
  <contextRules>
    <structureObjectRuleGroup>
      <structureObjectRule id="R-FORBIDDEN" brSeverityLevel="brsl01">
        <objectPath allowedObjectFlag="0">//forbidden</objectPath>
        <objectUse>forbidden must not be present.</objectUse>
        <brDecisionRef brDecisionIdentNumber="BRDP-0001"/>
      </structureObjectRule>
      <structureObjectRule id="R-DISCOURAGED" brSeverityLevel="brsl02">
        <objectPath allowedObjectFlag="0">//discouraged</objectPath>
        <objectUse>discouraged should not be present, but only warns.</objectUse>
      </structureObjectRule>
      <structureObjectRule id="R-CODE">
        <objectPath allowedObjectFlag="2">//coded/@code</objectPath>
        <objectUse>coded/@code must be aa or two digits.</objectUse>
        <objectValue valueForm="single" valueAllowed="aa"/>
        <objectValue valueForm="pattern" valueAllowed="[0-9]{2}"/>
      </structureObjectRule>
    </structureObjectRuleGroup>
  </contextRules>
</brex>
"""

BROKEN_BREX_CONTENT = """<brex>
  <contextRules>
    <structureObjectRuleGroup>
      <structureObjectRule id="R-BROKEN">
        <objectPath allowedObjectFlag="0">//[[[broken</objectPath>
        <objectUse>A rule whose objectPath does not compile.</objectUse>
      </structureObjectRule>
    </structureObjectRuleGroup>
  </contextRules>
</brex>
"""


def make_dm(body: str = "") -> str:
    return (
        '<dmodule xmlns:xsi="http://www.w3.org/2001/XMLSchema-instance" '
        f'xsi:noNamespaceSchemaLocation="{DMODULE_SCHEMA}">\n'
        f"{body}\n"
        "</dmodule>\n"
    )


@pytest.fixture
def brex_path(tmp_path):
    path = tmp_path / "brex.xml"
    path.write_text(BREX_CONTENT, encoding="utf-8")
    return str(path)


@pytest.fixture
def severity_levels_path(tmp_path):
    path = tmp_path / ".brseveritylevels"
    path.write_text(SEVERITY_LEVELS_CONTENT, encoding="utf-8")
    return str(path)


@pytest.fixture
def object_dir(tmp_path):
    objects = tmp_path / "objects"
    objects.mkdir()
    (objects / "clean.xml").write_text(make_dm('  <coded code="aa"/>'), encoding="utf-8")
    (objects / "bad.xml").write_text(
        make_dm('  <forbidden/>\n  <discouraged/>\n  <coded code="zz"/>'), encoding="utf-8"
    )
    return str(objects)


def _batch_workbook(tmp_path, object_dir, brex_path, severity_levels_path):
    checker = BrexChecker()
    checker.set_xml_dir(object_dir)
    checker.override_brex_list([brex_path])
    checker.set_severity_levels_path(severity_levels_path)
    result = checker.validate()

    path = str(tmp_path / "report.xlsx")
    assert checker.to_excel_report(result, path) == path
    return load_workbook(path), result


def _single_workbook(tmp_path, brex_path, body):
    xml_path = tmp_path / "object.xml"
    xml_path.write_text(make_dm(body), encoding="utf-8")

    checker = BrexChecker()
    checker.set_xml(str(xml_path))
    checker.override_brex_list([brex_path])
    result = checker.validate()

    path = str(tmp_path / "report.xlsx")
    checker.to_excel_report(result, path)
    return load_workbook(path)


def _rows(worksheet, header_row=1):
    """Body rows of a sheet as {header: value} dicts."""
    headers = [cell.value for cell in worksheet[header_row]]
    return [
        dict(zip(headers, [cell.value for cell in row]))
        for row in worksheet.iter_rows(min_row=header_row + 1)
    ]


# ---------------------------------------------------------------------------
# Sheets and content
# ---------------------------------------------------------------------------

def test_batch_run_produces_summary_and_violations_sheets(tmp_path, object_dir, brex_path,
                                                          severity_levels_path):
    workbook, _ = _batch_workbook(tmp_path, object_dir, brex_path, severity_levels_path)

    assert workbook.sheetnames[:2] == ["Summary", "Violations"]


def test_violations_sheet_carries_one_row_per_violation(tmp_path, object_dir, brex_path,
                                                        severity_levels_path):
    workbook, result = _batch_workbook(tmp_path, object_dir, brex_path, severity_levels_path)
    rows = _rows(workbook["Violations"])

    assert len(rows) == 3
    assert {row["Document"] for row in rows} == {"bad.xml"}
    by_path = {row["Object path"]: row for row in rows}
    assert by_path["//forbidden"]["Status"] == "Error"
    assert by_path["//forbidden"]["Rule ID"] == "R-FORBIDDEN"
    assert by_path["//forbidden"]["BR decision"] == "BRDP-0001"
    assert by_path["//forbidden"]["Flag"] == "0 - not allowed"
    # A severity marked fail="no" is reported as a warning, not an error.
    assert by_path["//discouraged"]["Status"] == "Warning"
    # A value violation carries both what was found and what the BREX allows.
    assert "zz" in by_path["//coded/@code"]["Finding"]
    assert by_path["//coded/@code"]["Allowed (single)"] == "aa"
    assert by_path["//coded/@code"]["Allowed (pattern)"] == "[0-9]{2}"


def test_summary_sheet_carries_run_totals_and_a_row_per_document(tmp_path, object_dir,
                                                                 brex_path, severity_levels_path):
    workbook, _ = _batch_workbook(tmp_path, object_dir, brex_path, severity_levels_path)
    summary = workbook["Summary"]
    cells = {
        row[0].value: row[1].value
        for row in summary.iter_rows(min_row=1, max_col=2)
        if row[0].value is not None
    }

    assert summary["A1"].value == "BREX check report"
    assert cells["Documents checked"] == 2
    assert cells["Documents passed"] == 1
    assert cells["Documents failed"] == 1
    assert cells["Errors"] == 2
    assert cells["Warnings"] == 1
    assert cells["brsl01"] == 1

    documents_header = next(
        row for row in summary.iter_rows(min_col=1, max_col=4)
        if row[0].value == "Document" and row[1].value == "Status"
    )
    document_rows = _rows(summary, header_row=documents_header[0].row)
    by_document = {row["Document"]: row for row in document_rows}
    assert by_document["bad.xml"]["Status"] == "Failed"
    assert by_document["bad.xml"]["Errors"] == 2
    assert by_document["bad.xml"]["Warnings"] == 1
    assert by_document["clean.xml"]["Status"] == "Passed"


def test_single_object_run_is_reported_the_same_way(tmp_path, brex_path):
    workbook = _single_workbook(tmp_path, brex_path, "  <forbidden/>")
    rows = _rows(workbook["Violations"])

    assert len(rows) == 1
    assert rows[0]["Document"] == "object.xml"
    assert rows[0]["Object path"] == "//forbidden"


def test_clean_run_still_produces_a_headed_but_empty_violations_sheet(tmp_path, brex_path):
    workbook = _single_workbook(tmp_path, brex_path, '  <coded code="42"/>')
    violations = workbook["Violations"]

    assert violations["A1"].value == "Document"
    assert _rows(violations) == []
    assert workbook.sheetnames == ["Summary", "Violations"]


# ---------------------------------------------------------------------------
# Formatting
# ---------------------------------------------------------------------------

def test_header_row_is_styled_frozen_and_filterable(tmp_path, object_dir, brex_path,
                                                    severity_levels_path):
    workbook, _ = _batch_workbook(tmp_path, object_dir, brex_path, severity_levels_path)
    violations = workbook["Violations"]
    header = violations["A1"]

    assert header.font.bold is True
    assert header.font.color.rgb.endswith("FFFFFF")
    assert header.fill.fgColor.rgb.endswith("1F4E79")
    assert violations.freeze_panes == "A2"
    assert violations.auto_filter.ref.startswith("A1:")
    assert header.border.bottom.style == "thin"


def test_violation_rows_are_tinted_by_outcome(tmp_path, object_dir, brex_path,
                                              severity_levels_path):
    workbook, _ = _batch_workbook(tmp_path, object_dir, brex_path, severity_levels_path)
    violations = workbook["Violations"]
    status_column = [cell.value for cell in violations[1]].index("Status") + 1
    tints = {
        violations.cell(row=row[0].row, column=status_column).value: row[0].fill.fgColor.rgb
        for row in violations.iter_rows(min_row=2)
    }

    assert tints["Error"].endswith("FCE4E4")
    assert tints["Warning"].endswith("FFF2CC")


def test_every_body_cell_is_bordered_and_wrapped(tmp_path, brex_path):
    workbook = _single_workbook(tmp_path, brex_path, "  <forbidden/>")
    cell = workbook["Violations"]["A2"]

    assert cell.border.left.style == "thin"
    assert cell.alignment.wrap_text is True
    assert cell.alignment.vertical == "top"


def test_column_widths_are_set(tmp_path, brex_path):
    workbook = _single_workbook(tmp_path, brex_path, "  <forbidden/>")

    assert workbook["Violations"].column_dimensions["A"].width == 26


def test_a_very_long_node_snippet_is_truncated_to_a_legal_cell_value(tmp_path, brex_path):
    body = "  <forbidden>" + "".join(f"<child>{'x' * 80}</child>" for _ in range(60)) + "</forbidden>"
    xml_path = tmp_path / "object.xml"
    xml_path.write_text(make_dm(body), encoding="utf-8")

    checker = BrexChecker()
    checker.set_xml(str(xml_path))
    checker.override_brex_list([brex_path])
    # deep_copy_nodes embeds the whole subtree, which is what makes the
    # snippet long enough to need truncating.
    result = checker.validate(deep_copy_nodes=True)
    path = str(tmp_path / "report.xlsx")
    checker.to_excel_report(result, path)

    node = _rows(load_workbook(path)["Violations"])[0]["Node"]
    assert node.endswith(" [...]")
    assert len(node) <= 2010


# ---------------------------------------------------------------------------
# Conditional sheets
# ---------------------------------------------------------------------------

def test_xpath_errors_get_their_own_sheet(tmp_path):
    brex = tmp_path / "broken_brex.xml"
    brex.write_text(BROKEN_BREX_CONTENT, encoding="utf-8")
    workbook = _single_workbook(tmp_path, str(brex), "  <anything/>")

    assert "XPath errors" in workbook.sheetnames
    rows = _rows(workbook["XPath errors"])
    assert len(rows) == 1
    assert rows[0]["Object path"] == "//[[[broken"
    assert rows[0]["Error"]


def test_sns_violations_get_their_own_sheet(tmp_path):
    brex = tmp_path / "sns_brex.xml"
    brex.write_text(
        """<brex>
  <snsRules>
    <snsDescr>
      <snsSystem>
        <snsCode>21</snsCode>
        <snsTitle>Air conditioning</snsTitle>
      </snsSystem>
    </snsDescr>
  </snsRules>
</brex>
""",
        encoding="utf-8",
    )
    xml_path = tmp_path / "object.xml"
    xml_path.write_text(
        make_dm(
            "  <identAndStatusSection><dmAddress><dmIdent>"
            '<dmCode modelIdentCode="TEST" systemDiffCode="A" systemCode="99" '
            'subSystemCode="0" subSubSystemCode="0" assyCode="00" disassyCode="00" '
            'disassyCodeVariant="A" infoCode="000" infoCodeVariant="A" itemLocationCode="D"/>'
            "</dmIdent></dmAddress></identAndStatusSection>"
        ),
        encoding="utf-8",
    )

    checker = BrexChecker()
    checker.set_xml(str(xml_path))
    checker.override_brex_list([str(brex)])
    result = checker.validate(check_sns=True)
    path = str(tmp_path / "report.xlsx")
    checker.to_excel_report(result, path)

    workbook = load_workbook(path)
    assert "SNS" in workbook.sheetnames
    assert _rows(workbook["SNS"])[0]["Invalid value"] == "99"


def test_sheets_with_no_findings_are_omitted(tmp_path, brex_path):
    workbook = _single_workbook(tmp_path, brex_path, "  <forbidden/>")

    assert "SNS" not in workbook.sheetnames
    assert "Notations" not in workbook.sheetnames
    assert "XPath errors" not in workbook.sheetnames


def test_skipped_documents_are_reported_without_violations(tmp_path, brex_path):
    objects = tmp_path / "objects"
    objects.mkdir()
    (objects / "empty.xml").write_text("", encoding="utf-8")
    (objects / "bad.xml").write_text(make_dm("  <forbidden/>"), encoding="utf-8")

    checker = BrexChecker()
    checker.set_xml_dir(str(objects))
    checker.override_brex_list([brex_path])
    checker.set_ignore_empty(True)
    result = checker.validate()

    path = str(tmp_path / "report.xlsx")
    checker.to_excel_report(result, path)
    workbook = load_workbook(path)

    # Directory mode drops a skipped file from the results mapping entirely
    # (see `validate`), so it simply does not appear -- what matters is that
    # the report is still written and holds only the real violation.
    assert len(_rows(workbook["Violations"])) == 1
